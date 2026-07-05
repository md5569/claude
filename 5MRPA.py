# -*- coding: utf-8 -*-
"""
5MRPA v1.0  REV.5
설치경로: C:/secu/RPA7.py  (회사)  /  C:/Users/UserK/RPA7.py  (개인)
실행: python 5MRPA.py

필요 패키지:
  pip install pyautogui pynput PyQt6 Pillow opencv-python pywinauto pygetwindow
Gemma3 (Ollama): http://localhost:11434 에서 실행 중이어야 합니다

[변경 이력]
REV.1  - 5MRPA 최종 통합본 (녹화/재생 안정화, DPI보정, 화면보호기 대응,
         글로벌 ESC 중단, AI 프로세스 최적화, PROC 노드별 반복 등 반영)
REV.2  - FLOW 우측 AI 패널 4개 버튼 → 2개로 통폐합
         ("프로세스 점검"=오류자동점검+AI검토 통합, "AI 최적화"=기존 유지)
         N회 AI 호출로 불안정했던 "프로세스 설명 자동 생성" 제거
         (동일 기능은 블록 더블클릭 편집창에서 안정적으로 제공됨)
REV.3  - TRIGGER 에 "폴더 신규 파일" / "메일 수신(Outlook)" 트리거 추가
         폴더 신규 파일: 확장자 필터, 파일명 패턴(와일드카드) 필터 지원,
                        표준 라이브러리만 사용 - 폐쇄망에서도 동작
         메일 수신: Outlook 데스크톱 앱과 로컬 COM 통신 (인터넷 불필요),
                    제목/발신자 포함 필터 지원, pywin32 없으면 안내 후 비활성
REV.4  - 외부 코드리뷰 검증 후 실제 버그 2건 수정
         ① FlowPlayWorker._exec_act 메서드 부재로 엑셀 반복재생 시
           AttributeError 발생 → 공통 실행 메서드로 분리해 정상화
         ② _fallback_parse 정규식 리스트 콤마 누락으로 패턴 오조합
           → 패턴 분리 수정 (「」 괄호 텍스트 입력 인식 정상화)
REV.5  - 2차 코드리뷰 검증 후 실제 이슈 3건 수정
         ① CoordHighlighter 에 WA_DeleteOnClose 누락 - close() 해도
           destroyed 시그널이 발생하지 않아 _HIGHLIGHTER_REFS 가 영구
           누적되는 메모리 누수 확인 및 수정 (리뷰 지적보다 심각한 상태였음)
         ② subprocess.run 4곳 중 3곳에 CREATE_NO_WINDOW 누락 확인,
           전체 적용하여 콘솔창 깜빡임 제거
         ③ 다중 모니터 DPI 대응 - 기존엔 항상 주모니터 DPI만 사용해
           배율이 다른 보조 모니터에서 좌표가 어긋날 수 있었음.
           MonitorFromPoint+GetDpiForMonitor 로 좌표별 실제 모니터의
           DPI 를 조회하도록 개선 (_get_dpi_scale_at 신설)
"""

APP_VERSION = "v1.0"
APP_REV     = "REV.5"
APP_NAME    = "5MRPA"

import sys
import os
import re
import io as _io
import json
import time
import threading
import datetime
import urllib.request
import urllib.error
from pathlib import Path

# ── Windows 콘솔 UTF-8 강제 ──
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'buffer'):
    sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ── DPI 설정 ──
os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
os.environ["QT_ENABLE_HIGHDPI_SCALING"]   = "1"

# =============================================
#  패키지 자가진단
# =============================================
_REQUIRED = {"pyautogui": "pyautogui", "PyQt6": "PyQt6", "PIL": "Pillow"}
_MISSING  = []
for _mod, _pkg in _REQUIRED.items():
    try: __import__(_mod)
    except ImportError: _MISSING.append(_pkg)

if _MISSING:
    try:
        from PyQt6.QtWidgets import QApplication, QMessageBox
        _app = QApplication(sys.argv)
        msg  = QMessageBox()
        msg.setWindowTitle("패키지 설치 필요")
        msg.setText("아래 패키지가 없습니다.\n\n"
                    + "\n".join(f"  - {p}" for p in _MISSING)
                    + "\n\nPowerShell:\n"
                    + f"pip install {' '.join(_MISSING)}")
        msg.exec()
    except Exception:
        print("필수 패키지 없음:", _MISSING)
    sys.exit(1)

# ── 선택적 패키지 ──
try:
    from pynput import mouse, keyboard
    from pynput.mouse import Button
    PYNPUT_OK = True
except ImportError:
    PYNPUT_OK = False

try:
    import cv2
    import numpy as np
    CV2_OK = True
except ImportError:
    CV2_OK = False

try:
    import pygetwindow as gw
    GW_OK = True
except ImportError:
    GW_OK = False

try:
    from pywinauto import Application as WinApp
    from pywinauto.findwindows import ElementNotFoundError
    WINAUTO_OK = True
except ImportError:
    WINAUTO_OK = False

try:
    import win32com.client as _win32com_client
    OUTLOOK_OK = True
except ImportError:
    OUTLOOK_OK = False

import pyautogui
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QListWidget, QListWidgetItem, QSpinBox,
    QDoubleSpinBox, QLineEdit, QTextEdit, QMessageBox, QFrame,
    QCheckBox, QComboBox, QScrollArea, QProgressBar, QDialog,
    QGridLayout, QStackedWidget, QInputDialog, QSizePolicy,
    QSplitter, QAbstractScrollArea, QSystemTrayIcon, QMenu
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, QTimer, QSize, QRect, QPoint,
    QPointF, QRectF, QPropertyAnimation, QEasingCurve
)
from PyQt6.QtGui import (
    QColor, QFont, QPainter, QBrush, QPen, QLinearGradient,
    QPixmap, QKeySequence, QShortcut, QPainterPath, QFontMetrics,
    QIcon, QRadialGradient, QImage, QAction
)

pyautogui.FAILSAFE = True
pyautogui.PAUSE    = 0.01

# ── DPI 스케일 보정 ──────────────────────────────────────────────────────────
# Windows 고DPI(125%/150%/200% 등) 환경에서 pynput은 물리 픽셀 좌표를 반환하고
# pyautogui도 물리 픽셀 기준으로 클릭한다.
# 반면 Qt/화면 배율 설정에 따라 논리 픽셀과 물리 픽셀이 달라진다.
# 녹화 PC 와 재생 PC 의 DPI 배율이 다르면 좌표가 어긋남.
# → 해결: 좌표를 항상 "논리 픽셀(배율 1.0 기준)" 으로 정규화해서 저장하고,
#          재생 시 현재 PC 의 DPI 배율을 곱해서 실제 좌표로 변환한다.

def _get_dpi_scale() -> float:
    """주 모니터의 DPI 배율 반환 (100% = 1.0, 150% = 1.5, 200% = 2.0). 좌표 미지정시 폴백용."""
    try:
        # ctypes 로 실제 물리 DPI 읽기 (가장 정확)
        import ctypes
        hdc = ctypes.windll.user32.GetDC(0)
        dpi = ctypes.windll.gdi32.GetDeviceCaps(hdc, 88)  # LOGPIXELSX
        ctypes.windll.user32.ReleaseDC(0, hdc)
        return dpi / 96.0
    except Exception:
        try:
            # QApplication fallback
            from PyQt6.QtWidgets import QApplication
            app = QApplication.instance()
            if app:
                return app.primaryScreen().devicePixelRatio()
        except Exception:
            pass
    return 1.0

def _get_dpi_scale_at(x: int, y: int) -> float:
    """
    다중 모니터 대응: 지정 좌표가 위치한 모니터의 DPI 배율을 반환한다.
    모니터마다 배율이 다른 환경(예: 주모니터 100% + 보조모니터 150%)에서
    _get_dpi_scale()(항상 주모니터 기준)만 쓰면 보조 모니터 좌표가 어긋남.
    MonitorFromPoint + GetDpiForMonitor(Windows 8.1+)로 해당 지점의
    실제 DPI 를 조회하고, 실패 시 _get_dpi_scale() 로 폴백.
    """
    try:
        import ctypes

        class _POINT(ctypes.Structure):
            _fields_ = [("x", ctypes.c_long), ("y", ctypes.c_long)]

        MONITOR_DEFAULTTONEAREST = 2
        MDT_EFFECTIVE_DPI = 0

        pt   = _POINT(int(x), int(y))
        hmon = ctypes.windll.user32.MonitorFromPoint(pt, MONITOR_DEFAULTTONEAREST)
        if not hmon:
            return _get_dpi_scale()

        dpi_x = ctypes.c_uint(); dpi_y = ctypes.c_uint()
        hr = ctypes.windll.shcore.GetDpiForMonitor(
            hmon, MDT_EFFECTIVE_DPI, ctypes.byref(dpi_x), ctypes.byref(dpi_y))
        if hr == 0 and dpi_x.value > 0:   # S_OK
            return dpi_x.value / 96.0
    except Exception:
        pass
    return _get_dpi_scale()

def _phys_to_logical(x: int, y: int) -> tuple:
    """물리 픽셀 → 논리 픽셀 (녹화 시 저장용). 클릭 지점 모니터의 실제 DPI 사용."""
    s = _get_dpi_scale_at(x, y)
    if s <= 1.0:
        return x, y
    return round(x / s), round(y / s)

def _logical_to_phys(x: int, y: int) -> tuple:
    """
    논리 픽셀 → 물리 픽셀 (재생 시 pyautogui 전달용).
    2단계 보정: 1) 주모니터 기준 배율로 대략적 물리좌표 추정
               2) 그 지점이 실제 속한 모니터의 정확한 DPI 로 재계산
    (모니터 경계 부근이 아니면 1단계만으로도 대부분 정확함)
    """
    s0 = _get_dpi_scale()
    if s0 <= 1.0:
        approx_x, approx_y = x, y
    else:
        approx_x, approx_y = round(x * s0), round(y * s0)
    s = _get_dpi_scale_at(approx_x, approx_y)
    if s <= 1.0:
        return x, y
    return round(x * s), round(y * s)

# 경로 설정 - 실행 위치 자동 감지
_HERE    = Path(sys.argv[0]).parent
SAVE_DIR = _HERE / "rpa_macros"
SAVE_DIR.mkdir(parents=True, exist_ok=True)
SNAPSHOT_DIR = SAVE_DIR / "snapshots"   # 녹화 시 클릭 위치 주변 자동 캡처 저장
SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)

# 첫 실행 시 샘플 플로우 자동 생성
def _init_sample_flows():
    """샘플 플로우 파일이 없으면 자동 생성"""
    sample_path = SAVE_DIR / "샘플_메모장자동화.json"
    if sample_path.exists(): return
    import json as _json
    sample = {
        "name": "샘플_메모장자동화",
        "created": "2025-01-01T00:00:00",
        "count": 6,
        "actions": [
            {"type": "hotkey",  "keys": ["win","r"],
             "delay": 0.5,  "label": "실행창 열기",
             "ai_desc": "Windows 실행 창을 엽니다"},
            {"type": "type",    "text": "notepad",
             "delay": 0.5,  "label": "메모장 입력",
             "ai_desc": "실행창에 notepad 를 입력합니다"},
            {"type": "key",     "key": "enter",
             "delay": 1.0,  "label": "실행",
             "ai_desc": "Enter 키로 메모장을 실행합니다"},
            {"type": "wait",    "seconds": 1,
             "delay": 0.0,  "label": "로딩 대기",
             "ai_desc": "메모장이 열릴 때까지 1초 대기합니다"},
            {"type": "type",    "text": "안녕하세요! RPA 자동화 테스트입니다.",
             "delay": 0.3,  "label": "텍스트 입력",
             "ai_desc": "메모장에 텍스트를 입력합니다"},
            {"type": "hotkey",  "keys": ["ctrl","s"],
             "delay": 0.5,  "label": "저장",
             "ai_desc": "Ctrl+S 로 파일을 저장합니다"},
        ]
    }
    with open(sample_path, "w", encoding="utf-8") as f:
        _json.dump(sample, f, ensure_ascii=False, indent=2)

_init_sample_flows()

# 보안정책 환경 자동 감지
# SentinelOne 설치 여부로 회사 PC 판단
# WA_TranslucentBackground: 보안정책 있으면 비활성 (COM 차단)
def _detect_secure_env() -> bool:
    """회사 보안정책 환경 감지 - True 면 보안정책 적용"""
    import os
    # 방법1: SentinelOne 서비스 확인
    sentinel_paths = [
        r"C:\Program Files\SentinelOne",
        r"C:\Program Files (x86)\SentinelOne",
    ]
    for p in sentinel_paths:
        if os.path.exists(p):
            return True
    # 방법2: 회사 PC 사용자명 확인
    username = os.environ.get("USERNAME", "").upper()
    if username == "HHI":
        return True
    return False

SECURE_ENV = _detect_secure_env()   # True = 회사PC, False = 개인PC

OLLAMA_URL = "http://127.0.0.1:11434"

# ── AI 모델명: PC(회사망/방산망/개인PC)마다 설치된 LLM이 다를 수 있으므로
#    하드코딩 대신 실행 시점에 자동 감지한다.
#    1) ai_model.txt 가 있으면 그 값을 그대로 사용 (수동 지정, 최우선)
#    2) 없으면 Ollama에 실제 설치된 모델 목록을 조회해 자동 선택
#    3) 둘 다 실패하면 기존 기본값(gemma3:4b)으로 폴백 (오프라인 상태에서도 앱이 죽지 않도록)
_MODEL_CFG_FILE = Path(os.path.expanduser("~")) / ".hd_easyrpa_ai_model.txt"

def _detect_ollama_model() -> str:
    # 1) 수동 지정 파일 우선
    try:
        if _MODEL_CFG_FILE.exists():
            saved = _MODEL_CFG_FILE.read_text(encoding="utf-8").strip()
            if saved:
                return saved
    except Exception:
        pass
    # 2) 설치된 모델 자동 조회
    try:
        with urllib.request.urlopen(
                f"{OLLAMA_URL}/api/tags", timeout=3) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        names = [m.get("name", "") for m in data.get("models", []) if m.get("name")]
        if names:
            # gemma 계열을 우선 선호 (기존 프롬프트가 gemma 기준으로 작성됨)
            preferred = [n for n in names if "gemma" in n.lower()]
            return (preferred or names)[0]
    except Exception:
        pass
    # 3) 폴백 기본값 (Ollama 미실행 상태에서도 변수는 존재해야 함)
    return "gemma3:4b"

OLLAMA_MODEL = _detect_ollama_model()

def set_ollama_model(name: str):
    """사용자가 설정 화면 등에서 모델을 수동 지정할 때 사용."""
    global OLLAMA_MODEL
    OLLAMA_MODEL = name
    try:
        _MODEL_CFG_FILE.write_text(name, encoding="utf-8")
    except Exception:
        pass

# CoordHighlighter GC 방지용 전역 참조
_HIGHLIGHTER_REFS = []

# =============================================
#  디자인 토큰 (라이트 테마)
# =============================================
C = {
    # 배경
    "bg0":      "#F4F6FB",   # 앱 전체 배경
    "bg1":      "#FFFFFF",   # 카드/패널
    "bg2":      "#F8FAFD",   # 메인 패널
    "bg3":      "#FFFFFF",
    "bg4":      "#EDF0F7",   # 호버
    "border":   "#E2E8F0",
    "border2":  "#EDF2F7",

    # 사이드바 - 딥 네이비 그라데이션
    "side_top":  "#1E2B5E",   # 사이드바 상단
    "side_bot":  "#16213A",   # 사이드바 하단
    "side_act":  "#4F8EF7",   # 활성 버튼 강조색

    # 브랜드
    "brand":    "#4F8EF7",   # 밝은 블루 (메인 액센트)
    "brand2":   "#2563EB",   # 진한 블루

    # 기능 컬러 - 선명하고 모던하게
    "rec":      "#F43F5E",   # 로즈 레드
    "rec_dim":  "#FFF1F2",
    "play":     "#10B981",   # 에메랄드 그린
    "play_dim": "#ECFDF5",
    "stop":     "#F59E0B",   # 앰버
    "stop_dim": "#FFFBEB",
    "tool":     "#8B5CF6",   # 바이올렛
    "tool_dim": "#F5F3FF",
    "ai":       "#06B6D4",   # 시안 (AI 컬러)
    "ai_dim":   "#ECFEFF",

    # 텍스트
    "t1":       "#0F172A",   # 거의 블랙
    "t2":       "#475569",   # 슬레이트
    "t3":       "#94A3B8",   # 연한 슬레이트

    # 상태
    "ok":       "#10B981",
    "warn":     "#F59E0B",
    "err":      "#F43F5E",
}

GLOBAL_QSS = f"""
* {{ font-family: 'Malgun Gothic', sans-serif; font-size: 13px; color: {C['t1']}; }}
QMainWindow, QWidget {{ background: {C['bg2']}; }}
QScrollArea {{ border: none; background: transparent; }}
QScrollBar:vertical {{
    background: {C['bg2']}; width: 7px; border-radius: 4px; margin: 2px 0;
}}
QScrollBar::handle:vertical {{
    background: {C['border']}; border-radius: 4px; min-height: 28px;
}}
QScrollBar::handle:vertical:hover {{ background: {C['brand']}; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{
    background: {C['bg2']}; height: 7px; border-radius: 4px;
}}
QScrollBar::handle:horizontal {{
    background: {C['border']}; border-radius: 4px;
}}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{ width: 0; }}
QToolTip {{
    background: {C['bg1']}; color: {C['t1']};
    border: 1px solid {C['border']}; border-radius: 6px; padding: 5px 10px;
}}
QMessageBox {{ background: {C['bg1']}; color: {C['t1']}; }}
QInputDialog {{ background: {C['bg1']}; color: {C['t1']}; }}
"""

# =============================================
#  AI 엔진 (Ollama 로컬 LLM)
# =============================================
class GemmaEngine:
    """Ollama 로컬 LLM 연동 - 폐쇄망 지원 (설치된 모델 자동 감지)"""

    @staticmethod
    def _ask(prompt: str, timeout: int = 30,
             num_predict: int = 256, _retry_no_think: bool = True) -> str:
        body_dict = {
            "model":  OLLAMA_MODEL,
            "prompt": prompt,
            "stream": False,
            "think":  False,   # 추론(thinking) 모델(qwen3.x, gemma4 등) 대응:
                                # think 토큰이 응답을 다 차지해서 실제 답변이
                                # 안 나오거나 잘리는 문제를 막기 위해 명시적으로 끔
            "options": {
                "temperature": 0.1,
                "num_predict": num_predict,
                "stop": ["```", "---", "\n\n\n"],
            }
        }
        payload = json.dumps(body_dict, ensure_ascii=False).encode("utf-8")
        import socket as _socket
        _prev_timeout = _socket.getdefaulttimeout()
        try:
            # urlopen(timeout=) 만으로는 일부 환경(프록시/특정 네트워크 스택)에서
            # read 타임아웃이 적용되지 않는 사례가 있어 소켓 기본 타임아웃도 함께 강제
            _socket.setdefaulttimeout(timeout)
            req = urllib.request.Request(
                f"{OLLAMA_URL}/api/generate",
                data    = payload,
                headers = {"Content-Type": "application/json"},
                method  = "POST")
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                text = data.get("response", "").strip()
                # think:false 를 무시하는 일부 모델/버전 대비 - <think> 블록 강제 제거
                if "<think>" in text:
                    text = re.sub(r"<think>.*?</think>", "", text,
                                   flags=re.DOTALL).strip()
                return text
        except urllib.error.HTTPError as e:
            # 구버전 Ollama가 "think" 파라미터 자체를 모르는 경우 -> 빼고 1회 재시도
            if e.code == 400 and _retry_no_think:
                body_dict.pop("think", None)
                try:
                    req2 = urllib.request.Request(
                        f"{OLLAMA_URL}/api/generate",
                        data=json.dumps(body_dict, ensure_ascii=False).encode("utf-8"),
                        headers={"Content-Type": "application/json"},
                        method="POST")
                    with urllib.request.urlopen(req2, timeout=timeout) as resp:
                        data = json.loads(resp.read().decode("utf-8"))
                        text = data.get("response", "").strip()
                        if "<think>" in text:
                            text = re.sub(r"<think>.*?</think>", "", text,
                                           flags=re.DOTALL).strip()
                        return text
                except Exception:
                    pass
            # 서버는 응답했지만 오류(모델 미존재 등) - "오프라인"이 아니라 실제 원인 표시
            try:
                body = e.read().decode("utf-8", "ignore")
                err_detail = json.loads(body).get("error", body)
            except Exception:
                err_detail = str(e)
            if e.code == 404 or "not found" in str(err_detail).lower():
                return (f"[AI 오류] 모델 '{OLLAMA_MODEL}' 을 찾을 수 없습니다.\n"
                        f"'ollama list' 로 설치된 모델명을 확인 후, "
                        f"맞으면 'ollama pull {OLLAMA_MODEL}' 실행하세요.")
            return f"[AI 오류 {e.code}] {err_detail}"
        except urllib.error.URLError as e:
            # 진짜 연결 실패 (서버 미실행/포트 차단 등)
            return (f"[AI 오프라인] {OLLAMA_URL} 접속 실패. "
                    f"'ollama serve' 실행 여부와 포트(11434)를 확인하세요. ({e.reason})")
        except TimeoutError:
            return "[AI 오류] 응답 시간 초과 - 모델이 너무 느리거나 PC 리소스 부족"
        except json.JSONDecodeError:
            return "[AI 오류] Ollama 응답을 해석할 수 없습니다 (응답 형식 오류)"
        except Exception as e:
            return f"[AI 오류] {type(e).__name__}: {e}"
        finally:
            _socket.setdefaulttimeout(_prev_timeout)

    @staticmethod
    def warmup():
        """
        모델을 미리 메모리에 로드해 콜드스타트(첫 호출 수십초 지연)를 완화.
        실패해도 무시 - 어차피 실제 사용 시 정상 timeout 로직으로 처리됨.
        """
        try:
            GemmaEngine._ask("ok", timeout=90, num_predict=5)
        except Exception:
            pass

    @staticmethod
    def is_online() -> bool:
        """Ollama 서버 접속 + 설치된 모델 존재 여부 확인 (PC마다 모델명이 달라도 동작)"""
        try:
            with urllib.request.urlopen(
                    f"{OLLAMA_URL}/api/tags", timeout=4) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            models = [m.get("name", "") for m in data.get("models", [])]
            if not models:
                return False
            # 현재 OLLAMA_MODEL 이 실제 설치 목록에 없으면 자동 재감지
            global OLLAMA_MODEL
            base = OLLAMA_MODEL.split(":")[0]
            if not any(OLLAMA_MODEL == m or m.startswith(base + ":") for m in models):
                OLLAMA_MODEL = _detect_ollama_model()
            return True
        except Exception:
            return False

    @staticmethod
    def describe_block(action: dict) -> str:
        t = action.get("type", "")
        prompts = {
            "click": (
                f"다음 RPA 자동화 동작을 초보자가 이해할 수 있는 한 줄 한국어 설명으로 바꿔주세요.\n"
                f"동작: 화면 좌표 ({action.get('x')}, {action.get('y')}) 를 "
                f"{action.get('button','left')} 클릭\n"
                f"예시 출력: '화면 위쪽 메뉴 버튼 클릭'\n설명만 출력하세요."
            ),
            "key": (
                f"RPA 동작을 한 줄 한국어로 설명하세요.\n"
                f"동작: 키보드 [{action.get('key','')}] 입력\n설명만 출력하세요."
            ),
            "type": (
                f"RPA 동작을 한 줄 한국어로 설명하세요.\n"
                f"동작: 텍스트 입력 '{action.get('text','')}'\n설명만 출력하세요."
            ),
            "wait": (
                f"RPA 동작을 한 줄 한국어로 설명하세요.\n"
                f"동작: {action.get('seconds',1)}초 대기\n설명만 출력하세요."
            ),
            "image_click": (
                f"RPA 동작을 한 줄 한국어로 설명하세요.\n"
                f"동작: 화면에서 '{action.get('label','')}' 이미지를 찾아 클릭\n설명만 출력하세요."
            ),
        }
        prompt = prompts.get(t, f"RPA 동작 '{t}'을 한 줄 한국어로 설명하세요.")
        return GemmaEngine._ask(prompt, timeout=45)

    @staticmethod
    def natural_to_blocks(text: str) -> list:
        # 프롬프트 최대한 간결하게 (토큰 절약 → 빠른 응답)
        prompt = (
            "RPA JSON only. No explanation.\n"
            "Types: click(x,y,button,delay) key(key,delay) "
            "type(text,delay) hotkey(keys[],delay) wait(seconds,delay)\n"
            "Rules: unknown coords=x:500,y:400. Korean text=type. delay=0.5\n"
            f"Request: {text}\n"
            "JSON array:"
        )
        raw = GemmaEngine._ask(prompt, timeout=45, num_predict=300)

        if raw.startswith("[AI"):
            # AI 오프라인 또는 오류 → fallback
            blocks = GemmaEngine._fallback_parse(text)
            return blocks

        # 디버그: AI 응답 로그 (첫 100자)
        import sys
        print(f"[AI 응답] {raw[:150]}", file=sys.stderr)

        # JSON 파싱 - 여러 방식 시도
        # 방법1: 전체가 JSON
        try:
            start = raw.find("[")
            end   = raw.rfind("]") + 1
            if start >= 0 and end > start:
                blocks = json.loads(raw[start:end])
                if isinstance(blocks, list) and blocks:
                    for b in blocks:
                        if "delay" not in b:
                            b["delay"] = 0.5
                    return blocks
        except Exception:
            pass

        # 방법2: 코드블록 안에 있는 경우
        try:
            import re
            m = re.search(r'```(?:json)?\s*(\[.*?\])\s*```',
                          raw, re.DOTALL)
            if m:
                blocks = json.loads(m.group(1))
                if isinstance(blocks, list):
                    return blocks
        except Exception:
            pass

        # 방법3: fallback
        blocks = GemmaEngine._fallback_parse(text)
        return blocks

    @staticmethod
    def _fallback_parse(text: str) -> list:
        """
        AI 응답 파싱 실패 시 규칙 기반으로 직접 생성.
        자주 쓰는 패턴을 키워드로 감지.
        """
        import re
        blocks = []
        t = text.lower()

        # 메모장 실행
        if any(k in t for k in ["메모장", "notepad"]):
            blocks.append({"type":"hotkey","keys":["win","r"],"delay":0.5})
            blocks.append({"type":"type","text":"notepad","delay":0.5})
            blocks.append({"type":"key","key":"enter","delay":1.5})

        # 엑셀 실행
        if any(k in t for k in ["엑셀", "excel"]):
            blocks.append({"type":"hotkey","keys":["win","r"],"delay":0.5})
            blocks.append({"type":"type","text":"excel","delay":0.5})
            blocks.append({"type":"key","key":"enter","delay":2.0})

        # 텍스트 입력 패턴 ("바보", '바보', 입력: 바보, 「바보」)
        for pattern in [r'"(.+?)"', r"'(.+?)'",
                        r'[입력|타이핑|써|작성][:\s]+([^\s,\.]+)',
                        r'「(.+?)」']:
            m = re.search(pattern, text)
            if m:
                txt = m.group(1)
                if txt and len(txt) < 100:
                    blocks.append({"type":"type","text":txt,"delay":0.5})
                    break

        # 바탕화면 저장 (다른 이름으로 저장)
        if any(k in t for k in ["바탕화면", "desktop", "저장"]):
            blocks.append({"type":"hotkey","keys":["ctrl","shift","s"],"delay":1.0})
            blocks.append({"type":"wait","seconds":1,"delay":0})
            # 파일명 감지
            fname_m = re.search(r'([가-힣a-zA-Z0-9_]+)\.(txt|docx|xlsx|csv)', text)
            if fname_m:
                blocks.append({"type":"type",
                               "text": f"C:\\Users\\HHI\\Desktop\\{fname_m.group(0)}",
                               "delay":0.5})
            else:
                blocks.append({"type":"type",
                               "text":"C:\\Users\\HHI\\Desktop\\문서.txt",
                               "delay":0.5})
            blocks.append({"type":"key","key":"enter","delay":0.3})

        # 그냥 저장 (Ctrl+S)
        elif any(k in t for k in ["저장", "save"]):
            blocks.append({"type":"hotkey","keys":["ctrl","s"],"delay":0.5})

        # 닫기
        if any(k in t for k in ["닫기", "close", "종료"]):
            blocks.append({"type":"hotkey","keys":["alt","f4"],"delay":0.5})
            if any(k in t for k in ["저장 안", "저장안", "no"]):
                blocks.append({"type":"key","key":"n","delay":0.3})

        # 엔터/확인
        if any(k in t for k in ["확인", "누르기"]) and not blocks:
            blocks.append({"type":"key","key":"enter","delay":0.3})

        return blocks

    @staticmethod
    def review_flow(actions: list) -> str:
        summary = []
        for i, a in enumerate(actions, 1):
            t = a.get("type","")
            if t == "click":
                summary.append(f"{i}. 클릭({a.get('x')},{a.get('y')})")
            elif t == "key":
                summary.append(f"{i}. 키입력[{a.get('key','')}]")
            elif t == "type":
                summary.append(f"{i}. 텍스트입력'{a.get('text','')[:20]}'")
            elif t == "wait":
                summary.append(f"{i}. 대기{a.get('seconds',1)}초")
            elif t == "image_click":
                summary.append(f"{i}. 이미지클릭[{a.get('label','')}]")
            else:
                summary.append(f"{i}. {t}")

        prompt = f"""당신은 RPA 자동화 전문가입니다.
아래 자동화 플로우를 검토하고 초보자가 이해할 수 있는 한국어로 조언해주세요.

플로우:
{chr(10).join(summary)}

다음 항목을 검토하세요:
1. 불안정할 수 있는 단계 (로딩 없이 바로 클릭 등)
2. 개선 제안 (대기 추가, 이미지 인식 활용 등)
3. 전체 플로우 요약

3~5문장으로 간결하게 한국어로 답하세요."""
        return GemmaEngine._ask(prompt, timeout=60)

    @staticmethod
    def explain_error(errors: list) -> str:
        if not errors:
            return "오류가 없습니다. 플로우가 정상입니다."
        err_txt = "\n".join(f"- {e}" for e in errors)
        prompt = f"""RPA 자동화 오류를 초보자가 이해할 수 있게 설명하고 해결 방법을 알려주세요.

오류 목록:
{err_txt}

각 오류마다 원인과 해결방법을 한국어로 설명하세요. 쉽고 친절하게."""
        return GemmaEngine._ask(prompt, timeout=45)

    @staticmethod
    def optimize_flow(actions: list) -> dict:
        """
        AI 프로세스 최적화.
        녹화 도중 생긴 '의미없는 시간'(망설임, 멈춤, 중복 동작)을 찾아내고
        delay 값을 줄인 최적화된 actions 리스트를 반환한다.
        AI(Ollama)가 꺼져 있어도 규칙기반(rule-based)으로 항상 동작하도록
        fallback을 갖춘다 - AI 오프라인이어도 이 기능 자체는 죽지 않음.
        반환: {"actions": [...], "report": "...", "saved_seconds": float, "removed": int}
        """
        if not actions:
            return {"actions": [], "report": "최적화할 동작이 없습니다.",
                     "saved_seconds": 0.0, "removed": 0}

        import copy
        opt = copy.deepcopy(actions)
        removed = 0
        saved = 0.0

        # ── 규칙 1: delay 캡 (사람이 멈춰서 생긴 비정상적으로 긴 대기 단축)
        #    실제 화면 전환/로딩으로 필요한 대기는 살리고, 단순 '망설임'만 줄인다.
        DELAY_CAP = 1.2          # 이보다 긴 delay는 망설임으로 간주해 캡 적용
        MIN_KEEP  = 0.3          # 최소한의 자연스러운 간격은 유지

        i = 0
        while i < len(opt):
            a = opt[i]
            d = a.get("delay", 0) or 0
            if d > DELAY_CAP:
                saved += (d - DELAY_CAP)
                a["delay"] = DELAY_CAP
            elif 0 <= d < MIN_KEEP and i > 0:
                # 너무 빠른 간격은 그대로 둠 (재생 안정성을 위해 줄이지 않음)
                pass
            i += 1

        # ── 규칙 2: 동일 좌표 연속 클릭 중복 제거
        #    (더블클릭 의도가 아닌, 손이 미끄러져 같은 곳을 또 누른 경우)
        cleaned = []
        for idx, a in enumerate(opt):
            if (a.get("type") == "click" and cleaned and
                    cleaned[-1].get("type") == "click" and
                    cleaned[-1].get("x") == a.get("x") and
                    cleaned[-1].get("y") == a.get("y") and
                    a.get("button") == cleaned[-1].get("button") and
                    (a.get("delay") or 0) < 0.15):
                # 같은 자리 0.15초 이내 재클릭 -> 중복으로 보고 제거
                removed += 1
                saved += (a.get("delay") or 0)
                continue
            cleaned.append(a)
        opt = cleaned

        # ── 규칙 3: 연속된 같은 키 반복입력 중 비정상적으로 긴 delay만 압축
        #    (텍스트 자체는 보존, 타이밍만 최적화 - 이미 규칙1에서 처리됨)

        rule_report = (
            f"중복 클릭 {removed}건 제거, 비정상 대기시간 약 {saved:.1f}초 단축 "
            f"(규칙 기반 1차 최적화)"
        )

        # ── AI(Ollama)로 추가 검토 - 오프라인이면 규칙기반 결과만 사용
        ai_report = ""
        if GemmaEngine.is_online():
            summary = []
            for i, a in enumerate(opt, 1):
                t = a.get("type", "")
                d = a.get("delay", 0)
                if t == "click":
                    summary.append(f"{i}.클릭({a.get('x')},{a.get('y')}) +{d:.1f}s")
                elif t == "type":
                    summary.append(f"{i}.입력'{a.get('text','')[:15]}' +{d:.1f}s")
                else:
                    summary.append(f"{i}.{t} +{d:.1f}s")
            prompt = (
                "다음은 RPA 자동화 플로우(이미 1차 규칙기반 최적화 완료)입니다.\n"
                + "\n".join(summary[:60]) +
                "\n\n추가로 비효율적이거나 불안정해 보이는 구간이 있으면 "
                "2~3문장으로 한국어로 간단히 조언하세요. 없으면 '추가 개선사항 없음'이라고 답하세요."
            )
            ai_report = GemmaEngine._ask(prompt, timeout=45)
            if ai_report.startswith("[AI"):
                ai_report = ""  # AI 오류 메시지는 보고서에 노출하지 않음

        report = rule_report + (f"\n\nAI 추가 검토: {ai_report}" if ai_report else "")

        return {
            "actions": opt,
            "report": report,
            "saved_seconds": round(saved, 2),
            "removed": removed,
        }


# =============================================
#  공통 커스텀 위젯
# =============================================

class GlowButton(QPushButton):
    def __init__(self, text, accent=None, parent=None):
        super().__init__(text, parent)
        self._accent   = accent or C['brand']
        self._glow     = 0
        self._pulsing  = False
        self._timer    = QTimer(self)
        self._dir      = 1
        self._timer.timeout.connect(self._tick)
        self.setCursor(Qt.CursorShape.PointingHandCursor)

    def start_pulse(self):
        self._pulsing = True
        self._timer.start(40)

    def stop_pulse(self):
        self._pulsing = False
        self._timer.stop()
        self._glow = 0
        self.update()

    def _tick(self):
        self._glow += self._dir * 8
        if self._glow >= 120: self._dir = -1
        if self._glow <= 0:   self._dir = 1
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        r  = self.rect().adjusted(2, 2, -2, -2)
        ac = QColor(self._accent)
        if not self.isEnabled():
            painter.setPen(QPen(QColor(C['border']), 1))
            painter.setBrush(QBrush(QColor(C['bg4'])))
            painter.drawRoundedRect(r, 10, 10)
            painter.setPen(QColor(C['t3']))
            painter.setFont(self.font())
            painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, self.text())
            painter.end(); return
        if self._glow > 0 or self._pulsing:
            # 호버: 채운 그라데이션 버튼
            grad = QLinearGradient(r.left(), r.top(), r.left(), r.bottom())
            grad.setColorAt(0, ac.lighter(118))
            grad.setColorAt(1, ac.darker(108))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QBrush(grad))
            painter.drawRoundedRect(r, 10, 10)
            # 미묘한 inner highlight
            hl = QColor(255,255,255,40)
            painter.setBrush(QBrush(hl))
            painter.drawRoundedRect(r.adjusted(0,0,0,-r.height()//2), 10, 10)
            painter.setPen(QColor("#FFFFFF"))
        else:
            # 기본: 연한 배경 + 테두리
            bg = QColor(ac); bg.setAlpha(10)
            painter.setPen(QPen(ac, 1.5))
            painter.setBrush(QBrush(bg))
            painter.drawRoundedRect(r, 10, 10)
            painter.setPen(ac)
        f = QFont(self.font()); f.setBold(True)
        painter.setFont(f)
        painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, self.text())
        painter.end()

    def enterEvent(self, e):
        if not self._pulsing: self._glow = 60; self.update()
        super().enterEvent(e)

    def leaveEvent(self, e):
        if not self._pulsing: self._glow = 0; self.update()
        super().leaveEvent(e)


class NavButton(QPushButton):
    def __init__(self, icon_text, label, accent, parent=None):
        super().__init__(parent)
        self._icon_text = icon_text
        self._label     = label
        self._accent    = accent
        self._active    = False
        self.setCheckable(True)
        self.setFixedHeight(64)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setToolTip(label)

    def setActive(self, val):
        self._active = val
        self.setChecked(val)
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        r  = self.rect()
        ac = QColor(self._accent)

        # 배경
        if self._active:
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QBrush(QColor(255, 255, 255, 28)))
            painter.drawRoundedRect(r.adjusted(4, 2, -4, -2), 10, 10)
            # 왼쪽 강조 바
            painter.setBrush(QBrush(ac))
            painter.drawRoundedRect(QRect(0, r.height()//5, 3, r.height()*3//5), 2, 2)
        elif self.underMouse():
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QBrush(QColor(255, 255, 255, 12)))
            painter.drawRoundedRect(r.adjusted(4, 2, -4, -2), 10, 10)

        # 아이콘 텍스트 - 고정 크기 11pt, 굵게
        icon_font = QFont("Malgun Gothic", 11)
        icon_font.setBold(True)
        icon_font.setLetterSpacing(QFont.SpacingType.AbsoluteSpacing, 1.5)
        painter.setFont(icon_font)
        painter.setPen(QColor(ac) if self._active else QColor(255, 255, 255, 215))
        painter.drawText(
            QRect(r.x(), r.y() + 8, r.width(), r.height() - 26),
            Qt.AlignmentFlag.AlignCenter, self._icon_text)

        # 레이블 - 고정 크기 8pt
        lf = QFont("Malgun Gothic", 8)
        lf.setBold(self._active)
        painter.setFont(lf)
        painter.setPen(QColor(ac) if self._active else QColor(255, 255, 255, 170))
        painter.drawText(
            QRect(r.x(), r.bottom() - 18, r.width(), 18),
            Qt.AlignmentFlag.AlignCenter, self._label)
        painter.end()


class SectionHeader(QLabel):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.setStyleSheet(f"""
            color: {C['brand']}; font-size: 11px; font-weight: 900;
            letter-spacing: 2px; padding: 0 0 8px 2px;
            border-bottom: 2px solid {C['brand']}; background: transparent;
        """)


class StyledInput(QLineEdit):
    def __init__(self, placeholder="", parent=None):
        super().__init__(parent)
        self.setPlaceholderText(placeholder)
        self.setStyleSheet(f"""
            QLineEdit {{
                background: {C['bg1']}; border: 1.5px solid {C['border']};
                border-radius: 8px; color: {C['t1']};
                padding: 8px 12px; font-size: 13px;
            }}
            QLineEdit:focus {{ border-color: {C['brand']}; background: #FAFCFF; }}
        """)


class StyledSpin(QSpinBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QSpinBox {{
                background: {C['bg1']}; border: 1.5px solid {C['border']};
                border-radius: 8px; color: {C['t1']}; padding: 6px 10px;
            }}
            QSpinBox:focus {{ border-color: {C['brand']}; }}
            QSpinBox::up-button, QSpinBox::down-button {{
                background: {C['bg4']}; border: none; width: 20px; border-radius: 4px;
            }}
        """)


class StyledDSpin(QDoubleSpinBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QDoubleSpinBox {{
                background: {C['bg1']}; border: 1.5px solid {C['border']};
                border-radius: 8px; color: {C['t1']}; padding: 6px 10px;
            }}
            QDoubleSpinBox:focus {{ border-color: {C['brand']}; }}
            QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {{
                background: {C['bg4']}; border: none; width: 20px; border-radius: 4px;
            }}
        """)


class StyledCombo(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QComboBox {{
                background: {C['bg1']}; border: 1.5px solid {C['border']};
                border-radius: 8px; color: {C['t1']}; padding: 6px 12px;
            }}
            QComboBox:focus {{ border-color: {C['brand']}; }}
            QComboBox::drop-down {{ border: none; width: 28px; }}
            QComboBox QAbstractItemView {{
                background: {C['bg1']}; border: 1px solid {C['border']};
                color: {C['t1']}; selection-background-color: {C['brand']};
                selection-color: #FFFFFF; padding: 4px;
            }}
        """)



# =============================================
#  QMessageBox 대체 - COM 충돌 방지
#  순수 PyQt6 구현 (SentinelOne 보안정책 대응)
# =============================================
def _msg(parent, title: str, text: str,
         kind: str = "info") -> bool:
    """
    QMessageBox 대체 다이얼로그.
    kind: "info" / "warn" / "error" / "question"
    question 은 True(Yes)/False(No) 반환.
    """
    dlg = QDialog(parent)
    dlg.setWindowTitle(title)
    dlg.setMinimumWidth(360)
    dlg.setStyleSheet(f"""
        QDialog{{background:{C['bg1']};}}
        QLabel{{background:transparent;color:{C['t1']};}}
    """)
    lay = QVBoxLayout(dlg)
    lay.setContentsMargins(24,20,24,18); lay.setSpacing(14)

    # 아이콘 + 메시지 행
    msg_row = QHBoxLayout(); msg_row.setSpacing(14)
    icons = {"info":"i","warn":"!","error":"X","question":"?"}
    colors = {"info":C['brand'],"warn":C['stop'],
              "error":C['rec'],"question":C['brand']}
    ic = QLabel(icons.get(kind,"i"))
    ic.setStyleSheet(f"font-size:28px;color:{colors.get(kind,C['brand'])};")
    ic.setFixedWidth(36)
    msg_lbl = QLabel(text)
    msg_lbl.setWordWrap(True)
    msg_lbl.setStyleSheet(f"font-size:13px;color:{C['t1']};")
    msg_row.addWidget(ic); msg_row.addWidget(msg_lbl,1)
    lay.addLayout(msg_row)

    # 버튼 행
    btn_row = QHBoxLayout(); btn_row.setSpacing(8)
    btn_row.addStretch()
    result = [False]

    if kind == "question":
        btn_no  = GlowButton("아니오", C['t2'])
        btn_yes = GlowButton("예",      colors[kind])
        btn_no.setFixedHeight(36);  btn_no.setFixedWidth(80)
        btn_yes.setFixedHeight(36); btn_yes.setFixedWidth(80)
        btn_no.clicked.connect(dlg.reject)
        btn_yes.clicked.connect(lambda: (result.__setitem__(0,True), dlg.accept()))
        btn_row.addWidget(btn_no); btn_row.addWidget(btn_yes)
    else:
        btn_ok = GlowButton("확인", colors[kind])
        btn_ok.setFixedHeight(36); btn_ok.setFixedWidth(80)
        btn_ok.clicked.connect(dlg.accept)
        btn_row.addWidget(btn_ok)

    lay.addLayout(btn_row)
    dlg.exec()
    return result[0]

def msg_info(parent, title, text):     _msg(parent, title, text, "info")
def msg_warn(parent, title, text):     _msg(parent, title, text, "warn")
def msg_error(parent, title, text):    _msg(parent, title, text, "error")
def msg_ask(parent, title, text)->bool: return _msg(parent, title, text, "question")


def _type_with_clipboard(text: str):
    """
    한글/특수문자 포함 텍스트 입력.
    pyautogui.typewrite 는 영어만 지원 → 클립보드 붙여넣기 방식 사용.
    QApplication.clipboard() 는 COM 충돌 위험 → PowerShell 경유 (보안정책 대응)
    """
    import subprocess, uuid
    try:
        # 매 호출마다 고유 파일명 사용 (반복재생/연속호출 시 파일 경합 방지)
        tmp_file = SAVE_DIR / f"_clip_tmp_{uuid.uuid4().hex[:8]}.txt"
        with open(tmp_file, "w", encoding="utf-8") as f:
            f.write(text)
        result = subprocess.run(
            ['powershell', '-NoProfile', '-Command',
             f"Get-Content -Path '{tmp_file}' -Raw -Encoding UTF8 | Set-Clipboard"],
            capture_output=True, timeout=3,
            creationflags=subprocess.CREATE_NO_WINDOW
                if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
        try: tmp_file.unlink(missing_ok=True)
        except Exception: pass
        if result.returncode != 0:
            # 클립보드 설정 실패 시 붙여넣기를 시도하지 않음(엉뚱한 내용 붙여넣기 방지)
            return
    except Exception:
        return
    time.sleep(0.2)   # 클립보드 반영 대기
    pyautogui.hotkey('ctrl', 'v')
    # 텍스트 길이에 비례한 안정화 대기 (짧은 텍스트 0.1초, 긴 텍스트 최대 0.4초)
    # 대상 앱이 붙여넣기를 완전히 처리하기 전에 다음 액션이 오면 입력이 씹힘
    wait = min(0.1 + len(text) * 0.005, 0.4)
    time.sleep(wait)


def _is_korean(text: str) -> bool:
    """한글 포함 여부 확인"""
    return any(ord(c) >= 0xAC00 and ord(c) <= 0xD7A3 or
               ord(c) >= 0x1100 and ord(c) <= 0x11FF or
               ord(c) >= 0x3130 and ord(c) <= 0x318F for c in text)


# =============================================
#  두벌식 영문 → 한글 변환
#  ──────────────────────────────────────
#  pynput 같은 OS 후크 라이브러리는 IME 가 변환하기 전의 raw 영문 키만
#  잡아낸다. 즉 사용자가 한글로 "안녕" 을 쳐도 후크는 "dkssud" 으로 받음.
#  → 녹화 시 IME 가 한글 모드면 입력 시퀀스를 두벌식 매핑으로 한글 음절로
#     변환해서 저장 → 재생 시 클립보드로 한글 그대로 입력
# =============================================
_DUBEOL_MAP = {
    'q':'ㅂ','w':'ㅈ','e':'ㄷ','r':'ㄱ','t':'ㅅ',
    'y':'ㅛ','u':'ㅕ','i':'ㅑ','o':'ㅐ','p':'ㅔ',
    'a':'ㅁ','s':'ㄴ','d':'ㅇ','f':'ㄹ','g':'ㅎ',
    'h':'ㅗ','j':'ㅓ','k':'ㅏ','l':'ㅣ',
    'z':'ㅋ','x':'ㅌ','c':'ㅊ','v':'ㅍ','b':'ㅠ',
    'n':'ㅜ','m':'ㅡ',
    # 쌍자음/이중모음 (Shift+키)
    'Q':'ㅃ','W':'ㅉ','E':'ㄸ','R':'ㄲ','T':'ㅆ','O':'ㅒ','P':'ㅖ',
}
_HCHO = ['ㄱ','ㄲ','ㄴ','ㄷ','ㄸ','ㄹ','ㅁ','ㅂ','ㅃ','ㅅ','ㅆ','ㅇ','ㅈ','ㅉ','ㅊ','ㅋ','ㅌ','ㅍ','ㅎ']
_HJUNG = ['ㅏ','ㅐ','ㅑ','ㅒ','ㅓ','ㅔ','ㅕ','ㅖ','ㅗ','ㅘ','ㅙ','ㅚ','ㅛ','ㅜ','ㅝ','ㅞ','ㅟ','ㅠ','ㅡ','ㅢ','ㅣ']
_HJONG = ['','ㄱ','ㄲ','ㄳ','ㄴ','ㄵ','ㄶ','ㄷ','ㄹ','ㄺ','ㄻ','ㄼ','ㄽ','ㄾ','ㄿ','ㅀ','ㅁ','ㅂ','ㅄ','ㅅ','ㅆ','ㅇ','ㅈ','ㅊ','ㅋ','ㅌ','ㅍ','ㅎ']
_HVOWEL = set(_HJUNG)
_COMPLEX_JUNG = {('ㅗ','ㅏ'):'ㅘ',('ㅗ','ㅐ'):'ㅙ',('ㅗ','ㅣ'):'ㅚ',
                 ('ㅜ','ㅓ'):'ㅝ',('ㅜ','ㅔ'):'ㅞ',('ㅜ','ㅣ'):'ㅟ',('ㅡ','ㅣ'):'ㅢ'}
_COMPLEX_JONG = {('ㄱ','ㅅ'):'ㄳ',('ㄴ','ㅈ'):'ㄵ',('ㄴ','ㅎ'):'ㄶ',
                 ('ㄹ','ㄱ'):'ㄺ',('ㄹ','ㅁ'):'ㄻ',('ㄹ','ㅂ'):'ㄼ',
                 ('ㄹ','ㅅ'):'ㄽ',('ㄹ','ㅌ'):'ㄾ',('ㄹ','ㅍ'):'ㄿ',
                 ('ㄹ','ㅎ'):'ㅀ',('ㅂ','ㅅ'):'ㅄ'}
_SPLIT_JONG = {v: k for k, v in _COMPLEX_JONG.items()}

def _compose_hangul(cho, jung, jong=''):
    """초/중/종성 자모 → 한글 음절. 실패하면 자모 그대로 반환."""
    try:
        return chr(0xAC00 + _HCHO.index(cho)*588 + _HJUNG.index(jung)*28 +
                   (_HJONG.index(jong) if jong else 0))
    except Exception:
        return (cho or '') + (jung or '') + (jong or '')

def _eng_to_kor(text: str) -> str:
    """
    두벌식 영문 시퀀스 → 한글 음절 변환.
    예) 'dkssud' → '안녕',  'spdlqj' → '네이버'
    한글 자모로 변환 안 되는 문자는 그대로 통과.
    """
    if not text:
        return text
    jamos = [_DUBEOL_MAP.get(c, c) for c in text]
    result = []
    cho = jung = jong = None

    def _flush():
        nonlocal cho, jung, jong
        if cho and jung:
            result.append(_compose_hangul(cho, jung, jong or ''))
        elif cho:
            result.append(cho)
        elif jung:
            result.append(jung)
        cho = jung = jong = None

    for j in jamos:
        if j in _HVOWEL:
            # 모음
            if cho is None:
                _flush(); result.append(j)
            elif jung is None:
                jung = j
            else:
                if jong is None:
                    if (jung, j) in _COMPLEX_JUNG:
                        jung = _COMPLEX_JUNG[(jung, j)]
                    else:
                        _flush(); result.append(j)
                else:
                    # 종성을 다음 음절 초성으로 분리
                    if jong in _SPLIT_JONG:
                        first, second = _SPLIT_JONG[jong]
                        result.append(_compose_hangul(cho, jung, first))
                        cho, jung, jong = second, j, None
                    else:
                        prev_jong = jong
                        result.append(_compose_hangul(cho, jung, ''))
                        cho, jung, jong = prev_jong, j, None
        elif j in _HCHO or j in _HJONG[1:]:
            # 자음
            if cho is None:
                if j in _HCHO: cho = j
                else: _flush(); result.append(j)
            elif jung is None:
                _flush()
                if j in _HCHO: cho = j
                else: result.append(j)
            else:
                if jong is None:
                    if j in _HJONG[1:]: jong = j
                    else:
                        _flush()
                        cho = j if j in _HCHO else None
                        if cho is None: result.append(j)
                else:
                    if (jong, j) in _COMPLEX_JONG:
                        jong = _COMPLEX_JONG[(jong, j)]
                    else:
                        _flush()
                        cho = j if j in _HCHO else None
                        if cho is None: result.append(j)
        else:
            # 자모 아닌 문자 (공백/숫자/기호 등)
            _flush(); result.append(j)
    _flush()
    return ''.join(result)


# =============================================
#  IME (한/영) 상태 감지 및 전환 유틸
# =============================================
def _get_ime_state() -> str:
    """현재 키보드 입력 언어 반환 ('korean' or 'english')"""
    try:
        import ctypes
        lang = ctypes.windll.user32.GetKeyboardLayout(0) & 0xFFFF
        return "korean" if lang == 0x0412 else "english"
    except Exception:
        return "english"


def _set_ime_state(target: str, retries: int = 3):
    """
    target: 'korean' or 'english'
    GetKeyboardLayout 은 포커스/환경에 따라 부정확할 수 있어
    전환 후 재확인하며 최대 retries 회까지 보정한다.
    """
    for _ in range(retries):
        current = _get_ime_state()
        if current == target:
            return
        pyautogui.press("hangul")
        time.sleep(0.25)   # IME 전환 완료까지 충분히 대기(기존 0.15초는 환경에 따라 부족)
    # retries 다 써도 안 맞으면 마지막 상태로 그냥 진행 (감지 자체가 부정확한 환경일 수 있음)


_last_ime_target = None  # (호환성 유지용 더미 - 더 이상 사용 안 함)

def _type_safe(text: str, ime_state: str = "english"):
    """
    한/영 상태와 무관하게 텍스트 입력.

    [중요] 이전 버전은 _set_ime_state 로 OS의 한/영 상태를 맞추려 했지만,
    - GetKeyboardLayout 은 워커 스레드 컨텍스트에서 부정확
    - pyautogui.press("hangul") 은 일부 PC에서 매핑이 달라 작동 안함
    - 한/영 키가 토글 안 되면 캐시와 실제 상태가 어긋나 글자가 반대 언어로 찍힘

    클립보드 붙여넣기(Ctrl+V) 자체는 OS의 IME 상태와 완전히 무관하게
    저장된 텍스트를 그대로 삽입한다. 따라서 IME 전환 로직 자체가 불필요하고,
    오히려 한/영 뒤섞임의 원인이었음 → 전환 로직 전부 제거.

    ime_state 인자는 하위 호환성을 위해 받지만 무시한다.
    """
    if not text:
        return
    _type_with_clipboard(text)


# =============================================
#  스마트 클릭 (자동 이미지 캡처 + 좌표 폴백)
# =============================================
def _capture_snapshot(x: int, y: int, size: int = 80) -> str:
    """
    클릭 위치 주변 size x size 픽셀을 캡처해 PNG 로 저장.
    반환: 저장된 파일 경로 (실패 시 빈 문자열)

    pyautogui.screenshot(region=) 사용. region 은 (left, top, w, h).
    화면 경계를 벗어나지 않도록 보정.
    """
    try:
        import uuid as _uuid
        half = size // 2
        left = max(0, x - half)
        top  = max(0, y - half)
        # 화면 크기 안에 들어오도록 제한
        try:
            sw, sh = pyautogui.size()
            if left + size > sw: left = max(0, sw - size)
            if top  + size > sh: top  = max(0, sh - size)
        except Exception:
            pass
        path = SNAPSHOT_DIR / f"snap_{int(time.time()*1000)}_{_uuid.uuid4().hex[:6]}.png"
        img = pyautogui.screenshot(region=(left, top, size, size))
        img.save(str(path))
        return str(path)
    except Exception:
        return ""


def _find_snapshot_on_screen(snapshot_path: str, confidence: float = 0.85):
    """
    저장된 스냅샷 이미지를 현재 화면에서 찾는다.
    반환: (center_x, center_y) 튜플 또는 None
    confidence 는 OpenCV 가 있을 때만 효과있음 (pyautogui 가 내부적으로 사용).
    """
    if not snapshot_path or not Path(snapshot_path).exists():
        return None
    try:
        # CV2 가 있으면 confidence 사용, 없으면 픽셀완전일치만
        if CV2_OK:
            loc = pyautogui.locateOnScreen(snapshot_path, confidence=confidence)
        else:
            loc = pyautogui.locateOnScreen(snapshot_path)
        if loc is None:
            return None
        # loc 은 Box(left, top, width, height) - 중앙 좌표 계산
        return (int(loc.left + loc.width/2), int(loc.top + loc.height/2))
    except Exception:
        return None


# =============================================
#  화면보호기 안전 대기
#  ──────────────────────────────────────
#  회사 보안정책상 5분 무입력 시 화면이 잠기는 환경 대응.
#  5분 이상 대기 시 4분 50초(290초) 단위로 쪼개고,
#  각 구간 사이에 마우스를 1px 움직여 실제 입력을 발생시킴.
#  (화면보호기 정책을 끄는 것이 아니라 자연스러운 입력으로 타이머 리셋)
# =============================================
_SAFE_WAIT_CHUNK = 290.0   # 4분 50초

def _safe_wait(seconds: float, running_check=None):
    """
    화면보호기에 안전한 대기.
    seconds: 총 대기 시간
    running_check: 호출 시 False 반환하면 대기 중단 (워커의 self.running 등)
    """
    remaining = float(seconds)
    while remaining > 0:
        if running_check and not running_check():
            return
        chunk = min(remaining, _SAFE_WAIT_CHUNK)
        # 1초 단위로 쪼개 중단 요청에 빠르게 반응
        end_t = time.time() + chunk
        while time.time() < end_t:
            if running_check and not running_check():
                return
            time.sleep(min(1.0, end_t - time.time()))
        remaining -= chunk
        if remaining > 0:
            # 다음 구간 전에 마우스 1px 이동 → 입력 발생 → 화면보호기 타이머 리셋
            try:
                cx, cy = pyautogui.position()
                pyautogui.moveTo(cx + 1, cy)
                pyautogui.moveTo(cx, cy)
            except Exception:
                pass


class _GlobalEscWatcher:
    """
    재생 중 글로벌 ESC 감지.
    Qt QShortcut 은 앱 창이 포커스를 가질 때만 동작하므로,
    창이 최소화된 상태(재생 중 기본)에서는 ESC 가 안 먹힌다.
    → pynput 글로벌 키보드 리스너로 OS 레벨에서 ESC 를 잡아
       어떤 상황에서도 재생을 중단할 수 있게 한다.

    사용법:
        watcher = _GlobalEscWatcher(on_esc=worker.stop)
        watcher.start()
        ... 재생 ...
        watcher.stop()
    """
    def __init__(self, on_esc):
        self._on_esc = on_esc
        self._listener = None

    def start(self):
        if not PYNPUT_OK:
            return
        try:
            from pynput import keyboard as _kb
            def _on_press(key):
                try:
                    if key == _kb.Key.esc:
                        try:
                            self._on_esc()
                        except Exception:
                            pass
                        return False  # 리스너 종료
                except Exception:
                    pass
                return True
            self._listener = _kb.Listener(on_press=_on_press)
            self._listener.daemon = True
            self._listener.start()
        except Exception:
            self._listener = None

    def stop(self):
        try:
            if self._listener:
                self._listener.stop()
                self._listener = None
        except Exception:
            pass


def _smart_click(x: int, y: int, snapshot_path: str = "",
                 button: str = "left", confidence: float = 0.85):
    """
    스마트 클릭: 스냅샷이 있으면 이미지 매칭 우선, 실패 시 좌표 폴백.
    좌표는 논리 픽셀 기준으로 받아서 현재 PC DPI 에 맞게 물리 픽셀로 변환.
    """
    target_x, target_y = x, y
    if snapshot_path:
        found = _find_snapshot_on_screen(snapshot_path, confidence)
        if found:
            target_x, target_y = found
            # 이미지 매칭 결과는 이미 물리 픽셀 → 변환 불필요
            pyautogui.click(target_x, target_y, button=button)
            return
    # 논리 → 물리 픽셀 변환 후 클릭
    px, py = _logical_to_phys(target_x, target_y)
    pyautogui.click(px, py, button=button)


def _minimize_all_windows(except_titles=None):
    """
    재생 직전 다른 창들을 최소화한다 (좌표 클릭이 엉뚱한 창에 가는 것 방지).
    except_titles: 최소화에서 제외할 창 제목 키워드 리스트.
    5MRPA 자체 창은 제외 (자기 자신을 최소화하면 카운트다운/상태창이 안 보임).
    """
    if not GW_OK:
        # 최후의 수단: Win+D 로 바탕화면 보기 (모든 창 최소화)
        try:
            pyautogui.hotkey('win', 'd')
            time.sleep(0.3)
        except Exception:
            pass
        return

    # 자기 자신(RPA 앱) 창은 제외
    SELF_KEYWORDS = ["5MRPA", "RPA", "Python"]
    except_titles = except_titles or []

    try:
        wins = gw.getAllWindows()
        for w in wins:
            try:
                title = (w.title or "").strip()
                if not title:
                    continue
                # 자기 자신 또는 사용자 지정 제외 창은 건너뜀
                if any(kw in title for kw in SELF_KEYWORDS):
                    continue
                if any(kw in title for kw in except_titles):
                    continue
                # 이미 최소화된 창은 건너뜀
                if hasattr(w, 'isMinimized') and w.isMinimized:
                    continue
                # 보이는 창만 최소화
                if hasattr(w, 'visible') and not w.visible:
                    continue
                try:
                    w.minimize()
                except Exception:
                    pass
            except Exception:
                continue
        time.sleep(0.5)  # 최소화 애니메이션 완료 대기
    except Exception:
        # 전체 실패 시 Win+D 폴백
        try:
            pyautogui.hotkey('win', 'd')
            time.sleep(0.3)
        except Exception:
            pass



class ExcelFileDialog(QDialog):
    """
    폴더 탐색 가능한 엑셀 파일 선택 다이얼로그.
    COM 없이 순수 PyQt6 구현.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.selected   = ""
        self._cur_dir   = Path.home()
        self.setWindowTitle("엑셀 파일 선택")
        self.setMinimumSize(620, 480)
        self.setStyleSheet(f"QDialog{{background:{C['bg1']};border:1.5px solid {C['border']};border-radius:12px;}}")
        self._build()
        self._refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(18,16,18,16); lay.setSpacing(10)

        # 경로 표시 + 상위 폴더 버튼
        nav = QHBoxLayout(); nav.setSpacing(8)
        btn_up = GlowButton("↑ 상위", C['brand'])
        btn_up.setFixedHeight(32); btn_up.setFixedWidth(70)
        btn_up.clicked.connect(self._go_up)
        self.path_lbl = QLabel("")
        self.path_lbl.setStyleSheet(
            f"color:{C['t2']};font-size:11px;background:transparent;")
        nav.addWidget(btn_up); nav.addWidget(self.path_lbl,1)
        lay.addLayout(nav)

        # 빠른 이동 버튼
        quick = QHBoxLayout(); quick.setSpacing(6)
        for label, path in [
            ("바탕화면", Path.home()/"Desktop"),
            ("문서",     Path.home()/"Documents"),
            ("C:\\",     Path("C:\\")),
            ("D:\\",     Path("D:\\")),
            ("C:\\secu", Path("C:\\secu")),
        ]:
            if path.exists():
                b = GlowButton(label, C['t2'])
                b.setFixedHeight(28)
                b.clicked.connect(lambda _, p=path: self._goto(p))
                quick.addWidget(b)
        quick.addStretch()
        lay.addLayout(quick)

        # 파일/폴더 목록
        self.file_list = QListWidget()
        self.file_list.setStyleSheet(f"""
            QListWidget{{background:{C['bg2']};border:1.5px solid {C['border']};
                border-radius:8px;color:{C['t1']};font-size:13px;
                padding:4px;outline:none;}}
            QListWidget::item{{padding:8px 12px;border-radius:6px;margin:1px;}}
            QListWidget::item:selected{{background:{C['brand']};color:#FFFFFF;}}
            QListWidget::item:hover{{background:{C['bg4']};}}
        """)
        self.file_list.itemDoubleClicked.connect(self._on_double)
        self.file_list.currentItemChanged.connect(self._on_select)
        lay.addWidget(self.file_list)

        # 선택 경로 표시
        self.sel_edit = StyledInput("선택된 파일 경로")
        self.sel_edit.setReadOnly(True)
        lay.addWidget(self.sel_edit)

        # 버튼
        btn_row = QHBoxLayout(); btn_row.setSpacing(8)
        btn_row.addStretch()
        bc = GlowButton("취소", C['rec'])
        bc.setFixedHeight(38); bc.setFixedWidth(80)
        bc.clicked.connect(self.reject)
        bo = GlowButton("선택", C['play'])
        bo.setFixedHeight(38); bo.setFixedWidth(80)
        bo.clicked.connect(self._accept)
        btn_row.addWidget(bc); btn_row.addWidget(bo)
        lay.addLayout(btn_row)

    def _refresh(self):
        self.file_list.clear()
        self.path_lbl.setText(str(self._cur_dir))
        try:
            items = sorted(self._cur_dir.iterdir(),
                           key=lambda p: (not p.is_dir(), p.name.lower()))
            for p in items:
                if p.is_dir():
                    item = QListWidgetItem(f"  [폴더]  {p.name}")
                    item.setForeground(QColor(C['stop']))
                elif p.suffix.lower() in ('.xlsx', '.xls', '.xlsm'):
                    item = QListWidgetItem(f"  [엑셀]  {p.name}")
                    item.setForeground(QColor(C['play']))
                else:
                    continue
                item.setData(Qt.ItemDataRole.UserRole, str(p))
                self.file_list.addItem(item)
        except Exception as e:
            self.file_list.addItem(QListWidgetItem(f"  오류: {e}"))

    def _goto(self, path: Path):
        if path.exists():
            self._cur_dir = path
            self._refresh()

    def _go_up(self):
        parent = self._cur_dir.parent
        if parent != self._cur_dir:
            self._cur_dir = parent
            self._refresh()

    def _on_select(self, item):
        if not item: return
        p = item.data(Qt.ItemDataRole.UserRole)
        if p and Path(p).is_file():
            self.sel_edit.setText(p)

    def _on_double(self, item):
        p = item.data(Qt.ItemDataRole.UserRole)
        if not p: return
        path = Path(p)
        if path.is_dir():
            self._cur_dir = path
            self._refresh()
        elif path.suffix.lower() in ('.xlsx', '.xls', '.xlsm'):
            self.sel_edit.setText(str(path))
            self._accept()

    def _accept(self):
        p = self.sel_edit.text().strip()
        if not p or not Path(p).exists():
            msg_warn(self, "선택 오류", "엑셀 파일을 선택하세요.")
            return
        self.selected = p
        self.accept()


# =============================================
#  PathDialog (QFileDialog 대체 - COM 블로킹 방지)
# =============================================
class PathDialog(QDialog):
    def __init__(self, mode="open", title="파일 선택",
                 start_dir=None, ext=".json", parent=None):
        super().__init__(parent)
        self._mode      = mode
        self._ext       = ext
        self._start_dir = Path(start_dir) if start_dir else SAVE_DIR
        self.selected   = ""
        self.setWindowTitle(title)
        self.setMinimumWidth(560)
        self.setMinimumHeight(400)
        self.setStyleSheet(f"QDialog {{ background:{C['bg1']}; border:1.5px solid {C['border']}; border-radius:12px; }}")
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20,18,20,18)
        lay.setSpacing(12)
        dir_lbl = QLabel(f"저장 폴더:  {self._start_dir}")
        dir_lbl.setStyleSheet(f"color:{C['t2']}; font-size:11px;")
        lay.addWidget(dir_lbl)
        if self._mode == "open":
            lay.addWidget(QLabel("파일 목록 (클릭해서 선택)"))
            self.file_list = QListWidget()
            self.file_list.setStyleSheet(f"""
                QListWidget {{ background:{C['bg2']}; border:1.5px solid {C['border']};
                    border-radius:8px; color:{C['t1']}; font-size:13px; padding:4px; outline:none; }}
                QListWidget::item {{ padding:8px 12px; border-radius:6px; margin:1px; }}
                QListWidget::item:selected {{ background:{C['brand']}; color:#FFFFFF; }}
                QListWidget::item:hover {{ background:{C['bg4']}; }}
            """)
            self.file_list.itemDoubleClicked.connect(self._on_double)
            self.file_list.currentItemChanged.connect(self._on_select)
            self._refresh_list()
            lay.addWidget(self.file_list)
        lbl_txt = "직접 경로 입력:" if self._mode == "open" else "저장할 파일명:"
        lay.addWidget(QLabel(lbl_txt))
        self.path_edit = StyledInput(str(self._start_dir / f"파일명{self._ext}"))
        if self._mode == "save":
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            self.path_edit.setText(str(self._start_dir / f"macro_{ts}{self._ext}"))
        lay.addWidget(self.path_edit)
        btn_row = QHBoxLayout()
        if self._mode == "open":
            br = GlowButton("새로고침", C['t2'])
            br.setFixedHeight(38); br.clicked.connect(self._refresh_list)
            btn_row.addWidget(br)
        btn_row.addStretch()
        bc = GlowButton("취소", C['rec']); bc.setFixedHeight(38); bc.setFixedWidth(80)
        bc.clicked.connect(self.reject)
        bo = GlowButton("열기" if self._mode=="open" else "저장", C['play'])
        bo.setFixedHeight(38); bo.setFixedWidth(80); bo.clicked.connect(self._accept)
        btn_row.addWidget(bc); btn_row.addWidget(bo)
        lay.addLayout(btn_row)

    def _refresh_list(self):
        self.file_list.clear()
        try:
            files = sorted(self._start_dir.glob(f"*{self._ext}"),
                           key=lambda p: p.stat().st_mtime, reverse=True)
            for fp in files:
                ts   = datetime.datetime.fromtimestamp(fp.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
                item = QListWidgetItem(f"  {fp.name}   {ts}")
                item.setData(Qt.ItemDataRole.UserRole, str(fp))
                self.file_list.addItem(item)
            if not files:
                self.file_list.addItem(QListWidgetItem("  저장된 파일이 없습니다"))
        except Exception as e:
            self.file_list.addItem(QListWidgetItem(f"  오류: {e}"))

    def _on_select(self, item):
        if item:
            fp = item.data(Qt.ItemDataRole.UserRole)
            if fp: self.path_edit.setText(fp)

    def _on_double(self, item):
        fp = item.data(Qt.ItemDataRole.UserRole)
        if fp: self.path_edit.setText(fp); self._accept()

    def _accept(self):
        p = self.path_edit.text().strip()
        if not p: return

        if self._mode == "open":
            if not Path(p).exists():
                msg_warn(self, "파일 없음", f"파일을 찾을 수 없습니다: {p}")
                return
        else:
            # save 모드: 파일명에서 특수문자/경로구분자 제거
            p_obj = Path(p)

            # 파일명만 입력한 경우 (경로 없음) → SAVE_DIR 에 저장
            if not p_obj.is_absolute():
                # \ / : * ? " < > | 제거
                import re
                safe_name = re.sub(r'[\\/:*?"<>|]', '_', p)
                p = str(self._start_dir / safe_name)
            else:
                # 절대경로인 경우 파일명만 정제
                import re
                safe_stem = re.sub(r'[\\/:*?"<>|]', '_', p_obj.stem)
                p = str(p_obj.parent / (safe_stem + p_obj.suffix))

            # 확장자 자동 보정
            if not p.endswith(self._ext):
                p += self._ext

            # 부모 디렉토리 존재 확인
            parent_dir = Path(p).parent
            if not parent_dir.exists():
                try:
                    parent_dir.mkdir(parents=True, exist_ok=True)
                except Exception:
                    # 부모 디렉토리 생성 실패 시 SAVE_DIR 로 폴백
                    p = str(self._start_dir / Path(p).name)

            self.path_edit.setText(p)

        self.selected = p
        self.accept()


def rpa_open_file(parent, title="파일 열기", start_dir=None, ext=".json"):
    dlg = PathDialog("open", title, start_dir, ext, parent)
    if dlg.exec() == QDialog.DialogCode.Accepted:
        return dlg.selected, ext
    return "", ""


def rpa_open_zip(parent, title="ZIP 파일 열기", start_dir=None):
    """ZIP 파일 전용 열기 대화상자"""
    dlg = PathDialog("open", title, start_dir, ".zip", parent)
    if dlg.exec() == QDialog.DialogCode.Accepted:
        return dlg.selected, ".zip"
    return "", ""


def rpa_save_file(parent, title="파일 저장", start_dir=None, ext=".json"):
    dlg = PathDialog("save", title, start_dir, ext, parent)
    if dlg.exec() == QDialog.DialogCode.Accepted:
        return dlg.selected, ext
    return "", ""



# =============================================
#  플로팅 카운트다운 캡처 창
#  화면 우측 상단에 떠서 카운트다운 후 마우스 좌표 캡처
# =============================================

class _CountdownCapture(QWidget):
    """
    화면 우측 상단에 반투명 플로팅 창을 띄우고
    카운트다운 후 마우스 위치를 캡처합니다.
    창이 작아서 마우스 이동에 방해되지 않습니다.
    """
    def __init__(self, seconds: int = 3,
                 on_done=None, on_cancel=None, parent=None):
        super().__init__(parent)
        self._seconds   = seconds
        self._remaining = seconds
        self._on_done   = on_done
        self._on_cancel = on_cancel

        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint |
            Qt.WindowType.Tool)
        self.setFixedSize(280, 110)

        # 화면 우측 상단 배치
        screen = QApplication.primaryScreen().availableGeometry()
        self.move(screen.right() - 300, screen.top() + 20)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 12, 16, 12)
        lay.setSpacing(8)

        # 안내 텍스트
        guide = QLabel("원하는 위치로 마우스를 이동하세요")
        guide.setAlignment(Qt.AlignmentFlag.AlignCenter)
        guide.setStyleSheet(f"""
            color: {C['t1']}; font-size: 12px; font-weight: bold;
            background: transparent;
        """)
        lay.addWidget(guide)

        # 카운트다운 숫자
        self.count_lbl = QLabel(str(seconds))
        self.count_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.count_lbl.setStyleSheet(f"""
            color: {C['rec']}; font-size: 36px; font-weight: 900;
            background: transparent;
        """)
        lay.addWidget(self.count_lbl)

        # 취소 버튼
        btn_cancel = QPushButton("취소 (ESC)")
        btn_cancel.setFixedHeight(26)
        btn_cancel.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_cancel.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {C['t3']};
                border: 1px solid {C['border']}; border-radius: 5px;
                font-size: 11px;
            }}
            QPushButton:hover {{ background: {C['bg4']}; }}
        """)
        btn_cancel.clicked.connect(self._cancel)
        lay.addWidget(btn_cancel)

        # 배경 스타일
        self.setStyleSheet(f"""
            QWidget {{
                background: {C['bg1']};
                border: 2px solid {C['rec']};
                border-radius: 12px;
            }}
        """)

        # 타이머
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(1000)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Escape:
            self._cancel()

    def _tick(self):
        self._remaining -= 1
        if self._remaining <= 0:
            self._timer.stop()
            self.count_lbl.setText("캡처!")
            self.count_lbl.setStyleSheet(f"""
                color: {C['play']}; font-size: 32px; font-weight: 900;
                background: transparent;
            """)
            # 100ms 후 캡처 (렌더링 완료 대기)
            QTimer.singleShot(150, self._capture)
        else:
            self.count_lbl.setText(str(self._remaining))
            # 색상 변화: 3=빨강 2=주황 1=초록
            colors = {3: C['rec'], 2: C['stop'], 1: C['play']}
            c = colors.get(self._remaining, C['rec'])
            self.count_lbl.setStyleSheet(f"""
                color: {c}; font-size: 36px; font-weight: 900;
                background: transparent;
            """)
            self.setStyleSheet(f"""
                QWidget {{
                    background: {C['bg1']};
                    border: 2px solid {c};
                    border-radius: 12px;
                }}
            """)

    def _capture(self):
        x, y = pyautogui.position()
        self.close()
        if self._on_done:
            self._on_done(x, y)

    def _cancel(self):
        self._timer.stop()
        self.close()
        if self._on_cancel:
            self._on_cancel()


# =============================================
#  블록 편집 다이얼로그
# =============================================
class BlockEditDialog(QDialog):
    """블록 클릭 시 열리는 편집 팝업 - 코드 없이 GUI 로 수정"""

    BLOCK_TYPES = ["click","key","type","hotkey","wait","image_click","scroll","move","cond_if","excel_read"]

    def __init__(self, action: dict, parent=None):
        super().__init__(parent)
        self.action = dict(action)
        self.setWindowTitle("블록 편집")
        self.setMinimumWidth(480)
        self.setStyleSheet(f"""
            QDialog {{ background:{C['bg1']}; }}
            QLabel {{ background:transparent; color:{C['t1']}; }}
        """)
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20,18,20,18)
        lay.setSpacing(12)

        # 동작 유형 선택
        type_row = QHBoxLayout()
        type_row.addWidget(QLabel("동작 유형:"))
        self.type_combo = StyledCombo()
        LABELS = {
            "click":       "클릭  (마우스 클릭)",
            "key":         "키 입력  (키보드)",
            "type":        "텍스트 입력  (타이핑)",
            "hotkey":      "단축키  (Ctrl+S 등)",
            "wait":        "대기  (N초 기다리기)",
            "image_click": "이미지 클릭  (화면에서 찾아 클릭)",
            "scroll":      "스크롤  (마우스 휠)",
            "move":        "마우스 이동",
            "cond_if":     "조건 분기  (이미지/좌표 조건으로 YES/NO 분기)",
            "excel_read":  "엑셀 읽기  (엑셀 A열에서 값을 읽어 반복 입력)",
        }
        for k, v in LABELS.items():
            self.type_combo.addItem(v, k)
        cur = self.action.get("type","click")
        idx = list(LABELS.keys()).index(cur) if cur in LABELS else 0
        self.type_combo.setCurrentIndex(idx)
        self.type_combo.currentIndexChanged.connect(self._refresh_fields)
        type_row.addWidget(self.type_combo)
        lay.addLayout(type_row)

        # 동적 필드 영역
        self.fields_frame = QFrame()
        self.fields_frame.setStyleSheet(f"""
            QFrame {{ background:{C['bg2']}; border:1.5px solid {C['border']};
                     border-radius:10px; padding:4px; }}
        """)
        self.fields_lay = QGridLayout(self.fields_frame)
        self.fields_lay.setContentsMargins(14,12,14,12)
        self.fields_lay.setSpacing(10)
        lay.addWidget(self.fields_frame)

        # 공통 딜레이
        delay_row = QHBoxLayout()
        delay_row.addWidget(QLabel("이전 동작 후 대기:"))
        self.delay_spin = StyledDSpin()
        self.delay_spin.setRange(0, 60)
        self.delay_spin.setValue(self.action.get("delay", 0.5))
        self.delay_spin.setSuffix("  초")
        self.delay_spin.setFixedWidth(130)
        delay_row.addWidget(self.delay_spin)
        delay_row.addStretch()
        lay.addLayout(delay_row)

        # AI 설명 표시
        self.ai_lbl = QLabel("")
        self.ai_lbl.setStyleSheet(f"""
            color:{C['ai']}; font-size:11px;
            background:{C['ai_dim']}; border:1px solid {C['ai']};
            border-radius:6px; padding:6px 10px;
        """)
        self.ai_lbl.setWordWrap(True)
        self.ai_lbl.setVisible(False)
        lay.addWidget(self.ai_lbl)

        # 버튼
        btn_row = QHBoxLayout()
        btn_ai = GlowButton("AI 설명 생성", C['ai'])
        btn_ai.setFixedHeight(36); btn_ai.clicked.connect(self._gen_ai_desc)

        self.btn_preview = GlowButton("위치 미리보기 (누르는 동안)", C['brand'])
        self.btn_preview.setFixedHeight(36)
        self.btn_preview.setToolTip("버튼을 누르는 동안 실제 화면에서 위치 표시")
        # 좌표가 없는 타입은 비활성
        t0 = self.action.get("type","")
        self.btn_preview.setEnabled(t0 in ("click","scroll","move","image_click"))
        self.type_combo.currentIndexChanged.connect(self._update_preview_btn)
        self.btn_preview.pressed.connect(self._preview_location)
        self.btn_preview.released.connect(hide_realtime_marker)

        btn_cancel = GlowButton("취소", C['rec'])
        btn_cancel.setFixedHeight(36); btn_cancel.setFixedWidth(80)
        btn_cancel.clicked.connect(self.reject)
        btn_ok = GlowButton("확인", C['play'])
        btn_ok.setFixedHeight(36); btn_ok.setFixedWidth(80)
        btn_ok.clicked.connect(self._accept)
        btn_row.addWidget(btn_ai)
        btn_row.addWidget(self.btn_preview)
        btn_row.addStretch()
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

        self._refresh_fields()

    def _clear_fields(self):
        while self.fields_lay.count():
            item = self.fields_lay.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        self._field_widgets = {}

    def _add_field(self, row, label, widget):
        lbl = QLabel(label)
        lbl.setStyleSheet(f"color:{C['t2']}; background:transparent;")
        self.fields_lay.addWidget(lbl, row, 0)
        self.fields_lay.addWidget(widget, row, 1)
        self._field_widgets[label] = widget

    def _refresh_fields(self):
        self._clear_fields()
        t = self.type_combo.currentData()

        if t == "click":
            self.w_x = StyledSpin(); self.w_x.setRange(0,9999)
            self.w_x.setValue(self.action.get("x",500))
            self.w_y = StyledSpin(); self.w_y.setRange(0,9999)
            self.w_y.setValue(self.action.get("y",400))
            self.w_btn = StyledCombo()
            self.w_btn.addItems(["left (왼쪽)", "right (오른쪽)", "double (더블)"])
            btn_map = {"left":0,"right":1,"double":2}
            self.w_btn.setCurrentIndex(btn_map.get(self.action.get("button","left"),0))
            self._add_field(0, "X 좌표:", self.w_x)
            self._add_field(1, "Y 좌표:", self.w_y)
            self._add_field(2, "클릭 종류:", self.w_btn)
            # 현재 위치 버튼
            btn_pos = GlowButton("현재 마우스 위치 가져오기 (3초)", C['brand'])
            btn_pos.setFixedHeight(34)
            btn_pos.clicked.connect(self._get_pos)
            self.fields_lay.addWidget(btn_pos, 3, 0, 1, 2)
            # 힌트
            hint = QLabel("클릭 후 3초 안에 원하는 아이콘/버튼 위에 마우스를 올려두세요")
            hint.setStyleSheet(f"color:{C['stop']}; font-size:11px; font-weight:bold; background:transparent;")
            self.fields_lay.addWidget(hint, 4, 0, 1, 2)

        elif t == "key":
            self.w_key = StyledInput("예: enter, tab, esc, ctrl, f5, delete")
            self.w_key.setText(self.action.get("key",""))
            self._add_field(0, "키 이름:", self.w_key)
            common_keys = QLabel(
                "자주 쓰는 키: enter / tab / esc / ctrl / alt / shift\n"
                "           f1~f12 / delete / backspace / up / down / left / right")
            common_keys.setStyleSheet(f"color:{C['t3']}; font-size:11px; background:transparent;")
            self.fields_lay.addWidget(common_keys, 1, 0, 1, 2)

        elif t == "type":
            self.w_text = QTextEdit()
            self.w_text.setFixedHeight(80)
            self.w_text.setPlainText(self.action.get("text",""))
            self.w_text.setStyleSheet(f"""
                QTextEdit {{ background:{C['bg1']}; border:1.5px solid {C['border']};
                    border-radius:8px; color:{C['t1']}; padding:6px; }}
                QTextEdit:focus {{ border-color:{C['brand']}; }}
            """)
            self._add_field(0, "입력할 텍스트:", self.w_text)

            # 두벌식 변환 버튼 - 영문으로 저장된 한글 의도 텍스트를 바로 변환
            btn_row2 = QHBoxLayout()
            btn_kor = GlowButton("두벌식 → 한글 변환", C['tool'])
            btn_kor.setFixedHeight(30)
            btn_kor.setToolTip(
                "영문으로 잘못 녹화된 텍스트를 한글로 변환합니다.\n"
                "예) spdlqj → 네이버,  audchsaktwlq → 명촌맛집")
            def _do_kor_convert():
                raw = self.w_text.toPlainText()
                if not raw: return
                converted = _eng_to_kor(raw)
                if converted != raw:
                    self.w_text.setPlainText(converted)
            btn_kor.clicked.connect(_do_kor_convert)
            btn_row2.addWidget(btn_kor)
            btn_row2.addStretch()
            self.fields_lay.addLayout(btn_row2, 1, 0, 1, 2)

            hint = QLabel("한글이 영문으로 녹화됐다면 '두벌식 → 한글 변환' 버튼 클릭")
            hint.setStyleSheet(f"color:{C['t3']}; font-size:11px; background:transparent;")
            self.fields_lay.addWidget(hint, 2, 0, 1, 2)

        elif t == "hotkey":
            keys = self.action.get("keys", ["ctrl","s"])
            self.w_key1 = StyledInput("예: ctrl")
            self.w_key1.setText(keys[0] if len(keys)>0 else "ctrl")
            self.w_key2 = StyledInput("예: s")
            self.w_key2.setText(keys[1] if len(keys)>1 else "")
            self.w_key3 = StyledInput("예: shift (선택)")
            self.w_key3.setText(keys[2] if len(keys)>2 else "")
            self._add_field(0, "키 1:", self.w_key1)
            self._add_field(1, "키 2:", self.w_key2)
            self._add_field(2, "키 3 (선택):", self.w_key3)
            examples = QLabel("예시: ctrl+s(저장), ctrl+z(취소), ctrl+c(복사), alt+f4(닫기)")
            examples.setStyleSheet(f"color:{C['t3']}; font-size:11px; background:transparent;")
            self.fields_lay.addWidget(examples, 3, 0, 1, 2)

        elif t == "wait":
            self.w_sec = StyledDSpin()
            self.w_sec.setRange(0.1, 60)
            self.w_sec.setValue(self.action.get("seconds", 2.0))
            self.w_sec.setSuffix("  초")
            self._add_field(0, "대기 시간:", self.w_sec)
            hint = QLabel("프로그램 로딩, 화면 전환 등 기다려야 할 때 사용")
            hint.setStyleSheet(f"color:{C['t3']}; font-size:11px; background:transparent;")
            self.fields_lay.addWidget(hint, 1, 0, 1, 2)

        elif t == "image_click":
            self.w_label = StyledInput("예: 저장버튼, 확인팝업")
            self.w_label.setText(self.action.get("label",""))
            self.w_path  = StyledInput(str(SAVE_DIR / "이미지파일.png"))
            self.w_path.setText(self.action.get("image_path",""))
            self.w_conf  = StyledDSpin()
            self.w_conf.setRange(0.5, 1.0); self.w_conf.setValue(self.action.get("confidence",0.85))
            self.w_conf.setSingleStep(0.05)
            # 대기 시간 옵션
            self.w_wait_img = StyledDSpin()
            self.w_wait_img.setRange(0, 60)
            self.w_wait_img.setValue(self.action.get("wait_timeout", 0))
            self.w_wait_img.setSuffix("  초  (0 = 대기 없음)")
            self.w_wait_img.setToolTip(
                "이미지가 나타날 때까지 최대 N초 대기\n"
                "0이면 즉시 찾고 없으면 넘어감\n"
                "로딩이 필요한 버튼은 3~10초 설정 권장")
            self._add_field(0, "이미지 이름:", self.w_label)
            self._add_field(1, "이미지 파일 경로:", self.w_path)
            self._add_field(2, "인식 정확도:", self.w_conf)
            self._add_field(3, "나타날 때까지 대기:", self.w_wait_img)
            btn_browse = GlowButton("이미지 파일 선택", C['brand'])
            btn_browse.setFixedHeight(34)
            btn_browse.clicked.connect(self._browse_img)
            self.fields_lay.addWidget(btn_browse, 4, 0, 1, 2)

        elif t == "scroll":
            self.w_x = StyledSpin(); self.w_x.setRange(0,9999)
            self.w_x.setValue(self.action.get("x",500))
            self.w_y = StyledSpin(); self.w_y.setRange(0,9999)
            self.w_y.setValue(self.action.get("y",400))
            self.w_dy = StyledSpin(); self.w_dy.setRange(-20,20)
            self.w_dy.setValue(self.action.get("dy",3))
            self._add_field(0, "X 좌표:", self.w_x)
            self._add_field(1, "Y 좌표:", self.w_y)
            self._add_field(2, "스크롤 양 (양수=위, 음수=아래):", self.w_dy)

        elif t == "move":
            self.w_x = StyledSpin(); self.w_x.setRange(0,9999)
            self.w_x.setValue(self.action.get("x",500))
            self.w_y = StyledSpin(); self.w_y.setRange(0,9999)
            self.w_y.setValue(self.action.get("y",400))
            self._add_field(0, "X 좌표:", self.w_x)
            self._add_field(1, "Y 좌표:", self.w_y)
            btn_pos = GlowButton("현재 마우스 위치 가져오기 (3초)", C['brand'])
            btn_pos.setFixedHeight(34); btn_pos.clicked.connect(self._get_pos)
            self.fields_lay.addWidget(btn_pos, 2, 0, 1, 2)

        elif t == "excel_read":
            # 엑셀 파일 경로
            self.w_xl_path = StyledInput("예: C:\\secu\\input.xlsx")
            self.w_xl_path.setText(self.action.get("excel_path",""))
            self.w_xl_path.setReadOnly(True)
            btn_xl = GlowButton("파일 선택", C['play'])
            btn_xl.setFixedHeight(34)
            btn_xl.clicked.connect(self._pick_excel)
            self._add_field(0, "엑셀 파일:", self.w_xl_path)
            self.fields_lay.addWidget(btn_xl, 1, 0, 1, 2)

            self.w_xl_sheet = StyledInput("Sheet1")
            self.w_xl_sheet.setText(self.action.get("sheet_name","Sheet1"))
            self._add_field(2, "시트 이름:", self.w_xl_sheet)

            self.w_xl_col = StyledInput("A")
            self.w_xl_col.setText(self.action.get("column","A"))
            self.w_xl_col.setFixedWidth(60)
            self._add_field(3, "컬럼:", self.w_xl_col)

            self.w_xl_start = StyledSpin()
            self.w_xl_start.setRange(1, 9999)
            self.w_xl_start.setValue(self.action.get("start_row", 1))
            self._add_field(4, "시작 행:", self.w_xl_start)

            hint = QLabel(
                "엑셀 A열 값을 위에서부터 읽어 플로우를 반복 실행합니다\n"
                "예: A1=P160, A2=P170 → 각각 조회 조건으로 입력")
            hint.setStyleSheet(
                f"color:{C['t3']};font-size:11px;background:transparent;")
            hint.setWordWrap(True)
            self.fields_lay.addWidget(hint, 5, 0, 1, 2)

        elif t == "cond_if":
            # 조건 이미지
            cond_lbl = QLabel("조건 유형:")
            cond_lbl.setStyleSheet(f"color:{C['t2']};background:transparent;")
            self.w_cond_type = StyledCombo()
            self.w_cond_type.addItems([
                "이미지가 화면에 있으면 YES",
                "이미지가 화면에 없으면 YES",
            ])
            cur_ct = self.action.get("cond_type", "image_exist")
            self.w_cond_type.setCurrentIndex(
                0 if cur_ct == "image_exist" else 1)
            self.fields_lay.addWidget(cond_lbl, 0, 0)
            self.fields_lay.addWidget(self.w_cond_type, 0, 1)

            self.w_cond_img = StyledInput("조건 이미지 파일 경로 (.png)")
            self.w_cond_img.setText(self.action.get("cond_img", ""))
            self.w_cond_img.setReadOnly(True)
            btn_img = GlowButton("이미지 선택", C['ai'])
            btn_img.setFixedHeight(34)
            btn_img.clicked.connect(self._pick_cond_img)
            self.fields_lay.addWidget(QLabel("조건 이미지:"), 1, 0)
            self.fields_lay.addWidget(self.w_cond_img, 1, 1)
            self.fields_lay.addWidget(btn_img, 2, 0, 1, 2)

            self.w_cond_conf = StyledDSpin()
            self.w_cond_conf.setRange(0.5, 1.0)
            self.w_cond_conf.setValue(self.action.get("cond_conf", 0.80))
            self.w_cond_conf.setSingleStep(0.05)
            self._add_field(3, "인식 정확도:", self.w_cond_conf)

            # YES/NO 분기 매크로 선택
            self.w_yes_path = StyledInput("YES 일 때 실행할 매크로")
            self.w_yes_path.setText(self.action.get("yes_macro", ""))
            self.w_yes_path.setReadOnly(True)
            btn_yes = GlowButton("YES 매크로", C['play'])
            btn_yes.setFixedHeight(34)
            btn_yes.clicked.connect(lambda: self._pick_branch_macro("yes"))
            self.fields_lay.addWidget(QLabel("YES 매크로:"), 4, 0)
            self.fields_lay.addWidget(self.w_yes_path, 4, 1)
            self.fields_lay.addWidget(btn_yes, 5, 0, 1, 2)

            self.w_no_path = StyledInput("NO 일 때 실행할 매크로 (없으면 다음 프로세스)")
            self.w_no_path.setText(self.action.get("no_macro", ""))
            self.w_no_path.setReadOnly(True)
            btn_no = GlowButton("NO 매크로 (선택)", C['stop'])
            btn_no.setFixedHeight(34)
            btn_no.clicked.connect(lambda: self._pick_branch_macro("no"))
            self.fields_lay.addWidget(QLabel("NO 매크로:"), 6, 0)
            self.fields_lay.addWidget(self.w_no_path, 6, 1)
            self.fields_lay.addWidget(btn_no, 7, 0, 1, 2)

            hint = QLabel(
                "YES: 이미지 발견 시 지정 매크로 실행 후 다음 프로세스로\n"
                "NO : 이미지 없을 때 지정 매크로 실행 (없으면 다음 프로세스)")
            hint.setStyleSheet(
                f"color:{C['t3']};font-size:11px;background:transparent;")
            hint.setWordWrap(True)
            self.fields_lay.addWidget(hint, 8, 0, 1, 2)

    def _get_pos(self):
        """
        플로팅 카운트다운 창을 화면 우측 상단에 띄우고
        3초 후 마우스 위치를 캡처합니다.
        """
        sender_btn = self.sender()
        if sender_btn:
            sender_btn.setEnabled(False)
            sender_btn.setText("3초 후 캡처...")

        # self 에 참조 보관 (GC 방지 - 지역변수면 즉시 소멸)
        self._countdown = _CountdownCapture(
            seconds=3,
            on_done=self._on_pos_captured,
            on_cancel=lambda: self._restore_pos_btn(sender_btn)
        )
        self._countdown.show()
        self._countdown.raise_()

    def _on_pos_captured(self, x, y, sender_btn_ref=None):
        """카운트다운 완료 후 좌표 수신"""
        if hasattr(self, 'w_x'): self.w_x.setValue(x)
        if hasattr(self, 'w_y'): self.w_y.setValue(y)
        # 버튼 복원 - sender 직접 찾기
        for child in self.findChildren(GlowButton):
            if "가져오기" in child.text() or "캡처" in child.text():
                child.setText("현재 마우스 위치 가져오기 (3초)")
                child.setEnabled(True)
                break

    def _restore_pos_btn(self, btn):
        if btn:
            btn.setText("현재 마우스 위치 가져오기 (3초)")
            btn.setEnabled(True)

    def _browse_img(self):
        fp, _ = rpa_open_file(self, "이미지 선택", SAVE_DIR, ".png")
        if fp and hasattr(self, 'w_path'):
            self.w_path.setText(fp)

    def _pick_excel(self):
        dlg = ExcelFileDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.selected:
            if hasattr(self, 'w_xl_path'):
                self.w_xl_path.setText(dlg.selected)

    def _pick_cond_img(self):
        fp, _ = rpa_open_file(self, "조건 이미지 선택", SAVE_DIR, ".png")
        if fp and hasattr(self, 'w_cond_img'):
            self.w_cond_img.setText(fp)

    def _pick_branch_macro(self, branch: str):
        fp, _ = rpa_open_file(self, f"{branch.upper()} 매크로 선택", SAVE_DIR)
        if branch == "yes" and fp and hasattr(self, 'w_yes_path'):
            self.w_yes_path.setText(fp)
        elif branch == "no" and fp and hasattr(self, 'w_no_path'):
            self.w_no_path.setText(fp)

    def _gen_ai_desc(self):
        act = self._build_action()
        self.ai_lbl.setVisible(True)
        self.ai_lbl.setText("AI 생성 중... (5~40초 소요)")
        self.ai_lbl.setStyleSheet(f"""
            color:{C['stop']}; font-size:11px;
            background:{C['stop_dim']}; border:1px solid {C['stop']};
            border-radius:6px; padding:6px 10px;
        """)
        sender = self.sender()
        if sender: sender.setEnabled(False)
        def _w():
            desc = GemmaEngine.describe_block(act)
            def _done():
                self.ai_lbl.setText(f"AI: {desc}")
                self.ai_lbl.setStyleSheet(f"""
                    color:{C['ai']}; font-size:11px;
                    background:{C['ai_dim']}; border:1px solid {C['ai']};
                    border-radius:6px; padding:6px 10px;
                """)
                if sender: sender.setEnabled(True)
            QTimer.singleShot(0, _done)
        threading.Thread(target=_w, daemon=True).start()

    def _update_preview_btn(self):
        t = self.type_combo.currentData()
        self.btn_preview.setEnabled(t in ("click","scroll","move","image_click"))

    def _preview_location(self):
        """현재 편집 중인 값으로 즉시 마커 표시"""
        act = self._build_action()
        show_realtime_marker(act)

    def _build_action(self) -> dict:
        t = self.type_combo.currentData()
        act = {"type": t, "delay": self.delay_spin.value()}
        if self.action.get("label"):
            act["label"] = self.action["label"]
        if self.action.get("ai_desc"):
            act["ai_desc"] = self.action["ai_desc"]

        if t == "click":
            act.update({"x": self.w_x.value(), "y": self.w_y.value(),
                        "button": ["left","right","double"][self.w_btn.currentIndex()]})
        elif t == "key":
            act["key"] = self.w_key.text().strip()
        elif t == "type":
            act["text"] = self.w_text.toPlainText()
        elif t == "hotkey":
            keys = [k for k in [self.w_key1.text().strip(),
                                  self.w_key2.text().strip(),
                                  self.w_key3.text().strip()] if k]
            act["keys"] = keys
        elif t == "wait":
            act["seconds"] = self.w_sec.value()
        elif t == "image_click":
            act.update({"label":       self.w_label.text().strip(),
                        "image_path":  self.w_path.text().strip(),
                        "confidence":  self.w_conf.value()})
        elif t == "scroll":
            act.update({"x": self.w_x.value(), "y": self.w_y.value(),
                        "dy": self.w_dy.value()})
        elif t == "move":
            act.update({"x": self.w_x.value(), "y": self.w_y.value()})
        elif t == "excel_read":
            act.update({
                "excel_path": self.w_xl_path.text().strip(),
                "sheet_name": self.w_xl_sheet.text().strip() or "Sheet1",
                "column":     self.w_xl_col.text().strip().upper() or "A",
                "start_row":  self.w_xl_start.value(),
            })
        elif t == "cond_if":
            cond_types = ["image_exist", "image_not_exist"]
            act.update({
                "cond_type": cond_types[self.w_cond_type.currentIndex()],
                "cond_img":  self.w_cond_img.text().strip(),
                "cond_conf": self.w_cond_conf.value(),
                "yes_macro": self.w_yes_path.text().strip(),
                "no_macro":  self.w_no_path.text().strip(),
            })
        return act

    def _accept(self):
        self.action = self._build_action()
        self.accept()


# =============================================
#  플로우 블록 위젯
# =============================================

class CoordHighlighter(QWidget):
    """
    전체화면 좌표 하이라이터.
    WA_TranslucentBackground 미사용 (보안정책 대응)
    스크린샷을 배경으로 직접 그려서 반투명 효과 구현.
    """
    def __init__(self, x: int, y: int, label: str = "",
                 color: str = "#F43F5E", duration_ms: int = 2500,
                 parent=None):
        super().__init__(parent)
        self._x        = x
        self._y        = y
        self._label    = label
        self._color    = QColor(color)
        self._radius   = 0
        self._max_r    = 52
        self._alpha    = 255
        self._phase    = "grow"
        self._bg       = None
        self._ss_buf   = None

        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint)
        # close() 호출 시 실제로 C++ 객체까지 파괴되도록 설정.
        # 이게 없으면 close()는 위젯을 숨기기만 하고 destroyed 시그널이
        # 발생하지 않아 _HIGHLIGHTER_REFS 에서 영원히 제거되지 않는
        # 메모리 누수가 발생함 (장시간 사용 시 누적).
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
        if not SECURE_ENV:
            self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setCursor(Qt.CursorShape.ArrowCursor)

        # 스크린샷 배경
        try:
            ss           = pyautogui.screenshot()
            self._ss_buf = ss.tobytes("raw", "RGB")
            qimg         = QImage(
                self._ss_buf, ss.width, ss.height,
                ss.width * 3, QImage.Format.Format_RGB888)
            self._bg = QPixmap.fromImage(qimg)
        except Exception:
            self._bg = None

        # 전체 화면에 맞게 배치
        screen = QApplication.primaryScreen().geometry()
        self.setGeometry(screen)
        self.activateWindow()
        self.raise_()

        # 애니메이션 타이머
        self._anim_timer = QTimer(self)
        self._anim_timer.timeout.connect(self._animate)
        self._anim_timer.start(16)

        # 자동 종료
        QTimer.singleShot(duration_ms, self.close)

    def _animate(self):
        if self._phase == "grow":
            self._radius = min(self._radius + 4, self._max_r)
            if self._radius >= self._max_r:
                self._phase = "hold"
        elif self._phase == "hold":
            pass   # hold 는 QTimer.singleShot 으로 fade 트리거
        elif self._phase == "fade":
            self._alpha = max(self._alpha - 10, 0)
            if self._alpha <= 0:
                self._anim_timer.stop()
                self.close()
                return
        self.update()

    def start_fade(self):
        self._phase = "fade"

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Escape:
            self.close()

    def mousePressEvent(self, event):
        self.close()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # ── 배경: 스크린샷 + 반투명 어두운 오버레이 ──
        if self._bg and not self._bg.isNull():
            painter.drawPixmap(0, 0, self._bg)
        else:
            painter.fillRect(self.rect(), QColor(30, 30, 30))
        painter.fillRect(self.rect(), QColor(0, 0, 0, 110))

        x, y = self._x, self._y
        r    = self._radius
        ac   = QColor(self._color)
        ac.setAlpha(self._alpha)

        # ── 중심 주변 원형 밝히기 (스크린샷 복원) ──
        if self._bg and not self._bg.isNull() and r > 0:
            from PyQt6.QtGui import QRegion
            clear_r = r + 40
            painter.setClipRegion(QRegion(
                x - clear_r, y - clear_r, clear_r * 2, clear_r * 2,
                QRegion.RegionType.Ellipse))
            painter.drawPixmap(0, 0, self._bg)
            painter.setClipping(False)
            border_c = QColor(self._color)
            border_c.setAlpha(180)
            painter.setPen(QPen(border_c, 2))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawEllipse(x - clear_r, y - clear_r, clear_r * 2, clear_r * 2)

        # ── 파동 원 (4겹 - 더 선명하게) ──
        for i, (radius_mult, alpha_mult, width) in enumerate([
            (2.2, 0.15, 1.5),
            (1.7, 0.30, 2.0),
            (1.3, 0.55, 2.5),
            (1.0, 0.90, 3.5),
        ]):
            wave_r  = int(r * radius_mult)
            if wave_r < 2: continue
            wave_ac = QColor(ac)
            wave_ac.setAlpha(int(self._alpha * alpha_mult))
            painter.setPen(QPen(wave_ac, width))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawEllipse(x - wave_r, y - wave_r, wave_r*2, wave_r*2)

        # ── 채운 중심 원 (크게) ──
        center_c = QColor(ac); center_c.setAlpha(int(self._alpha * 0.85))
        painter.setPen(QPen(QColor(255,255,255,int(self._alpha*0.9)), 2))
        painter.setBrush(QBrush(center_c))
        painter.drawEllipse(x - 14, y - 14, 28, 28)
        # 중심점
        white_c = QColor(255,255,255,self._alpha)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(white_c))
        painter.drawEllipse(x - 4, y - 4, 8, 8)

        # ── 크로스헤어 (굵고 선명하게) ──
        ch_solid = QPen(ac, 2.5)
        painter.setPen(ch_solid)
        painter.drawLine(x - 80, y, x - 20, y)
        painter.drawLine(x + 20, y, x + 80, y)
        painter.drawLine(x, y - 80, x, y - 20)
        painter.drawLine(x, y + 20, x, y + 80)
        # 크로스헤어 끝 화살표
        arrow_c = QColor(ac); arrow_c.setAlpha(int(self._alpha*0.6))
        painter.setPen(QPen(arrow_c, 1.5))
        for dx,dy in [(-80,0),(80,0),(0,-80),(0,80)]:
            painter.drawLine(x+dx, y+dy, x+dx+(4 if dx>0 else -4 if dx<0 else 0), y+dy-3)
            painter.drawLine(x+dx, y+dy, x+dx+(4 if dx>0 else -4 if dx<0 else 0), y+dy+3)

        # ── 좌표 & 레이블 텍스트 ──
        coord_txt = f"({x}, {y})"
        lbl_txt   = f"  {self._label}  " if self._label else ""

        txt_x = x + 20
        txt_y = y - 20
        # 화면 경계 보정
        screen = QApplication.primaryScreen().geometry()
        if txt_x + 160 > screen.width():  txt_x = x - 180
        if txt_y - 50  < 0:               txt_y = y + 60

        # ── 정보 박스 (더 크고 선명하게) ──
        fm    = QFontMetrics(QFont("Malgun Gothic", 12))
        lines = []
        if lbl_txt:   lines.append(("lbl",  lbl_txt.strip()))
        lines.append(("coord", coord_txt))
        lines.append(("hint",  "클릭 또는 ESC 로 닫기"))
        max_w = max(fm.horizontalAdvance(t) for _, t in lines) + 32
        box_h = len(lines) * 22 + 18
        box_rect = QRect(txt_x - 6, txt_y - 28, max_w, box_h)

        # 박스 배경 (어둡고 불투명)
        box_bg = QColor(15, 15, 15, int(230 * self._alpha / 255))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(box_bg))
        painter.drawRoundedRect(box_rect, 10, 10)

        # 박스 테두리 (액센트 컬러)
        border_box = QColor(self._color); border_box.setAlpha(int(self._alpha * 0.8))
        painter.setPen(QPen(border_box, 2))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRoundedRect(box_rect, 10, 10)

        # 텍스트 렌더링
        ac2 = QColor(self._color); ac2.setAlpha(self._alpha)
        ty_cursor = txt_y - 8
        for kind, txt in lines:
            if kind == "lbl":
                painter.setFont(QFont("Malgun Gothic", 12, QFont.Weight.Bold))
                painter.setPen(ac2)
            elif kind == "coord":
                painter.setFont(QFont("Consolas", 13, QFont.Weight.Bold))
                painter.setPen(QColor(255, 255, 255, self._alpha))
            else:
                painter.setFont(QFont("Malgun Gothic", 9))
                painter.setPen(QColor(180, 180, 180, int(self._alpha * 0.65)))
            painter.drawText(txt_x + 10, ty_cursor, txt)
            ty_cursor += 22

        painter.end()


def show_coord_highlight(action: dict):
    """
    PRV 버튼 클릭시 실제 화면에서 위치 가시화.
    2초 카운트다운 후 표시 (화면 전환 시간 확보).
    전역 참조 리스트로 GC 방지.
    """
    t = action.get("type", "")
    COLORS = {
        "click":       "#F43F5E",   # 강렬한 레드
        "scroll":      "#4F8EF7",   # 블루
        "move":        "#10B981",   # 그린
        "image_click": "#06B6D4",   # 시안
    }

    def _launch_highlighter(x, y, lbl, color, duration=3000):
        """메인 스레드에서 안전하게 하이라이터 생성"""
        h = CoordHighlighter(x, y, lbl, color, duration)
        _HIGHLIGHTER_REFS.append(h)   # GC 방지
        h.destroyed.connect(lambda: _HIGHLIGHTER_REFS.remove(h)
                            if h in _HIGHLIGHTER_REFS else None)
        h.show()
        h.raise_()
        h.activateWindow()
        QTimer.singleShot(int(duration * 0.75), h.start_fade)

    if t in ("click", "scroll", "move"):
        x   = action.get("x", 0)
        y   = action.get("y", 0)
        lbl = {
            "click":  f"{action.get('button','left').upper()} 클릭",
            "scroll": f"스크롤  dy={action.get('dy', 0):+d}",
            "move":   "마우스 이동",
        }.get(t, t)
        color = COLORS.get(t, "#F43F5E")
        # 2초 카운트다운 후 표시 (사용자가 화면 전환할 시간)
        QTimer.singleShot(2000, lambda: _launch_highlighter(x, y, lbl, color))

    elif t == "image_click":
        img_path = action.get("image_path", "")
        label    = action.get("label", "이미지")
        if CV2_OK and img_path and Path(img_path).exists():
            def _find():
                try:
                    loc = pyautogui.locateOnScreen(
                        img_path,
                        confidence=action.get("confidence", 0.85))
                    if loc:
                        cx = int(loc.left + loc.width  / 2)
                        cy = int(loc.top  + loc.height / 2)
                        QTimer.singleShot(0, lambda: _launch_highlighter(
                            cx, cy, f"이미지: {label}",
                            COLORS["image_click"], 3500))
                    else:
                        sc = QApplication.primaryScreen().geometry()
                        QTimer.singleShot(0, lambda: _launch_highlighter(
                            sc.width()//2, sc.height()//2,
                            f"이미지 못 찾음: {label}", "#F59E0B", 2000))
                except Exception as e:
                    sc = QApplication.primaryScreen().geometry()
                    QTimer.singleShot(0, lambda: _launch_highlighter(
                        sc.width()//2, sc.height()//2,
                        f"오류: {e}", "#F59E0B", 2000))
            QTimer.singleShot(2000, lambda: threading.Thread(
                target=_find, daemon=True).start())
        else:
            sc = QApplication.primaryScreen().geometry()
            QTimer.singleShot(2000, lambda: _launch_highlighter(
                sc.width()//2, sc.height()//2,
                f"이미지 파일 없음: {label}", "#F59E0B", 2000))

# =============================================
#  CoordMarker - 실시간 좌표 마커 (작은 플로팅 위젯)
#  전체화면 오버레이 없음 - 해당 좌표에 아이콘만 표시
# =============================================

class CoordMarker(QWidget):
    """
    지정 좌표 위에 작은 마커 위젯을 띄움.
    배경 어둡게 하지 않음 - 마커만 화면 위에 떠 있음.
    PRV 버튼을 누르는 동안만 표시.
    """
    def __init__(self, x: int, y: int, label: str,
                 type_icon: str, color: str, parent=None):
        super().__init__(parent)
        self._x      = x
        self._y      = y
        self._label  = label
        self._icon   = type_icon
        self._color  = QColor(color)
        self._pulse  = 0
        self._dir    = 1

        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint |
            Qt.WindowType.Tool)
        if not SECURE_ENV:
            self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating)
        self.setFixedSize(140, 90)

        # 화면 경계 보정
        screen = QApplication.primaryScreen().geometry()
        wx = x - 70
        wy = y - 80
        if wx < 0:                wx = x + 10
        if wx + 140 > screen.width():  wx = x - 150
        if wy < 0:                wy = y + 10
        if wy + 90 > screen.height():  wy = y - 100
        self.move(wx, wy)

        # 펄스 애니메이션
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(30)

    def _tick(self):
        self._pulse += self._dir * 6
        if self._pulse >= 100: self._dir = -1
        if self._pulse <= 0:   self._dir =  1
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        w, h  = self.width(), self.height()
        ac    = QColor(self._color)
        pulse = self._pulse / 100.0

        # 배경을 투명하게 (WA_TranslucentBackground 없으므로 직접 처리)
        painter.fillRect(self.rect(), QColor(0, 0, 0, 0))

        # ── 말풍선 배경 ──
        bubble_h = 52
        bubble_r = QRect(0, 0, w, bubble_h)

        # 그림자 효과 (살짝 아래로)
        shadow = QColor(0, 0, 0, 60)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(shadow))
        painter.drawRoundedRect(bubble_r.adjusted(3, 3, 3, 3), 10, 10)

        # 말풍선 본체
        bg = QColor(15, 15, 15, 230)
        painter.setBrush(QBrush(bg))
        painter.setPen(QPen(ac, 2))
        painter.drawRoundedRect(bubble_r.adjusted(0, 0, -1, -1), 10, 10)

        # 말풍선 꼬리 (아래쪽 삼각형 → 좌표 방향)
        tail = QPainterPath()
        tail.moveTo(w//2 - 8, bubble_h - 1)
        tail.lineTo(w//2,     bubble_h + 14)
        tail.lineTo(w//2 + 8, bubble_h - 1)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(QColor(15, 15, 15, 230)))
        painter.drawPath(tail)
        # 꼬리 테두리
        tail_border = QPainterPath()
        tail_border.moveTo(w//2 - 9, bubble_h)
        tail_border.lineTo(w//2,     bubble_h + 15)
        tail_border.lineTo(w//2 + 9, bubble_h)
        painter.setPen(QPen(ac, 2))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawPath(tail_border)

        # ── 아이콘 뱃지 ──
        badge_r = QRect(8, 8, 34, 34)
        badge_grad = QLinearGradient(8, 8, 42, 42)
        badge_grad.setColorAt(0, ac.lighter(130))
        badge_grad.setColorAt(1, ac)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(badge_grad))
        painter.drawRoundedRect(badge_r, 8, 8)

        painter.setFont(QFont("Malgun Gothic", 9, QFont.Weight.Bold))
        painter.setPen(QColor(255, 255, 255, 240))
        painter.drawText(badge_r, Qt.AlignmentFlag.AlignCenter, self._icon)

        # ── 레이블 & 좌표 ──
        painter.setFont(QFont("Malgun Gothic", 9, QFont.Weight.Bold))
        painter.setPen(QColor(ac))
        painter.drawText(QRect(48, 6, w - 52, 20),
                         Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
                         self._label)

        painter.setFont(QFont("Consolas", 10, QFont.Weight.Bold))
        painter.setPen(QColor(255, 255, 255, 220))
        painter.drawText(QRect(48, 26, w - 52, 20),
                         Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
                         f"({self._x}, {self._y})")

        # ── 펄스 원 (좌표 위치 - 마커 아래) ──
        dot_x = w // 2
        dot_y = bubble_h + 14 + 8   # 꼬리 끝 + 여유
        pulse_r = int(8 + pulse * 6)
        pulse_c = QColor(ac); pulse_c.setAlpha(int(220 - pulse * 140))
        painter.setPen(QPen(pulse_c, 2.5))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawEllipse(dot_x - pulse_r, dot_y - pulse_r,
                            pulse_r * 2, pulse_r * 2)

        # 중심 점
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(QColor(ac)))
        painter.drawEllipse(dot_x - 5, dot_y - 5, 10, 10)
        painter.setBrush(QBrush(QColor(255, 255, 255, 230)))
        painter.drawEllipse(dot_x - 2, dot_y - 2, 4, 4)

        painter.end()

    def closeEvent(self, event):
        self._timer.stop()
        super().closeEvent(event)


# 전역 마커 참조 (GC 방지)
_MARKER_REF = [None]

def show_realtime_marker(action: dict):
    """PRV 버튼 누를 때 즉시 마커 표시"""
    t = action.get("type", "")
    if t not in ("click", "scroll", "move", "image_click"):
        return

    COLORS = {
        "click":       C['rec'],
        "scroll":      C['brand'],
        "move":        C['play'],
        "image_click": C['ai'],
    }
    LABELS = {
        "click":  f"{action.get('button','left').upper()} 클릭",
        "scroll": f"스크롤 {action.get('dy',0):+d}",
        "move":   "마우스 이동",
        "image_click": f"이미지: {action.get('label','')}",
    }
    meta = BLOCK_META.get(t, {"icon": "?"})

    if t == "image_click":
        # 이미지는 화면에서 찾아야 하므로 백그라운드 탐색
        img_path = action.get("image_path", "")
        if CV2_OK and img_path and Path(img_path).exists():
            def _find():
                try:
                    loc = pyautogui.locateOnScreen(
                        img_path, confidence=action.get("confidence", 0.85))
                    x = int(loc.left + loc.width/2)  if loc else 0
                    y = int(loc.top  + loc.height/2) if loc else 0
                    lbl = f"이미지: {action.get('label','')}" if loc else "이미지 못 찾음"
                    color = COLORS["image_click"] if loc else C['stop']
                    QTimer.singleShot(0, lambda: _show(x, y, lbl, meta["icon"], color))
                except Exception:
                    pass
            threading.Thread(target=_find, daemon=True).start()
        return

    x   = action.get("x", 0)
    y   = action.get("y", 0)
    lbl = LABELS.get(t, t)
    _show(x, y, lbl, meta["icon"], COLORS.get(t, C['rec']))


def _show(x, y, label, icon, color):
    """마커 생성 및 전역 참조 보관"""
    hide_realtime_marker()   # 기존 마커 닫기
    m = CoordMarker(x, y, label, icon, color)
    _MARKER_REF[0] = m
    m.show()
    m.raise_()


def hide_realtime_marker():
    """PRV 버튼 뗄 때 마커 닫기"""
    if _MARKER_REF[0] is not None:
        try:
            _MARKER_REF[0].close()
        except Exception:
            pass
        _MARKER_REF[0] = None


BLOCK_META = {
    "click":       {"icon": "CLK",  "color": C['brand'],  "label": "마우스 클릭"},
    "key":         {"icon": "KEY",  "color": C['play'],   "label": "키 입력"},
    "type":        {"icon": "TXT",  "color": C['play'],   "label": "텍스트"},
    "hotkey":      {"icon": "HOT",  "color": C['tool'],   "label": "단축키"},
    "wait":        {"icon": "WAIT", "color": C['stop'],   "label": "대기"},
    "image_click": {"icon": "IMG",  "color": C['ai'],     "label": "이미지 클릭"},
    "scroll":      {"icon": "SCR",  "color": C['brand'],  "label": "스크롤"},
    "move":        {"icon": "MOV",  "color": C['t2'],     "label": "이동"},
    "cond_if":     {"icon": "IF",   "color": "#F59E0B",   "label": "조건 분기"},
    "excel_read":  {"icon": "XLS",  "color": "#22C55E",   "label": "엑셀 읽기"},
}

class FlowBlock(QFrame):
    """플로우 캔버스의 블록 하나"""
    edit_requested   = pyqtSignal(int)
    delete_requested = pyqtSignal(int)
    move_up          = pyqtSignal(int)
    move_down        = pyqtSignal(int)

    def __init__(self, index: int, action: dict, parent=None):
        super().__init__(parent)
        self.index  = index
        self.action = action
        self._error = False
        self._running = False
        self._done   = False
        self.setFixedHeight(80)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._build()

    def _build(self):
        lay = QHBoxLayout(self)
        lay.setContentsMargins(10, 6, 10, 6)
        lay.setSpacing(10)

        t    = self.action.get("type","click")
        meta = BLOCK_META.get(t, {"icon":"?","color":C['t2'],"label":t})
        ac   = meta["color"]

        # 번호 + 아이콘 뱃지
        badge = QLabel(f"{self.index+1}\n{meta['icon']}")
        badge.setFixedSize(48, 48)
        badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        bg_ac = QColor(ac)
        badge.setStyleSheet(f"""
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 {QColor(ac).lighter(120).name()}, stop:1 {ac});
            color: #FFFFFF;
            border-radius: 12px;
            font-size: 9px; font-weight: 900;
            line-height: 14px;
        """)
        lay.addWidget(badge)

        # 텍스트 영역
        txt_col = QVBoxLayout()
        txt_col.setSpacing(2)

        # 메인 설명
        main_txt = self._get_main_text()
        self.main_lbl = QLabel(main_txt)
        self.main_lbl.setStyleSheet(f"color:{C['t1']}; font-size:13px; font-weight:bold; background:transparent;")
        txt_col.addWidget(self.main_lbl)

        # AI 설명 또는 서브 텍스트
        ai_desc = self.action.get("ai_desc","")
        sub_txt = ai_desc if ai_desc else self._get_sub_text()
        self.sub_lbl = QLabel(sub_txt)
        self.sub_lbl.setStyleSheet(
            f"color:{C['ai']}; font-size:11px; background:transparent;" if ai_desc
            else f"color:{C['t3']}; font-size:11px; background:transparent;")
        txt_col.addWidget(self.sub_lbl)

        lay.addLayout(txt_col)
        lay.addStretch()

        # 딜레이 표시
        delay_lbl = QLabel(f"+{self.action.get('delay',0.5):.1f}s")
        delay_lbl.setStyleSheet(f"color:{C['t3']}; font-size:10px; background:transparent;")
        lay.addWidget(delay_lbl)

        # 실행 상태 아이콘
        self.status_lbl = QLabel("")
        self.status_lbl.setFixedWidth(20)
        self.status_lbl.setStyleSheet("background:transparent; font-size:12px;")
        lay.addWidget(self.status_lbl)

        # 조작 버튼
        btn_col = QVBoxLayout()
        btn_col.setSpacing(2)

        # 미리보기 버튼 (좌표 가시화) - 좌표가 있는 타입만 활성
        t_type = self.action.get("type","")
        has_coord = t_type in ("click","scroll","move","image_click")
        prv_btn = QPushButton("PRV")
        prv_btn.setFixedSize(30, 14)
        prv_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        prv_btn.setToolTip("실제 화면에서 이 블록의 위치를 표시합니다")
        prv_btn.setEnabled(has_coord)
        ac_col  = C['rec'] if has_coord else 'transparent'
        prv_btn.setStyleSheet(f"""
            QPushButton {{
                background: {ac_col};
                color: #FFFFFF;
                border: none;
                border-radius: 3px; font-size: 8px; font-weight: bold;
            }}
            QPushButton:hover {{
                background: {C['rec']};
                color: #FFFFFF;
            }}
            QPushButton:pressed {{
                background: {C['rec_dim']};
                color: {C['rec']};
            }}
            QPushButton:disabled {{
                background: transparent;
                color: {C['t3']}; border: 1px solid {C['t3']};
            }}
        """)
        act_copy = dict(self.action)
        prv_btn.setToolTip("누르는 동안 실제 화면에서 위치 표시")
        # pressed: 마커 표시 / released: 마커 숨김
        prv_btn.pressed.connect(lambda a=act_copy: show_realtime_marker(a))
        prv_btn.released.connect(hide_realtime_marker)
        btn_col.addWidget(prv_btn)

        for symbol, sig in [("UP", self.move_up), ("DN", self.move_down),
                              ("ED", self.edit_requested), ("DEL", self.delete_requested)]:
            colors_map = {"UP":C['brand'],"DN":C['brand'],
                          "ED":C['stop'],"DEL":C['rec']}
            b = QPushButton(symbol)
            b.setFixedSize(30, 14)
            b.setCursor(Qt.CursorShape.PointingHandCursor)
            b.setStyleSheet(f"""
                QPushButton {{
                    background: transparent; color: {colors_map[symbol]};
                    border: 1px solid {colors_map[symbol]};
                    border-radius: 3px; font-size: 8px; font-weight: bold;
                }}
                QPushButton:hover {{
                    background: {colors_map[symbol]}; color: #FFFFFF;
                }}
            """)
            idx = self.index
            if symbol == "UP":  b.clicked.connect(lambda _, i=idx: self.move_up.emit(i))
            elif symbol == "DN":b.clicked.connect(lambda _, i=idx: self.move_down.emit(i))
            elif symbol == "ED":b.clicked.connect(lambda _, i=idx: self.edit_requested.emit(i))
            else:               b.clicked.connect(lambda _, i=idx: self.delete_requested.emit(i))
            btn_col.addWidget(b)
        lay.addLayout(btn_col)

    def _get_main_text(self) -> str:
        a = self.action
        t = a.get("type","")
        if t == "click":
            return f"마우스 클릭  ({a.get('x',0)}, {a.get('y',0)})  {a.get('button','left')}"
        elif t == "key":
            return f"키 입력  [{a.get('key','')}]"
        elif t == "type":
            txt = a.get('text','')
            return f"텍스트 입력  \"{txt[:20]}{'...' if len(txt)>20 else ''}\""
        elif t == "hotkey":
            return f"단축키  {' + '.join(a.get('keys',[]))}"
        elif t == "wait":
            return f"대기  {a.get('seconds',1)}초"
        elif t == "image_click":
            return f"이미지 클릭  [{a.get('label','?')}]"
        elif t == "scroll":
            return f"스크롤  ({a.get('x',0)}, {a.get('y',0)})  {'위' if a.get('dy',1)>0 else '아래'}"
        elif t == "move":
            return f"마우스 이동  ({a.get('x',0)}, {a.get('y',0)})"
        elif t == "cond_if":
            ctype = a.get("cond_type","image_exist")
            img   = Path(a.get("cond_img","")).name if a.get("cond_img") else "미설정"
            label = "이미지 있으면" if ctype=="image_exist" else "이미지 없으면"
            return f"조건 분기  [{label}]  {img}"
        return t

    def _get_sub_text(self) -> str:
        a = self.action
        t = a.get("type","")
        if t == "click":
            return f"이전 동작 후 {a.get('delay',0.5):.1f}초 대기 후 클릭"
        elif t == "key":
            key_hints = {
                "enter":"확인/실행","tab":"다음 칸 이동","esc":"취소/닫기",
                "ctrl":"컨트롤 키","delete":"삭제","f5":"새로고침"
            }
            return key_hints.get((a.get("key") or "").lower(), "키보드 입력")
        elif t == "type":
            return "키보드로 텍스트 직접 타이핑"
        elif t == "hotkey":
            hotkey_hints = {
                "ctrl+s":"저장","ctrl+z":"실행취소","ctrl+c":"복사",
                "ctrl+v":"붙여넣기","ctrl+a":"전체선택","alt+f4":"닫기"
            }
            k = "+".join(a.get("keys") or []).lower()
            return hotkey_hints.get(k, "단축키 실행")
        elif t == "wait":
            return "프로그램 로딩 / 화면 전환 대기"
        elif t == "image_click":
            return f"화면에서 이미지 찾아 클릭  (정확도: {a.get('confidence',0.85):.0%})"
        return ""

    def set_error(self, has_error: bool):
        self._error = has_error
        self._refresh_style()

    def set_running(self, running: bool):
        self._running = running
        self.status_lbl.setText(">" if running else "")
        self._refresh_style()

    def set_done(self, done: bool):
        self._done = done
        self.status_lbl.setText("OK" if done else "")
        self._refresh_style()

    def _refresh_style(self):
        t    = self.action.get("type","click")
        meta = BLOCK_META.get(t, {"color":C['t2']})
        ac   = meta["color"]
        if self._error:
            self.setStyleSheet(f"""
                QFrame {{
                    background: {C['rec_dim']};
                    border: 2px solid {C['err']};
                    border-left: 5px solid {C['err']};
                    border-radius: 12px;
                }}
            """)
        elif self._running:
            self.setStyleSheet(f"""
                QFrame {{
                    background: {C['stop_dim']};
                    border: 2px solid {C['stop']};
                    border-left: 5px solid {C['stop']};
                    border-radius: 12px;
                }}
            """)
        elif self._done:
            self.setStyleSheet(f"""
                QFrame {{
                    background: {C['play_dim']};
                    border: 1.5px solid {C['play']};
                    border-left: 5px solid {C['play']};
                    border-radius: 12px;
                }}
            """)
        else:
            self.setStyleSheet(f"""
                QFrame {{
                    background: {C['bg1']};
                    border: 1px solid {C['border']};
                    border-left: 5px solid {ac};
                    border-radius: 12px;
                }}
                QFrame:hover {{
                    background: {C['bg4']};
                    border: 1px solid {ac};
                    border-left: 5px solid {ac};
                }}
            """)

    def mouseDoubleClickEvent(self, event):
        self.edit_requested.emit(self.index)


# =============================================
#  플로우 캔버스
# =============================================
class FlowCanvas(QWidget):
    """블록들을 수직으로 배치하는 스크롤 가능한 캔버스"""
    block_edited   = pyqtSignal()
    ai_insert_req  = pyqtSignal(int)   # AI 삽입 위치 요청 (after_index)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._actions  : list = []
        self._blocks   : list = []
        self._layout   = QVBoxLayout(self)
        self._layout.setContentsMargins(16, 16, 16, 16)
        self._layout.setSpacing(0)
        self.setStyleSheet(f"background: {C['bg0']};")

    # ── 데이터 관리 ──
    def set_block_done_proxy(self, index: int):
        """블록 실행 완료 시각 업데이트"""
        if 0 <= index < len(self._blocks):
            self._blocks[index].set_running(False)
            self._blocks[index].set_done(True)

    def set_block_running(self, index: int):
        """스텝 실행 중 블록 하이라이트"""
        for b in self._blocks:
            b.set_running(False)
        if 0 <= index < len(self._blocks):
            self._blocks[index].set_running(True)

    def set_block_done(self, index: int):
        """스텝 실행 완료 블록 표시"""
        if 0 <= index < len(self._blocks):
            self._blocks[index].set_running(False)
            self._blocks[index].set_done(True)

    def set_actions(self, actions: list):
        self._actions = [dict(a) for a in actions]
        self._rebuild()

    def get_actions(self) -> list:
        return [dict(a) for a in self._actions]

    def _rebuild(self):
        # 기존 위젯 제거
        while self._layout.count():
            item = self._layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        self._blocks.clear()

        # START 노드
        self._layout.addWidget(self._make_terminal("START", C['play']))
        self._layout.addWidget(self._make_arrow())

        for i, action in enumerate(self._actions):
            blk = FlowBlock(i, action)
            blk.edit_requested.connect(self._on_edit)
            blk.delete_requested.connect(self._on_delete)
            blk.move_up.connect(self._on_move_up)
            blk.move_down.connect(self._on_move_down)
            blk._refresh_style()
            self._blocks.append(blk)
            self._layout.addWidget(blk)

            # [+] 삽입 버튼
            add_btn = self._make_add_btn(i)
            self._layout.addWidget(add_btn)

        # END 노드
        self._layout.addWidget(self._make_terminal("END", C['rec']))
        self._layout.addStretch()

    def _make_terminal(self, text, color):
        f = QFrame()
        f.setFixedHeight(38)
        f.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 {color}, stop:1 {QColor(color).darker(120).name()});
                border-radius: 19px;
                margin: 0 80px;
            }}
        """)
        l = QHBoxLayout(f)
        lbl = QLabel(text)
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl.setStyleSheet("color:#FFFFFF; font-weight:900; font-size:12px; letter-spacing:3px; background:transparent;")
        l.addWidget(lbl)
        return f

    def _make_arrow(self):
        lbl = QLabel(":")
        lbl.setFixedHeight(14)
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl.setStyleSheet(f"color:{C['border']}; font-size:18px; font-weight:900; background:transparent; letter-spacing:2px;")
        return lbl

    def _make_add_btn(self, after_index: int):
        w   = QWidget()
        w.setFixedHeight(28)
        w.setStyleSheet("background:transparent;")
        lay = QHBoxLayout(w)
        lay.setContentsMargins(40, 2, 40, 2)
        lay.setSpacing(4)

        # + 프로세스 추가 버튼 - 클릭 시 선택 메뉴
        btn = QPushButton("+ 플로우 추가")
        btn.setFixedHeight(22)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {C['brand']};
                border: 1.5px dashed {C['brand']}; border-radius: 6px;
                font-size: 11px; font-weight: bold;
                padding: 0 8px;
            }}
            QPushButton:hover {{
                background: {C['brand']}; color: #FFFFFF;
                border-style: solid;
            }}
        """)

        def _show_menu(checked=False, i=after_index):
            menu = QMenu(btn)
            menu.setStyleSheet(f"""
                QMenu {{
                    background: {C['bg1']};
                    border: 1.5px solid {C['border']};
                    border-radius: 10px;
                    padding: 4px;
                    font-size: 13px;
                }}
                QMenu::item {{
                    padding: 8px 20px;
                    border-radius: 6px;
                    color: {C['t1']};
                }}
                QMenu::item:selected {{
                    background: {C['brand']};
                    color: #FFFFFF;
                }}
                QMenu::separator {{
                    height: 1px;
                    background: {C['border']};
                    margin: 4px 8px;
                }}
            """)
            act_normal = QAction("  직접 추가  (타입 선택)", menu)
            act_normal.setToolTip("클릭/키입력/텍스트 등 직접 설정")

            act_ai     = QAction("  AI 로 프로세스 만들기", menu)
            act_ai.setToolTip("자연어로 입력하면 AI 가 프로세스 자동 생성")

            menu.addAction(act_normal)
            menu.addSeparator()
            menu.addAction(act_ai)

            chosen = menu.exec(btn.mapToGlobal(btn.rect().bottomLeft()))
            if chosen == act_normal:
                self._on_insert(i + 1)
            elif chosen == act_ai:
                self.ai_insert_req.emit(i + 1)

        btn.clicked.connect(_show_menu)
        lay.addWidget(btn)
        return w

    # ── 이벤트 ──
    def _on_edit(self, index: int):
        if index < 0 or index >= len(self._actions): return
        dlg = BlockEditDialog(self._actions[index], self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self._actions[index] = dlg.action
            self._rebuild()
            self.block_edited.emit()

    def _on_delete(self, index: int):
        if index < 0 or index >= len(self._actions): return
        self._actions.pop(index)
        self._rebuild()
        self.block_edited.emit()

    def _on_move_up(self, index: int):
        if index <= 0: return
        self._actions[index-1], self._actions[index] = \
            self._actions[index], self._actions[index-1]
        self._rebuild()
        self.block_edited.emit()

    def _on_move_down(self, index: int):
        if index >= len(self._actions)-1: return
        self._actions[index], self._actions[index+1] = \
            self._actions[index+1], self._actions[index]
        self._rebuild()
        self.block_edited.emit()

    def _on_insert(self, at_index: int):
        new_action = {"type":"click","x":500,"y":400,
                      "button":"left","delay":0.5}
        dlg = BlockEditDialog(new_action, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self._actions.insert(at_index, dlg.action)
            self._rebuild()
            self.block_edited.emit()

    # ── 상태 표시 ──
    def highlight_running(self, index: int):
        for i, blk in enumerate(self._blocks):
            blk.set_running(i == index)
            if i < index: blk.set_done(True)

    def clear_highlights(self):
        for blk in self._blocks:
            blk.set_running(False)
            blk.set_done(False)

    def mark_errors(self, error_indices: list):
        for i, blk in enumerate(self._blocks):
            blk.set_error(i in error_indices)


# =============================================
#  플로우 재생 워커
# =============================================
class FlowPlayWorker(QThread):
    block_started   = pyqtSignal(int)
    block_done      = pyqtSignal(int)
    status_update   = pyqtSignal(str)
    finished_play   = pyqtSignal()

    def __init__(self, actions, speed=1.0, minimize_others=False):
        super().__init__()
        self.actions         = actions
        self.speed           = speed
        self.minimize_others = minimize_others
        self.running         = False

    def _exec_act(self, act: dict):
        """
        단일 액션 실행 (click/key/type/hotkey/wait/image_click/scroll/move).
        메인 재생 루프와 excel_read 반복 루프가 공통으로 사용.
        """
        t = act.get("type","")
        if t == "click":
            btn = act.get("button","left")
            x = act.get("x",0); y = act.get("y",0)
            snap = act.get("snapshot","")
            if btn == "double":
                dx, dy_ = _logical_to_phys(x, y)
                pyautogui.doubleClick(dx, dy_)
            else:
                _smart_click(x, y, snapshot_path=snap, button=btn,
                             confidence=act.get("confidence",0.85))
        elif t == "key":
            k = act.get("key") or ""
            if k:
                if k.startswith("Key."): k = k.replace("Key.","")
                pyautogui.press(k)
        elif t == "type":
            txt = act.get("text") or ""
            if txt: _type_safe(txt, act.get("ime_state", "english"))
        elif t == "hotkey":
            keys = act.get("keys") or []
            if keys:
                if keys == ["hangul"]:
                    pass  # 한/영 키는 재생 안함
                else:
                    pyautogui.hotkey(*keys)
        elif t == "wait":
            _safe_wait(float(act.get("seconds",1)),
                       running_check=lambda: self.running)
        elif t == "image_click":
            if CV2_OK:
                img     = act.get("image_path","")
                conf    = act.get("confidence", 0.85)
                timeout = act.get("wait_timeout", 0)
                if img and Path(img).exists():
                    loc = None
                    if timeout > 0:
                        elapsed = 0
                        while elapsed < timeout and self.running:
                            try:
                                loc = pyautogui.locateOnScreen(
                                    img, confidence=conf)
                                if loc: break
                            except Exception:
                                pass
                            time.sleep(0.5)
                            elapsed += 0.5
                        if not loc:
                            self.status_update.emit(
                                f"[대기 타임아웃] {timeout}초 내 이미지 미발견: "
                                f"{act.get('label','')}")
                    else:
                        try:
                            loc = pyautogui.locateOnScreen(
                                img, confidence=conf)
                        except Exception:
                            loc = None
                    if loc:
                        pyautogui.click(
                            int(loc.left+loc.width/2),
                            int(loc.top+loc.height/2))
                    elif timeout == 0:
                        self.status_update.emit(
                            f"[주의] 이미지 못찾음: {act.get('label','')}")
        elif t == "scroll":
            sx, sy = _logical_to_phys(act.get("x",0), act.get("y",0))
            pyautogui.scroll(act.get("dy",3), x=sx, y=sy)
        elif t == "move":
            mx, my = _logical_to_phys(act.get("x",0), act.get("y",0))
            pyautogui.moveTo(mx, my, duration=0.2)
        # excel_read 는 중첩 반복을 막기 위해 여기서 처리하지 않음
        # (엑셀값으로 반복되는 하위 액션에 또 엑셀읽기가 있으면 무시)

    def run(self):
        self.running = True
        _esc_watcher = _GlobalEscWatcher(on_esc=self.stop)
        _esc_watcher.start()
        # 재생 시작 전 다른 창 최소화 (옵션)
        if self.minimize_others:
            self.status_update.emit("다른 창 최소화 중...")
            _minimize_all_windows()
            time.sleep(0.3)
        for i, act in enumerate(self.actions):
            if not self.running: break
            if not isinstance(act, dict): continue
            self.block_started.emit(i)
            self.status_update.emit(f"실행 중: {i+1}/{len(self.actions)} 블록")
            try:
                # 액션 타입별 최소 delay 보장 (빠른 타이핑/연속동작 시 안정성)
                t_check = act.get("type","")
                raw_delay = act.get("delay", 0.5)
                if t_check == "type":
                    min_delay = 0.15
                elif t_check == "click":
                    min_delay = 0.08
                elif t_check == "hotkey":
                    min_delay = 0.1
                else:
                    min_delay = 0.03
                scaled = raw_delay / self.speed
                time.sleep(max(scaled, min_delay))
            except Exception:
                time.sleep(0.1)
            try:
                t = act.get("type","")
                if t == "excel_read":
                    # 엑셀에서 값 읽어 반복 실행
                    xl_path   = act.get("excel_path","")
                    sheet     = act.get("sheet_name","Sheet1")
                    col       = act.get("column","A")
                    start_row = act.get("start_row", 1)
                    if not xl_path or not Path(xl_path).exists():
                        self.status_update.emit(f"[엑셀오류] 파일 없음: {xl_path}")
                    else:
                        try:
                            import openpyxl as _opxl
                            wb  = _opxl.load_workbook(xl_path, data_only=True)
                            ws  = wb[sheet] if sheet in wb.sheetnames else wb.active
                            # A열(또는 지정 컬럼)에서 값 읽기
                            col_idx = ord(col.upper()) - ord('A') + 1
                            values  = []
                            for row in ws.iter_rows(
                                    min_row=start_row,
                                    min_col=col_idx,
                                    max_col=col_idx):
                                v = row[0].value
                                if v is not None:
                                    values.append(str(v).strip())
                            wb.close()
                            self.status_update.emit(
                                f"엑셀 {len(values)}개 값 로드: {values[:3]}...")
                            # 현재 블록 이후 플로우를 값만큼 반복 실행
                            remaining = self.actions[i+1:]
                            for vi, val in enumerate(values):
                                if not self.running: break
                                self.status_update.emit(
                                    f"엑셀 반복 {vi+1}/{len(values)}: {val}")
                                # 변수 저장
                                self._excel_var = val
                                for sub_act in remaining:
                                    if not self.running: break
                                    if not isinstance(sub_act, dict): continue
                                    # type 타입이고 text 가 {{VALUE}} 면 엑셀값 치환
                                    sa = dict(sub_act)
                                    if sa.get("type") == "type":
                                        sa["text"] = sa.get(
                                            "text","").replace(
                                            "{{VALUE}}", val)
                                    time.sleep(max(
                                        sa.get("delay",0.3)/self.speed,0.01))
                                    self._exec_act(sa)
                            # 반복 완료 후 루프 종료
                            self.running = False
                        except Exception as e:
                            self.status_update.emit(f"[엑셀오류] {e}")
                else:
                    self._exec_act(act)
            except pyautogui.FailSafeException:
                self.status_update.emit("FAILSAFE: 마우스 모서리 감지 - 중지")
                self.running = False
                break
            except Exception as e:
                self.status_update.emit(f"[오류] 블록 {i+1}: {e}")
            self.block_done.emit(i)
        self.running = False
        _esc_watcher.stop()
        self.status_update.emit("재생 완료!")
        self.finished_play.emit()

    def stop(self):
        self.running = False



# =============================================
#  페이지 3: [FLOW] 플로우 에디터 (핵심)
# =============================================
class FlowEditorPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._worker    = None
        self._ai_online = False
        self._insert_at = None
        self._step_idx  = -1   # 스텝 실행 현재 위치 (-1 = 미시작)
        self._build()
        # AI 상태 비동기 확인 - 빌드 완료 후 500ms 딜레이
        QTimer.singleShot(500, lambda: threading.Thread(
            target=self._check_ai, daemon=True).start())
        # 샘플 플로우 자동 로드 (캔버스 비어있을 때)
        QTimer.singleShot(800, self._load_sample_if_empty)
        # ESC 키 → 재생 중지 (ApplicationShortcut: 어디서든 동작)
        self._esc_sc = QShortcut(QKeySequence("Escape"), self)
        self._esc_sc.setContext(
            Qt.ShortcutContext.ApplicationShortcut)
        self._esc_sc.activated.connect(self._on_flow_esc)

    def _load_sample_if_empty(self):
        """플로우가 비어있으면 샘플 자동 로드"""
        if len(self.canvas.get_actions()) > 0: return
        sample_path = SAVE_DIR / "샘플_메모장자동화.json"
        if not sample_path.exists(): return
        try:
            with open(sample_path, encoding="utf-8") as f:
                d = json.load(f)
            actions = (d.get("actions",[]) if isinstance(d,dict) else d)
            if actions:
                self.canvas.set_actions(actions)
                self.flow_status.setText(
                    f"샘플 플로우 로드됨 ({len(actions)}개) - "
                    "자유롭게 수정하세요!")
        except Exception:
            pass

    def _check_ai(self):
        ok = GemmaEngine.is_online()
        self._ai_online = ok
        if ok:
            QTimer.singleShot(0, self._set_ai_status_on)
        else:
            QTimer.singleShot(0, self._set_ai_status_off)
            QTimer.singleShot(30000, lambda: threading.Thread(
                target=self._check_ai, daemon=True).start())

    def _set_ai_status_on(self):
        if hasattr(self, 'ai_status_lbl'):
            self.ai_status_lbl.setText(f"AI ON ({OLLAMA_MODEL})")
            self.ai_status_lbl.setStyleSheet(
                "color:#2F9E44; font-size:11px; font-weight:bold; background:transparent;")

    def _set_ai_status_off(self):
        if hasattr(self, 'ai_status_lbl'):
            self.ai_status_lbl.setText("AI OFF")
            self.ai_status_lbl.setStyleSheet(
                f"color:{C['rec']}; font-size:11px; background:transparent;")

    def _update_ai_status(self, ok: bool):
        if ok: self._set_ai_status_on()
        else:  self._set_ai_status_off()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── 툴바 ──
        toolbar = QFrame()
        toolbar.setFixedHeight(52)
        toolbar.setStyleSheet(f"""
            QFrame {{
                background: {C['bg1']};
                border-bottom: 1.5px solid {C['border']};
            }}
        """)
        tb = QHBoxLayout(toolbar)
        tb.setContentsMargins(16, 0, 16, 0)
        tb.setSpacing(8)

        tb.addWidget(SectionHeader("FLOW EDITOR  /  플로우 에디터"))
        tb.addStretch()

        # AI 상태
        self.ai_status_lbl = QLabel("AI 확인 중...")
        self.ai_status_lbl.setStyleSheet(f"color:{C['t3']}; font-size:11px; background:transparent;")
        tb.addWidget(self.ai_status_lbl)

        # 수동 재확인 버튼
        btn_ai_check = QPushButton("재확인")
        btn_ai_check.setFixedHeight(24)
        btn_ai_check.setFixedWidth(52)
        btn_ai_check.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_ai_check.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {C['brand']};
                border: 1px solid {C['brand']}; border-radius: 4px;
                font-size: 10px;
            }}
            QPushButton:hover {{ background: {C['bg4']}; }}
        """)
        btn_ai_check.clicked.connect(lambda: (
            self.ai_status_lbl.setText("확인 중..."),
            threading.Thread(target=self._check_ai, daemon=True).start()
        ))
        tb.addWidget(btn_ai_check)
        tb.addSpacing(8)

        for label, color, fn in [
            ("불러오기",   C['brand'],  self._load),
            ("저장",       C['play'],   self._save),
        ]:
            b = GlowButton(label, color)
            b.setFixedHeight(34); b.setFixedWidth(80)
            b.clicked.connect(fn)
            tb.addWidget(b)

        root.addWidget(toolbar)

        # ── 메인 영역: 팔레트 | 캔버스 | AI패널 ──
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setStyleSheet("QSplitter { background: transparent; }")

        # ── 좌측: 블록 팔레트 ──
        palette = QFrame()
        palette.setFixedWidth(150)
        palette.setStyleSheet(f"""
            QFrame {{
                background: {C['bg1']};
                border-right: 1px solid {C['border']};
            }}
        """)
        pal_lay = QVBoxLayout(palette)
        pal_lay.setContentsMargins(8, 12, 8, 12)
        pal_lay.setSpacing(6)

        pal_title = QLabel("플로우 추가")
        pal_title.setStyleSheet(f"color:{C['brand']}; font-weight:900; font-size:11px; background:transparent;")
        pal_lay.addWidget(pal_title)

        PALETTE_ITEMS = [
            ("CLK",  "클릭",       "click",        C['brand']),
            ("KEY",  "키 입력",    "key",          C['play']),
            ("TXT",  "텍스트",     "type",         C['play']),
            ("HOT",  "단축키",     "hotkey",       C['tool']),
            ("WAIT", "대기",       "wait",         C['stop']),
            ("IMG",  "이미지",     "image_click",  C['ai']),
            ("SCR",  "스크롤",     "scroll",       C['brand']),
            ("IF",   "조건 분기",  "cond_if",      "#F59E0B"),
            ("XLS",  "엑셀 읽기",  "excel_read",   "#22C55E"),
        ]
        for icon, label, btype, color in PALETTE_ITEMS:
            btn = QPushButton(f"{icon}  {label}")
            btn.setFixedHeight(40)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {C['bg1']};
                    color: {color};
                    border: none;
                    border-left: 4px solid {color};
                    border-radius: 8px;
                    font-size: 11px; font-weight: bold;
                    text-align: left;
                    padding-left: 10px;
                }}
                QPushButton:hover {{
                    background: {color}18;
                    border-left: 4px solid {color};
                }}
                QPushButton:pressed {{
                    background: {color}35;
                }}
            """)
            btn.clicked.connect(lambda _, bt=btype: self._palette_add(bt))
            pal_lay.addWidget(btn)

        pal_lay.addStretch()

        # 재생 컨트롤
        pal_lay.addWidget(QLabel(""))
        play_title = QLabel("재생")
        play_title.setStyleSheet(f"color:{C['play']}; font-weight:900; font-size:11px; background:transparent;")
        pal_lay.addWidget(play_title)

        self.speed_combo = StyledCombo()
        self.speed_combo.addItems(["0.5x","1.0x","1.5x","2.0x"])
        self.speed_combo.setCurrentIndex(1)
        pal_lay.addWidget(self.speed_combo)

        # 재생 전 다른 창 최소화 옵션
        self.chk_min_others = QCheckBox("배경창 최소화")
        self.chk_min_others.setChecked(True)
        self.chk_min_others.setToolTip(
            "재생 전 다른 창을 모두 최소화 (좌표 클릭 오작동 방지)")
        self.chk_min_others.setStyleSheet(
            f"color:{C['t2']}; font-size:11px; background:transparent;")
        pal_lay.addWidget(self.chk_min_others)

        # RPA 자체 창도 최소화 옵션 (재생 중 RPA UI 가 화면 일부 가리는 것 방지)
        self.chk_min_self = QCheckBox("RPA 창도 최소화")
        self.chk_min_self.setChecked(True)
        self.chk_min_self.setToolTip(
            "재생 중 5MRPA 창을 최소화하여 가리는 영역 제거")
        self.chk_min_self.setStyleSheet(
            f"color:{C['t2']}; font-size:11px; background:transparent;")
        pal_lay.addWidget(self.chk_min_self)

        self.btn_play = GlowButton("PLAY", C['play'])
        self.btn_play.setFixedHeight(38)
        self.btn_play.clicked.connect(self._play)
        self.btn_stop_play = GlowButton("STOP", C['rec'])
        self.btn_stop_play.setFixedHeight(38)
        self.btn_stop_play.setEnabled(False)
        self.btn_stop_play.clicked.connect(self._stop_play)
        self.btn_step = GlowButton("STEP >", C['stop'])
        self.btn_step.setFixedHeight(38)
        self.btn_step.setToolTip("한 스텝씩 정방향 실행")
        self.btn_step.clicked.connect(self._step_play)
        self.btn_step_back = GlowButton("< BACK", C['tool'])
        self.btn_step_back.setFixedHeight(38)
        self.btn_step_back.setToolTip("한 스텝씩 역방향 실행")
        self.btn_step_back.clicked.connect(self._step_back)
        pal_lay.addWidget(self.btn_play)
        pal_lay.addWidget(self.btn_stop_play)
        pal_lay.addWidget(self.btn_step)
        pal_lay.addWidget(self.btn_step_back)

        splitter.addWidget(palette)

        # ── 중앙: 플로우 캔버스 (스크롤) ──
        canvas_wrap = QWidget()
        canvas_wrap.setStyleSheet(f"background:{C['bg0']};")
        cw_lay = QVBoxLayout(canvas_wrap)
        cw_lay.setContentsMargins(0, 0, 0, 0)
        cw_lay.setSpacing(0)

        # 캔버스 안내 툴팁 바
        hint_bar = QFrame()
        hint_bar.setFixedHeight(32)
        hint_bar.setStyleSheet(f"""
            QFrame {{
                background: {C['ai_dim']};
                border-bottom: 1px solid {C['border']};
            }}
        """)
        hb_lay = QHBoxLayout(hint_bar)
        hb_lay.setContentsMargins(14, 0, 14, 0)
        hint_lbl = QLabel("  [PRV] 클릭 = 실제 화면에서 위치 표시  |  플로우 더블클릭 = 편집  |  편집 팝업의 [위치 미리보기]로도 확인")
        hint_lbl.setStyleSheet(
            f"color:{C['ai']}; font-size:11px; background:transparent; font-weight:bold;")
        hb_lay.addWidget(hint_lbl)
        cw_lay.addWidget(hint_bar)

        canvas_scroll = QScrollArea()
        canvas_scroll.setWidgetResizable(True)
        canvas_scroll.setStyleSheet(f"""
            QScrollArea {{
                background: {C['bg0']};
                border: none;
            }}
        """)
        self.canvas = FlowCanvas()
        self.canvas.block_edited.connect(self._on_canvas_edited)
        self.canvas.ai_insert_req.connect(self._on_ai_insert_req)
        canvas_scroll.setWidget(self.canvas)
        cw_lay.addWidget(canvas_scroll)
        splitter.addWidget(canvas_wrap)

        # ── 우측: AI 패널 ──
        ai_panel = QFrame()
        ai_panel.setFixedWidth(260)
        ai_panel.setStyleSheet(f"""
            QFrame {{
                background: #FAFCFF;
                border-left: 1px solid {C['border']};
            }}
        """)
        ai_lay = QVBoxLayout(ai_panel)
        ai_lay.setContentsMargins(12, 12, 12, 12)
        ai_lay.setSpacing(10)

        ai_title = QLabel(f"AI 도우미  ({OLLAMA_MODEL})")
        ai_title.setStyleSheet(f"""
            color:{C['ai']}; font-weight:900; font-size:12px; background:transparent;
            padding:6px 10px; border-radius:8px;
            background: {C['ai_dim']};
            border: 1px solid {C['ai']};
        """)
        ai_lay.addWidget(ai_title)

        # 자연어 → 블록
        nl_lbl = QLabel("AI 로 플로우 만들기")
        nl_lbl.setStyleSheet(f"color:{C['t1']}; font-weight:bold; font-size:12px; background:transparent;")
        ai_lay.addWidget(nl_lbl)

        self.nl_input = QTextEdit()
        self.nl_input.setFixedHeight(80)
        self.nl_input.setPlaceholderText(
            "예:\n엑셀 저장하고 닫아줘\n메모장 열고 안녕하세요 입력해줘")
        self.nl_input.setStyleSheet(f"""
            QTextEdit {{
                background: {C['bg2']}; border: 1.5px solid {C['border']};
                border-radius: 8px; color: {C['t1']}; padding: 6px; font-size:12px;
            }}
            QTextEdit:focus {{ border-color: {C['ai']}; }}
        """)
        ai_lay.addWidget(self.nl_input)

        btn_nl = GlowButton("플로우 자동 생성", C['ai'])
        btn_nl.setFixedHeight(36)
        btn_nl.clicked.connect(self._nl_to_blocks)
        ai_lay.addWidget(btn_nl)

        # 구분선
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet(f"color:{C['border']};")
        ai_lay.addWidget(line)

        # 프로세스 점검 (오류 자동 점검 + AI 검토 통합 - 규칙기반 항상 정확, AI는 1회만 보조)
        btn_check = GlowButton("프로세스 점검", C['stop'])
        btn_check.setFixedHeight(36)
        btn_check.setToolTip("좌표/설정 오류를 즉시 확인하고, AI가 해설·개선점을 덧붙입니다")
        btn_check.clicked.connect(self._process_check)
        ai_lay.addWidget(btn_check)

        # AI 프로세스 최적화 (의미없는 대기/중복 동작 제거 - 규칙기반이 항상 동작, AI는 보조)
        btn_optimize = GlowButton("AI 최적화", C['ai'])
        btn_optimize.setFixedHeight(36)
        btn_optimize.setToolTip("의미없는 대기시간 단축, 중복 클릭 제거")
        btn_optimize.clicked.connect(self._ai_optimize)
        ai_lay.addWidget(btn_optimize)

        # 결과 표시
        res_lbl = QLabel("결과")
        res_lbl.setStyleSheet(f"color:{C['t2']}; font-weight:bold; font-size:11px; background:transparent;")
        ai_lay.addWidget(res_lbl)

        self.ai_result = QTextEdit()
        self.ai_result.setReadOnly(True)
        self.ai_result.setStyleSheet(f"""
            QTextEdit {{
                background: #FAFCFF;
                border: 1.5px solid {C['border']};
                border-radius: 10px;
                color: {C['t1']};
                padding: 10px;
                font-size: 12px;
                line-height: 1.5;
            }}
        """)
        self.ai_result.setPlaceholderText(
            "AI 분석 결과가 여기에 표시됩니다.\n\n"
            "  프로세스 점검  --  좌표/설정 오류 즉시 확인 + AI 해설·개선 제안\n"
            "  AI 최적화     --  의미없는 대기/중복 동작 제거\n\n"
            "  ※ 블록별 상세 설명은 블록을 더블클릭해 편집창에서 생성할 수 있습니다.")
        ai_lay.addWidget(self.ai_result)

        # 상태 표시
        self.flow_status = QLabel("대기 중")
        self.flow_status.setWordWrap(True)
        self.flow_status.setStyleSheet(f"""
            color:{C['t2']}; font-size:11px;
            background:{C['bg2']}; border:1px solid {C['border']};
            border-radius:6px; padding:6px 8px;
        """)
        ai_lay.addWidget(self.flow_status)

        splitter.addWidget(ai_panel)
        splitter.setSizes([150, 560, 260])

        root.addWidget(splitter)

    # ── 팔레트에서 블록 추가 ──
    def _palette_add(self, btype: str):
        defaults = {
            "click":       {"type":"click","x":500,"y":400,"button":"left","delay":0.5},
            "key":         {"type":"key","key":"enter","delay":0.3},
            "type":        {"type":"type","text":"","delay":0.5},
            "hotkey":      {"type":"hotkey","keys":["ctrl","s"],"delay":0.3},
            "wait":        {"type":"wait","seconds":2,"delay":0},
            "image_click": {"type":"image_click","label":"","image_path":"","confidence":0.85,"delay":1.0},
            "scroll":      {"type":"scroll","x":500,"y":400,"dy":3,"delay":0.3},
        }
        act = defaults.get(btype, {"type":btype,"delay":0.5})
        dlg = BlockEditDialog(act, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            acts = self.canvas.get_actions()
            acts.append(dlg.action)
            self.canvas.set_actions(acts)
            self.flow_status.setText(f"프로세스 추가됨: {btype}")

    # ── 불러오기 / 저장 ──
    def _load(self):
        fp, _ = rpa_open_file(self, "매크로 불러오기", SAVE_DIR)
        if not fp: return
        try:
            with open(fp, encoding="utf-8") as f:
                d = json.load(f)
            acts = d.get("actions", d) if isinstance(d, dict) else d
            self.canvas.set_actions(acts)
            self.flow_status.setText(f"불러오기 완료: {len(acts)}개 프로세스")
            self.ai_result.setPlainText(
                f"'{Path(fp).stem}' 플로우를 불러왔습니다.\n"
                f"총 {len(acts)}개 블록.\n\n"
                "프로세스 점검을 실행해보세요.")
        except Exception as e:
            msg_error(self, "오류", str(e))

    def _save(self):
        acts = self.canvas.get_actions()
        if not acts:
            msg_warn(self, "저장 실패", "플로우가 없습니다."); return
        fp, _ = rpa_save_file(self, "플로우 저장", SAVE_DIR)
        if not fp: return
        # 같은 파일 이미 존재하면 덮어쓰기 확인
        if Path(fp).exists():
            if not msg_ask(self, "덮어쓰기 확인",
                           f"'{Path(fp).name}' 이 이미 있습니다.\n덮어쓸까요?"):
                return
        data = {
            "name":    Path(fp).stem,
            "created": datetime.datetime.now().isoformat(),
            "count":   len(acts),
            "actions": acts,
        }
        with open(fp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self.flow_status.setText(f"저장 완료: {Path(fp).name}")

    # ── 재생 ──
    def _play(self):
        acts = self.canvas.get_actions()
        if not acts:
            msg_warn(self, "재생 실패", "프로세스가 없습니다."); return
        speed_map = {"0.5x":0.5,"1.0x":1.0,"1.5x":1.5,"2.0x":2.0}
        speed = speed_map.get(self.speed_combo.currentText(), 1.0)

        self.canvas.clear_highlights()
        self._worker = FlowPlayWorker(acts, speed=speed,
                                       minimize_others=self.chk_min_others.isChecked())
        self._worker.block_started.connect(self.canvas.highlight_running)
        self._worker.block_done.connect(self.canvas.set_block_done_proxy
                                        if hasattr(self.canvas,'set_block_done_proxy')
                                        else lambda i: None)
        self._worker.status_update.connect(self.flow_status.setText)
        self._worker.finished_play.connect(self._on_play_done)

        def _go():
            # RPA 자체 창 최소화 (옵션) - 메인스레드에서 직접 처리
            # _minimize_all_windows 는 RPA 창을 제외하므로 여기서 별도 처리
            if self.chk_min_self.isChecked():
                w = self.window()
                if w: w.showMinimized()
            self._worker.start()
            self.btn_play.setEnabled(False)
            self.btn_stop_play.setEnabled(True)

        self.flow_status.setText("5초 후 재생 시작... 대상 창으로 이동하세요!")
        # 카운트다운 플로팅 창
        self._flow_countdown = _RecCountdown(
            seconds=5,
            hint_text="플로우 재생 준비 중... 대상 창으로 이동하세요")
        self._flow_countdown.show()
        self._flow_countdown.raise_()
        QTimer.singleShot(5000, _go)

    def _stop_play(self):
        if self._worker: self._worker.stop()
        self.btn_play.setEnabled(True)
        self.btn_stop_play.setEnabled(False)
        self.canvas.clear_highlights()
        # RPA 창 복원
        win = self.window()
        if win and win.isMinimized():
            win.showNormal()
            win.activateWindow()
            win.raise_()
        self.flow_status.setText("재생 중지됨")

    def _on_flow_esc(self):
        """ESC 키 -> 플로우 재생/스텝 중지"""
        if self._worker and self._worker.running:
            self._stop_play()
            self.flow_status.setText("ESC 로 중지됨")
        self._step_reset()

    def _step_reset(self):
        self._step_idx = -1
        self.btn_step.setText("STEP >")
        self.btn_step_back.setText("< BACK")
        self.canvas.clear_highlights()

    def _step_exec(self, idx: int):
        acts  = self.canvas.get_actions()
        total = len(acts)
        if not acts or not (0 <= idx < total): return
        act = acts[idx]

        # UI 업데이트
        self.canvas.clear_highlights()
        self.canvas.set_block_running(idx)
        self.flow_status.setText(
            f"[STEP {idx+1}/{total}]  "
            f"{act.get('label', act.get('type',''))}  "
            f"| STEP > 다음  /  < BACK 이전")
        self.btn_step.setText(
            f"STEP > ({idx+2}/{total})" if idx+1 < total else "STEP > (완료)")
        self.btn_step_back.setText(
            f"< BACK ({idx}/{total})" if idx > 0 else "< BACK (처음)")

        # 버튼 잠금 (실행 중 중복 클릭 방지)
        self.btn_step.setEnabled(False)
        self.btn_step_back.setEnabled(False)

        # 스레드 없이 메인스레드에서 직접 실행
        # (QThread 생명주기 문제를 원천 차단 - 가장 안정적인 방식)
        QApplication.processEvents()  # 하이라이트 먼저 화면에 그리기
        self._run_step_action(act)
        QTimer.singleShot(50, lambda: self._on_step_done(idx))

    def _run_step_action(self, act: dict):
        """단일 액션을 메인스레드에서 직접 실행 (스텝 모드 전용)"""
        t = act.get("type", "")
        try:
            time.sleep(max(act.get("delay", 0.3), 0.05))
            if t == "click":
                pyautogui.click(act.get("x", 0), act.get("y", 0),
                                button=act.get("button", "left"))
            elif t == "key":
                k = act.get("key") or ""
                if k: pyautogui.press(k.replace("Key.", ""))
            elif t == "type":
                txt = act.get("text") or ""
                if txt: _type_safe(txt, act.get("ime_state", "english"))
            elif t == "hotkey":
                keys = act.get("keys") or []
                if keys: pyautogui.hotkey(*keys)
            elif t == "wait":
                _safe_wait(float(act.get("seconds", 1)))
            elif t == "scroll":
                pyautogui.scroll(act.get("dy", 3),
                                 x=act.get("x", 0), y=act.get("y", 0))
            elif t == "move":
                pyautogui.moveTo(act.get("x", 0), act.get("y", 0),
                                 duration=0.2)
        except pyautogui.FailSafeException:
            self.flow_status.setText("FAILSAFE: 마우스 모서리 감지 - 중지")
        except Exception as e:
            self.flow_status.setText(f"[스텝오류] {e}")

    def _on_step_done(self, idx: int):
        """스텝 실행 완료 시 호출"""
        self.canvas.set_block_done(idx)
        # 버튼 잠금 해제
        self.btn_step.setEnabled(True)
        self.btn_step_back.setEnabled(True)
        acts  = self.canvas.get_actions()
        total = len(acts)
        self.flow_status.setText(
            f"[STEP {idx+1}/{total}] 완료  "
            f"| STEP > 다음  /  < BACK 이전")

    def _step_play(self):
        """정방향 한 스텝"""
        acts = self.canvas.get_actions()
        if not acts:
            self.flow_status.setText("실행할 플로우가 없습니다."); return
        if self._step_idx < 0:
            self._step_idx = 0
        elif self._step_idx < len(acts) - 1:
            self._step_idx += 1
        else:
            self.flow_status.setText(
                f"모든 스텝 완료! ({len(acts)}개) -- STEP > 로 처음부터")
            self._step_reset()
            return
        self._step_exec(self._step_idx)

    def _step_back(self):
        """역방향 한 스텝"""
        acts = self.canvas.get_actions()
        if not acts: return
        if self._step_idx <= 0:
            self.flow_status.setText("첫 번째 스텝입니다.")
            self._step_idx = 0
            return
        self._step_idx -= 1
        self._step_exec(self._step_idx)

    def _on_play_done(self):
        self.btn_play.setEnabled(True)
        self.btn_stop_play.setEnabled(False)
        # RPA 창 복원 (재생 전 최소화했다면)
        win = self.window()
        if win and win.isMinimized():
            win.showNormal()
            win.activateWindow()
            win.raise_()
        acts = len(self.canvas.get_actions())
        LogEngine.add("프로세스", "플로우 에디터", "성공",
                      f"{acts}개 프로세스 실행 완료")
        # 트레이 알림
        if win and hasattr(win, '_tray'):
            win._tray.show_message(
                "5MRPA - 프로세스 완료",
                f"{acts}개 프로세스 실행이 완료되었습니다.",
                duration_ms=3000)

    def _on_canvas_edited(self):
        n = len(self.canvas.get_actions())
        self.flow_status.setText(f"편집됨 - 총 {n}개 프로세스")

    # ── AI 기능 ──
    def _focus_ai_input(self):
        """팔레트 AI 버튼 → AI 패널 입력창으로 포커스"""
        self.nl_input.setFocus()
        self.nl_input.selectAll()
        # AI 패널 강조 표시
        self.ai_result.setPlainText(
            "아래 입력창에 원하는 동작을 입력하고\n"
            "[플로우 자동 생성] 버튼을 누르세요.\n\n"
            "예시:\n"
            "  메모장 열고 안녕하세요 입력\n"
            "  엑셀 저장하고 닫기\n"
            "  ctrl+c 복사 후 메모장에 붙여넣기")
        self._insert_at = None   # 맨 끝에 추가

    def _on_ai_insert_req(self, at_index: int):
        """캔버스 [AI 생성] 버튼 → 특정 위치에 삽입 예약"""
        self._insert_at = at_index
        self.nl_input.setFocus()
        self.nl_input.selectAll()
        n = len(self.canvas.get_actions())
        pos_str = (f"{at_index}번 프로세스 다음" if at_index < n
                   else "맨 끝")
        self.ai_result.setPlainText(
            f"삽입 위치: {pos_str}\n\n"
            "원하는 동작을 입력하고\n"
            "[플로우 자동 생성] 버튼을 누르세요.")
        self.flow_status.setText(f"AI 생성 위치: {pos_str}")

    def _set_ai_loading(self, msg: str):
        self.ai_result.setPlainText(msg)
        self.flow_status.setText("AI 처리 중...")

    def _nl_to_blocks(self):
        text = self.nl_input.toPlainText().strip()
        if not text:
            msg_info(self, "입력 없음", "자연어 요청을 입력하세요."); return

        self._set_ai_loading(
            f"AI 프로세스 자동 생성 중...\n"
            f"'{text[:30]}{'...' if len(text)>30 else ''}'\n\n"
            f"AI 모델: {OLLAMA_MODEL}\n"
            "응답 시간: 모델/PC 성능에 따라 5~40초 소요될 수 있습니다\n"
            "(응답 없으면 키워드 자동 생성으로 대체)")
        self.ai_result.setStyleSheet(f"""
            QTextEdit {{
                background:#FFFBEB; border:1.5px solid {C['stop']};
                border-radius:10px; color:{C['t1']}; padding:10px; font-size:12px;
            }}
        """)

        # 버튼 비활성화 + 경과 시간 카운터
        sender_btn = self.sender()
        if sender_btn: sender_btn.setEnabled(False)
        self._nl_elapsed = 0

        def _tick():
            self._nl_elapsed += 1
            self.flow_status.setText(
                f"AI 처리 중... {self._nl_elapsed}초 경과")
        self._nl_timer = QTimer()
        self._nl_timer.timeout.connect(_tick)
        self._nl_timer.start(1000)

        def _w():
            try:
                blocks = GemmaEngine.natural_to_blocks(text)
            except Exception as e:
                blocks = []
                QTimer.singleShot(0, lambda err=e: (
                    self.ai_result.setPlainText(f"AI 호출 오류:\n{err}"),
                    self.flow_status.setText("AI 오류 발생")))
                if hasattr(self, '_nl_timer'): self._nl_timer.stop()
                if sender_btn: QTimer.singleShot(0, lambda: sender_btn.setEnabled(True))
                return

            def _done():
                # 타이머 정리
                if hasattr(self, '_nl_timer'):
                    self._nl_timer.stop()
                if sender_btn: sender_btn.setEnabled(True)

                if blocks:
                    try:
                        acts = self.canvas.get_actions()
                        # 삽입 위치 처리
                        insert_at = getattr(self, '_insert_at', None)
                        if insert_at is not None and 0 <= insert_at <= len(acts):
                            for j, b in enumerate(blocks):
                                acts.insert(insert_at + j, b)
                        else:
                            acts.extend(blocks)   # 기본: 맨 끝
                        self._insert_at = None   # 초기화
                        self.canvas.set_actions(acts)
                    except Exception as e:
                        self.ai_result.setPlainText(f"캔버스 오류:\n{e}")
                        self.flow_status.setText("오류 발생")
                        return
                    has_ai_result = any(
                        b.get("type") in ("click","scroll","image_click")
                        or len(b.get("text","")) > 2
                        for b in blocks)
                    note = "" if has_ai_result else (
                        "\n\n[규칙 기반 생성] AI 응답 파싱 실패로 키워드 자동 생성.\n"
                        "프로세스를 확인하고 필요하면 수정하세요.")
                    self.ai_result.setStyleSheet(f"""
                        QTextEdit {{
                            background:#FAFCFF; border:1.5px solid {C['border']};
                            border-radius:10px; color:{C['t1']}; padding:10px; font-size:12px;
                        }}
                    """)
                    self.ai_result.setPlainText(
                        f"총 {len(blocks)}개 플로우를 생성했습니다!\n\n"
                        + "\n".join(
                            f"{i+1}. {b.get('type','')} - "
                            f"{b.get('label','') or b.get('key','') or b.get('text','') or str(b.get('keys','')) or ''}"
                            for i, b in enumerate(blocks))
                        + "\n\n프로세스를 클릭해서 세부 내용을 확인하고 수정하세요."
                        + note)
                    self.flow_status.setText(
                        f"프로세스 자동 생성 완료 - {len(blocks)}개 플로우")
                else:
                    self.ai_result.setStyleSheet(f"""
                        QTextEdit {{
                            background:#FFF1F2; border:1.5px solid {C['rec']};
                            border-radius:10px; color:{C['t1']}; padding:10px; font-size:12px;
                        }}
                    """)
                    self.ai_result.setPlainText(
                        "프로세스를 생성하지 못했습니다.\n\n"
                        "원인:\n"
                        f"  1. AI({OLLAMA_MODEL}) 응답 시간 초과\n"
                        "  2. AI 응답 형식 오류\n\n"
                        "해결:\n"
                        "  - 더 짧고 구체적으로 입력해보세요\n"
                        "  - 예: '메모장 열기' (한 가지씩)\n"
                        "  - ollama serve 실행 상태 확인\n"
                        "  - 잠시 후 다시 시도")
                    self.flow_status.setText("생성 실패 - 다시 시도해주세요")
            QTimer.singleShot(0, _done)
        threading.Thread(target=_w, daemon=True).start()

    def _ai_optimize(self):
        acts = self.canvas.get_actions()
        if not acts:
            msg_info(self, "프로세스 없음", "프로세스를 먼저 만들어주세요."); return
        self._set_ai_loading("의미없는 대기시간/중복 동작을 분석 중입니다...\n잠시 기다려주세요.")
        def _w():
            result = GemmaEngine.optimize_flow(acts)
            def _done():
                self.canvas.set_actions(result["actions"])
                self.ai_result.setStyleSheet(f"""
                    QTextEdit {{
                        background:#FAFCFF; border:1.5px solid {C['border']};
                        border-radius:10px; color:{C['t1']}; padding:10px; font-size:12px;
                    }}
                """)
                self.ai_result.setPlainText(
                    "AI 프로세스 최적화 완료\n"
                    + "-"*30 + "\n"
                    + result["report"]
                    + f"\n\n총 단계: {len(acts)} -> {len(result['actions'])}개")
                self.flow_status.setText(
                    f"최적화 완료 - 약 {result['saved_seconds']:.1f}초 단축, "
                    f"{result['removed']}건 중복 제거")
            QTimer.singleShot(0, _done)
        threading.Thread(target=_w, daemon=True).start()

    def _process_check(self):
        """
        프로세스 점검 (구 '오류 자동 점검' + 'AI 프로세스 검토' 통합).
        1) 규칙 기반 오류 검사 - 즉시·항상 정확 (AI 불필요)
        2) AI 는 최대 1회만 호출:
           - 오류가 있으면 → 원인/해결법 해설
           - 오류가 없으면 → 개선 제안 (구 '검토' 기능 대체)
        """
        acts = self.canvas.get_actions()
        if not acts:
            msg_info(self, "프로세스 없음", "프로세스를 먼저 만들어주세요."); return

        errors        = []
        error_indices = []
        for i, a in enumerate(acts):
            t = a.get("type","")
            if t == "click":
                if a.get("x",0) == 0 and a.get("y",0) == 0:
                    errors.append(f"블록 {i+1}: 클릭 좌표가 (0,0)입니다. 실제 좌표를 설정하세요.")
                    error_indices.append(i)
            elif t == "type":
                if not a.get("text","").strip():
                    errors.append(f"블록 {i+1}: 텍스트 입력 내용이 비어있습니다.")
                    error_indices.append(i)
            elif t == "image_click":
                img_path = a.get("image_path","")
                if not img_path or not Path(img_path).exists():
                    errors.append(f"블록 {i+1}: 이미지 파일이 없습니다. ({img_path or '경로 없음'})")
                    error_indices.append(i)
            elif t == "hotkey":
                if not a.get("keys"):
                    errors.append(f"블록 {i+1}: 단축키가 설정되지 않았습니다.")
                    error_indices.append(i)
            elif t == "cond_if":
                if not a.get("cond_img"):
                    errors.append(f"블록 {i+1}: 조건 분기 이미지가 설정되지 않았습니다.")
                    error_indices.append(i)
                elif not Path(a.get("cond_img","")).exists():
                    errors.append(f"블록 {i+1}: 조건 이미지 파일이 없습니다.")
                    error_indices.append(i)
            elif t == "wait":
                if a.get("seconds",1) < 0.1:
                    errors.append(f"블록 {i+1}: 대기 시간이 너무 짧습니다 (0.1초 이상 권장).")
                    error_indices.append(i)

            if i > 0 and a.get("delay",0.5) < 0.05:
                prev = acts[i-1]
                if prev.get("type") in ["click","key"]:
                    errors.append(f"블록 {i+1}: 이전 동작과 간격이 너무 짧습니다 (delay {a['delay']}초).")

        self.canvas.mark_errors(error_indices)

        if errors:
            self._set_ai_loading(f"{len(errors)}개 오류 발견. AI가 해결 방법을 설명 중...")
            def _w():
                explanation = GemmaEngine.explain_error(errors)
                def _done():
                    self.ai_result.setStyleSheet(f"""
                        QTextEdit {{
                            background:#FAFCFF; border:1.5px solid {C['border']};
                            border-radius:10px; color:{C['t1']}; padding:10px; font-size:12px;
                        }}
                    """)
                    self.ai_result.setPlainText(
                        f"오류 {len(errors)}개 발견\n"
                        + "-"*30 + "\n"
                        + "\n".join(errors)
                        + "\n\nAI 해설:\n"
                        + explanation)
                    self.flow_status.setText(f"오류 {len(errors)}개 - 빨간 프로세스 확인")
                QTimer.singleShot(0, _done)
            threading.Thread(target=_w, daemon=True).start()
        else:
            # 오류 없음 → AI 로 개선 제안 1회 (구 'AI 프로세스 검토' 대체)
            self._set_ai_loading("오류 없음. AI 가 개선점을 검토 중입니다...")
            def _w():
                review = GemmaEngine.review_flow(acts)
                def _done():
                    self.ai_result.setStyleSheet(f"""
                        QTextEdit {{
                            background:#FAFCFF; border:1.5px solid {C['border']};
                            border-radius:10px; color:{C['t1']}; padding:10px; font-size:12px;
                        }}
                    """)
                    self.ai_result.setPlainText(
                        "오류 없음 - 정상\n"
                        + "-"*30 + "\n"
                        + "AI 검토 의견:\n"
                        + review)
                    self.flow_status.setText("프로세스 점검 완료 - 오류 없음")
                QTimer.singleShot(0, _done)
            threading.Thread(target=_w, daemon=True).start()


# =============================================
#  워커 스레드 (녹화/재생)
# =============================================
class RecordWorker(QThread):
    action_recorded = pyqtSignal(dict)
    status_update   = pyqtSignal(str)

    def __init__(self, smart_capture: bool = True):
        super().__init__()
        self.recording = False
        self.actions   = []
        self.last_time = None
        self._ml = None; self._kl = None
        self._lock = threading.Lock()      # last_time/actions 동시접근 보호
        self._stopping = False             # stop() 이 명시적으로 호출됐는지 여부
        self._listener_error = None        # 리스너 콜백에서 잡힌 예외 보관
        self._cached_ime = "english"       # IME 상태 캐시(후크스레드에서 ctypes 직접호출 금지)
        self.smart_capture = smart_capture # 클릭 위치 자동 캡처(스마트 클릭) 활성화 여부

    def run(self):
        if not PYNPUT_OK:
            self.status_update.emit("err:pynput"); return

        self.recording = True
        self.last_time = time.time()
        try:
            self._cached_ime = _get_ime_state()  # 워커 스레드에서 1회만 호출(안전)
        except Exception:
            self._cached_ime = "english"
        self.status_update.emit("recording")

        # ── 콜백 내부 예외가 절대 리스너/스레드를 죽이지 않도록
        #    모든 콜백을 try/except 로 감싼다.
        #    pynput 은 콜백에서 예외가 발생하면 리스너를 조용히 중단시키고,
        #    이는 마우스/키보드 입력이 갑자기 먹통이 되거나
        #    앱 전체가 응답불가(강제종료처럼 보임) 상태가 되는 핵심 원인이다.

        def _record_action(act):
            """공통 액션 기록 + emit. 예외가 나도 절대 위로 전파하지 않음."""
            try:
                with self._lock:
                    self.actions.append(act)
                self.action_recorded.emit(act)
            except Exception as e:
                self._listener_error = e

        def on_click(x, y, button, pressed):
            try:
                if not self.recording:
                    return False
                if pressed:
                    now = time.time()
                    with self._lock:
                        delay = round(now - (self.last_time or now), 3)
                        self.last_time = now
                    # 물리 픽셀 → 논리 픽셀 변환 (DPI 스케일 보정)
                    lx, ly = _phys_to_logical(x, y)
                    act = {"type":"click","x":lx,"y":ly,
                           "button":"right" if button==Button.right else "left",
                           "delay":delay}
                    _record_action(act)
                    # 스마트 클릭: 클릭 위치 주변 자동 캡처
                    # 후크 콜백을 막지 않도록 백그라운드 스레드로 위임
                    # (캡처는 100ms 정도 걸릴 수 있어 콜백에서 직접 하면
                    #  pynput 후크 타임아웃 위험)
                    if self.smart_capture:
                        def _snap(_x=x, _y=y, _act=act):
                            try:
                                p = _capture_snapshot(_x, _y, size=80)
                                if p:
                                    _act["snapshot"] = p
                            except Exception:
                                pass
                        threading.Thread(target=_snap, daemon=True).start()
            except Exception as e:
                # 콜백 내부 예외는 절대 리스너를 죽이지 않고 계속 진행
                self._listener_error = e
            return True

        def on_scroll(x, y, dx, dy):
            try:
                if not self.recording:
                    return False
                now = time.time()
                with self._lock:
                    delay = round(now - (self.last_time or now), 3)
                    self.last_time = now
                lx, ly = _phys_to_logical(x, y)
                act = {"type":"scroll","x":lx,"y":ly,"dy":dy,"delay":delay}
                _record_action(act)
            except Exception as e:
                self._listener_error = e
            return True

        def on_key_press(key):
            try:
                if not self.recording:
                    return False
                try:
                    k = key.char
                    if k is None: k = str(key)
                except Exception:
                    k = str(key)

                if k == "Key.esc":
                    self.recording = False
                    return False

                # 한글 입력 감지 - 연속 한글은 type 으로 묶기
                if k and len(k) == 1 and ord(k) > 127:
                    now = time.time()
                    with self._lock:
                        delay = round(now - (self.last_time or now), 3)
                        self.last_time = now
                    ime = self._cached_ime  # ctypes 직접호출 금지(후크스레드 크래시 원인)
                    with self._lock:
                        merge = (self.actions and
                                 self.actions[-1].get("type") == "type" and
                                 delay < 1.2)
                        if merge:
                            self.actions[-1]["text"] += k
                            last_act = self.actions[-1]
                        else:
                            last_act = {"type":"type","text":k,
                                        "ime_state": ime, "delay":delay}
                            self.actions.append(last_act)
                    self.action_recorded.emit(last_act)
                    return True

                # 한/영 키 처리 (가상키 코드는 PC/키보드에 따라 다름)
                # <21> = VK_HANGUL (0x15), <25> = VK_HANJA (0x19), <179> = 미디어키 매핑
                k_str = str(key)
                if "hangul" in k_str.lower() or k_str in (
                    "Key.hangul", "Key.hangul_hanja",
                    "<21>", "<25>", "<179>", "Key.media_volume_mute"):
                    now = time.time()
                    with self._lock:
                        delay = round(now - (self.last_time or now), 3)
                        self.last_time = now
                    act = {"type":"hotkey","keys":["hangul"],"delay":delay}
                    _record_action(act)
                    # 한/영 전환키 입력시 캐시 토글 (실제 OS 호출 없이 추정)
                    self._cached_ime = "korean" if self._cached_ime == "english" else "english"
                    return True

                now = time.time()
                with self._lock:
                    delay = round(now - (self.last_time or now), 3)
                    self.last_time = now

                # 단일 영문자면 type 으로 묶기
                if k and len(k) == 1 and ord(k) < 128 and k.isprintable():
                    # 현재 IME 캐시 상태에 따라 ime_state 결정
                    # (한/영 키 토글로 추적된 self._cached_ime 사용)
                    cur_ime = self._cached_ime
                    with self._lock:
                        # 같은 ime 상태인 직전 type 액션과 묶음 (delay 1.2초 이내)
                        merge = (self.actions and
                                 self.actions[-1].get("type") == "type" and
                                 self.actions[-1].get("ime_state") == cur_ime and
                                 delay < 1.2)
                        if merge:
                            # 영문 raw 키 누적 (한글이면 마지막에 두벌식 변환)
                            self.actions[-1]["_raw"] = self.actions[-1].get("_raw","") + k
                            if cur_ime == "korean":
                                # 누적된 raw 키 전체를 한글로 재변환 (조합형이라 마지막 글자가 합쳐질 수 있음)
                                self.actions[-1]["text"] = _eng_to_kor(
                                    self.actions[-1]["_raw"])
                            else:
                                self.actions[-1]["text"] = self.actions[-1]["_raw"]
                            last_act = self.actions[-1]
                        else:
                            if cur_ime == "korean":
                                last_act = {"type":"type",
                                            "text": _eng_to_kor(k),
                                            "_raw": k,
                                            "ime_state":"korean",
                                            "delay":delay}
                            else:
                                last_act = {"type":"type","text":k,"_raw":k,
                                            "ime_state":"english","delay":delay}
                            self.actions.append(last_act)
                    self.action_recorded.emit(last_act)
                else:
                    act = {"type":"key","key":k,"delay":delay}
                    _record_action(act)
            except Exception as e:
                # 키 콜백에서 어떤 예외가 나도 리스너를 살려둔다
                self._listener_error = e
            return True

        try:
            from pynput import mouse as _m, keyboard as _k
            with _m.Listener(on_click=on_click, on_scroll=on_scroll) as ml, \
                 _k.Listener(on_press=on_key_press) as kl:
                self._ml = ml; self._kl = kl
                ml.join()
                kl.stop()  # 마우스 리스너가 끝나면 키보드 리스너도 함께 정리
        except Exception as e:
            # 리스너 생성/실행 자체가 실패한 경우
            # (예: 보안 소프트웨어가 글로벌 후크를 차단하는 경우 등)
            self.recording = False
            self.status_update.emit(f"err:listener:{e}")
            return

        self.recording = False

        if self._listener_error is not None:
            # 콜백 도중 예외가 있었지만 녹화 자체는 끝까지 진행된 경우 -> 경고만 표시
            self.status_update.emit(
                f"warn:{len(self.actions)}:{self._listener_error}")
        else:
            self.status_update.emit(f"done:{len(self.actions)}")

    def stop(self):
        """메인 스레드에서 안전하게 녹화 중지를 요청한다."""
        self._stopping = True
        self.recording = False
        try:
            if self._ml: self._ml.stop()
        except Exception:
            pass
        try:
            if self._kl: self._kl.stop()
        except Exception:
            pass


class PlayWorker(QThread):
    progress_update = pyqtSignal(int, int)
    status_update   = pyqtSignal(str)
    finished_play   = pyqtSignal()

    def __init__(self, actions, repeat=1, speed=1.0, minimize_others=False):
        super().__init__()
        self.actions         = actions
        self.repeat          = repeat
        self.speed           = speed
        self.minimize_others = minimize_others
        self.running         = False

    def run(self):
        self.running = True
        # 글로벌 ESC 감시 시작 (창 최소화 상태에서도 ESC 로 중단 가능)
        _esc_watcher = _GlobalEscWatcher(on_esc=self.stop)
        _esc_watcher.start()
        # 재생 시작 전 다른 창 최소화 (옵션) - 좌표 클릭이 엉뚱한 창에 가는 것 방지
        if self.minimize_others:
            self.status_update.emit("다른 창 최소화 중...")
            _minimize_all_windows()
            time.sleep(0.3)
        # ── 재생 전 최적화: 연속 type 액션을 같은 ime_state 기준으로 병합
        #    녹화 시 천천히 타이핑하면 글자마다 별도 청크가 생기는데,
        #    재생 시 매 청크마다 PowerShell(클립보드)+IME전환이 호출되면
        #    느리고 불안정 → 미리 합쳐서 한 번에 붙여넣기
        merged = []
        for act in self.actions:
            if (act.get("type") == "type" and merged and
                    merged[-1].get("type") == "type" and
                    merged[-1].get("ime_state") == act.get("ime_state", "english")):
                merged[-1]["text"] = merged[-1].get("text","") + act.get("text","")
                # delay 는 첫 청크의 값만 유지 (이후 청크의 delay 는 타이핑 간격이라 무의미)
            else:
                import copy
                merged.append(copy.deepcopy(act))
        self.actions = merged
        total        = len(self.actions)
        sig_interval = max(1, min(10, total//20 or 1))
        for r in range(self.repeat):
            if not self.running: break
            self.status_update.emit(f"playing:{r+1}:{self.repeat}")
            for i, act in enumerate(self.actions):
                if not self.running: break
                # act 가 dict 가 아니면 스킵
                if not isinstance(act, dict): continue
                try:
                    # 액션 타입별 최소 delay 보장
                    # - 너무 빠른 녹화에서도 재생이 안정되게 자동 보정
                    # - speed 배율과 무관하게 floor 적용 (배율 가속으로도 깨지지 않음)
                    t_check = act.get("type","")
                    raw_delay = act.get("delay", 0.1)
                    if t_check == "type":
                        # 텍스트 입력 직전: 이전 동작(클릭 등) 안정화 시간 필요
                        min_delay = 0.15
                    elif t_check == "click":
                        # 클릭 직전: 이전 동작 완료 보장
                        min_delay = 0.08
                    elif t_check == "hotkey":
                        min_delay = 0.1
                    else:
                        min_delay = 0.03
                    scaled = raw_delay / self.speed
                    time.sleep(max(scaled, min_delay))
                except Exception:
                    time.sleep(0.1)
                try:
                    t = act.get("type","")
                    if t == "click":
                        x = act.get("x",0); y = act.get("y",0)
                        _smart_click(x, y,
                                     snapshot_path=act.get("snapshot",""),
                                     button=act.get("button","left"),
                                     confidence=act.get("confidence",0.85))
                    elif t == "scroll":
                        sx, sy = _logical_to_phys(act.get("x",0), act.get("y",0))
                        pyautogui.scroll(act.get("dy",1), x=sx, y=sy)
                    elif t == "key":
                        k = act.get("key") or ""
                        if k:
                            if k.startswith("Key."):
                                pyautogui.press(k.replace("Key.",""))
                            else:
                                pyautogui.press(k)
                    elif t == "type":
                        txt = act.get("text") or ""
                        if txt:
                            ime = act.get("ime_state", "english")
                            _type_safe(txt, ime)
                    elif t == "hotkey":
                        keys = act.get("keys") or []
                        if keys:
                            if keys == ["hangul"]:
                                pass
                            else:
                                pyautogui.hotkey(*keys)
                    elif t == "move":
                        mx, my = _logical_to_phys(act.get("x",0), act.get("y",0))
                        pyautogui.moveTo(mx, my, duration=0.2)
                    elif t == "wait":
                        _safe_wait(float(act.get("seconds",1)),
                                   running_check=lambda: self.running)
                    elif t == "image_click" and CV2_OK:
                        img     = act.get("image_path","")
                        conf    = act.get("confidence", 0.85)
                        timeout = act.get("wait_timeout", 0)
                        if img and Path(img).exists():
                            loc = None
                            if timeout > 0:
                                # 나타날 때까지 최대 timeout 초 대기
                                import time as _t
                                elapsed = 0
                                while elapsed < timeout and self.running:
                                    try:
                                        loc = pyautogui.locateOnScreen(
                                            img, confidence=conf)
                                        if loc: break
                                    except Exception:
                                        pass
                                    _t.sleep(0.5)
                                    elapsed += 0.5
                                if not loc:
                                    self.status_update.emit(
                                        f"[대기 타임아웃] {timeout}초 내 이미지 미발견: "
                                        f"{act.get('label','')}")
                            else:
                                try:
                                    loc = pyautogui.locateOnScreen(
                                        img, confidence=conf)
                                except Exception:
                                    loc = None
                            if loc:
                                pyautogui.click(
                                    int(loc.left+loc.width/2),
                                    int(loc.top+loc.height/2))
                    elif t == "cond_if":
                        cond_type = act.get("cond_type","image_exist")
                        cond_img  = act.get("cond_img","")
                        found     = False
                        if CV2_OK and cond_img and Path(cond_img).exists():
                            try:
                                loc   = pyautogui.locateOnScreen(
                                    cond_img,
                                    confidence=act.get("cond_conf",0.80))
                                found = loc is not None
                            except Exception:
                                found = False
                        is_yes     = (found if cond_type=="image_exist"
                                      else not found)
                        macro_path = (act.get("yes_macro","") if is_yes
                                      else act.get("no_macro",""))
                        if macro_path and Path(macro_path).exists():
                            try:
                                with open(macro_path,encoding="utf-8") as f:
                                    d = json.load(f)
                                branch_acts = (d.get("actions",[])
                                    if isinstance(d,dict) else d)
                                for ba in branch_acts:
                                    if not self.running: break
                                    if not isinstance(ba, dict): continue
                                    time.sleep(max(
                                        ba.get("delay",0.3)/self.speed,0.01))
                                    bt = ba.get("type","")
                                    if bt=="click":
                                        pyautogui.click(
                                            ba.get("x",0), ba.get("y",0),
                                            button=ba.get("button","left"))
                                    elif bt=="key":
                                        k2 = ba.get("key") or ""
                                        if k2: pyautogui.press(k2)
                                    elif bt=="type":
                                        tx = ba.get("text") or ""
                                        if tx: _type_safe(tx, ba.get("ime_state","english"))
                                    elif bt=="hotkey":
                                        ks = ba.get("keys") or []
                                        if ks: pyautogui.hotkey(*ks)
                                    elif bt=="wait":
                                        _safe_wait(float(ba.get("seconds",1)), running_check=lambda: self.running)
                            except Exception as e:
                                self.status_update.emit(f"[분기오류] {e}")
                        self.status_update.emit(
                            f"조건분기: {'YES' if is_yes else 'NO'}")
                except pyautogui.FailSafeException:
                    # 마우스 모서리 FAILSAFE - 재생 중지
                    self.status_update.emit("FAILSAFE: 마우스 모서리 감지 - 중지")
                    self.running = False
                    break
                except Exception as e:
                    self.status_update.emit(f"err:{e}")
                if (i+1) % sig_interval == 0 or (i+1) == total:
                    self.progress_update.emit(i+1, total)
        self.running = False
        _esc_watcher.stop()
        self.status_update.emit("done")
        self.finished_play.emit()

    def stop(self): self.running = False


# =============================================
#  StatCard
# =============================================
class StatCard(QFrame):
    def __init__(self, title, value="0", accent=None, parent=None):
        super().__init__(parent)
        self._accent = accent or C['brand']
        self.setMinimumHeight(72)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14,10,14,10)
        lay.setSpacing(2)
        self.val_lbl = QLabel(value)
        self.val_lbl.setStyleSheet(
            f"color:{self._accent}; font-size:20px; font-weight:900; background:transparent;")
        self.ttl_lbl = QLabel(title)
        self.ttl_lbl.setStyleSheet(
            f"color:{C['t2']}; font-size:11px; background:transparent;")
        lay.addWidget(self.val_lbl)
        lay.addWidget(self.ttl_lbl)

    def setValue(self, v): self.val_lbl.setText(str(v))

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        r  = self.rect().adjusted(0,0,-1,-1)
        ac = QColor(self._accent)
        painter.setPen(QPen(QColor(C['border']),1))
        painter.setBrush(QBrush(QColor(C['bg1'])))
        painter.drawRoundedRect(r,12,12)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(ac))
        painter.drawRoundedRect(QRect(0, 8, 4, r.height()-16), 2, 2)
        painter.end()
        super().paintEvent(event)



# =============================================
#  녹화 시작 카운트다운 플로팅 창
# =============================================
class _RecCountdown(QWidget):
    """
    3초 카운트다운 플로팅 창.
    녹화/재생/플로우 시작 전 공용.
    """
    def __init__(self, parent=None, hint_text="녹화 준비 중... 대상 창으로 이동하세요", seconds=3):
        super().__init__(parent)
        self._total = seconds
        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint |
            Qt.WindowType.Tool)
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating)
        self.setFixedSize(260, 160)

        screen = QApplication.primaryScreen().geometry()
        self.move(
            screen.center().x() - 130,
            screen.center().y() - 80)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(6)

        hint = QLabel(hint_text)
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hint.setStyleSheet(
            "color:#FFFFFF; font-size:11px; font-weight:500;"
            "background:transparent;")
        hint.setWordWrap(True)
        lay.addWidget(hint)

        # 카운트 숫자 (초기값을 self._total 로 - 하드코딩 시 첫 숫자가 어긋남)
        self.count_lbl = QLabel(str(seconds))
        self.count_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.count_lbl.setStyleSheet(
            "color:#FFFFFF; font-size:64px; font-weight:900;"
            "background:transparent;")
        lay.addWidget(self.count_lbl)

        # 하단 텍스트
        self.sub_lbl = QLabel("초 후 녹화 시작")
        self.sub_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.sub_lbl.setStyleSheet(
            "color:rgba(255,255,255,0.8); font-size:13px; font-weight:700;"
            "background:transparent;")
        lay.addWidget(self.sub_lbl)

        self._count = self._total
        self._timer = QTimer()
        self._timer.timeout.connect(self._tick)
        self._timer.start(1000)

    def _tick(self):
        self._count -= 1
        if self._count <= 0:
            self._timer.stop()
            self.count_lbl.setText("●")
            self.sub_lbl.setText("녹화 시작!")
            self.count_lbl.setStyleSheet(
                "color:#F43F5E; font-size:52px; font-weight:900;"
                "background:transparent;")
            QTimer.singleShot(600, self.close)
        else:
            self.count_lbl.setText(str(self._count))

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        # 반투명 다크 배경
        painter.setBrush(QBrush(QColor(15, 23, 42, 220)))
        painter.setPen(QPen(QColor("#F43F5E"), 2))
        painter.drawRoundedRect(self.rect().adjusted(1,1,-1,-1), 18, 18)
        painter.end()


# =============================================
#  RecordPage
# =============================================
class RecordPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.worker  = None
        self.actions = []
        self._cnt    = {"click":0,"scroll":0,"key":0}
        self.smart_capture = True   # 스마트 클릭(자동 이미지 캡처) 기본 ON
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24,20,24,20)
        root.setSpacing(14)
        root.addWidget(SectionHeader("RECORD  /  녹화"))

        # 상태 카드
        self.state_card = QFrame()
        self.state_card.setMinimumHeight(110)
        self.state_card.setStyleSheet(f"""
            QFrame {{ background:{C['bg1']}; border:1.5px solid {C['border']};
                     border-radius:14px; }}
        """)
        sc = QHBoxLayout(self.state_card)
        sc.setContentsMargins(20,0,20,0)
        self.state_icon = QLabel("REC")
        self.state_icon.setStyleSheet(
            f"color:{C['rec']}; font-size:24px; font-weight:900; background:transparent;")
        state_info = QVBoxLayout()
        self.state_title = QLabel("대기 중")
        self.state_title.setStyleSheet(
            f"color:{C['t1']}; font-size:16px; font-weight:900; background:transparent;")
        self.state_sub = QLabel("녹화 버튼을 누르면 마우스/키보드 동작이 기록됩니다")
        self.state_sub.setStyleSheet(
            f"color:{C['t2']}; font-size:12px; background:transparent;")
        self.state_sub.setWordWrap(True)
        state_info.addWidget(self.state_title)
        state_info.addWidget(self.state_sub)
        self.big_count = QLabel("0")
        self.big_count.setStyleSheet(
            f"color:{C['rec']}; font-size:38px; font-weight:900; background:transparent;")
        self.big_count.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sc.addWidget(self.state_icon)
        sc.addSpacing(16)
        sc.addLayout(state_info)
        sc.addStretch()
        sc.addWidget(self.big_count)
        root.addWidget(self.state_card)

        # 통계 카드
        stats = QHBoxLayout()
        self.card_click  = StatCard("클릭",    "0", C['brand'])
        self.card_scroll = StatCard("스크롤",  "0", C['tool'])
        self.card_key    = StatCard("키 입력", "0", C['stop'])
        self.card_time   = StatCard("경과",    "0s",C['t2'])
        for c in [self.card_click,self.card_scroll,self.card_key,self.card_time]:
            stats.addWidget(c)
        root.addLayout(stats)

        # 스마트 클릭 옵션
        self.chk_smart = QCheckBox(
            "스마트 클릭  (클릭 위치 자동 캡처 → 재생 시 이미지 인식 우선, 실패 시 좌표 폴백)")
        self.chk_smart.setChecked(True)
        self.chk_smart.setStyleSheet(f"color:{C['t2']}; padding:4px;")
        self.chk_smart.toggled.connect(
            lambda v: setattr(self, 'smart_capture', v))
        root.addWidget(self.chk_smart)

        # 버튼
        btn_row = QHBoxLayout(); btn_row.setSpacing(8)
        self.btn_rec = GlowButton("  [REC]  녹화 시작", C['rec'])
        self.btn_rec.setFixedHeight(48)
        self.btn_rec.setFont(QFont("Malgun Gothic",13,QFont.Weight.Bold))
        self.btn_rec.clicked.connect(self.start_recording)

        self.btn_stop = GlowButton("  [STOP]  중지", C['stop'])
        self.btn_stop.setFixedHeight(48); self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.stop_recording)

        self.btn_to_flow = GlowButton("  플로우 에디터로 보내기", C['ai'])
        self.btn_to_flow.setFixedHeight(48)
        self.btn_to_flow.setToolTip("녹화 결과를 프로세스 에디터에서 시각적으로 편집합니다")
        self.btn_to_flow.clicked.connect(self._send_to_flow)

        self.btn_save = GlowButton("  저장", C['play'])
        self.btn_save.setFixedHeight(48)
        self.btn_save.clicked.connect(self.save_macro)

        btn_row.addWidget(self.btn_rec,3)
        btn_row.addWidget(self.btn_stop,2)
        btn_row.addWidget(self.btn_to_flow,3)
        btn_row.addWidget(self.btn_save,2)
        root.addLayout(btn_row)

        # 이름 입력
        name_row = QHBoxLayout()
        name_row.addWidget(QLabel("저장 이름:"))
        self.name_edit = StyledInput("예: 로그인_자동화")
        name_row.addWidget(self.name_edit)
        root.addLayout(name_row)

        # 로그 헤더 + 삭제 버튼
        log_row = QHBoxLayout()
        log_lbl = QLabel("녹화 로그")
        log_lbl.setStyleSheet(f"color:{C['t2']}; font-size:11px; font-weight:bold;")
        log_row.addWidget(log_lbl)
        log_row.addStretch()
        btn_del_one = GlowButton("선택 삭제", C['stop'])
        btn_del_one.setFixedHeight(26); btn_del_one.setFixedWidth(80)
        btn_del_one.clicked.connect(self._del_selected)
        btn_del_all = GlowButton("전체 삭제", C['rec'])
        btn_del_all.setFixedHeight(26); btn_del_all.setFixedWidth(80)
        btn_del_all.clicked.connect(self._del_all)
        log_row.addWidget(btn_del_one)
        log_row.addWidget(btn_del_all)
        root.addLayout(log_row)
        self.action_list = QListWidget()
        self.action_list.setStyleSheet(f"""
            QListWidget {{ background:{C['bg1']}; border:1.5px solid {C['border']};
                border-radius:10px; padding:4px; outline:none; }}
            QListWidget::item {{ padding:6px 12px; border-radius:6px; margin:1px 2px;
                font-family:Consolas,monospace; font-size:12px; color:{C['t1']}; }}
            QListWidget::item:selected {{ background:{C['brand']}; color:#FFFFFF; }}
            QListWidget::item:hover {{ background:{C['bg4']}; }}
        """)
        root.addWidget(self.action_list)

        # 상태 표시 라벨
        self.status_lbl = QLabel("대기 중")
        self.status_lbl.setStyleSheet(f"""
            color:{C['t2']}; font-size:12px;
            background:{C['bg1']}; border:1px solid {C['border']};
            border-radius:8px; padding:6px 12px;
        """)
        root.addWidget(self.status_lbl)

        self._elapsed = 0
        self._elapsed_timer = QTimer()
        self._elapsed_timer.timeout.connect(
            lambda: (setattr(self,'_elapsed',self._elapsed+1),
                     self.card_time.setValue(f"{self._elapsed}s")))

    def _del_selected(self):
        row = self.action_list.currentRow()
        if row < 0: return
        self.action_list.takeItem(row)
        if row < len(self.actions):
            self.actions.pop(row)
        self.big_count.setText(str(len(self.actions)))

    def _del_all(self):
        self.actions.clear()
        self.action_list.clear()
        self._cnt = {"click":0,"scroll":0,"key":0,"type":0,"hotkey":0}
        self._update_stats()

    def start_recording(self):
        if not PYNPUT_OK:
            self.status_lbl.setText("패키지 없음 - pip install pynput"); return
        self.actions.clear(); self.action_list.clear()
        self._cnt = {"click":0,"scroll":0,"key":0,"type":0,"hotkey":0}
        self._elapsed = 0; self._update_stats()
        self.worker = RecordWorker(smart_capture=self.smart_capture)
        self.worker.action_recorded.connect(self._on_action)
        self.worker.status_update.connect(self._on_status)
        self.worker.finished.connect(self._on_finished)
        self.btn_rec.setEnabled(False); self.btn_stop.setEnabled(True)
        self._set_state("wait","5초 후 녹화 시작...","대상 창으로 이동하세요")

        # 5초 카운트다운 플로팅 창
        self._rec_countdown = _RecCountdown(seconds=5)
        self._rec_countdown.show()
        self._rec_countdown.raise_()

        QTimer.singleShot(1000, lambda: self._set_state("wait","4초...",""))
        QTimer.singleShot(2000, lambda: self._set_state("wait","3초...",""))
        QTimer.singleShot(3000, lambda: self._set_state("wait","2초...",""))
        QTimer.singleShot(4000, lambda: self._set_state("wait","1초...",""))
        QTimer.singleShot(5000, lambda: (
            self.worker.start(),
            self._elapsed_timer.start(1000),
            self._set_state("rec","녹화 중...","ESC 또는 [STOP] 으로 중지")))

    def stop_recording(self):
        if self.worker:
            self.worker.stop()
            self.worker.wait(2000)   # 리스너 정리 대기, 최대 2초
        self._elapsed_timer.stop()
        self.btn_rec.stop_pulse(); self.btn_rec.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self._set_state("idle","녹화 완료",f"총 {len(self.actions)}개 동작")

    def _on_action(self, act):
        # type 묶음(연속 타이핑)은 emit이 매 글자마다 와도 리스트엔 갱신만, 신규 항목은 1개만 추가
        is_update = (act.get("type") == "type" and
                     self.actions and self.actions[-1] is act)
        if not is_update:
            self.actions.append(act)
        t = act.get("type", "")
        if t in self._cnt: self._cnt[t] += 1
        n = len(self.actions)
        try:
            if t == "click":
                txt=f"  [클릭]   ({act['x']:4d},{act['y']:4d}) {act.get('button','left')} +{act.get('delay',0):.2f}s"
                c=C['brand']
            elif t == "scroll":
                txt=f"  [스크롤] ({act['x']:4d},{act['y']:4d}) dy={act.get('dy',0)} +{act.get('delay',0):.2f}s"
                c=C['tool']
            elif t == "type":
                txt=f"  [입력]   \"{act.get('text','')}\" +{act.get('delay',0):.2f}s"
                c=C['ok']
            elif t == "hotkey":
                keys = act.get("keys", [])
                txt=f"  [핫키]   [{'+'.join(keys)}] +{act.get('delay',0):.2f}s"
                c=C['stop']
            elif t == "key":
                txt=f"  [키]     [{act.get('key','?')}] +{act.get('delay',0):.2f}s"
                c=C['stop']
            else:
                txt=f"  [{t}]  +{act.get('delay',0):.2f}s"
                c=C['t2']
        except Exception:
            txt=f"  [{t}]"
            c=C['t2']

        if is_update and self.action_list.count() > 0:
            # 마지막 항목 텍스트만 갱신 (새 항목 추가 안 함)
            item = self.action_list.item(self.action_list.count()-1)
            item.setText(f"{n:03d} {txt}")
            item.setForeground(QColor(c))
        else:
            item = QListWidgetItem(f"{n:03d} {txt}")
            item.setForeground(QColor(c))
            self.action_list.addItem(item)

        if n % 50 == 0: self.action_list.scrollToBottom()
        self.big_count.setText(str(n))
        self._update_stats()

    def _on_status(self, msg):
        if msg=="recording":
            self.btn_rec.start_pulse()
        elif msg.startswith("done:"):
            self._set_state("idle","녹화 완료",f"총 {msg.split(':')[1]}개")
        elif msg.startswith("warn:"):
            parts = msg.split(":", 2)
            n = parts[1] if len(parts) > 1 else "?"
            self._set_state("idle","녹화 완료(경고)",f"총 {n}개 - 일부 입력 처리 중 오류 발생")
        elif msg.startswith("err:listener"):
            self._set_state("idle","녹화 실패","입력 후크 시작 불가 (보안 SW 차단 가능)")
            self.btn_rec.setEnabled(True); self.btn_stop.setEnabled(False)
        elif msg == "err:pynput":
            self._set_state("idle","녹화 불가","pynput 미설치")

    def _on_finished(self):
        self._elapsed_timer.stop()
        self.btn_rec.stop_pulse()
        self.btn_rec.setEnabled(True); self.btn_stop.setEnabled(False)

    def _set_state(self, mode, title, sub):
        colors = {"rec":(C['rec'],C['rec_dim'],C['rec']),
                  "wait":(C['stop'],C['stop_dim'],C['stop']),
                  "idle":(C['t2'],C['bg1'],C['border'])}
        ic, bg, bd = colors.get(mode, colors['idle'])
        self.state_card.setStyleSheet(
            f"QFrame{{background:{bg};border:1.5px solid {bd};border-radius:14px;}}")
        self.state_title.setText(title); self.state_sub.setText(sub)
        self.state_icon.setStyleSheet(
            f"color:{ic}; font-size:24px; font-weight:900; background:transparent;")

    def _update_stats(self):
        self.card_click.setValue(self._cnt["click"])
        self.card_scroll.setValue(self._cnt["scroll"])
        self.card_key.setValue(self._cnt["key"] + self._cnt.get("type",0) + self._cnt.get("hotkey",0))

    def _send_to_flow(self):
        """녹화된 동작을 프로세스 에디터로 전달"""
        if not self.actions:
            self.status_lbl.setText("먼저 녹화하세요."); return
        win = self.window()
        if hasattr(win, 'send_to_flow'):
            win.send_to_flow(list(self.actions))
        # QMessageBox 없이 상태 표시만 (창 숨김 방지)
        self.status_lbl.setText(
            f"프로세스 에디터로 {len(self.actions)}개 전송 완료 → [PROC] 탭으로 이동")

    def save_macro(self):
        if not self.actions:
            self.status_lbl.setText("녹화된 동작이 없습니다."); return

        # 파일명: 입력값 없으면 자동 생성, 특수문자 제거
        raw  = self.name_edit.text().strip()
        name = re.sub(r'[\\/:*?"<>|]', '_', raw) if raw else \
               f"macro_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        if not name:
            name = f"macro_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"

        # 파일명 입력창에 반영 (저장된 이름 표시)
        self.name_edit.setText(name)

        fpath = SAVE_DIR / f"{name}.json"
        # 같은 파일 이미 존재하면 덮어쓰기 확인
        if fpath.exists():
            if not msg_ask(self, "덮어쓰기 확인",
                           f"'{name}.json' 이 이미 있습니다.\n덮어쓸까요?"):
                return
        try:
            data = {"name": name,
                    "created": datetime.datetime.now().isoformat(),
                    "count": len(self.actions),
                    "actions": self.actions}
            with open(fpath, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            LogEngine.add("녹화", name, "성공",
                          f"{len(self.actions)}개 프로세스 저장")
            self.status_lbl.setText(f"저장 완료: {fpath.name}")
        except Exception as e:
            self.status_lbl.setText(f"저장 실패: {e}")


# =============================================
#  PlayPage (간소화)
# =============================================
class PlayPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.worker  = None
        self.actions = []
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24,20,24,20); root.setSpacing(14)
        root.addWidget(SectionHeader("PLAY  /  재생"))

        # 파일 로드 카드
        load_card = QFrame()
        load_card.setStyleSheet(f"""
            QFrame {{ background:{C['bg1']}; border:1.5px dashed {C['border']};
                     border-radius:14px; }}
        """)
        lc = QVBoxLayout(load_card); lc.setContentsMargins(20,14,20,14); lc.setSpacing(8)
        self.file_lbl = QLabel("매크로 파일을 선택하세요")
        self.file_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.file_lbl.setStyleSheet(f"color:{C['t2']}; font-size:14px; background:transparent;")
        self.file_meta = QLabel("")
        self.file_meta.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.file_meta.setStyleSheet(f"color:{C['t3']}; font-size:11px; background:transparent;")
        btn_load = GlowButton("  [OPEN]  파일 열기", C['brand'])
        btn_load.setFixedHeight(38); btn_load.clicked.connect(self.load_file)
        lc.addWidget(self.file_lbl); lc.addWidget(self.file_meta); lc.addWidget(btn_load)
        root.addWidget(load_card)

        # 설정
        cfg = QFrame()
        cfg.setStyleSheet(f"""
            QFrame {{ background:{C['bg1']}; border:1.5px solid {C['border']};
                     border-radius:14px; }}
        """)
        cg = QGridLayout(cfg); cg.setContentsMargins(20,14,20,14); cg.setSpacing(10)
        cg.addWidget(QLabel("반복 횟수"),0,0)
        self.repeat_spin = StyledSpin(); self.repeat_spin.setRange(1,9999)
        self.repeat_spin.setValue(1); self.repeat_spin.setSuffix("  회")
        cg.addWidget(self.repeat_spin,0,1)
        cg.addWidget(QLabel("재생 속도"),0,2)
        self.speed_combo = StyledCombo()
        self.speed_combo.addItems(["0.5x  (천천히)","1.0x  (보통)","1.5x  (빠르게)","2.0x  (최고속)"])
        self.speed_combo.setCurrentIndex(1)
        cg.addWidget(self.speed_combo,0,3)
        self.chk_cd = QCheckBox("시작 전 3초 카운트다운")
        self.chk_cd.setChecked(True); self.chk_cd.setStyleSheet(f"color:{C['t2']};")
        cg.addWidget(self.chk_cd,1,0,1,2)

        self.chk_hide = QCheckBox("재생 중 창 숨기기 (블라인드 모드)")
        self.chk_hide.setChecked(False)
        self.chk_hide.setStyleSheet(f"color:{C['t2']};")
        cg.addWidget(self.chk_hide,2,0,1,2)

        self.chk_min_others = QCheckBox("재생 전 다른 창 모두 최소화 (좌표 오작동 방지)")
        self.chk_min_others.setChecked(True)  # 기본 켜짐 - 안정성↑
        self.chk_min_others.setStyleSheet(f"color:{C['t2']};")
        cg.addWidget(self.chk_min_others,3,0,1,2)
        root.addWidget(cfg)

        # 재생 버튼
        br = QHBoxLayout(); br.setSpacing(10)
        self.btn_play = GlowButton("  [PLAY]  재생 시작", C['play'])
        self.btn_play.setFixedHeight(52); self.btn_play.setEnabled(False)
        self.btn_play.setFont(QFont("Malgun Gothic",13,QFont.Weight.Bold))
        self.btn_play.clicked.connect(self.start_play)
        self.btn_stop = GlowButton("  [STOP]  중지", C['stop'])
        self.btn_stop.setFixedHeight(52); self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.stop_play)
        br.addWidget(self.btn_play,3); br.addWidget(self.btn_stop,1)

        # ESC 키로 재생 중지 (ApplicationShortcut: 어디서든 동작)
        self._esc_shortcut = QShortcut(QKeySequence("Escape"), self)
        self._esc_shortcut.setContext(
            Qt.ShortcutContext.ApplicationShortcut)
        self._esc_shortcut.activated.connect(self._on_esc)
        root.addLayout(br)

        # 진행 바
        self.progress = QProgressBar()
        self.progress.setFixedHeight(8); self.progress.setTextVisible(False)
        self.progress.setStyleSheet(f"""
            QProgressBar {{ background:{C['border']}; border:none; border-radius:4px; }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 {C['brand']}, stop:1 {C['tool']});
                border-radius:4px;
            }}
        """)
        root.addWidget(self.progress)

        self.status_lbl = QLabel("대기 중")
        self.status_lbl.setStyleSheet(f"""
            color:{C['t2']}; padding:8px 12px;
            background:{C['bg1']}; border:1px solid {C['border']}; border-radius:8px;
        """)
        root.addWidget(self.status_lbl)

        # 미리보기
        prev_lbl = QLabel("동작 미리보기")
        prev_lbl.setStyleSheet(f"color:{C['t2']}; font-size:11px; font-weight:bold;")
        root.addWidget(prev_lbl)
        self.preview = QListWidget()
        self.preview.setStyleSheet(f"""
            QListWidget {{ background:{C['bg1']}; border:1.5px solid {C['border']};
                border-radius:10px; padding:4px; outline:none; }}
            QListWidget::item {{ padding:7px 12px; border-radius:6px; margin:1px 2px;
                font-family:Consolas,monospace; font-size:12px; color:{C['t1']}; }}
            QListWidget::item:selected {{ background:{C['brand']}; color:#FFFFFF; }}
            QListWidget::item:hover {{ background:{C['bg4']}; }}
        """)
        root.addWidget(self.preview)

    _SPEED = {"0.5x  (천천히)":0.5,"1.0x  (보통)":1.0,
               "1.5x  (빠르게)":1.5,"2.0x  (최고속)":2.0}

    def load_file(self):
        fpath,_ = rpa_open_file(self,"매크로 열기",SAVE_DIR)
        if not fpath: return
        try:
            with open(fpath,encoding="utf-8") as f: d = json.load(f)
            # JSON 이 list 직접이면 그대로, dict 이면 actions 키 추출
            if isinstance(d, list):
                self.actions = d
                name = Path(fpath).stem
            else:
                self.actions = (d.get("actions",[]) if isinstance(d,dict) else d)
                name = d.get("name", Path(fpath).stem)
            self.file_lbl.setText(f"  {name}")
            self.file_lbl.setStyleSheet(
                f"color:{C['play']}; font-size:14px; font-weight:bold; background:transparent;")
            self.file_meta.setText(f"{len(self.actions)}개 동작  |  {Path(fpath).name}")
            self._loaded_name = name

            # 두벌식 변환 가능한 type 액션 자동 감지
            # (예: 한글 모드에서 친 "spdlqj" 같은 영문이 ime_state="english"로
            #  저장된 구버전 매크로 → 사용자가 원하면 한글로 일괄 변환)
            self._suggest_hangul_convert()

            self.btn_play.setEnabled(True)
            self._build_preview()
        except Exception as e:
            if hasattr(self, 'status_lbl'):
                self.status_lbl.setText(f"파일 오류: {e}")

    def _suggest_hangul_convert(self):
        """
        영문으로 저장된 type 액션 중 두벌식 변환 시 한글이 되는 항목이 있으면
        사용자에게 일괄 변환을 제안한다. (구버전 매크로 호환)
        """
        candidates = []  # (index, original, converted) 튜플
        for i, a in enumerate(self.actions):
            if not isinstance(a, dict): continue
            if a.get("type") != "type": continue
            text = a.get("text","")
            # 이미 한글이 포함되어 있으면 변환 대상 아님
            if _is_korean(text): continue
            # 한글로 변환했을 때 한글이 충분히 나오는지 확인
            if len(text) < 2: continue
            converted = _eng_to_kor(text)
            kor_chars = sum(1 for c in converted
                            if 0xAC00 <= ord(c) <= 0xD7A3)
            # 변환 결과 절반 이상이 한글 음절이면 한글 의도로 추정
            if kor_chars >= max(1, len(text) // 2):
                candidates.append((i, text, converted))

        if not candidates:
            return

        # 사용자 확인
        preview = "\n".join(
            f'  "{orig}"  →  "{conv}"' for _, orig, conv in candidates[:5])
        more = f"\n  ... 외 {len(candidates)-5}건" if len(candidates) > 5 else ""
        ok = msg_ask(self, "한글 변환 감지",
                     f"이 매크로에 한글로 의도된 영문 입력이 {len(candidates)}건 발견되었습니다.\n\n"
                     f"미리보기:\n{preview}{more}\n\n"
                     f"한글로 일괄 변환할까요?\n"
                     f"(녹화 당시 한/영 키 인식이 안 됐을 때 사용)")
        if not ok:
            return

        # 변환 적용
        for i, _, conv in candidates:
            self.actions[i]["text"] = conv
            self.actions[i]["ime_state"] = "korean"
        if hasattr(self, 'status_lbl'):
            self.status_lbl.setText(f"한글 자동 변환 {len(candidates)}건 적용됨")

    def _build_preview(self):
        self.preview.clear(); self.preview.setUpdatesEnabled(False)
        icons  = {"click":"[클릭]","scroll":"[스크롤]","key":"[키]",
                  "type":"[입력]","wait":"[대기]","hotkey":"[단축키]",
                  "image_click":"[이미지]"}
        colors = {"click":C['brand'],"scroll":C['tool'],"key":C['stop'],
                  "type":C['play'],"wait":C['t2'],"hotkey":C['brand'],
                  "image_click":C['ai']}
        sliced = self.actions[:300]
        batches = [sliced[i:i+50] for i in range(0,len(sliced),50)]
        self._prev_offset = 0
        def _add():
            if not batches:
                self.preview.setUpdatesEnabled(True)
                if len(self.actions)>300:
                    self.preview.addItem(QListWidgetItem(f"  ... 외 {len(self.actions)-300}개"))
                return
            batch = batches.pop(0)
            for j,a in enumerate(batch):
                if not isinstance(a, dict): continue
                t=a.get("type",""); lbl=icons.get(t,"[?]")
                if t=="click": det=f"({a['x']},{a['y']}) {a.get('button','left')}"
                elif t=="scroll": det=f"({a['x']},{a['y']}) dy={a['dy']}"
                elif t=="key": det=f"[{a['key']}]"
                elif t=="type": det=a.get('text','')[:25]
                elif t=="wait": det=f"{a.get('seconds',1)}초"
                elif t=="image_click": det=a.get('label','')
                else: det=""
                idx=self._prev_offset+j+1
                item=QListWidgetItem(f"  {idx:03d}  {lbl}  {det}  +{a.get('delay',0):.2f}s")
                item.setForeground(QColor(colors.get(t,C['t2'])))
                self.preview.addItem(item)
            self._prev_offset+=len(batch)
            QTimer.singleShot(0,_add)
        QTimer.singleShot(0,_add)

    def start_play(self):
        if not self.actions: return
        speed  = self._SPEED.get(self.speed_combo.currentText(),1.0)
        repeat = self.repeat_spin.value()
        def _run():
            self.worker = PlayWorker(self.actions, repeat=repeat, speed=speed,
                                     minimize_others=self.chk_min_others.isChecked())
            self.worker.progress_update.connect(self._on_prog)
            self.worker.status_update.connect(self._on_stat)
            self.worker.finished_play.connect(self._on_done)
            self.worker.start()
            self.btn_play.setEnabled(False); self.btn_stop.setEnabled(True)
            # 재생 시작 시 모든 창 최소화 (Win+M)
            win = self.window()
            if win:
                if self.chk_hide.isChecked() and hasattr(win, '_tray'):
                    win.hide()
                else:
                    win.showMinimized()
            # 모든 창 최소화
            pyautogui.hotkey('win', 'm')
            time.sleep(0.2)
        if self.chk_cd.isChecked():
            self.status_lbl.setText("5초 후 재생...")
            QTimer.singleShot(1000,lambda:self.status_lbl.setText("4초 후..."))
            QTimer.singleShot(2000,lambda:self.status_lbl.setText("3초 후..."))
            QTimer.singleShot(3000,lambda:self.status_lbl.setText("2초 후..."))
            QTimer.singleShot(4000,lambda:self.status_lbl.setText("1초 후..."))
            # 카운트다운 플로팅 창
            self._play_countdown = _RecCountdown(
                seconds=5,
                hint_text="재생 준비 중... 대상 창으로 이동하세요")
            self._play_countdown.show()
            self._play_countdown.raise_()
            QTimer.singleShot(5000,_run)
        else: _run()

    def stop_play(self):
        if self.worker: self.worker.stop()
        self.btn_play.setEnabled(True); self.btn_stop.setEnabled(False)
        self.status_lbl.setText("중지됨")
        name = getattr(self, '_loaded_name', '알 수 없음')
        LogEngine.add("재생", name, "중지", "사용자 중지")

    def _on_esc(self):
        """ESC 키 → 재생 중이면 즉시 중지"""
        if self.worker and self.worker.running:
            self.stop_play()
            self.status_lbl.setText("ESC 로 중지됨")

    def _on_prog(self,cur,total):
        self.progress.setMaximum(total); self.progress.setValue(cur)
        self.status_lbl.setText(f"재생 중...  {cur}/{total}  ({int(cur/total*100)}%)")

    def _on_stat(self,msg):
        if msg.startswith("playing:"): _, c, t = msg.split(":"); self.status_lbl.setText(f"재생 중 ({c}/{t}회)")
        elif msg=="done": self.status_lbl.setText("재생 완료!")
        elif msg.startswith("err:"): self.status_lbl.setText(f"오류: {msg[4:]}")

    def _on_done(self):
        self.btn_play.setEnabled(True); self.btn_stop.setEnabled(False)
        name = getattr(self, '_loaded_name', '알 수 없음')
        LogEngine.add("재생", name, "성공",
                      f"{len(self.actions)}개 프로세스 완료")
        win = self.window()
        if win:
            if self.chk_hide.isChecked() and hasattr(win, '_tray'):
                win._tray.show_window()
            else:
                win.showNormal()
                win.raise_()
            if hasattr(win, '_tray'):
                self._on_play_notify()

    def _on_play_notify(self):
        win = self.window()
        if not win or not hasattr(win, '_tray'): return
        acts = len(self.actions)
        win._tray.show_message(
            "5MRPA - 재생 완료",
            f"{acts}개 프로세스 실행이 완료되었습니다.",
            duration_ms=3000)




# =============================================
#  실행 로그 엔진 (전역 싱글톤)
# =============================================
LOG_FILE = SAVE_DIR / "execution_log.json"

class LogEngine:
    """
    전역 실행 로그 관리.
    모든 페이지에서 LogEngine.add() 로 기록.
    JSON 파일로 영구 저장.
    """
    _listeners = []   # UI 업데이트 콜백 목록

    @classmethod
    def add(cls, action: str, target: str,
            result: str, detail: str = ""):
        """
        로그 항목 추가.
        action : 녹화/재생/스케줄/이미지/프로세스
        target : 매크로 이름 또는 파일명
        result : 성공/실패/중지
        detail : 추가 정보
        """
        entry = {
            "ts":     datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": action,
            "target": target,
            "result": result,
            "detail": detail,
        }
        # 파일에 추가
        logs = cls._load()
        logs.insert(0, entry)   # 최신이 맨 위
        logs = logs[:500]       # 최대 500건 유지
        cls._save(logs)
        # UI 콜백 알림
        for cb in list(cls._listeners):
            try:
                QTimer.singleShot(0, lambda e=entry, c=cb: c(e))
            except Exception:
                pass

    @classmethod
    def _load(cls) -> list:
        try:
            if LOG_FILE.exists():
                with open(LOG_FILE, encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return []

    @classmethod
    def _save(cls, logs: list):
        try:
            with open(LOG_FILE, "w", encoding="utf-8") as f:
                json.dump(logs, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    @classmethod
    def get_all(cls) -> list:
        return cls._load()

    @classmethod
    def clear(cls):
        cls._save([])

    @classmethod
    def register(cls, callback):
        """UI 콜백 등록 - 새 로그 추가 시 호출됨"""
        if callback not in cls._listeners:
            cls._listeners.append(callback)

    @classmethod
    def unregister(cls, callback):
        if callback in cls._listeners:
            cls._listeners.remove(callback)

    @classmethod
    def export_csv(cls, path: str):
        """CSV 로 내보내기"""
        logs = cls._load()
        try:
            with open(path, "w", encoding="utf-8-sig") as f:
                f.write("날짜시간,동작,대상,결과,상세\n")
                for e in logs:
                    f.write(
                        f"{e['ts']},"
                        f"{e['action']},"
                        f"{e['target']},"
                        f"{e['result']},"
                        f"{e.get('detail','')}\n")
            return True
        except Exception:
            return False


# =============================================
#  페이지: [LOG] 실행 로그
# =============================================

class LogPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build()
        # 실시간 업데이트 콜백 등록
        LogEngine.register(self._on_new_entry)

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(12)

        root.addWidget(SectionHeader("LOG  /  실행 로그"))

        # 통계 카드 행
        stat_row = QHBoxLayout(); stat_row.setSpacing(10)
        self.card_total   = StatCard("전체 실행",  "0", C['brand'])
        self.card_success = StatCard("성공",       "0", C['play'])
        self.card_fail    = StatCard("실패",       "0", C['rec'])
        self.card_today   = StatCard("오늘",       "0", C['stop'])
        for c in [self.card_total, self.card_success,
                  self.card_fail, self.card_today]:
            stat_row.addWidget(c)
        root.addLayout(stat_row)

        # 필터 + 검색 행
        filter_row = QHBoxLayout(); filter_row.setSpacing(8)
        filter_row.addWidget(QLabel("필터:"))

        self.filter_combo = StyledCombo()
        self.filter_combo.addItems([
            "전체", "성공만", "실패만",
            "재생", "스케줄", "녹화", "이미지", "프로세스"
        ])
        self.filter_combo.setFixedWidth(120)
        self.filter_combo.currentIndexChanged.connect(self._refresh)

        self.search_edit = StyledInput("검색...")
        self.search_edit.textChanged.connect(self._refresh)

        btn_refresh = GlowButton("새로고침", C['brand'])
        btn_refresh.setFixedHeight(36); btn_refresh.setFixedWidth(90)
        btn_refresh.clicked.connect(self._refresh)

        btn_export = GlowButton("CSV 내보내기", C['play'])
        btn_export.setFixedHeight(36)
        btn_export.clicked.connect(self._export)

        btn_clear = GlowButton("로그 삭제", C['rec'])
        btn_clear.setFixedHeight(36)
        btn_clear.clicked.connect(self._clear)

        filter_row.addWidget(self.filter_combo)
        filter_row.addWidget(self.search_edit)
        filter_row.addStretch()
        filter_row.addWidget(btn_refresh)
        filter_row.addWidget(btn_export)
        filter_row.addWidget(btn_clear)
        root.addLayout(filter_row)

        # 로그 테이블 (QListWidget 으로 구현)
        self.log_list = QListWidget()
        self.log_list.setStyleSheet(f"""
            QListWidget {{
                background: {C['bg1']};
                border: 1.5px solid {C['border']};
                border-radius: 12px;
                padding: 4px;
                outline: none;
                font-family: 'Consolas', monospace;
                font-size: 12px;
            }}
            QListWidget::item {{
                padding: 8px 14px;
                border-radius: 6px;
                margin: 1px 2px;
                border-bottom: 1px solid {C['border2']};
            }}
            QListWidget::item:selected {{
                background: {C['bg4']};
                color: {C['brand']};
            }}
            QListWidget::item:hover {{
                background: {C['bg4']};
            }}
        """)
        root.addWidget(self.log_list)

        # 상세 정보
        self.detail_lbl = QLabel("로그 항목을 클릭하면 상세 정보가 표시됩니다.")
        self.detail_lbl.setStyleSheet(f"""
            color: {C['t2']}; font-size: 12px;
            background: {C['bg1']};
            border: 1px solid {C['border']};
            border-radius: 8px; padding: 8px 12px;
        """)
        self.detail_lbl.setWordWrap(True)
        root.addWidget(self.detail_lbl)
        self.log_list.currentItemChanged.connect(self._on_select)

        self._refresh()

    def _refresh(self):
        logs  = LogEngine.get_all()
        filt  = self.filter_combo.currentText()
        query = self.search_edit.text().strip().lower()

        # 필터 적용
        filtered = []
        for e in logs:
            if filt == "성공만"  and e["result"] != "성공": continue
            if filt == "실패만"  and e["result"] != "실패": continue
            if filt in ("재생","스케줄","녹화","이미지","프로세스"):
                if filt not in e["action"]: continue
            if query and query not in (
                e["ts"] + e["action"] + e["target"] +
                e["result"] + e.get("detail","")).lower():
                continue
            filtered.append(e)

        # 통계 업데이트
        today = datetime.date.today().strftime("%Y-%m-%d")
        total   = len(logs)
        success = sum(1 for e in logs if e["result"] == "성공")
        fail    = sum(1 for e in logs if e["result"] == "실패")
        t_today = sum(1 for e in logs if e["ts"].startswith(today))

        self.card_total.setValue(total)
        self.card_success.setValue(success)
        self.card_fail.setValue(fail)
        self.card_today.setValue(t_today)

        # 리스트 렌더링
        self.log_list.clear()
        RESULT_COLORS = {
            "성공": C['play'],
            "실패": C['rec'],
            "중지": C['stop'],
            "실행중": C['brand'],
        }
        ACTION_ICONS = {
            "재생": ">",
            "스케줄": "#",
            "녹화": "@",
            "이미지": "&",
            "프로세스": "*",
        }
        for e in filtered:
            icon = next((v for k,v in ACTION_ICONS.items()
                        if k in e["action"]), "-")
            color = RESULT_COLORS.get(e["result"], C['t2'])

            # 결과 뱃지
            result_badge = f"[{e['result']}]"
            text = (f"  {e['ts']}   "
                    f"{icon} {e['action']:<8}  "
                    f"{result_badge:<6}  "
                    f"{e['target']}")

            item = QListWidgetItem(text)
            item.setForeground(QColor(color))
            item.setData(Qt.ItemDataRole.UserRole, e)
            self.log_list.addItem(item)

        # 결과 없음
        if not filtered:
            item = QListWidgetItem("  로그가 없습니다.")
            item.setForeground(QColor(C['t3']))
            self.log_list.addItem(item)

    def _on_select(self, item):
        if not item: return
        e = item.data(Qt.ItemDataRole.UserRole)
        if not e: return
        detail = e.get("detail", "없음")
        self.detail_lbl.setText(
            f"날짜시간: {e['ts']}   |   "
            f"동작: {e['action']}   |   "
            f"대상: {e['target']}   |   "
            f"결과: {e['result']}   |   "
            f"상세: {detail}")
        # 결과에 따라 색상
        c = {
            "성공": C['play'], "실패": C['rec'],
            "중지": C['stop']
        }.get(e["result"], C['t2'])
        self.detail_lbl.setStyleSheet(f"""
            color: {c}; font-size: 12px;
            background: {C['bg1']};
            border: 1.5px solid {c};
            border-radius: 8px; padding: 8px 12px;
        """)

    def _on_new_entry(self, entry: dict):
        """실시간 새 로그 수신 - 맨 위에 추가"""
        self._refresh()
        # 통계 카드 잠깐 강조
        result = entry.get("result","")
        if result == "성공":
            self.card_success.val_lbl.setStyleSheet(
                f"color:{C['play']}; font-size:22px; font-weight:900;"
                "background:transparent;")
        elif result == "실패":
            self.card_fail.val_lbl.setStyleSheet(
                f"color:{C['rec']}; font-size:22px; font-weight:900;"
                "background:transparent;")

    def _export(self):
        fp, _ = rpa_save_file(self, "CSV 내보내기", SAVE_DIR, ".csv")
        if fp:
            ok = LogEngine.export_csv(fp)
            if ok:
                msg_info(self, "완료", f"CSV 내보내기 완료:\n{fp}")
            else:
                msg_error(self, "오류", "CSV 저장 실패")

    def _clear(self):
        r = msg_ask(self, "로그 삭제", "전체 로그를 삭제할까요?\n이 작업은 되돌릴 수 없습니다.")
        if r == True:
            LogEngine.clear()
            self._refresh()

    def hideEvent(self, event):
        """페이지 숨겨질 때 콜백 유지 (unregister 안 함)"""
        super().hideEvent(event)

    def closeEvent(self, event):
        LogEngine.unregister(self._on_new_entry)
        super().closeEvent(event)

# =============================================
#  스케줄 데이터 저장 경로
# =============================================
SCHEDULE_FILE = SAVE_DIR / "schedules.json"


def load_schedules() -> list:
    try:
        if SCHEDULE_FILE.exists():
            with open(SCHEDULE_FILE, encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return []


def save_schedules(schedules: list):
    with open(SCHEDULE_FILE, "w", encoding="utf-8") as f:
        json.dump(schedules, f, ensure_ascii=False, indent=2)


# =============================================
#  스케줄러 엔진 (백그라운드 스레드)
# =============================================
class SchedulerEngine(QThread):
    """
    백그라운드에서 1분마다 스케줄 체크.
    실행 시간이 되면 해당 매크로 파일을 재생.
    """
    schedule_triggered = pyqtSignal(str, str)   # (name, macro_path)
    status_update      = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.running   = True
        self.schedules = []

    def set_schedules(self, schedules: list):
        self.schedules = [s for s in schedules if s.get("enabled", True)]

    def run(self):
        while self.running:
            now = datetime.datetime.now()
            for s in self.schedules:
                try:
                    self._check(s, now)
                except Exception:
                    pass
            time.sleep(15)   # 15초마다 체크

    def _check(self, s: dict, now: datetime.datetime):
        if not s.get("enabled", True):
            return

        mode = s.get("mode", "daily")   # daily / weekday / once / interval

        if mode == "daily":
            # 매일 지정 시간
            t = s.get("time", "09:00")
            h, m = map(int, t.split(":"))
            if now.hour == h and now.minute == m:
                last = s.get("last_run", "")
                today = now.strftime("%Y-%m-%d")
                if last != today:
                    s["last_run"] = today
                    save_schedules(load_schedules())
                    self.schedule_triggered.emit(s["name"], s["macro_path"])

        elif mode == "weekday":
            # 특정 요일 지정 시간
            days = s.get("days", [0,1,2,3,4])   # 0=월 ~ 6=일
            if now.weekday() not in days:
                return
            t = s.get("time", "09:00")
            h, m = map(int, t.split(":"))
            if now.hour == h and now.minute == m:
                last = s.get("last_run", "")
                today = now.strftime("%Y-%m-%d")
                if last != today:
                    s["last_run"] = today
                    save_schedules(load_schedules())
                    self.schedule_triggered.emit(s["name"], s["macro_path"])

        elif mode == "once":
            # 지정 날짜+시간 1회
            dt_str = s.get("datetime", "")
            if not dt_str:
                return
            target = datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
            if (now.year  == target.year  and
                now.month == target.month and
                now.day   == target.day   and
                now.hour  == target.hour  and
                now.minute == target.minute):
                if not s.get("done", False):
                    s["done"] = True
                    save_schedules(load_schedules())
                    self.schedule_triggered.emit(s["name"], s["macro_path"])

        elif mode == "interval":
            # N분 간격
            interval_min = s.get("interval_min", 60)
            last_str = s.get("last_run_dt", "")
            if last_str:
                last_dt = datetime.datetime.fromisoformat(last_str)
                elapsed = (now - last_dt).total_seconds() / 60
                if elapsed < interval_min:
                    return
            s["last_run_dt"] = now.isoformat()
            save_schedules(load_schedules())
            self.schedule_triggered.emit(s["name"], s["macro_path"])

    def stop(self):
        self.running = False


# =============================================
#  페이지 7: [SCHED] 스케줄러
# =============================================
class SchedulerPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._schedules = load_schedules()
        self._engine    = SchedulerEngine()
        self._engine.schedule_triggered.connect(self._on_trigger)
        self._engine.set_schedules(self._schedules)
        self._engine.start()
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(14)

        root.addWidget(SectionHeader("SCHEDULER  /  스케줄 자동 실행"))

        # 안내
        guide = QLabel(
            "  지정한 시간에 매크로를 자동으로 실행합니다.\n"
            "  프로그램이 실행 중이어야 스케줄이 동작합니다.")
        guide.setStyleSheet(f"""
            color:{C['t2']}; font-size:12px;
            background:{C['bg1']};
            border:1px solid {C['border']};
            border-left:4px solid {C['brand']};
            border-radius:8px; padding:10px;
        """)
        guide.setWordWrap(True)
        root.addWidget(guide)

        # ── 스케줄 추가 카드 ──
        add_card = QFrame()
        add_card.setStyleSheet(f"""
            QFrame {{
                background:{C['bg1']};
                border:1.5px solid {C['border']};
                border-radius:14px;
            }}
        """)
        ag = QGridLayout(add_card)
        ag.setContentsMargins(18, 14, 18, 14)
        ag.setSpacing(10)

        # 이름
        ag.addWidget(QLabel("스케줄 이름:"), 0, 0)
        self.name_edit = StyledInput("예: 매일 보고서 작성")
        ag.addWidget(self.name_edit, 0, 1, 1, 3)

        # 매크로 파일
        ag.addWidget(QLabel("매크로 파일:"), 1, 0)
        self.macro_path_edit = StyledInput("파일 경로")
        self.macro_path_edit.setReadOnly(True)
        self.macro_name_lbl = QLabel("선택 없음")
        self.macro_name_lbl.setStyleSheet(f"color:{C['t3']}; font-size:11px; background:transparent;")
        btn_pick = GlowButton("파일 선택", C['brand'])
        btn_pick.setFixedHeight(34)
        btn_pick.clicked.connect(self._pick_macro)
        ag.addWidget(self.macro_path_edit, 1, 1, 1, 2)
        ag.addWidget(btn_pick, 1, 3)
        ag.addWidget(self.macro_name_lbl, 2, 1, 1, 3)

        # 실행 모드
        ag.addWidget(QLabel("실행 방식:"), 3, 0)
        self.mode_combo = StyledCombo()
        self.mode_combo.addItems([
            "매일  (지정 시간마다)",
            "특정 요일  (요일 + 시간)",
            "1회만  (날짜 + 시간)",
            "반복 간격  (N분마다)",
        ])
        self.mode_combo.currentIndexChanged.connect(self._on_mode_change)
        ag.addWidget(self.mode_combo, 3, 1, 1, 3)

        # 시간 설정 영역 (동적)
        self.time_frame = QFrame()
        self.time_frame.setStyleSheet("QFrame{background:transparent;}")
        self.time_lay = QHBoxLayout(self.time_frame)
        self.time_lay.setContentsMargins(0, 0, 0, 0)
        self.time_lay.setSpacing(10)
        ag.addWidget(self.time_frame, 4, 0, 1, 4)

        self._build_time_widgets("daily")

        # 추가 버튼
        btn_add = GlowButton("+ 스케줄 추가", C['play'])
        btn_add.setFixedHeight(40)
        btn_add.clicked.connect(self._add_schedule)
        ag.addWidget(btn_add, 5, 0, 1, 4)

        root.addWidget(add_card)

        # ── 스케줄 목록 ──
        list_hdr = QHBoxLayout()
        list_lbl = QLabel("등록된 스케줄")
        list_lbl.setStyleSheet(f"color:{C['t1']}; font-weight:bold; font-size:13px;")
        list_hdr.addWidget(list_lbl)
        list_hdr.addStretch()
        btn_ref = GlowButton("새로고침", C['brand'])
        btn_ref.setFixedHeight(30); btn_ref.setFixedWidth(80)
        btn_ref.clicked.connect(self._refresh_list)
        list_hdr.addWidget(btn_ref)
        root.addLayout(list_hdr)

        self.sched_list = QListWidget()
        self.sched_list.setStyleSheet(f"""
            QListWidget {{
                background:{C['bg1']};
                border:1.5px solid {C['border']};
                border-radius:12px; padding:4px; outline:none;
            }}
            QListWidget::item {{
                padding:10px 14px; border-radius:8px; margin:2px 3px;
                font-size:13px; color:{C['t1']};
            }}
            QListWidget::item:selected {{
                background:{C['brand']}; color:#FFFFFF;
            }}
            QListWidget::item:hover {{ background:{C['bg4']}; }}
        """)
        root.addWidget(self.sched_list)

        # 제어 버튼
        ctrl_row = QHBoxLayout(); ctrl_row.setSpacing(8)
        self.btn_toggle = GlowButton("ON / OFF 전환", C['stop'])
        self.btn_toggle.setFixedHeight(38)
        self.btn_toggle.clicked.connect(self._toggle_schedule)
        btn_del = GlowButton("삭제", C['rec'])
        btn_del.setFixedHeight(38)
        btn_del.clicked.connect(self._delete_schedule)
        btn_run_now = GlowButton("지금 즉시 실행", C['play'])
        btn_run_now.setFixedHeight(38)
        btn_run_now.clicked.connect(self._run_now)
        ctrl_row.addWidget(self.btn_toggle)
        ctrl_row.addWidget(btn_del)
        ctrl_row.addStretch()
        ctrl_row.addWidget(btn_run_now)
        root.addLayout(ctrl_row)

        # 상태
        self.status_lbl = QLabel("스케줄러 실행 중...")
        self.status_lbl.setStyleSheet(f"""
            color:{C['t2']}; font-size:12px;
            background:{C['bg1']}; border:1px solid {C['border']};
            border-radius:8px; padding:8px 12px;
        """)
        root.addWidget(self.status_lbl)

        # 실행 로그
        log_lbl = QLabel("실행 로그")
        log_lbl.setStyleSheet(f"color:{C['t2']}; font-size:11px; font-weight:bold;")
        root.addWidget(log_lbl)

        self.log_list = QListWidget()
        self.log_list.setMaximumHeight(120)
        self.log_list.setStyleSheet(f"""
            QListWidget {{
                background:{C['bg1']}; border:1px solid {C['border']};
                border-radius:8px; padding:4px; outline:none;
                font-family:Consolas,monospace; font-size:11px;
            }}
            QListWidget::item {{ padding:4px 10px; color:{C['t2']}; }}
            QListWidget::item:hover {{ background:{C['bg4']}; }}
        """)
        root.addWidget(self.log_list)

        self._refresh_list()

    def _build_time_widgets(self, mode: str):
        # 기존 위젯 제거
        while self.time_lay.count():
            item = self.time_lay.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        if mode == "daily":
            self.time_lay.addWidget(QLabel("실행 시간:"))
            self.w_hour = StyledSpin()
            self.w_hour.setRange(0, 23); self.w_hour.setValue(9)
            self.w_hour.setSuffix(" 시"); self.w_hour.setFixedWidth(80)
            self.w_min  = StyledSpin()
            self.w_min.setRange(0, 59); self.w_min.setValue(0)
            self.w_min.setSuffix(" 분"); self.w_min.setFixedWidth(80)
            self.time_lay.addWidget(self.w_hour)
            self.time_lay.addWidget(self.w_min)
            self.time_lay.addStretch()

        elif mode == "weekday":
            self.time_lay.addWidget(QLabel("요일:"))
            self._day_cbs = []
            for i, day in enumerate(["월","화","수","목","금","토","일"]):
                cb = QCheckBox(day)
                cb.setChecked(i < 5)
                cb.setStyleSheet(f"color:{C['t1']};")
                self._day_cbs.append(cb)
                self.time_lay.addWidget(cb)
            self.time_lay.addSpacing(16)
            self.time_lay.addWidget(QLabel("시간:"))
            self.w_hour = StyledSpin()
            self.w_hour.setRange(0,23); self.w_hour.setValue(9)
            self.w_hour.setSuffix(" 시"); self.w_hour.setFixedWidth(80)
            self.w_min  = StyledSpin()
            self.w_min.setRange(0,59); self.w_min.setValue(0)
            self.w_min.setSuffix(" 분"); self.w_min.setFixedWidth(80)
            self.time_lay.addWidget(self.w_hour)
            self.time_lay.addWidget(self.w_min)
            self.time_lay.addStretch()

        elif mode == "once":
            self.time_lay.addWidget(QLabel("날짜:"))
            today = datetime.date.today()
            self.w_year  = StyledSpin()
            self.w_year.setRange(2024,2030); self.w_year.setValue(today.year)
            self.w_year.setSuffix(" 년"); self.w_year.setFixedWidth(90)
            self.w_month = StyledSpin()
            self.w_month.setRange(1,12); self.w_month.setValue(today.month)
            self.w_month.setSuffix(" 월"); self.w_month.setFixedWidth(75)
            self.w_day   = StyledSpin()
            self.w_day.setRange(1,31); self.w_day.setValue(today.day)
            self.w_day.setSuffix(" 일"); self.w_day.setFixedWidth(75)
            self.w_hour  = StyledSpin()
            self.w_hour.setRange(0,23); self.w_hour.setValue(9)
            self.w_hour.setSuffix(" 시"); self.w_hour.setFixedWidth(80)
            self.w_min   = StyledSpin()
            self.w_min.setRange(0,59); self.w_min.setValue(0)
            self.w_min.setSuffix(" 분"); self.w_min.setFixedWidth(80)
            for w in [self.w_year,self.w_month,self.w_day,
                      QLabel("시간:"),self.w_hour,self.w_min]:
                self.time_lay.addWidget(w)
            self.time_lay.addStretch()

        elif mode == "interval":
            self.time_lay.addWidget(QLabel("실행 간격:"))
            self.w_interval = StyledSpin()
            self.w_interval.setRange(1, 1440)
            self.w_interval.setValue(60)
            self.w_interval.setSuffix(" 분마다")
            self.w_interval.setFixedWidth(120)
            self.time_lay.addWidget(self.w_interval)
            hint = QLabel("(프로그램 실행 중일 때만 동작)")
            hint.setStyleSheet(f"color:{C['t3']}; font-size:11px; background:transparent;")
            self.time_lay.addWidget(hint)
            self.time_lay.addStretch()

    def _on_mode_change(self, idx):
        modes = ["daily", "weekday", "once", "interval"]
        self._build_time_widgets(modes[idx])

    def _pick_macro(self):
        fp, _ = rpa_open_file(self, "매크로 파일 선택", SAVE_DIR)
        if fp:
            self.macro_path_edit.setText(fp)
            try:
                with open(fp, encoding="utf-8") as f:
                    d = json.load(f)
                name = d.get("name", Path(fp).stem)
                count = len((d.get("actions",[]) if isinstance(d,dict) else d))
                self.macro_name_lbl.setText(
                    f"{name}  ({count}개 프로세스)")
                self.macro_name_lbl.setStyleSheet(
                    f"color:{C['play']}; font-size:11px; font-weight:bold; background:transparent;")
            except Exception:
                self.macro_name_lbl.setText(Path(fp).name)

    def _add_schedule(self):
        name = self.name_edit.text().strip()
        path = self.macro_path_edit.text().strip()
        if not name:
            msg_warn(self, "입력 오류", "스케줄 이름을 입력하세요.")
            return
        if not path or not Path(path).exists():
            msg_warn(self, "입력 오류", "매크로 파일을 선택하세요.")
            return

        modes = ["daily", "weekday", "once", "interval"]
        mode  = modes[self.mode_combo.currentIndex()]

        s = {
            "name":       name,
            "macro_path": path,
            "mode":       mode,
            "enabled":    True,
            "created":    datetime.datetime.now().isoformat(),
            "last_run":   "",
        }

        if mode in ("daily", "weekday"):
            h = self.w_hour.value()
            m = self.w_min.value()
            s["time"] = f"{h:02d}:{m:02d}"
            if mode == "weekday":
                s["days"] = [i for i,cb in enumerate(self._day_cbs)
                             if cb.isChecked()]
        elif mode == "once":
            s["datetime"] = (
                f"{self.w_year.value():04d}-"
                f"{self.w_month.value():02d}-"
                f"{self.w_day.value():02d} "
                f"{self.w_hour.value():02d}:"
                f"{self.w_min.value():02d}")
            s["done"] = False
        elif mode == "interval":
            s["interval_min"] = self.w_interval.value()
            s["last_run_dt"]  = ""

        self._schedules.append(s)
        save_schedules(self._schedules)
        self._engine.set_schedules(self._schedules)
        self._refresh_list()
        self.name_edit.clear()
        self.macro_path_edit.clear()
        self.macro_name_lbl.setText("선택 없음")
        self.status_lbl.setText(f"스케줄 추가됨: {name}")
        self._add_log(f"[등록] {name}  ({mode})")

    def _refresh_list(self):
        self._schedules = load_schedules()
        self.sched_list.clear()
        MODE_KR = {
            "daily":    "매일",
            "weekday":  "요일 반복",
            "once":     "1회",
            "interval": "간격",
        }
        DAY_KR = ["월","화","수","목","금","토","일"]
        for s in self._schedules:
            mode = s.get("mode","daily")
            enabled = s.get("enabled", True)

            if mode == "daily":
                detail = f"{s.get('time','?')} 실행"
            elif mode == "weekday":
                days = s.get("days", [])
                day_str = "".join(DAY_KR[d] for d in days if d < 7)
                detail = f"{day_str}  {s.get('time','?')} 실행"
            elif mode == "once":
                detail = f"{s.get('datetime','?')}  {'(완료)' if s.get('done') else ''}"
            elif mode == "interval":
                detail = f"{s.get('interval_min',60)}분마다"
            else:
                detail = ""

            status_icon = "ON" if enabled else "OFF"
            last = s.get("last_run","") or s.get("last_run_dt","")
            last_str = f"  | 마지막: {last[:16]}" if last else ""

            text = (f"  {'[ON] ' if enabled else '[OFF]'}  "
                    f"{s['name']}  |  {MODE_KR.get(mode,mode)}  {detail}{last_str}")
            item = QListWidgetItem(text)
            c = C['play'] if enabled else C['t3']
            item.setForeground(QColor(c))
            self.sched_list.addItem(item)

        n_on = sum(1 for s in self._schedules if s.get("enabled", True))
        self.status_lbl.setText(
            f"스케줄러 실행 중  |  총 {len(self._schedules)}개  |  활성 {n_on}개")

    def _toggle_schedule(self):
        row = self.sched_list.currentRow()
        if row < 0 or row >= len(self._schedules):
            return
        s = self._schedules[row]
        s["enabled"] = not s.get("enabled", True)
        save_schedules(self._schedules)
        self._engine.set_schedules(self._schedules)
        self._refresh_list()
        state = "ON" if s["enabled"] else "OFF"
        self._add_log(f"[{state}] {s['name']}")

    def _delete_schedule(self):
        row = self.sched_list.currentRow()
        if row < 0 or row >= len(self._schedules):
            return
        name = self._schedules[row]["name"]
        r = msg_ask(self, "삭제 확인", f"'{name}' 스케줄을 삭제할까요?")
        if r == True:
            self._schedules.pop(row)
            save_schedules(self._schedules)
            self._engine.set_schedules(self._schedules)
            self._refresh_list()
            self._add_log(f"[삭제] {name}")

    def _run_now(self):
        row = self.sched_list.currentRow()
        if row < 0 or row >= len(self._schedules):
            msg_info(self, "선택 없음", "실행할 스케줄을 선택하세요.")
            return
        s = self._schedules[row]
        if not s.get("macro_path") or not Path(s["macro_path"]).exists():
            self._add_log(f"[오류] {s['name']}: 매크로 파일 없음")
            self.status_lbl.setText("오류: 매크로 파일을 찾을 수 없습니다")
            return

        # 3초 카운트다운
        self._add_log(f"[즉시실행 예약] {s['name']} - 3초 후 시작")
        self.status_lbl.setText(f"3초 후 실행: {s['name']}")
        QTimer.singleShot(1000, lambda: self.status_lbl.setText(
            f"2초 후 실행: {s['name']}"))
        QTimer.singleShot(2000, lambda: self.status_lbl.setText(
            f"1초 후 실행: {s['name']}"))

        # 카운트다운 플로팅 창
        self._sched_countdown = _RecCountdown(
            hint_text=f"스케줄 실행 준비 중...\n{s['name']}")
        self._sched_countdown.show()
        self._sched_countdown.raise_()

        QTimer.singleShot(3000, lambda: (
            self._add_log(f"[즉시실행] {s['name']}"),
            self.status_lbl.setText(f"실행 중: {s['name']}"),
            self._execute_macro(s["name"], s["macro_path"])))

    def _on_trigger(self, name: str, path: str):
        """스케줄 엔진에서 실행 신호 수신"""
        self._add_log(f"[자동실행] {name}")
        self.status_lbl.setText(f"자동 실행 중: {name}")
        self._execute_macro(name, path)
        self._refresh_list()

    def _execute_macro(self, name: str, path: str):
        try:
            with open(path, encoding="utf-8") as f:
                d = json.load(f)
            acts = (d.get("actions",[]) if isinstance(d,dict) else d)
            if not acts:
                self._add_log(f"[오류] {name}: 프로세스 없음")
                return

            def _go():
                w = PlayWorker(acts, 1, 1.0)
                w.start()
                w.wait()
                ts = datetime.datetime.now().strftime("%H:%M:%S")
                LogEngine.add("스케줄", name, "성공",
                              f"스케줄 자동 실행 완료 {ts}")
                QTimer.singleShot(0, lambda: self._add_log(
                    f"[완료] {name}  {ts}"))
                QTimer.singleShot(0, lambda: self.status_lbl.setText(
                    f"완료: {name}  ({ts})"))

            threading.Thread(target=_go, daemon=True).start()
        except Exception as e:
            self._add_log(f"[오류] {name}: {e}")

    def _add_log(self, msg: str):
        ts   = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        text = f"  {ts}  {msg}"
        item = QListWidgetItem(text)
        if "[오류]" in msg:
            item.setForeground(QColor(C['rec']))
        elif "[완료]" in msg:
            item.setForeground(QColor(C['play']))
        elif "[자동실행]" in msg:
            item.setForeground(QColor(C['brand']))
        else:
            item.setForeground(QColor(C['t2']))
        self.log_list.insertItem(0, item)
        if self.log_list.count() > 50:
            self.log_list.takeItem(self.log_list.count() - 1)

    def closeEvent(self, event):
        self._engine.stop()
        super().closeEvent(event)


# =============================================
#  프로세스 공유 유틸리티
# =============================================
import zipfile as _zipfile
import shutil as _shutil


def export_macro_zip(macro_path: str, out_dir: str = None) -> str:
    """
    매크로 JSON + 참조 이미지 파일을 ZIP 으로 패키징.
    동료에게 파일 하나만 전달하면 바로 사용 가능.
    """
    macro_p  = Path(macro_path)
    out_dir  = Path(out_dir) if out_dir else macro_p.parent
    zip_path = out_dir / (macro_p.stem + "_share.zip")

    with open(macro_path, encoding="utf-8") as f:
        data = json.load(f)

    acts = data.get("actions", [])

    with _zipfile.ZipFile(zip_path, "w", _zipfile.ZIP_DEFLATED) as zf:
        # 매크로 JSON
        zf.write(macro_path, macro_p.name)
        # 참조 이미지 수집
        added = set()
        for a in acts:
            for key in ["image_path", "cond_img", "yes_macro", "no_macro"]:
                fp = a.get(key, "")
                if fp and Path(fp).exists() and fp not in added:
                    zf.write(fp, Path(fp).name)
                    added.add(fp)

    return str(zip_path)


def import_macro_zip(zip_path: str, out_dir: str = None) -> str:
    """
    공유 ZIP 파일을 SAVE_DIR 에 풀고 경로를 보정.
    """
    out_dir = Path(out_dir) if out_dir else SAVE_DIR

    with _zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(out_dir)

    # JSON 파일 찾기
    with _zipfile.ZipFile(zip_path, "r") as zf:
        names = zf.namelist()
    json_files = [n for n in names if n.endswith(".json")
                  and not n.endswith("_share.zip")]
    if not json_files:
        raise ValueError("ZIP 안에 매크로 JSON 파일이 없습니다.")

    macro_file = out_dir / json_files[0]
    # 경로 보정: 이미지 파일 경로를 out_dir 기준으로 업데이트
    with open(macro_file, encoding="utf-8") as f:
        data = json.load(f)

    acts = data.get("actions", [])
    for a in acts:
        for key in ["image_path", "cond_img"]:
            fp = a.get(key, "")
            if fp:
                fname  = Path(fp).name
                new_fp = out_dir / fname
                if new_fp.exists():
                    a[key] = str(new_fp)

    with open(macro_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return str(macro_file)



# =============================================
#  윈도우 자동화 - 공통 리스트 위젯
# =============================================
class ActionList(QListWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QListWidget {{
                background: {C['bg1']};
                border: 1.5px solid {C['border']};
                border-radius: 10px;
                padding: 4px; outline: none;
            }}
            QListWidget::item {{
                padding: 8px 12px; border-radius: 6px; margin: 1px 2px;
                font-family: Consolas, monospace; font-size: 12px;
                color: {C['t1']};
            }}
            QListWidget::item:selected {{
                background: {C['brand']}; color: #FFFFFF;
            }}
            QListWidget::item:hover {{ background: {C['bg4']}; }}
        """)


# =============================================
#  윈도우 자동화 페이지
# =============================================
class WinAutoPage(QWidget):
    """
    pywinauto / pygetwindow 로 창 제목, 메뉴, 버튼 텍스트 기반 자동화.
    좌표 무관 - 텍스트/컨트롤 이름으로 찾아 클릭.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self._win_actions = []
        self._running     = False
        self._build()

    def _build(self):
        from PyQt6.QtWidgets import QTabWidget
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(14)
        root.addWidget(SectionHeader("WINDOW  /  윈도우 텍스트 자동화"))

        # 패키지 상태 배너
        status_row = QHBoxLayout()
        for pkg, name in [(GW_OK,"pygetwindow"), (WINAUTO_OK,"pywinauto")]:
            lbl = QLabel(f"  {name}: {'OK' if pkg else '미설치'}  ")
            lbl.setStyleSheet(f"""
                color: {'#2F9E44' if pkg else C['rec']};
                font-size: 11px; font-weight: bold;
                background: {'#ECFDF5' if pkg else '#FFF1F2'};
                border: 1px solid {'#2F9E44' if pkg else C['rec']};
                border-radius: 6px; padding: 4px 8px;
            """)
            status_row.addWidget(lbl)
        if not WINAUTO_OK or not GW_OK:
            inst_lbl = QLabel(
                "  미설치 패키지:  "
                "pip install pywinauto pygetwindow")
            inst_lbl.setStyleSheet(
                f"color:{C['stop']};font-size:12px;background:transparent;")
            status_row.addWidget(inst_lbl)
        status_row.addStretch()
        root.addLayout(status_row)

        # 탭 구조
        inner_tabs = QTabWidget()
        inner_tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                background:{C['bg2']};border:1px solid {C['border']};
                border-radius:8px;
            }}
            QTabBar::tab {{
                background:{C['bg1']};color:{C['t2']};
                padding:7px 16px;border-radius:5px 5px 0 0;
                margin-right:2px;font-weight:bold;
            }}
            QTabBar::tab:selected {{
                background:{C['bg2']};color:{C['brand']};
                border-bottom:2px solid {C['brand']};
            }}
        """)

        # ── 탭1: 창 탐색기 ──
        win_tab = QWidget()
        wt_lay  = QVBoxLayout(win_tab)
        wt_lay.setContentsMargins(14,14,14,14); wt_lay.setSpacing(10)

        wt_hdr = QHBoxLayout()
        wt_hdr.addWidget(QLabel("현재 열린 창 목록"))
        btn_ref = GlowButton("새로고침", C['brand'])
        btn_ref.setFixedHeight(30); btn_ref.setFixedWidth(80)
        btn_ref.clicked.connect(self._refresh_windows)
        wt_hdr.addStretch(); wt_hdr.addWidget(btn_ref)
        wt_lay.addLayout(wt_hdr)

        self.win_list = ActionList()
        self.win_list.setFixedHeight(160)
        wt_lay.addWidget(self.win_list)

        # 창 조작 버튼
        win_btn_row = QHBoxLayout(); win_btn_row.setSpacing(8)
        for lbl, c, fn in [
            ("앞으로 가져오기", C['brand'], self._win_activate),
            ("최대화",          C['play'],  self._win_maximize),
            ("최소화",          C['stop'],  self._win_minimize),
            ("창 닫기",         C['rec'],   self._win_close),
        ]:
            b = GlowButton(lbl, c); b.setFixedHeight(36)
            b.clicked.connect(fn); win_btn_row.addWidget(b)
        wt_lay.addLayout(win_btn_row)

        # 컨트롤 탐색
        ctrl_hdr = QHBoxLayout()
        ctrl_hdr.addWidget(QLabel("선택한 창의 컨트롤 탐색"))
        btn_find_ctrl = GlowButton("컨트롤 목록 가져오기", C['tool'])
        btn_find_ctrl.setFixedHeight(30)
        btn_find_ctrl.clicked.connect(self._get_controls)
        ctrl_hdr.addStretch(); ctrl_hdr.addWidget(btn_find_ctrl)
        wt_lay.addLayout(ctrl_hdr)

        self.ctrl_list = ActionList()
        self.ctrl_list.setFixedHeight(130)
        wt_lay.addWidget(self.ctrl_list)

        btn_click_ctrl = GlowButton("[CLICK]  선택한 컨트롤 클릭", C['play'])
        btn_click_ctrl.setFixedHeight(38)
        btn_click_ctrl.clicked.connect(self._click_control)
        wt_lay.addWidget(btn_click_ctrl)

        self.win_status = QLabel("대기 중")
        self.win_status.setStyleSheet(f"""
            color:{C['t2']};padding:6px 12px;
            background:{C['bg1']};border:1px solid {C['border']};
            border-radius:8px;
        """)
        wt_lay.addWidget(self.win_status)
        inner_tabs.addTab(win_tab, "  창 탐색  ")

        # ── 탭2: 액션 빌더 ──
        builder_tab = QWidget()
        bl_lay = QVBoxLayout(builder_tab)
        bl_lay.setContentsMargins(14,14,14,14); bl_lay.setSpacing(10)

        hint = QLabel("윈도우 자동화 동작을 순서대로 쌓아서 실행합니다.")
        hint.setStyleSheet(f"color:{C['t2']};background:transparent;")
        bl_lay.addWidget(hint)

        form_card = QFrame()
        form_card.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};border:1.5px solid {C['border']};
                border-radius:10px;}}
        """)
        fg = QGridLayout(form_card)
        fg.setContentsMargins(14,12,14,12); fg.setSpacing(8)

        fg.addWidget(QLabel("동작 유형:"), 0, 0)
        self.act_type = StyledCombo()
        self.act_type.addItems([
            "창 활성화  (제목으로 찾기)",
            "버튼 클릭  (텍스트로 찾기)",
            "메뉴 클릭  (메뉴 경로)",
            "텍스트 입력  (컨트롤에 타이핑)",
            "대기  (초)",
            "창 최대화",
            "창 최소화",
        ])
        self.act_type.currentIndexChanged.connect(self._update_form)
        fg.addWidget(self.act_type, 0, 1, 1, 3)

        fg.addWidget(QLabel("창 제목:"), 1, 0)
        self.f_wintitle = StyledInput(
            "예: 메모장  /  Excel  /  Chrome  (일부만 입력해도 됩니다)")
        fg.addWidget(self.f_wintitle, 1, 1, 1, 3)

        fg.addWidget(QLabel("대상/값:"), 2, 0)
        self.f_target = StyledInput(
            "버튼명 / 메뉴경로(메뉴>서브메뉴) / 입력텍스트 / 초")
        fg.addWidget(self.f_target, 2, 1, 1, 3)

        fg.addWidget(QLabel("대기:"), 3, 0)
        self.f_delay = StyledDSpin()
        self.f_delay.setRange(0, 30); self.f_delay.setValue(0.5)
        self.f_delay.setSuffix(" 초"); self.f_delay.setFixedWidth(110)
        fg.addWidget(self.f_delay, 3, 1)

        btn_add_act = GlowButton("동작 추가 +", C['play'])
        btn_add_act.setFixedHeight(36)
        btn_add_act.clicked.connect(self._add_action)
        fg.addWidget(btn_add_act, 3, 2, 1, 2)
        bl_lay.addWidget(form_card)

        # 힌트 박스
        self._hint_box = QLabel("  동작 유형을 선택하면 입력 예시가 표시됩니다.")
        self._hint_box.setStyleSheet(f"""
            QLabel{{color:{C['brand']};font-size:11px;background:transparent;
                padding:4px 8px;border-left:3px solid {C['brand']};}}
        """)
        bl_lay.addWidget(self._hint_box)

        # 액션 목록
        act_hdr = QHBoxLayout()
        act_hdr.addWidget(QLabel("동작 순서"))
        btn_del = GlowButton("선택 삭제", C['rec'])
        btn_del.setFixedHeight(28); btn_del.setFixedWidth(80)
        btn_del.clicked.connect(self._del_action)
        btn_clr = GlowButton("전체 삭제", C['t3'])
        btn_clr.setFixedHeight(28); btn_clr.setFixedWidth(80)
        btn_clr.clicked.connect(lambda: (
            self._win_actions.clear(), self.action_list_w.clear()))
        act_hdr.addStretch()
        act_hdr.addWidget(btn_del); act_hdr.addWidget(btn_clr)
        bl_lay.addLayout(act_hdr)

        self.action_list_w = ActionList()
        self.action_list_w.setFixedHeight(150)
        bl_lay.addWidget(self.action_list_w)

        # 저장/불러오기/실행
        exec_row = QHBoxLayout(); exec_row.setSpacing(8)
        btn_save = GlowButton("시퀀스 저장",     C['brand'])
        btn_load = GlowButton("시퀀스 불러오기", C['t2'])
        self.btn_exec      = GlowButton("[PLAY]  시퀀스 실행", C['play'])
        self.btn_stop_exec = GlowButton("[STOP]  중지",        C['stop'])
        for b in [btn_save, btn_load, self.btn_exec, self.btn_stop_exec]:
            b.setFixedHeight(40)
        self.btn_stop_exec.setEnabled(False)
        btn_save.clicked.connect(self._save_seq)
        btn_load.clicked.connect(self._load_seq)
        self.btn_exec.clicked.connect(self._exec_seq)
        self.btn_stop_exec.clicked.connect(lambda: setattr(self,'_running',False))
        exec_row.addWidget(btn_save); exec_row.addWidget(btn_load)
        exec_row.addStretch()
        exec_row.addWidget(self.btn_exec); exec_row.addWidget(self.btn_stop_exec)
        bl_lay.addLayout(exec_row)

        self.exec_status = QLabel("대기 중")
        self.exec_status.setStyleSheet(f"""
            color:{C['t2']};padding:8px 14px;
            background:{C['bg1']};border:1px solid {C['border']};
            border-radius:8px;
        """)
        bl_lay.addWidget(self.exec_status)
        inner_tabs.addTab(builder_tab, "  액션 빌더  ")

        root.addWidget(inner_tabs)
        QTimer.singleShot(500, self._refresh_windows)

    def _update_form(self, idx):
        hints = [
            ("창 제목으로 창을 앞으로 가져옵니다",  "예: 메모장  /  Excel"),
            ("텍스트로 버튼을 찾아 클릭합니다",      "예: 확인  /  저장  /  취소"),
            ("메뉴를 경로로 클릭합니다",              "예: 파일>저장  /  편집>복사"),
            ("텍스트 박스에 타이핑합니다",            "예: 홍길동  /  2024-01-01"),
            ("지정 시간 동안 대기합니다",             "예: 2  (초 단위 숫자)"),
            ("창을 최대화합니다",                     "예: 메모장  /  Excel"),
            ("창을 최소화합니다",                     "예: 메모장  /  Excel"),
        ]
        colors = [C['brand'],C['play'],C['tool'],C['stop'],C['t2'],C['brand'],C['brand']]
        if 0 <= idx < len(hints):
            self.f_wintitle.setPlaceholderText(hints[idx][0])
            self.f_target.setPlaceholderText(hints[idx][1])
            c = colors[idx]
            self._hint_box.setStyleSheet(f"""
                QLabel{{color:{c};font-size:11px;background:transparent;
                    padding:4px 8px;border-left:3px solid {c};}}
            """)
            self._hint_box.setText(
                f"  {hints[idx][0]}  |  대상: {hints[idx][1]}")

    # ── 창 탐색 ──
    def _refresh_windows(self):
        self.win_list.clear()
        if not GW_OK:
            self.win_list.addItem(QListWidgetItem(
                "  pygetwindow 미설치  --  pip install pygetwindow"))
            return
        try:
            wins = gw.getAllWindows()
            for w in wins:
                if w.title.strip():
                    item = QListWidgetItem(f"  {w.title}")
                    item.setForeground(QColor(C['t1']))
                    self.win_list.addItem(item)
        except Exception as e:
            self.win_list.addItem(QListWidgetItem(f"오류: {e}"))

    def _get_win_title(self):
        item = self.win_list.currentItem()
        return item.text().strip() if item else None

    def _do_win_action(self, act, title):
        if not GW_OK: return
        def _w():
            try:
                wins = gw.getWindowsWithTitle(title)
                if not wins:
                    QTimer.singleShot(0, lambda: self.win_status.setText(
                        f"창을 찾지 못함: {title}")); return
                w = wins[0]
                if act == "activate": w.activate()
                elif act == "maximize": w.maximize()
                elif act == "minimize": w.minimize()
                elif act == "close":    w.close()
                QTimer.singleShot(0, lambda: self.win_status.setText(
                    f"[OK] {act}: {title}"))
            except Exception as e:
                QTimer.singleShot(0, lambda: self.win_status.setText(
                    f"오류: {e}"))
        threading.Thread(target=_w, daemon=True).start()

    def _win_activate(self):
        t = self._get_win_title()
        if t: self._do_win_action("activate", t)

    def _win_maximize(self):
        t = self._get_win_title()
        if t: self._do_win_action("maximize", t)

    def _win_minimize(self):
        t = self._get_win_title()
        if t: self._do_win_action("minimize", t)

    def _win_close(self):
        t = self._get_win_title()
        if not t: return
        r = msg_ask(self, "창 닫기", f"'{t}' 창을 닫을까요?")
        if r == True:
            self._do_win_action("close", t)

    def _get_controls(self):
        self.ctrl_list.clear()
        item = self.win_list.currentItem()
        if not item:
            self.ctrl_list.addItem(QListWidgetItem("  창을 먼저 선택하세요"))
            return
        title = item.text().strip()
        if not WINAUTO_OK:
            self.ctrl_list.addItem(QListWidgetItem(
                "  pywinauto 미설치  --  pip install pywinauto"))
            return
        self.win_status.setText("컨트롤 목록 가져오는 중...")
        def _w():
            try:
                app   = WinApp(backend="uia").connect(title_re=f".*{title}.*")
                dlg   = app.top_window()
                ctrls = dlg.descendants(control_type="Button")
                QTimer.singleShot(0, self.ctrl_list.clear)
                for c in ctrls[:50]:
                    try:
                        txt = c.window_text()
                        if txt.strip():
                            it = QListWidgetItem(f"  [Button]  {txt}")
                            it.setForeground(QColor(C['brand']))
                            QTimer.singleShot(0, lambda i=it:
                                self.ctrl_list.addItem(i))
                    except Exception:
                        pass
                menus = dlg.descendants(control_type="MenuItem")
                for m in menus[:30]:
                    try:
                        txt = m.window_text()
                        if txt.strip():
                            it = QListWidgetItem(f"  [Menu]    {txt}")
                            it.setForeground(QColor(C['tool']))
                            QTimer.singleShot(0, lambda i=it:
                                self.ctrl_list.addItem(i))
                    except Exception:
                        pass
                QTimer.singleShot(0, lambda: self.win_status.setText(
                    f"컨트롤 로드 완료: {title}"))
            except Exception as e:
                QTimer.singleShot(0, lambda: self.win_status.setText(
                    f"오류: {e}"))
        threading.Thread(target=_w, daemon=True).start()

    def _click_control(self):
        item     = self.ctrl_list.currentItem()
        win_item = self.win_list.currentItem()
        if not item or not win_item:
            msg_warn(self, "선택 없음", "창과 컨트롤을 모두 선택하세요.")
            return
        ctrl_txt  = item.text().strip().split("]")[-1].strip()
        win_title = win_item.text().strip()
        if not WINAUTO_OK:
            msg_warn(self, "미설치", "pip install pywinauto"); return
        def _w():
            try:
                app = WinApp(backend="uia").connect(title_re=f".*{win_title}.*")
                dlg = app.top_window()
                dlg[ctrl_txt].click()
                QTimer.singleShot(0, lambda: self.win_status.setText(
                    f"[OK] 클릭: {ctrl_txt}"))
            except Exception as e:
                QTimer.singleShot(0, lambda: self.win_status.setText(
                    f"오류: {e}"))
        threading.Thread(target=_w, daemon=True).start()

    # ── 액션 빌더 ──
    def _add_action(self):
        entry = {
            "type":     self.act_type.currentText(),
            "wintitle": self.f_wintitle.text().strip(),
            "target":   self.f_target.text().strip(),
            "delay":    self.f_delay.value(),
        }
        self._win_actions.append(entry)
        n   = len(self._win_actions)
        txt = (f"  {n:02d}.  {entry['type']}"
               + (f"  |  창: {entry['wintitle']}" if entry['wintitle'] else "")
               + (f"  |  대상: {entry['target']}"  if entry['target']   else "")
               + f"  |  +{entry['delay']:.1f}s")
        item = QListWidgetItem(txt)
        item.setForeground(QColor(C['play']))
        self.action_list_w.addItem(item)

    def _del_action(self):
        row = self.action_list_w.currentRow()
        if row < 0: return
        self._win_actions.pop(row)
        self.action_list_w.takeItem(row)
        for i in range(self.action_list_w.count()):
            it = self.action_list_w.item(i)
            it.setText(f"  {i+1:02d}." + it.text()[5:])

    def _save_seq(self):
        if not self._win_actions:
            msg_warn(self, "없음", "동작이 없습니다."); return
        fp, _ = rpa_save_file(self, "시퀀스 저장", SAVE_DIR)
        if fp:
            with open(fp, "w", encoding="utf-8") as f:
                json.dump({"type":"win_sequence",
                           "created":datetime.datetime.now().isoformat(),
                           "actions":self._win_actions},
                          f, ensure_ascii=False, indent=2)
            self.exec_status.setText(f"저장 완료: {Path(fp).name}")
            LogEngine.add("윈도우자동화", Path(fp).stem, "성공", "시퀀스 저장")

    def _load_seq(self):
        fp, _ = rpa_open_file(self, "시퀀스 불러오기", SAVE_DIR)
        if not fp: return
        try:
            with open(fp, encoding="utf-8") as f:
                data = json.load(f)
            self._win_actions = data.get("actions", [])
            self.action_list_w.clear()
            for i, a in enumerate(self._win_actions, 1):
                txt = (f"  {i:02d}.  {a['type']}"
                       + (f"  |  창: {a['wintitle']}" if a.get('wintitle') else "")
                       + (f"  |  대상: {a['target']}"  if a.get('target')   else "")
                       + f"  |  +{a.get('delay',0):.1f}s")
                item = QListWidgetItem(txt)
                item.setForeground(QColor(C['play']))
                self.action_list_w.addItem(item)
            self.exec_status.setText(f"불러오기 완료: {len(self._win_actions)}개")
        except Exception as e:
            msg_error(self, "오류", str(e))

    def _exec_seq(self):
        if not self._win_actions:
            msg_warn(self, "없음", "동작 목록이 비어 있습니다."); return
        self._running = True
        self.btn_exec.setEnabled(False)
        self.btn_stop_exec.setEnabled(True)
        actions = list(self._win_actions)

        def _worker():
            for act in actions:
                if not self._running: break
                time.sleep(act.get("delay",0.5))
                t        = act["type"]
                wintitle = act.get("wintitle","")
                target   = act.get("target","")
                sf = lambda msg: QTimer.singleShot(
                    0, lambda m=msg: self.exec_status.setText(m))
                try:
                    if "활성화" in t:
                        if GW_OK:
                            wins = gw.getWindowsWithTitle(wintitle)
                            if wins: wins[0].activate()
                            sf(f"[OK] 활성화: {wintitle}")
                    elif "버튼 클릭" in t:
                        if WINAUTO_OK:
                            app = WinApp(backend="uia").connect(
                                title_re=f".*{wintitle}.*")
                            app.top_window()[target].click()
                            sf(f"[OK] 버튼 클릭: {target}")
                    elif "메뉴 클릭" in t:
                        if WINAUTO_OK:
                            app = WinApp(backend="uia").connect(
                                title_re=f".*{wintitle}.*")
                            path = [p.strip() for p in target.split(">")]
                            app.top_window().menu_select("->".join(path))
                            sf(f"[OK] 메뉴: {target}")
                    elif "텍스트 입력" in t:
                        if WINAUTO_OK:
                            app = WinApp(backend="uia").connect(
                                title_re=f".*{wintitle}.*")
                            app.top_window().type_keys(
                                target, with_spaces=True)
                            sf(f"[OK] 입력: {target}")
                        else:
                            _type_with_clipboard(target)
                            sf(f"[OK] 입력(fallback): {target}")
                    elif "대기" in t:
                        sec = float(target) if target else 1.0
                        time.sleep(sec); sf(f"[OK] 대기 {sec}초")
                    elif "최대화" in t:
                        if GW_OK:
                            wins = gw.getWindowsWithTitle(wintitle)
                            if wins: wins[0].maximize()
                            sf(f"[OK] 최대화: {wintitle}")
                    elif "최소화" in t:
                        if GW_OK:
                            wins = gw.getWindowsWithTitle(wintitle)
                            if wins: wins[0].minimize()
                            sf(f"[OK] 최소화: {wintitle}")
                except Exception as e:
                    sf(f"[오류] {t}: {e}")

            self._running = False
            QTimer.singleShot(0, lambda: (
                self.btn_exec.setEnabled(True),
                self.btn_stop_exec.setEnabled(False),
                self.exec_status.setText("시퀀스 실행 완료"),
                LogEngine.add("윈도우자동화","시퀀스","성공",
                              f"{len(actions)}개 완료")))

        threading.Thread(target=_worker, daemon=True).start()
        self.exec_status.setText("시퀀스 실행 중...")

# =============================================
#  빠른 도구 페이지
# =============================================
class QuickPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._track_timer = QTimer()
        self._track_timer.timeout.connect(self._upd_pos)
        self._click_running = False
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(14)
        root.addWidget(SectionHeader("TOOLS  /  빠른 도구"))

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        inner = QWidget(); inner.setStyleSheet("background:transparent;")
        lay   = QVBoxLayout(inner); lay.setSpacing(14)

        # ── 자동 반복 클릭 ──
        lay.addWidget(self._section(
            "자동 반복 클릭",
            "같은 위치를 여러 번 반복 클릭합니다"))
        click_card = self._card()
        cg = QGridLayout(click_card); cg.setSpacing(10)

        cg.addWidget(QLabel("X 좌표"), 0, 0)
        self.cx = StyledSpin(); self.cx.setRange(0,9999); self.cx.setValue(500)
        cg.addWidget(self.cx, 0, 1)
        cg.addWidget(QLabel("Y 좌표"), 0, 2)
        self.cy = StyledSpin(); self.cy.setRange(0,9999); self.cy.setValue(500)
        cg.addWidget(self.cy, 0, 3)

        cg.addWidget(QLabel("반복"), 1, 0)
        self.cc = StyledSpin()
        self.cc.setRange(1,99999); self.cc.setValue(10); self.cc.setSuffix(" 회")
        cg.addWidget(self.cc, 1, 1)
        cg.addWidget(QLabel("간격"), 1, 2)
        self.ci = StyledDSpin()
        self.ci.setRange(0.05,60); self.ci.setValue(0.5); self.ci.setSuffix(" 초")
        cg.addWidget(self.ci, 1, 3)

        cg.addWidget(QLabel("버튼"), 2, 0)
        self.click_btn_combo = StyledCombo()
        self.click_btn_combo.addItems(["왼쪽 클릭","오른쪽 클릭","더블 클릭"])
        cg.addWidget(self.click_btn_combo, 2, 1)

        btn_row = QHBoxLayout()
        btn_get = GlowButton("현재 마우스 위치 (3초)", C['brand'])
        btn_get.setFixedHeight(38)
        btn_get.clicked.connect(self._get_pos)
        self.btn_ac = GlowButton("[PLAY]  자동 클릭 시작", C['play'])
        self.btn_ac.setFixedHeight(38)
        self.btn_ac.clicked.connect(self._run_click)
        self.btn_stop_click = GlowButton("[STOP]  중지", C['rec'])
        self.btn_stop_click.setFixedHeight(38)
        self.btn_stop_click.setEnabled(False)
        self.btn_stop_click.clicked.connect(self._stop_click)
        btn_row.addWidget(btn_get)
        btn_row.addWidget(self.btn_ac)
        btn_row.addWidget(self.btn_stop_click)
        cg.addLayout(btn_row, 3, 0, 1, 4)
        lay.addWidget(click_card)

        # ── 텍스트 자동 입력 ──
        lay.addWidget(self._section(
            "텍스트 자동 입력",
            "3초 후 커서 위치에 텍스트를 자동으로 타이핑합니다"))
        type_card = self._card()
        tg = QVBoxLayout(type_card); tg.setSpacing(10)
        self.type_text = QTextEdit()
        self.type_text.setFixedHeight(70)
        self.type_text.setPlaceholderText("자동으로 입력할 텍스트를 여기에 쓰세요...")
        self.type_text.setStyleSheet(f"""
            QTextEdit{{background:{C['bg1']};border:1px solid {C['border']};
                border-radius:8px;color:{C['t1']};padding:8px;}}
            QTextEdit:focus{{border-color:{C['brand']};}}
        """)
        t_row = QHBoxLayout()
        t_row.addWidget(QLabel("속도"))
        self.ts = StyledDSpin()
        self.ts.setRange(0.01,1.0); self.ts.setValue(0.05)
        self.ts.setSuffix(" 초/글자"); self.ts.setFixedWidth(130)
        t_row.addWidget(self.ts)
        t_row.addStretch()
        btn_type = GlowButton("입력 시작 (3초 후)", C['tool'])
        btn_type.setFixedHeight(38)
        btn_type.clicked.connect(self._run_type)
        t_row.addWidget(btn_type)
        tg.addWidget(self.type_text)
        tg.addLayout(t_row)
        lay.addWidget(type_card)

        # ── 단축키 팔레트 ──
        lay.addWidget(self._section(
            "단축키 빠른 실행",
            "버튼 클릭 0.2초 후 단축키를 실행합니다"))
        hk_card = self._card()
        hg = QGridLayout(hk_card); hg.setSpacing(8)
        hotkeys = [
            ("복사",     "ctrl+c",   C['brand']),
            ("붙여넣기", "ctrl+v",   C['brand']),
            ("저장",     "ctrl+s",   C['play']),
            ("실행취소", "ctrl+z",   C['stop']),
            ("다시실행", "ctrl+y",   C['stop']),
            ("전체선택", "ctrl+a",   C['tool']),
            ("찾기",     "ctrl+f",   C['brand']),
            ("새창",     "ctrl+n",   C['t2']),
            ("닫기",     "ctrl+w",   C['rec']),
            ("인쇄",     "ctrl+p",   C['t2']),
            ("탭이동",   "ctrl+tab", C['tool']),
            ("Alt+F4",   "alt+f4",   C['rec']),
        ]
        for i,(lbl,keys,c) in enumerate(hotkeys):
            btn = QPushButton(f"{lbl}\n{keys}")
            btn.setFixedSize(110, 52)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(f"""
                QPushButton{{background:{C['bg1']};color:{c};
                    border:1px solid {c};border-radius:8px;
                    font-size:11px;font-weight:bold;}}
                QPushButton:hover{{background:{C['bg4']};}}
                QPushButton:pressed{{background:{c};color:#FFFFFF;}}
            """)
            btn.clicked.connect(lambda _,k=keys: self._hotkey(k))
            hg.addWidget(btn, i//6, i%6)
        lay.addWidget(hk_card)

        # ── 마우스 좌표 추적기 ──
        lay.addWidget(self._section(
            "마우스 좌표 추적기",
            "실시간 추적 ON 후 원하는 위치에 마우스를 올리세요"))
        pos_card = self._card()
        pg = QHBoxLayout(pos_card); pg.setContentsMargins(16,10,16,10)
        self.pos_disp = QLabel("X :  ----     Y :  ----")
        self.pos_disp.setStyleSheet(f"""
            color:{C['brand']}; font-size:22px; font-weight:900;
            font-family:Consolas,monospace; background:transparent;
        """)
        self.chk_track = QCheckBox("실시간 추적")
        self.chk_track.setStyleSheet(f"color:{C['t2']};font-weight:bold;")
        self.chk_track.stateChanged.connect(self._toggle_track)

        # 좌표 복사 버튼
        btn_copy_pos = GlowButton("좌표 복사", C['brand'])
        btn_copy_pos.setFixedHeight(34); btn_copy_pos.setFixedWidth(90)
        btn_copy_pos.clicked.connect(self._copy_pos)

        pg.addWidget(self.pos_disp)
        pg.addStretch()
        pg.addWidget(self.chk_track)
        pg.addWidget(btn_copy_pos)
        lay.addWidget(pos_card)

        lay.addStretch()
        scroll.setWidget(inner)
        root.addWidget(scroll)

    # ── 헬퍼 ──
    def _section(self, txt, desc=""):
        w = QWidget(); w.setStyleSheet("background:transparent;")
        hl = QHBoxLayout(w); hl.setContentsMargins(0,4,0,0); hl.setSpacing(10)
        lbl = QLabel(txt)
        lbl.setStyleSheet(
            f"color:{C['brand']};font-size:11px;font-weight:900;"
            f"letter-spacing:2px;background:transparent;")
        hl.addWidget(lbl)
        if desc:
            d = QLabel(f"-- {desc}")
            d.setStyleSheet(f"color:{C['t3']};font-size:11px;background:transparent;")
            hl.addWidget(d)
        hl.addStretch()
        return w

    def _card(self):
        f = QFrame()
        f.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};border:1.5px solid {C['border']};
                border-radius:12px;padding:4px;}}
        """)
        return f

    # ── 이벤트 ──
    def _get_pos(self):
        """3초 카운트다운 후 마우스 위치 캡처"""
        sender_btn = self.sender()
        if sender_btn: sender_btn.setEnabled(False)
        self._pos_countdown = _CountdownCapture(
            seconds=3,
            on_done=self._on_pos_captured,
            on_cancel=lambda: (
                sender_btn.setEnabled(True) if sender_btn else None))
        self._pos_countdown.show()
        self._pos_countdown.raise_()

    def _on_pos_captured(self, x, y):
        self.cx.setValue(x); self.cy.setValue(y)
        # 버튼 복원
        for btn in self.findChildren(GlowButton):
            if "마우스 위치" in btn.text():
                btn.setEnabled(True); break

    def _toggle_track(self, s):
        if s == Qt.CheckState.Checked.value:
            self._track_timer.start(80)
        else:
            self._track_timer.stop()

    def _upd_pos(self):
        x, y = pyautogui.position()
        self.pos_disp.setText(f"X :  {x:5d}     Y :  {y:5d}")

    def _copy_pos(self):
        x, y = pyautogui.position()
        try:
            import subprocess
            subprocess.run(
                ['powershell', '-NoProfile', '-Command',
                 f"Set-Clipboard -Value '{x}, {y}'"],
                capture_output=True, timeout=3,
                creationflags=subprocess.CREATE_NO_WINDOW
                    if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
        except Exception:
            pass
        self.pos_disp.setText(f"X :  {x:5d}     Y :  {y:5d}  (복사됨)")

    def _run_click(self):
        x   = self.cx.value()
        y   = self.cy.value()
        n   = self.cc.value()
        ivl = self.ci.value()
        btn_map = {"왼쪽 클릭":"left","오른쪽 클릭":"right","더블 클릭":"double"}
        btn = btn_map.get(self.click_btn_combo.currentText(), "left")
        self._click_running = True
        self.btn_ac.setEnabled(False)
        self.btn_stop_click.setEnabled(True)
        LogEngine.add("빠른도구", f"자동클릭({x},{y})", "실행중",
                      f"{n}회 간격{ivl}s")
        def _w():
            done = 0
            for _ in range(n):
                if not self._click_running: break
                if btn == "double":
                    pyautogui.doubleClick(x, y)
                else:
                    pyautogui.click(x, y, button=btn)
                done += 1
                time.sleep(ivl)
            self._click_running = False
            LogEngine.add("빠른도구", f"자동클릭({x},{y})", "성공",
                          f"{done}회 완료")
            QTimer.singleShot(0, lambda: (
                self.btn_ac.setEnabled(True),
                self.btn_stop_click.setEnabled(False),
                self.btn_ac.setText("[PLAY]  자동 클릭 시작")))
        self.btn_ac.setText(f"클릭 중... ({n}회)")
        threading.Thread(target=_w, daemon=True).start()

    def _stop_click(self):
        self._click_running = False
        self.btn_ac.setEnabled(True)
        self.btn_stop_click.setEnabled(False)
        self.btn_ac.setText("[PLAY]  자동 클릭 시작")

    def _run_type(self):
        text  = self.type_text.toPlainText()
        speed = self.ts.value()
        if not text:
            msg_info(self, "없음", "입력할 텍스트를 작성하세요.")
            return
        self._type_countdown = _CountdownCapture(
            seconds=3,
            on_done=lambda x,y: threading.Thread(
                target=lambda: _type_with_clipboard(text),
                daemon=True).start(),
            on_cancel=None)
        self._type_countdown.show()
        self._type_countdown.raise_()

    def _hotkey(self, k):
        keys = k.split("+")
        def _w(): time.sleep(0.2); pyautogui.hotkey(*keys)
        threading.Thread(target=_w, daemon=True).start()

# =============================================
#  ManagerPage (매크로 관리)
# =============================================
class ManagerPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._files = []
        self._build()
        self.refresh()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24,20,24,20); root.setSpacing(12)
        root.addWidget(SectionHeader("MACROS  /  매크로 관리"))

        sr = QHBoxLayout()
        self.search = StyledInput("검색...")
        self.search.textChanged.connect(self._filter)
        br = GlowButton("새로고침",C['brand']); br.setFixedWidth(90); br.setFixedHeight(36)
        br.clicked.connect(self.refresh)
        sr.addWidget(self.search); sr.addWidget(br)
        root.addLayout(sr)

        self.lst = QListWidget()
        self.lst.setStyleSheet(f"""
            QListWidget {{ background:{C['bg1']}; border:1.5px solid {C['border']};
                border-radius:12px; padding:4px; outline:none; }}
            QListWidget::item {{ padding:12px 16px; border-radius:8px; margin:2px 3px; font-size:13px; }}
            QListWidget::item:selected {{ background:{C['brand']}; color:#FFFFFF; border-left:none; }}
            QListWidget::item:hover {{ background:{C['bg4']}; }}
        """)
        self.lst.currentRowChanged.connect(self._on_select)
        self.lst.itemDoubleClicked.connect(lambda _: self._quick_run())
        root.addWidget(self.lst)

        ic = QFrame()
        ic.setMinimumHeight(70)
        ic.setStyleSheet(f"""
            QFrame {{ background:{C['bg1']}; border:1.5px solid {C['border']};
                     border-radius:10px; }}
        """)
        il = QVBoxLayout(ic); il.setContentsMargins(14,8,14,8)
        self.info_lbl = QLabel("매크로를 선택하면 정보가 표시됩니다.")
        self.info_lbl.setStyleSheet(f"color:{C['t2']}; font-family:Consolas,monospace; font-size:12px; background:transparent;")
        self.info_lbl.setWordWrap(True)
        il.addWidget(self.info_lbl)
        root.addWidget(ic)

        br2 = QHBoxLayout(); br2.setSpacing(8)
        for lbl,c,fn in [("PLAY  실행",C['play'],self._quick_run),
                          ("이름 변경",C['brand'],self._rename),
                          ("삭제",C['rec'],self._delete)]:
            b=GlowButton(lbl,c); b.setFixedHeight(38); b.clicked.connect(fn)
            br2.addWidget(b)
        br2.addStretch()
        # 공유 버튼
        btn_export = GlowButton("공유 ZIP 내보내기", C['tool'])
        btn_export.setFixedHeight(38)
        btn_export.setToolTip("매크로 + 이미지를 ZIP 으로 패키징해서 동료에게 공유")
        btn_export.clicked.connect(self._export_share)
        btn_import = GlowButton("ZIP 가져오기", C['tool'])
        btn_import.setFixedHeight(38)
        btn_import.setToolTip("동료에게 받은 공유 ZIP 파일 가져오기")
        btn_import.clicked.connect(self._import_share)
        br2.addWidget(btn_export)
        br2.addWidget(btn_import)
        root.addLayout(br2)

    def refresh(self):
        self._files = sorted(SAVE_DIR.glob("*.json"),
                             key=lambda p:p.stat().st_mtime,reverse=True)
        self._render(self._files)

    def _render(self,files):
        self.lst.clear()
        for fp in files:
            try:
                with open(fp,encoding="utf-8") as f: d=json.load(f)
                name=d.get("name",fp.stem); count=d.get("count",len(d.get("actions",[])))
                created=d.get("created","")[:16].replace("T"," ")
                item=QListWidgetItem(f"  {name}   ({count}개 동작)   {created}")
                item.setData(Qt.ItemDataRole.UserRole,fp)
                self.lst.addItem(item)
            except Exception: pass

    def _filter(self,txt):
        self._render([f for f in self._files if txt.lower() in f.stem.lower()])

    def _on_select(self,row):
        if row<0: return
        item=self.lst.item(row)
        if not item: return
        fp=item.data(Qt.ItemDataRole.UserRole)
        try:
            with open(fp,encoding="utf-8") as f: d=json.load(f)
            acts=d.get("actions",[]); clicks=sum(1 for a in acts if a["type"]=="click")
            keys=sum(1 for a in acts if a["type"]=="key")
            self.info_lbl.setText(
                f"이름: {d.get('name',fp.stem)}  |  생성: {d.get('created','')[:19]}  |  "
                f"총: {len(acts)}개  (클릭:{clicks}, 키:{keys})  |  {fp.name}")
        except Exception as e: self.info_lbl.setText(f"오류: {e}")

    def _quick_run(self):
        item=self.lst.currentItem()
        if not item: return
        fp=item.data(Qt.ItemDataRole.UserRole)
        try:
            with open(fp,encoding="utf-8") as f: d=json.load(f)
            acts=d.get("actions",[])
            if not acts: return
            def _go(): time.sleep(3); w=PlayWorker(acts,1,1.0); w.start(); w.wait()
            threading.Thread(target=_go,daemon=True).start()
            msg_info(self, "재생 시작", f"3초 후 '{d.get('name','')}' 재생\n대상 창으로 이동하세요.")
        except Exception as e: msg_error(self, "오류", str(e))

    def _rename(self):
        item=self.lst.currentItem()
        if not item: return
        fp=item.data(Qt.ItemDataRole.UserRole)
        new,ok=QInputDialog.getText(self,"이름 변경","새 이름:",text=fp.stem)
        if not ok or not new.strip(): return
        try:
            with open(fp,encoding="utf-8") as f: d=json.load(f)
            d["name"]=new.strip()
            np=fp.parent/f"{new.strip()}.json"
            with open(np,"w",encoding="utf-8") as f: json.dump(d,f,ensure_ascii=False,indent=2)
            if np!=fp: fp.unlink()
            self.refresh()
        except Exception as e: msg_error(self, "오류", str(e))

    def _delete(self):
        item=self.lst.currentItem()
        if not item: return
        fp=item.data(Qt.ItemDataRole.UserRole)
        r=msg_ask(self, "삭제 확인", f"'{fp.stem}' 을 삭제할까요?")
        if r==True: fp.unlink(missing_ok=True); self.refresh()

    def _export_share(self):
        item = self.lst.currentItem()
        if not item:
            msg_info(self, "선택 없음", "공유할 매크로를 선택하세요.")
            return
        fp = item.data(Qt.ItemDataRole.UserRole)
        try:
            zip_path = export_macro_zip(str(fp))
            msg_info(self, "내보내기 완료", f"공유 ZIP 파일이 생성됐습니다:\n{zip_path}\n\n"
                "이 파일을 동료에게 전달하세요.")
            LogEngine.add("공유", fp.stem, "성공", f"ZIP 내보내기: {zip_path}")
        except Exception as e:
            msg_error(self, "오류", str(e))

    def _import_share(self):
        fp, _ = rpa_open_file(self, "공유 ZIP 가져오기", SAVE_DIR, ".zip")
        if not fp: return
        try:
            macro_path = import_macro_zip(fp)
            self.refresh()
            msg_info(self, "가져오기 완료", f"매크로를 성공적으로 가져왔습니다:\n{macro_path}\n\n"
                "매크로 목록에서 확인하세요.")
            LogEngine.add("공유", Path(fp).stem, "성공", f"ZIP 가져오기: {macro_path}")
        except Exception as e:
            msg_error(self, "오류", str(e))




# =============================================
#  트리거 시스템
#  파일 변경 / 이미지 등장 / 프로세스 시작 감지
# =============================================
TRIGGER_FILE = SAVE_DIR / "triggers.json"


def load_triggers() -> list:
    try:
        if TRIGGER_FILE.exists():
            with open(TRIGGER_FILE, encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return []


def save_triggers(triggers: list):
    with open(TRIGGER_FILE, "w", encoding="utf-8") as f:
        json.dump(triggers, f, ensure_ascii=False, indent=2)


class TriggerEngine(QThread):
    """
    백그라운드에서 트리거 조건을 감시.
    조건 충족 시 매크로 자동 실행.
    """
    triggered = pyqtSignal(str, str)   # (name, macro_path)
    status_update = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.running  = True
        self.triggers = []
        self._file_mtimes  = {}   # path -> mtime
        self._folder_known = {}   # trigger_id -> set(filenames) - 폴더 감시용
        self._email_baseline = {} # trigger_id -> ReceivedTime (기준 시각, 이전 메일은 무시)
        self._email_next_check = {}  # trigger_id -> 다음 확인 가능 시각 (COM 부하 완화)
        self._outlook_warned = False

    def set_triggers(self, triggers: list):
        self.triggers = [t for t in triggers if t.get("enabled", True)]

    def run(self):
        while self.running:
            for t in self.triggers:
                try:
                    self._check(t)
                except Exception:
                    pass
            time.sleep(2)

    def _check(self, t: dict):
        kind = t.get("type", "")

        if kind == "file_change":
            path = t.get("watch_path", "")
            if not path or not Path(path).exists():
                return
            mtime = Path(path).stat().st_mtime
            key   = t["id"]
            last  = self._file_mtimes.get(key, 0)
            if last > 0 and mtime > last:
                self._file_mtimes[key] = mtime
                self.triggered.emit(t["name"], t["macro_path"])
                LogEngine.add("트리거", t["name"], "성공",
                              f"파일 변경 감지: {path}")
            else:
                self._file_mtimes[key] = mtime

        elif kind == "image_appear":
            img_path = t.get("image_path", "")
            if not CV2_OK or not img_path or not Path(img_path).exists():
                return
            try:
                loc = pyautogui.locateOnScreen(
                    img_path,
                    confidence=t.get("confidence", 0.80))
                if loc:
                    # 중복 실행 방지 (30초 쿨다운)
                    last_run = t.get("last_trigger_dt", "")
                    if last_run:
                        last_dt = datetime.datetime.fromisoformat(last_run)
                        if (datetime.datetime.now() - last_dt).seconds < 30:
                            return
                    t["last_trigger_dt"] = datetime.datetime.now().isoformat()
                    save_triggers(load_triggers())
                    self.triggered.emit(t["name"], t["macro_path"])
                    LogEngine.add("트리거", t["name"], "성공",
                                  f"이미지 감지: {Path(img_path).name}")
            except Exception:
                pass

        elif kind == "folder_new_file":
            self._check_folder_new_file(t)

        elif kind == "email_received":
            self._check_email_received(t)

        elif kind == "process_start":
            proc_name = t.get("process_name", "").lower()
            if not proc_name:
                return
            try:
                import subprocess
                result = subprocess.run(
                    ["powershell", "-NoProfile", "-Command",
                     f"Get-Process | Where-Object {{$_.Name -like '*{proc_name}*'}} | Select-Object -First 1 -ExpandProperty Name"],
                    capture_output=True, text=True, timeout=3,
                    creationflags=subprocess.CREATE_NO_WINDOW
                        if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
                found = result.stdout.strip()
                if found:
                    last_run = t.get("last_trigger_dt", "")
                    if last_run:
                        last_dt = datetime.datetime.fromisoformat(last_run)
                        if (datetime.datetime.now() - last_dt).seconds < 60:
                            return
                    t["last_trigger_dt"] = datetime.datetime.now().isoformat()
                    save_triggers(load_triggers())
                    self.triggered.emit(t["name"], t["macro_path"])
                    LogEngine.add("트리거", t["name"], "성공",
                                  f"프로세스 감지: {found}")
            except Exception:
                pass

    def _match_filename(self, fname: str, ext_filter: str, pattern_filter: str) -> bool:
        """확장자 필터 + 파일명 패턴 필터를 모두 통과해야 True"""
        import fnmatch
        if ext_filter:
            exts = [e.strip().lower() for e in ext_filter.split(",") if e.strip()]
            exts = [e if e.startswith(".") else "."+e for e in exts]
            if not any(fname.lower().endswith(e) for e in exts):
                return False
        if pattern_filter:
            if not fnmatch.fnmatch(fname.lower(), pattern_filter.strip().lower()):
                return False
        return True

    def _check_folder_new_file(self, t: dict):
        """
        특정 폴더에 새 파일이 추가되면 감지.
        폐쇄망 환경에서도 표준 라이브러리(os.listdir)만으로 동작 - 추가 설치 불필요.
        확장자 필터(ext_filter), 파일명 패턴 필터(pattern_filter, 와일드카드 * ? 지원)
        """
        folder = t.get("watch_path", "")
        if not folder or not Path(folder).is_dir():
            return
        key = t["id"]
        ext_filter     = t.get("ext_filter", "")
        pattern_filter = t.get("pattern_filter", "")

        try:
            current = set(os.listdir(folder))
        except Exception:
            return

        known = self._folder_known.get(key)
        if known is None:
            # 최초 확인: 기존 파일은 트리거 대상에서 제외 (베이스라인만 기록)
            self._folder_known[key] = current
            return

        new_files = current - known
        self._folder_known[key] = current
        if not new_files:
            return

        matched = sorted(f for f in new_files
                          if self._match_filename(f, ext_filter, pattern_filter))
        if matched:
            self.triggered.emit(t["name"], t["macro_path"])
            LogEngine.add("트리거", t["name"], "성공",
                          f"신규 파일 감지: {', '.join(matched[:5])}"
                          + (f" 외 {len(matched)-5}건" if len(matched) > 5 else ""))

    def _check_email_received(self, t: dict):
        """
        Outlook 데스크톱 앱(로컬 COM)으로 새 메일 수신 감지.
        인터넷 연결 불필요 - 로컬에 설치된 Outlook 앱과만 통신하므로 폐쇄망에서도 동작.
        단, 해당 PC에 Outlook 데스크톱 앱이 설치·로그인되어 있어야 함.
        제목 포함 필터(subject_filter), 발신자 포함 필터(sender_filter) 지원.
        """
        if not OUTLOOK_OK:
            if not self._outlook_warned:
                self._outlook_warned = True
                LogEngine.add("트리거", t["name"], "실패",
                              "pywin32 미설치 - 이메일 트리거를 사용할 수 없습니다. "
                              "pip install pywin32 필요")
            return

        key = t["id"]
        now_mono = time.time()
        # Outlook COM 부하 완화 - 15초에 한 번만 확인
        next_ok = self._email_next_check.get(key, 0)
        if now_mono < next_ok:
            return
        self._email_next_check[key] = now_mono + 15

        try:
            outlook = _win32com_client.Dispatch("Outlook.Application")
            ns      = outlook.GetNamespace("MAPI")
            inbox   = ns.GetDefaultFolder(6)  # 6 = olFolderInbox
            items   = inbox.Items
            items.Sort("[ReceivedTime]", True)  # 최신순 정렬
            if items.Count == 0:
                return
            latest = items.GetFirst()
            latest_time = latest.ReceivedTime

            baseline = self._email_baseline.get(key)
            if baseline is None:
                # 최초 확인: 현재 최신 메일 시각을 기준선으로 잡음 (기존 메일은 무시)
                self._email_baseline[key] = latest_time
                return

            if latest_time <= baseline:
                return  # 새 메일 없음

            # 기준시각 이후의 새 메일들을 모아 필터 적용
            subject_filter = t.get("subject_filter", "").strip().lower()
            sender_filter  = t.get("sender_filter", "").strip().lower()
            matched_subjects = []
            item = items.GetFirst()
            checked = 0
            while item is not None and checked < 30:  # 안전장치: 최대 30건만 확인
                try:
                    r_time = item.ReceivedTime
                except Exception:
                    break
                if r_time <= baseline:
                    break
                checked += 1
                try:
                    subj   = (item.Subject or "")
                    sender = (item.SenderName or "") + " " + (item.SenderEmailAddress or "")
                except Exception:
                    subj, sender = "", ""
                ok = True
                if subject_filter and subject_filter not in subj.lower():
                    ok = False
                if sender_filter and sender_filter not in sender.lower():
                    ok = False
                if ok:
                    matched_subjects.append(subj)
                item = items.GetNext()

            self._email_baseline[key] = latest_time

            if matched_subjects:
                self.triggered.emit(t["name"], t["macro_path"])
                preview = matched_subjects[0][:40]
                extra = f" 외 {len(matched_subjects)-1}건" if len(matched_subjects) > 1 else ""
                LogEngine.add("트리거", t["name"], "성공",
                              f"새 메일 수신: {preview}{extra}")
        except Exception as e:
            LogEngine.add("트리거", t["name"], "실패", f"Outlook 연결 오류: {e}")

    def _execute_macro(self, name: str, path: str):
        try:
            with open(path, encoding="utf-8") as f:
                d = json.load(f)
            acts = (d.get("actions",[]) if isinstance(d,dict) else d)
            if acts:
                w = PlayWorker(acts, 1, 1.0)
                w.start(); w.wait()
        except Exception as e:
            LogEngine.add("트리거", name, "실패", str(e))

    def stop(self):
        self.running = False


# =============================================
#  트리거 페이지
# =============================================
class TriggerPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._triggers = load_triggers()
        self._engine   = TriggerEngine()
        self._engine.triggered.connect(self._on_trigger)
        self._engine.set_triggers(self._triggers)
        self._engine.start()
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(14)
        root.addWidget(SectionHeader("TRIGGER  /  자동 트리거"))

        guide = QLabel(
            "  지정한 조건이 발생하면 매크로를 자동으로 실행합니다.\n"
            "  파일 변경 / 이미지 등장 / 프로그램 실행 / 폴더 신규 파일 / 메일 수신 감지 지원")
        guide.setStyleSheet(f"""
            color:{C['t2']}; font-size:12px;
            background:{C['bg1']};
            border:1px solid {C['border']};
            border-left:4px solid {C['ai']};
            border-radius:8px; padding:10px;
        """)
        guide.setWordWrap(True)
        root.addWidget(guide)

        # ── 트리거 추가 카드 ──
        add_card = QFrame()
        add_card.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};
                border:1.5px solid {C['border']};border-radius:14px;}}
        """)
        ag = QGridLayout(add_card)
        ag.setContentsMargins(18,14,18,14); ag.setSpacing(10)

        ag.addWidget(QLabel("트리거 이름:"), 0, 0)
        self.name_edit = StyledInput("예: 보고서 파일 변경 감지")
        ag.addWidget(self.name_edit, 0, 1, 1, 3)

        ag.addWidget(QLabel("트리거 유형:"), 1, 0)
        self.type_combo = StyledCombo()
        self.type_combo.addItems([
            "파일 변경  (파일이 수정되면 실행)",
            "이미지 등장  (화면에 이미지가 나타나면 실행)",
            "프로세스 시작  (특정 프로그램이 실행되면 실행)",
            "폴더 신규 파일  (폴더에 파일이 추가되면 실행)",
            "메일 수신  (Outlook 새 메일 도착 시 실행)",
        ])
        self.type_combo.currentIndexChanged.connect(self._on_type_change)
        ag.addWidget(self.type_combo, 1, 1, 1, 3)

        # 동적 설정 영역
        self.cfg_frame = QFrame()
        self.cfg_frame.setStyleSheet("QFrame{background:transparent;}")
        self.cfg_lay = QHBoxLayout(self.cfg_frame)
        self.cfg_lay.setContentsMargins(0,0,0,0); self.cfg_lay.setSpacing(8)
        ag.addWidget(self.cfg_frame, 2, 0, 1, 4)
        self._build_cfg(0)

        ag.addWidget(QLabel("매크로 파일:"), 3, 0)
        self.macro_edit = StyledInput("실행할 매크로 파일 경로")
        self.macro_edit.setReadOnly(True)
        btn_pick = GlowButton("선택", C['brand'])
        btn_pick.setFixedHeight(34); btn_pick.setFixedWidth(60)
        btn_pick.clicked.connect(self._pick_macro)
        ag.addWidget(self.macro_edit, 3, 1, 1, 2)
        ag.addWidget(btn_pick, 3, 3)

        self.macro_name_lbl = QLabel("")
        self.macro_name_lbl.setStyleSheet(
            f"color:{C['play']};font-size:11px;background:transparent;")
        ag.addWidget(self.macro_name_lbl, 4, 1, 1, 3)

        btn_add = GlowButton("+ 트리거 추가", C['ai'])
        btn_add.setFixedHeight(40)
        btn_add.clicked.connect(self._add_trigger)
        ag.addWidget(btn_add, 5, 0, 1, 4)
        root.addWidget(add_card)

        # ── 트리거 목록 ──
        list_hdr = QHBoxLayout()
        list_lbl = QLabel("등록된 트리거")
        list_lbl.setStyleSheet(
            f"color:{C['t1']};font-weight:bold;font-size:13px;")
        list_hdr.addWidget(list_lbl); list_hdr.addStretch()
        btn_ref = GlowButton("새로고침", C['brand'])
        btn_ref.setFixedHeight(30); btn_ref.setFixedWidth(80)
        btn_ref.clicked.connect(self._refresh_list)
        list_hdr.addWidget(btn_ref)
        root.addLayout(list_hdr)

        self.trig_list = QListWidget()
        self.trig_list.setStyleSheet(f"""
            QListWidget{{background:{C['bg1']};
                border:1.5px solid {C['border']};border-radius:12px;
                padding:4px;outline:none;}}
            QListWidget::item{{padding:10px 14px;border-radius:8px;
                margin:2px 3px;font-size:13px;color:{C['t1']};}}
            QListWidget::item:selected{{background:{C['ai']};color:#FFFFFF;}}
            QListWidget::item:hover{{background:{C['bg4']};}}
        """)
        root.addWidget(self.trig_list)

        ctrl_row = QHBoxLayout(); ctrl_row.setSpacing(8)
        btn_toggle = GlowButton("ON / OFF 전환", C['stop'])
        btn_toggle.setFixedHeight(38)
        btn_toggle.clicked.connect(self._toggle)
        btn_del = GlowButton("삭제", C['rec'])
        btn_del.setFixedHeight(38)
        btn_del.clicked.connect(self._delete)
        btn_test = GlowButton("지금 테스트 실행", C['play'])
        btn_test.setFixedHeight(38)
        btn_test.clicked.connect(self._test_run)
        ctrl_row.addWidget(btn_toggle)
        ctrl_row.addWidget(btn_del)
        ctrl_row.addStretch()
        ctrl_row.addWidget(btn_test)
        root.addLayout(ctrl_row)

        self.status_lbl = QLabel("트리거 엔진 실행 중...")
        self.status_lbl.setStyleSheet(f"""
            color:{C['t2']};font-size:12px;
            background:{C['bg1']};border:1px solid {C['border']};
            border-radius:8px;padding:8px 12px;
        """)
        root.addWidget(self.status_lbl)

        # 실행 로그
        log_lbl = QLabel("실행 로그")
        log_lbl.setStyleSheet(
            f"color:{C['t2']};font-size:11px;font-weight:bold;")
        root.addWidget(log_lbl)
        self.log_list = QListWidget()
        self.log_list.setMaximumHeight(110)
        self.log_list.setStyleSheet(f"""
            QListWidget{{background:{C['bg1']};border:1px solid {C['border']};
                border-radius:8px;padding:4px;outline:none;
                font-family:Consolas,monospace;font-size:11px;}}
            QListWidget::item{{padding:4px 10px;color:{C['t2']};}}
        """)
        root.addWidget(self.log_list)
        self._refresh_list()

    def _build_cfg(self, idx: int):
        while self.cfg_lay.count():
            item = self.cfg_lay.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        if idx == 0:   # 파일 변경
            self.cfg_lay.addWidget(QLabel("감시 파일/폴더:"))
            self.cfg_input = StyledInput("예: C:\\Users\\HHI\\Desktop\\report.xlsx")
            btn = GlowButton("선택", C['ai'])
            btn.setFixedHeight(34); btn.setFixedWidth(60)
            btn.clicked.connect(self._pick_watch_path)
            self.cfg_lay.addWidget(self.cfg_input)
            self.cfg_lay.addWidget(btn)
        elif idx == 1:  # 이미지 등장
            self.cfg_lay.addWidget(QLabel("감지 이미지:"))
            self.cfg_input = StyledInput("이미지 파일 경로 (.png)")
            self.cfg_input.setReadOnly(True)
            btn = GlowButton("선택", C['ai'])
            btn.setFixedHeight(34); btn.setFixedWidth(60)
            btn.clicked.connect(self._pick_img)
            self.cfg_lay.addWidget(self.cfg_input)
            self.cfg_lay.addWidget(btn)
        elif idx == 2:  # 프로세스 시작
            self.cfg_lay.addWidget(QLabel("프로세스 이름:"))
            self.cfg_input = StyledInput("예: notepad, excel, chrome")
            self.cfg_lay.addWidget(self.cfg_input)
        elif idx == 3:  # 폴더 신규 파일
            v = QVBoxLayout(); v.setSpacing(6)
            row1 = QHBoxLayout()
            row1.addWidget(QLabel("감시 폴더:"))
            self.cfg_input = StyledInput("예: C:\\Users\\HHI\\Desktop\\받은자료  (경로 직접 입력)")
            row1.addWidget(self.cfg_input)
            v.addLayout(row1)
            row2 = QHBoxLayout()
            row2.addWidget(QLabel("확장자 필터:"))
            self.ext_filter_input = StyledInput("예: .xlsx,.csv  (비우면 전체)")
            row2.addWidget(self.ext_filter_input)
            v.addLayout(row2)
            row3 = QHBoxLayout()
            row3.addWidget(QLabel("파일명 패턴:"))
            self.pattern_filter_input = StyledInput("예: report_*.xlsx  (비우면 전체)")
            row3.addWidget(self.pattern_filter_input)
            v.addLayout(row3)
            self.cfg_lay.addLayout(v)
        elif idx == 4:  # 메일 수신
            v = QVBoxLayout(); v.setSpacing(6)
            if not OUTLOOK_OK:
                warn = QLabel("⚠ pywin32 미설치 - 이 PC에서는 이메일 트리거를 사용할 수 없습니다.\n"
                              "(pip install pywin32 필요, Outlook 데스크톱 앱 설치도 필요)")
                warn.setStyleSheet(f"color:{C['stop']};font-size:11px;background:transparent;")
                warn.setWordWrap(True)
                v.addWidget(warn)
            row1 = QHBoxLayout()
            row1.addWidget(QLabel("제목 포함:"))
            self.subject_filter_input = StyledInput("예: 보고서  (비우면 전체 메일)")
            row1.addWidget(self.subject_filter_input)
            v.addLayout(row1)
            row2 = QHBoxLayout()
            row2.addWidget(QLabel("발신자 포함:"))
            self.sender_filter_input = StyledInput("예: hhi.co.kr  (비우면 전체 발신자)")
            row2.addWidget(self.sender_filter_input)
            v.addLayout(row2)
            self.cfg_input = None  # 이메일 트리거는 감시경로가 없으므로 별도 처리
            self.cfg_lay.addLayout(v)

    def _on_type_change(self, idx):
        self._build_cfg(idx)

    def _pick_watch_path(self):
        fp, _ = rpa_open_file(self, "감시 파일 선택", SAVE_DIR, ".json")
        if fp: self.cfg_input.setText(fp)

    def _pick_img(self):
        fp, _ = rpa_open_file(self, "감지 이미지 선택", SAVE_DIR, ".png")
        if fp: self.cfg_input.setText(fp)

    def _pick_macro(self):
        fp, _ = rpa_open_file(self, "매크로 선택", SAVE_DIR)
        if fp:
            self.macro_edit.setText(fp)
            try:
                with open(fp, encoding="utf-8") as f:
                    d = json.load(f)
                name = d.get("name", Path(fp).stem)
                cnt  = len((d.get("actions",[]) if isinstance(d,dict) else d))
                self.macro_name_lbl.setText(f"{name}  ({cnt}개 프로세스)")
            except Exception:
                self.macro_name_lbl.setText(Path(fp).name)

    def _add_trigger(self):
        name = self.name_edit.text().strip()
        path = self.macro_edit.text().strip()
        t_idx = self.type_combo.currentIndex()
        cfg  = getattr(self, 'cfg_input', None)
        cfg_val = cfg.text().strip() if cfg else ""

        if not name:
            msg_warn(self, "입력 오류", "트리거 이름을 입력하세요."); return
        if not path or not Path(path).exists():
            msg_warn(self, "입력 오류", "매크로 파일을 선택하세요."); return
        # 이메일 트리거(idx 4)는 감시경로가 없어 cfg_val 검사를 건너뜀
        if t_idx != 4 and not cfg_val:
            msg_warn(self, "입력 오류", "감지 대상을 설정하세요."); return
        if t_idx == 3 and not Path(cfg_val).is_dir():
            msg_warn(self, "입력 오류", "존재하는 폴더 경로를 입력하세요."); return
        if t_idx == 4 and not OUTLOOK_OK:
            msg_warn(self, "사용 불가", "이 PC에는 pywin32 가 설치되어 있지 않아 "
                     "이메일 트리거를 사용할 수 없습니다.\n(pip install pywin32)")
            return

        types = ["file_change", "image_appear", "process_start",
                 "folder_new_file", "email_received"]

        t = {
            "id":          _new_id(),
            "name":        name,
            "type":        types[t_idx],
            "macro_path":  path,
            "enabled":     True,
            "created":     datetime.datetime.now().isoformat(),
            "last_trigger_dt": "",
        }
        if t_idx == 0:   t["watch_path"]    = cfg_val
        elif t_idx == 1:
            t["image_path"]  = cfg_val
            t["confidence"]  = 0.80
        elif t_idx == 2: t["process_name"]  = cfg_val
        elif t_idx == 3:
            t["watch_path"]      = cfg_val
            t["ext_filter"]      = self.ext_filter_input.text().strip()
            t["pattern_filter"]  = self.pattern_filter_input.text().strip()
        elif t_idx == 4:
            t["subject_filter"]  = self.subject_filter_input.text().strip()
            t["sender_filter"]   = self.sender_filter_input.text().strip()

        self._triggers.append(t)
        save_triggers(self._triggers)
        self._engine.set_triggers(self._triggers)
        self._refresh_list()
        self.name_edit.clear()
        self.macro_edit.clear()
        self.macro_name_lbl.clear()
        if cfg: cfg.clear()
        self._add_log(f"[등록] {name}")
        self.status_lbl.setText(f"트리거 등록: {name}")

    def _refresh_list(self):
        self._triggers = load_triggers()
        self.trig_list.clear()
        TYPES = {
            "file_change":    "파일 변경",
            "image_appear":   "이미지 등장",
            "process_start":  "프로세스 시작",
            "folder_new_file":"폴더 신규 파일",
            "email_received": "메일 수신",
        }
        for t in self._triggers:
            enabled = t.get("enabled", True)
            icon    = "ON" if enabled else "OFF"
            c       = C['ai'] if enabled else C['t3']
            kind    = TYPES.get(t.get("type",""), "?")
            last    = t.get("last_trigger_dt", "")
            last_s  = f"  | 마지막: {last[:16]}" if last else ""
            item = QListWidgetItem(
                f"  [{icon}]  {t['name']}  |  {kind}{last_s}")
            item.setForeground(QColor(c))
            self.trig_list.addItem(item)
        n_on = sum(1 for t in self._triggers if t.get("enabled", True))
        self.status_lbl.setText(
            f"트리거 엔진 실행 중  |  총 {len(self._triggers)}개  |  활성 {n_on}개")

    def _toggle(self):
        row = self.trig_list.currentRow()
        if row < 0 or row >= len(self._triggers): return
        t = self._triggers[row]
        t["enabled"] = not t.get("enabled", True)
        save_triggers(self._triggers)
        self._engine.set_triggers(self._triggers)
        self._refresh_list()
        self._add_log(f"[{'ON' if t['enabled'] else 'OFF'}] {t['name']}")

    def _delete(self):
        row = self.trig_list.currentRow()
        if row < 0 or row >= len(self._triggers): return
        name = self._triggers[row]["name"]
        r = msg_ask(self, "삭제 확인", f"'{name}' 트리거를 삭제할까요?")
        if r == True:
            self._triggers.pop(row)
            save_triggers(self._triggers)
            self._engine.set_triggers(self._triggers)
            self._refresh_list()
            self._add_log(f"[삭제] {name}")

    def _test_run(self):
        row = self.trig_list.currentRow()
        if row < 0 or row >= len(self._triggers):
            msg_info(self, "선택 없음", "테스트할 트리거를 선택하세요."); return
        t = self._triggers[row]
        self._on_trigger(t["name"], t["macro_path"])

    def _on_trigger(self, name: str, path: str):
        self._add_log(f"[감지] {name}")
        self.status_lbl.setText(f"트리거 실행 중: {name}")
        try:
            with open(path, encoding="utf-8") as f:
                d = json.load(f)
            acts = (d.get("actions",[]) if isinstance(d,dict) else d)
            if acts:
                def _go():
                    w = PlayWorker(acts, 1, 1.0)
                    w.start(); w.wait()
                    ts = datetime.datetime.now().strftime("%H:%M:%S")
                    QTimer.singleShot(0, lambda: self._add_log(
                        f"[완료] {name}  {ts}"))
                    QTimer.singleShot(0, lambda: self.status_lbl.setText(
                        f"완료: {name}  ({ts})"))
                threading.Thread(target=_go, daemon=True).start()
        except Exception as e:
            self._add_log(f"[오류] {name}: {e}")

    def _add_log(self, msg: str):
        ts   = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        item = QListWidgetItem(f"  {ts}  {msg}")
        if "[감지]" in msg or "[완료]" in msg:
            item.setForeground(QColor(C['ai']))
        elif "[오류]" in msg:
            item.setForeground(QColor(C['rec']))
        else:
            item.setForeground(QColor(C['t2']))
        self.log_list.insertItem(0, item)
        if self.log_list.count() > 50:
            self.log_list.takeItem(self.log_list.count()-1)

    def closeEvent(self, event):
        self._engine.stop()
        super().closeEvent(event)


# =============================================
#  모니터링 대시보드 페이지
# =============================================
class DashboardPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._refresh_timer = QTimer()
        self._refresh_timer.timeout.connect(self._refresh)
        self._refresh_timer.start(3000)
        self._build()
        QTimer.singleShot(300, self._refresh)

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(12)
        root.addWidget(SectionHeader("DASHBOARD  /  실시간 현황"))

        # 통계 카드
        stat_row = QHBoxLayout(); stat_row.setSpacing(10)
        self.c_total   = StatCard("전체 실행",   "0", C['brand'])
        self.c_success = StatCard("성공",        "0", C['play'])
        self.c_fail    = StatCard("실패",        "0", C['rec'])
        self.c_today   = StatCard("오늘",        "0", C['stop'])
        self.c_sched   = StatCard("활성 스케줄", "0", C['tool'])
        self.c_trigger = StatCard("활성 트리거", "0", C['ai'])
        for c in [self.c_total, self.c_success, self.c_fail,
                  self.c_today, self.c_sched, self.c_trigger]:
            stat_row.addWidget(c)
        root.addLayout(stat_row)

        # 중단
        mid_row = QHBoxLayout(); mid_row.setSpacing(12)

        # 현재 실행 현황
        status_card = QFrame()
        status_card.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};border:1.5px solid {C['border']};
                border-radius:14px;}}
        """)
        sc_lay = QVBoxLayout(status_card)
        sc_lay.setContentsMargins(16,12,16,12); sc_lay.setSpacing(8)
        sc_t = QLabel("현재 실행 현황")
        sc_t.setStyleSheet(
            f"color:{C['brand']};font-size:12px;font-weight:900;background:transparent;")
        sc_lay.addWidget(sc_t)
        self.running_list = QListWidget()
        self.running_list.setStyleSheet(f"""
            QListWidget{{background:{C['bg2']};border:1px solid {C['border']};
                border-radius:8px;padding:4px;outline:none;font-size:12px;}}
            QListWidget::item{{padding:8px 12px;border-radius:6px;color:{C['t1']};}}
            QListWidget::item:hover{{background:{C['bg4']};}}
        """)
        sc_lay.addWidget(self.running_list)
        mid_row.addWidget(status_card, 1)

        # 최근 로그
        log_card = QFrame()
        log_card.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};border:1.5px solid {C['border']};
                border-radius:14px;}}
        """)
        lc_lay = QVBoxLayout(log_card)
        lc_lay.setContentsMargins(16,12,16,12); lc_lay.setSpacing(8)
        lc_t = QLabel("최근 실행 로그")
        lc_t.setStyleSheet(
            f"color:{C['brand']};font-size:12px;font-weight:900;background:transparent;")
        lc_lay.addWidget(lc_t)
        self.recent_log = QListWidget()
        self.recent_log.setStyleSheet(f"""
            QListWidget{{background:{C['bg2']};border:1px solid {C['border']};
                border-radius:8px;padding:4px;outline:none;
                font-family:Consolas,monospace;font-size:11px;}}
            QListWidget::item{{padding:5px 10px;color:{C['t2']};}}
            QListWidget::item:hover{{background:{C['bg4']};}}
        """)
        lc_lay.addWidget(self.recent_log)
        mid_row.addWidget(log_card, 1)
        root.addLayout(mid_row)

        # 하단
        bot_row = QHBoxLayout(); bot_row.setSpacing(12)

        # 스케줄 현황
        sched_card = QFrame()
        sched_card.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};border:1.5px solid {C['border']};
                border-radius:14px;}}
        """)
        sl = QVBoxLayout(sched_card)
        sl.setContentsMargins(16,12,16,12); sl.setSpacing(8)
        st = QLabel("스케줄 현황")
        st.setStyleSheet(
            f"color:{C['stop']};font-size:12px;font-weight:900;background:transparent;")
        sl.addWidget(st)
        self.sched_status_list = QListWidget()
        self.sched_status_list.setStyleSheet(f"""
            QListWidget{{background:{C['bg2']};border:1px solid {C['border']};
                border-radius:8px;padding:4px;outline:none;font-size:12px;}}
            QListWidget::item{{padding:7px 12px;border-radius:4px;color:{C['t1']};}}
            QListWidget::item:hover{{background:{C['bg4']};}}
        """)
        sl.addWidget(self.sched_status_list)
        bot_row.addWidget(sched_card, 1)

        # 트리거 현황
        trig_card = QFrame()
        trig_card.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};border:1.5px solid {C['border']};
                border-radius:14px;}}
        """)
        tl = QVBoxLayout(trig_card)
        tl.setContentsMargins(16,12,16,12); tl.setSpacing(8)
        tt = QLabel("트리거 현황")
        tt.setStyleSheet(
            f"color:{C['ai']};font-size:12px;font-weight:900;background:transparent;")
        tl.addWidget(tt)
        self.trig_status_list = QListWidget()
        self.trig_status_list.setStyleSheet(f"""
            QListWidget{{background:{C['bg2']};border:1px solid {C['border']};
                border-radius:8px;padding:4px;outline:none;font-size:12px;}}
            QListWidget::item{{padding:7px 12px;border-radius:4px;color:{C['t1']};}}
            QListWidget::item:hover{{background:{C['bg4']};}}
        """)
        tl.addWidget(self.trig_status_list)
        bot_row.addWidget(trig_card, 1)
        root.addLayout(bot_row)

        # 하단 갱신 표시
        ref_row = QHBoxLayout()
        ref_row.addStretch()
        self.last_update_lbl = QLabel("")
        self.last_update_lbl.setStyleSheet(
            f"color:{C['t3']};font-size:10px;background:transparent;")
        btn_ref = GlowButton("새로고침", C['brand'])
        btn_ref.setFixedHeight(32); btn_ref.setFixedWidth(90)
        btn_ref.clicked.connect(self._refresh)
        ref_row.addWidget(self.last_update_lbl)
        ref_row.addWidget(btn_ref)
        root.addLayout(ref_row)

    def _refresh(self):
        logs    = LogEngine.get_all()
        today   = datetime.date.today().strftime("%Y-%m-%d")
        total   = len(logs)
        success = sum(1 for e in logs if e["result"] == "성공")
        fail    = sum(1 for e in logs if e["result"] == "실패")
        t_today = sum(1 for e in logs if e["ts"].startswith(today))
        self.c_total.setValue(total)
        self.c_success.setValue(success)
        self.c_fail.setValue(fail)
        self.c_today.setValue(t_today)

        # 스케줄
        scheds = load_schedules()
        n_on   = sum(1 for s in scheds if s.get("enabled", True))
        self.c_sched.setValue(n_on)
        self.sched_status_list.clear()
        MODE_KR = {"daily":"매일","weekday":"요일","once":"1회","interval":"간격"}
        for s in scheds:
            enabled = s.get("enabled", True)
            c       = C['play'] if enabled else C['t3']
            mode    = MODE_KR.get(s.get("mode","daily"), "?")
            t_val   = s.get("time","") or f"{s.get('interval_min',0)}분"
            last    = s.get("last_run","") or s.get("last_run_dt","")
            last_s  = f"  마지막:{last[:16]}" if last else ""
            item = QListWidgetItem(
                f"  [{'ON' if enabled else 'OFF'}]  {s['name']}  {mode} {t_val}{last_s}")
            item.setForeground(QColor(c))
            self.sched_status_list.addItem(item)
        if not scheds:
            self.sched_status_list.addItem(QListWidgetItem("  등록된 스케줄 없음"))

        # 트리거
        triggers = load_triggers()
        n_trig   = sum(1 for t in triggers if t.get("enabled", True))
        self.c_trigger.setValue(n_trig)
        self.trig_status_list.clear()
        TRIG_KR = {
            "file_change":    "파일 변경",
            "image_appear":   "이미지 등장",
            "process_start":  "프로세스 시작",
        }
        for t in triggers:
            enabled = t.get("enabled", True)
            c       = C['ai'] if enabled else C['t3']
            kind    = TRIG_KR.get(t.get("type",""), "?")
            item    = QListWidgetItem(
                f"  [{'ON' if enabled else 'OFF'}]  {t['name']}  |  {kind}")
            item.setForeground(QColor(c))
            self.trig_status_list.addItem(item)
        if not triggers:
            self.trig_status_list.addItem(QListWidgetItem("  등록된 트리거 없음"))

        # 최근 로그
        self.recent_log.clear()
        RESULT_C = {"성공":C['play'],"실패":C['rec'],"중지":C['stop']}
        for e in logs[:15]:
            item = QListWidgetItem(
                f"  {e['ts'][5:]}  [{e['result']}]  {e['action']}  {e['target'][:20]}")
            item.setForeground(QColor(RESULT_C.get(e["result"], C['t2'])))
            self.recent_log.addItem(item)

        # 현재 실행 중
        self.running_list.clear()
        running = [e for e in logs[:5] if e.get("result") == "실행중"]
        if running:
            for e in running:
                item = QListWidgetItem(
                    f"  RUNNING  {e['action']}  {e['target']}")
                item.setForeground(QColor(C['brand']))
                self.running_list.addItem(item)
        else:
            item = QListWidgetItem("  현재 실행 중인 작업 없음")
            item.setForeground(QColor(C['t3']))
            self.running_list.addItem(item)

        ts = datetime.datetime.now().strftime("%H:%M:%S")
        self.last_update_lbl.setText(f"마지막 갱신: {ts}  (3초마다 자동 갱신)")

# =============================================
#  워크플로우 페이지
# =============================================

# =============================================
#  워크플로우 데이터 모델
# =============================================
import uuid as _uuid

WORKFLOW_FILE = SAVE_DIR / "workflows.json"

def _new_id():
    return _uuid.uuid4().hex[:8]

def load_workflows() -> list:
    try:
        if WORKFLOW_FILE.exists():
            with open(WORKFLOW_FILE, encoding="utf-8") as f:
                data = json.load(f)
                if data:
                    return data
    except Exception:
        pass
    # 첫 실행 시 샘플 프로세스 자동 생성
    sample = _make_sample_workflow()
    save_workflows([sample])
    return [sample]


def _make_sample_workflow() -> dict:
    """첫 실행용 샘플 프로세스"""
    def nid(): return _new_id()
    s_id = nid(); p1_id = nid(); p2_id = nid()
    w_id = nid(); p3_id = nid(); e_id  = nid()
    return {
        "id":   nid(),
        "name": "[샘플] 메모장 자동화",
        "nodes": [
            {"id": s_id,  "type": "start",   "name": "시작",
             "x": 300, "y": 80,  "macro_path": "", "macro_name": "",
             "wait_sec": 1, "cond_img": "", "on_fail": "stop"},
            {"id": p1_id, "type": "process", "name": "메모장 열기",
             "x": 300, "y": 200, "macro_path": "", "macro_name": "매크로 파일 연결 필요",
             "wait_sec": 1, "cond_img": "", "on_fail": "stop"},
            {"id": p2_id, "type": "process", "name": "텍스트 입력",
             "x": 300, "y": 320, "macro_path": "", "macro_name": "매크로 파일 연결 필요",
             "wait_sec": 1, "cond_img": "", "on_fail": "continue"},
            {"id": w_id,  "type": "wait",    "name": "잠시 대기",
             "x": 300, "y": 440, "macro_path": "", "macro_name": "",
             "wait_sec": 2, "cond_img": "", "on_fail": "stop"},
            {"id": p3_id, "type": "process", "name": "파일 저장",
             "x": 300, "y": 560, "macro_path": "", "macro_name": "매크로 파일 연결 필요",
             "wait_sec": 1, "cond_img": "", "on_fail": "stop"},
            {"id": e_id,  "type": "end",     "name": "종료",
             "x": 300, "y": 680, "macro_path": "", "macro_name": "",
             "wait_sec": 1, "cond_img": "", "on_fail": "stop"},
        ],
        "edges": [
            {"id": nid(), "src": s_id,  "dst": p1_id, "label": ""},
            {"id": nid(), "src": p1_id, "dst": p2_id, "label": ""},
            {"id": nid(), "src": p2_id, "dst": w_id,  "label": ""},
            {"id": nid(), "src": w_id,  "dst": p3_id, "label": ""},
            {"id": nid(), "src": p3_id, "dst": e_id,  "label": ""},
        ]
    }

def save_workflows(wfs: list):
    with open(WORKFLOW_FILE, "w", encoding="utf-8") as f:
        json.dump(wfs, f, ensure_ascii=False, indent=2)


# =============================================
#  워크플로우 노드 타입 정의
# =============================================
WF_NODE_TYPES = {
    "start":   {"label": "시작",      "color": C["play"],  "shape": "ellipse",  "icon": "S"},
    "end":     {"label": "종료",      "color": C["rec"],   "shape": "ellipse",  "icon": "E"},
    "process": {"label": "프로세스",  "color": C["brand"], "shape": "rect",     "icon": "P"},
    "wait":    {"label": "대기",      "color": C["stop"],  "shape": "rect",     "icon": "W"},
    "cond":    {"label": "조건",      "color": "#F59E0B",  "shape": "diamond",  "icon": "?"},
}

# 노드 크기
NODE_W  = 160
NODE_H  = 60
COND_W  = 120
COND_H  = 70


# =============================================
#  워크플로우 노드 편집 다이얼로그
# =============================================
class WFNodeDialog(QDialog):
    def __init__(self, node: dict, parent=None):
        super().__init__(parent)
        self.node = dict(node)
        self.setWindowTitle("노드 편집")
        self.setMinimumWidth(420)
        self.setStyleSheet(f"QDialog{{background:{C['bg1']};}} QLabel{{background:transparent;}}")
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20,18,20,18); lay.setSpacing(12)

        # 노드 타입
        row_type = QHBoxLayout()
        row_type.addWidget(QLabel("노드 타입:"))
        self.type_combo = StyledCombo()
        for k, v in WF_NODE_TYPES.items():
            self.type_combo.addItem(v["label"], k)
        cur = self.node.get("type","process")
        idx = list(WF_NODE_TYPES.keys()).index(cur) if cur in WF_NODE_TYPES else 0
        self.type_combo.setCurrentIndex(idx)
        self.type_combo.currentIndexChanged.connect(self._on_type_change)
        row_type.addWidget(self.type_combo)
        lay.addLayout(row_type)

        # 노드 이름
        row_name = QHBoxLayout()
        row_name.addWidget(QLabel("이름:"))
        self.name_edit = StyledInput("예: 보고서 작성, 이메일 발송")
        self.name_edit.setText(self.node.get("name",""))
        row_name.addWidget(self.name_edit)
        lay.addLayout(row_name)

        # 연결 매크로 (process 타입만)
        self.macro_frame = QFrame()
        self.macro_frame.setStyleSheet("QFrame{background:transparent;}")
        mf_lay = QHBoxLayout(self.macro_frame)
        mf_lay.setContentsMargins(0,0,0,0)
        mf_lay.addWidget(QLabel("매크로:"))
        self.macro_edit = StyledInput("연결할 매크로 파일")
        self.macro_edit.setReadOnly(True)
        self.macro_edit.setText(self.node.get("macro_path",""))
        btn_pick = GlowButton("선택", C["brand"])
        btn_pick.setFixedHeight(34); btn_pick.setFixedWidth(60)
        btn_pick.clicked.connect(self._pick_macro)
        self.macro_name_lbl = QLabel(self.node.get("macro_name",""))
        self.macro_name_lbl.setStyleSheet(f"color:{C['play']}; font-size:11px;")
        mf_lay.addWidget(self.macro_edit)
        mf_lay.addWidget(btn_pick)
        lay.addWidget(self.macro_frame)
        lay.addWidget(self.macro_name_lbl)

        # 반복 횟수 (process 타입만) - 이 노드만 N회 반복 실행
        self.repeat_frame = QFrame()
        self.repeat_frame.setStyleSheet("QFrame{background:transparent;}")
        rf_lay = QHBoxLayout(self.repeat_frame)
        rf_lay.setContentsMargins(0,0,0,0); rf_lay.setSpacing(8)
        rf_lay.addWidget(QLabel("반복 횟수:"))
        self.repeat_spin = StyledSpin()
        self.repeat_spin.setRange(1, 9999)
        self.repeat_spin.setValue(self.node.get("repeat", 1))
        self.repeat_spin.setSuffix("  회")
        self.repeat_spin.setToolTip("이 프로세스 블록만 N회 반복 실행")
        rf_lay.addWidget(self.repeat_spin)
        rf_lay.addWidget(QLabel("   간격:"))
        self.repeat_gap_spin = StyledDSpin()
        self.repeat_gap_spin.setRange(0, 600)
        self.repeat_gap_spin.setValue(self.node.get("repeat_gap", 0.0))
        self.repeat_gap_spin.setSuffix("  초")
        self.repeat_gap_spin.setToolTip("반복 사이 대기 시간")
        rf_lay.addWidget(self.repeat_gap_spin)
        rf_lay.addStretch()
        lay.addWidget(self.repeat_frame)

        # 대기 시간 (wait 타입만)
        self.wait_frame = QFrame()
        self.wait_frame.setStyleSheet("QFrame{background:transparent;}")
        wf_lay = QHBoxLayout(self.wait_frame)
        wf_lay.setContentsMargins(0,0,0,0)
        wf_lay.addWidget(QLabel("대기 시간:"))
        self.wait_spin = StyledDSpin()
        self.wait_spin.setRange(0.5, 3600)
        self.wait_spin.setValue(self.node.get("wait_sec", 5.0))
        self.wait_spin.setSuffix("  초")
        wf_lay.addWidget(self.wait_spin)
        wf_lay.addStretch()
        lay.addWidget(self.wait_frame)

        # 조건 (cond 타입만)
        self.cond_frame = QFrame()
        self.cond_frame.setStyleSheet("QFrame{background:transparent;}")
        cf_lay = QVBoxLayout(self.cond_frame)
        cf_lay.setContentsMargins(0,0,0,0); cf_lay.setSpacing(6)
        cf_lay.addWidget(QLabel("조건 (이미지 파일):"))
        cond_row = QHBoxLayout()
        self.cond_img_edit = StyledInput("조건 이미지 파일 경로")
        self.cond_img_edit.setText(self.node.get("cond_img",""))
        btn_cond_img = GlowButton("선택", C["ai"])
        btn_cond_img.setFixedHeight(34); btn_cond_img.setFixedWidth(60)
        btn_cond_img.clicked.connect(self._pick_cond_img)
        cond_row.addWidget(self.cond_img_edit); cond_row.addWidget(btn_cond_img)
        cf_lay.addLayout(cond_row)
        cf_lay.addWidget(QLabel("YES 연결 → 하단 왼쪽  /  NO 연결 → 하단 오른쪽"))
        lay.addWidget(self.cond_frame)

        # 실패 처리
        self.fail_frame = QFrame()
        self.fail_frame.setStyleSheet("QFrame{background:transparent;}")
        ff_lay = QHBoxLayout(self.fail_frame)
        ff_lay.setContentsMargins(0,0,0,0)
        ff_lay.addWidget(QLabel("실패 시:"))
        self.fail_combo = StyledCombo()
        self.fail_combo.addItems(["프로세스 중단", "다음 노드 진행", "재시도 (3회)"])
        fail_map = {"stop":0,"continue":1,"retry":2}
        self.fail_combo.setCurrentIndex(
            fail_map.get(self.node.get("on_fail","stop"), 0))
        ff_lay.addWidget(self.fail_combo)
        lay.addWidget(self.fail_frame)

        # 버튼
        btn_row = QHBoxLayout()
        btn_cancel = GlowButton("취소", C["rec"])
        btn_cancel.setFixedHeight(36); btn_cancel.setFixedWidth(80)
        btn_cancel.clicked.connect(self.reject)
        btn_ok = GlowButton("확인", C["play"])
        btn_ok.setFixedHeight(36); btn_ok.setFixedWidth(80)
        btn_ok.clicked.connect(self._accept)
        btn_row.addStretch()
        btn_row.addWidget(btn_cancel); btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

        self._on_type_change()

    def _on_type_change(self):
        t = self.type_combo.currentData()
        self.macro_frame.setVisible(t == "process")
        self.macro_name_lbl.setVisible(t == "process")
        self.repeat_frame.setVisible(t == "process")
        self.wait_frame.setVisible(t == "wait")
        self.cond_frame.setVisible(t == "cond")
        self.fail_frame.setVisible(t in ("process","cond"))

    def _pick_macro(self):
        fp, _ = rpa_open_file(self, "매크로 선택", SAVE_DIR)
        if fp:
            self.macro_edit.setText(fp)
            try:
                with open(fp, encoding="utf-8") as f:
                    d = json.load(f)
                name = d.get("name", Path(fp).stem)
                cnt  = len(d.get("actions",[]))
                self.macro_name_lbl.setText(f"{name}  ({cnt}개 프로세스)")
                self.node["macro_name"] = name
            except Exception:
                self.macro_name_lbl.setText(Path(fp).name)

    def _pick_cond_img(self):
        fp, _ = rpa_open_file(self, "조건 이미지 선택", SAVE_DIR, ".png")
        if fp: self.cond_img_edit.setText(fp)

    def _accept(self):
        t = self.type_combo.currentData()
        self.node["type"]    = t
        self.node["name"]    = self.name_edit.text().strip() or WF_NODE_TYPES[t]["label"]
        if t == "process":
            self.node["macro_path"]  = self.macro_edit.text().strip()
            self.node["repeat"]      = self.repeat_spin.value()
            self.node["repeat_gap"]  = self.repeat_gap_spin.value()
        elif t == "wait":
            self.node["wait_sec"] = self.wait_spin.value()
        elif t == "cond":
            self.node["cond_img"] = self.cond_img_edit.text().strip()
        fail_vals = ["stop","continue","retry"]
        self.node["on_fail"] = fail_vals[self.fail_combo.currentIndex()]
        self.accept()


# =============================================
#  워크플로우 캔버스 (QPainter 기반)
# =============================================
class WFCanvas(QWidget):
    """
    드래그로 노드 이동, 클릭으로 선택/편집,
    노드 포트 드래그로 연결선 그리기
    """
    def keyPressEvent(self, event):
        """DEL 키 → 선택된 노드 삭제"""
        if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace):
            if self._sel_id:
                self._save_undo()
                self.delete_node(self._sel_id)
        else:
            super().keyPressEvent(event)

    def focusInEvent(self, event):
        super().focusInEvent(event)


    node_selected  = pyqtSignal(dict)
    canvas_changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._nodes   : list = []
        self._edges   : list = []
        self._sel_id  : str  = ""
        self._drag_node = None
        self._drag_off  = QPoint()
        # 드래그 연결
        self._conn_src  = None      # 연결 시작 노드 id
        self._conn_end  = QPointF() # 연결선 끝점 (드래그 중)
        self._connecting = False    # 연결 드래그 중 여부
        # Ctrl+Z 실행 취소 스택 (최대 30단계)
        self._undo_stack: list = []
        self.setMinimumSize(800, 600)
        self.setStyleSheet(f"background:{C['bg0']};")
        self.setMouseTracking(True)

    # ── 실행 취소 ──
    def _save_undo(self):
        """현재 상태를 undo 스택에 저장"""
        import copy
        self._undo_stack.append(
            (copy.deepcopy(self._nodes), copy.deepcopy(self._edges)))
        if len(self._undo_stack) > 30:
            self._undo_stack.pop(0)

    def undo(self):
        """직전 상태로 복원"""
        if not self._undo_stack: return
        self._nodes, self._edges = self._undo_stack.pop()
        self._sel_id = ""
        self.update()
        self.canvas_changed.emit()

    # ── 데이터 ──
    def set_data(self, nodes, edges):
        self._nodes = [dict(n) for n in nodes]
        self._edges = [dict(e) for e in edges]
        self.update()

    def get_data(self):
        return [dict(n) for n in self._nodes], [dict(e) for e in self._edges]

    def add_node(self, ntype: str, x=None, y=None) -> dict:
        self._save_undo()
        cx = x if x else 200 + len(self._nodes)*30
        cy = y if y else 200 + (len(self._nodes)%4)*90
        node = {
            "id":   _new_id(),
            "type": ntype,
            "name": WF_NODE_TYPES[ntype]["label"],
            "x":    cx, "y": cy,
            "macro_path": "",
            "macro_name": "",
            "wait_sec":   5.0,
            "cond_img":   "",
            "on_fail":    "stop",
        }
        self._nodes.append(node)
        self.update()
        self.canvas_changed.emit()
        return node

    def delete_node(self, nid: str):
        self._save_undo()
        self._nodes = [n for n in self._nodes if n["id"] != nid]
        self._edges = [e for e in self._edges
                       if e["src"] != nid and e["dst"] != nid]
        if self._sel_id == nid: self._sel_id = ""
        self.update(); self.canvas_changed.emit()

    def add_edge(self, src_id, dst_id, label=""):
        self._save_undo()
        for e in self._edges:
            if e["src"] == src_id and e["dst"] == dst_id: return
        self._edges.append({
            "id": _new_id(), "src": src_id, "dst": dst_id, "label": label})
        self.update(); self.canvas_changed.emit()

    def delete_edge(self, eid: str):
        self._edges = [e for e in self._edges if e["id"] != eid]
        self.update(); self.canvas_changed.emit()

    # ── 좌표 헬퍼 ──
    def _node_rect(self, n: dict) -> QRect:
        t = n.get("type","process")
        if t == "cond":
            return QRect(int(n["x"]-COND_W//2), int(n["y"]-COND_H//2),
                         COND_W, COND_H)
        elif t in ("start","end"):
            r = 36
            return QRect(int(n["x"]-r), int(n["y"]-r), r*2, r*2)
        return QRect(int(n["x"]-NODE_W//2), int(n["y"]-NODE_H//2),
                     NODE_W, NODE_H)

    def _node_at(self, pos: QPoint):
        for n in reversed(self._nodes):
            if self._node_rect(n).contains(pos): return n
        return None

    def _port_hit(self, pos: QPoint, port="bottom", radius=12):
        """포트 근처 클릭 감지"""
        for n in reversed(self._nodes):
            pp = self._port_pos(n, port)
            if (QPointF(pos) - pp).manhattanLength() < radius:
                return n
        return None

    def _port_pos(self, n: dict, port="bottom") -> QPointF:
        r = self._node_rect(n)
        cx, cy = r.center().x(), r.center().y()
        if port == "top":    return QPointF(cx, r.top())
        if port == "bottom": return QPointF(cx, r.bottom())
        if port == "left":   return QPointF(r.left(), cy)
        if port == "right":  return QPointF(r.right(), cy)
        return QPointF(cx, cy)

    # ── 마우스 이벤트 ──
    def mousePressEvent(self, event):
        self.setFocus()   # DEL 키 동작을 위해 포커스 획득
        pos = event.pos()

        if event.button() == Qt.MouseButton.LeftButton:
            # 포트(하단) 클릭이면 연결 드래그 시작
            port_node = self._port_hit(pos, "bottom")
            if port_node:
                self._connecting = True
                self._conn_src   = port_node["id"]
                self._conn_end   = QPointF(pos)
                self.setCursor(Qt.CursorShape.CrossCursor)
                self.update()
                return

            n = self._node_at(pos)
            if n:
                self._sel_id    = n["id"]
                self._drag_node = n["id"]
                self._drag_off  = pos - QPoint(int(n["x"]), int(n["y"]))
                self.update()
            else:
                self._sel_id = ""
                self.update()

        elif event.button() == Qt.MouseButton.RightButton:
            n = self._node_at(pos)
            if n:
                self._sel_id = n["id"]
                self._show_node_menu(n, event.globalPosition().toPoint())

    def mouseMoveEvent(self, event):
        if self._connecting:
            self._conn_end = QPointF(event.pos())
            self.update()
            return
        if self._drag_node and event.buttons() & Qt.MouseButton.LeftButton:
            for n in self._nodes:
                if n["id"] == self._drag_node:
                    new_pos = event.pos() - self._drag_off
                    n["x"] = max(80, new_pos.x())
                    n["y"] = max(60, new_pos.y())
                    break
            self.update()

    def mouseReleaseEvent(self, event):
        if self._connecting:
            self._connecting = False
            self.setCursor(Qt.CursorShape.ArrowCursor)
            # 도착 노드 찾기 (상단 포트 or 노드 안)
            pos = event.pos()
            dst = self._port_hit(pos, "top") or self._node_at(pos)
            if dst and dst["id"] != self._conn_src:
                # cond 노드면 YES/NO 선택
                src_node = next((n for n in self._nodes
                                 if n["id"] == self._conn_src), None)
                label = ""
                if src_node and src_node.get("type") == "cond":
                    # 기존 엣지에서 YES/NO 자동 배정
                    existing = [e["label"] for e in self._edges
                                if e["src"] == self._conn_src]
                    label = "NO" if "YES" in existing else "YES"
                self.add_edge(self._conn_src, dst["id"], label)
            self._conn_src = None
            self.update()
            return
        if self._drag_node:
            self._save_undo()
            self.canvas_changed.emit()
        self._drag_node = None

    def mouseDoubleClickEvent(self, event):
        n = self._node_at(event.pos())
        if n: self._edit_node(n)

    def _show_node_menu(self, node: dict, gpos: QPoint):
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{ background:{C['bg1']}; border:1.5px solid {C['border']};
                border-radius:10px; padding:4px; font-size:13px; }}
            QMenu::item {{ padding:8px 20px; border-radius:6px; color:{C['t1']}; }}
            QMenu::item:selected {{ background:{C['brand']}; color:#FFFFFF; }}
            QMenu::separator {{ height:1px; background:{C['border']}; margin:4px 8px; }}
        """)
        act_edit = QAction("편집", menu)
        act_edit.triggered.connect(lambda: self._edit_node(node))
        act_conn = QAction("연결선 그리기", menu)
        act_conn.triggered.connect(lambda: self._start_connect(node))
        act_del  = QAction("삭제", menu)
        act_del.triggered.connect(lambda: self.delete_node(node["id"]))
        menu.addAction(act_edit)
        menu.addSeparator()
        menu.addAction(act_conn)
        menu.addSeparator()
        menu.addAction(act_del)
        menu.exec(gpos)

    def _edit_node(self, node: dict):
        dlg = WFNodeDialog(node, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            for i, n in enumerate(self._nodes):
                if n["id"] == node["id"]:
                    self._nodes[i] = dlg.node
                    break
            self.update(); self.canvas_changed.emit()

    def _start_connect(self, node: dict):
        # 연결 대상 선택 다이얼로그
        others = [n for n in self._nodes if n["id"] != node["id"]]
        if not others:
            msg_info(self, "없음", "연결할 다른 노드가 없습니다."); return
        dlg = QDialog(self)
        dlg.setWindowTitle("연결 대상 선택")
        dlg.setMinimumWidth(320)
        dlg.setStyleSheet(f"QDialog{{background:{C['bg1']};}} QLabel{{background:transparent;}}")
        dlay = QVBoxLayout(dlg)
        dlay.addWidget(QLabel("연결할 노드를 선택하세요:"))
        lst = QListWidget()
        lst.setStyleSheet(f"""
            QListWidget{{background:{C['bg2']};border:1px solid {C['border']};
                border-radius:8px;padding:4px;outline:none;}}
            QListWidget::item{{padding:8px 12px;border-radius:4px;}}
            QListWidget::item:selected{{background:{C['brand']};color:#FFFFFF;}}
        """)
        for n in others:
            item = QListWidgetItem(
                f"  {WF_NODE_TYPES.get(n['type'],{}).get('label','?')}  :  {n['name']}")
            item.setData(Qt.ItemDataRole.UserRole, n["id"])
            lst.addItem(item)
        dlay.addWidget(lst)

        # 조건 노드면 YES/NO 레이블 선택
        lbl_combo = None
        if node.get("type") == "cond":
            dlay.addWidget(QLabel("연결 레이블:"))
            lbl_combo = StyledCombo()
            lbl_combo.addItems(["YES", "NO"])
            dlay.addWidget(lbl_combo)

        btn_row = QHBoxLayout()
        btn_cancel = GlowButton("취소",C["rec"]); btn_cancel.setFixedHeight(34)
        btn_ok     = GlowButton("연결",C["play"]); btn_ok.setFixedHeight(34)
        btn_cancel.clicked.connect(dlg.reject)
        btn_ok.clicked.connect(dlg.accept)
        btn_row.addStretch(); btn_row.addWidget(btn_cancel); btn_row.addWidget(btn_ok)
        dlay.addLayout(btn_row)

        if dlg.exec() == QDialog.DialogCode.Accepted:
            item = lst.currentItem()
            if item:
                dst_id = item.data(Qt.ItemDataRole.UserRole)
                lbl    = lbl_combo.currentText() if lbl_combo else ""
                self.add_edge(node["id"], dst_id, lbl)

    # ── 페인팅 ──
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # 배경 격자
        painter.setPen(QPen(QColor(C["border"]), 1))
        grid = 40
        for gx in range(0, self.width(), grid):
            painter.drawLine(gx, 0, gx, self.height())
        for gy in range(0, self.height(), grid):
            painter.drawLine(0, gy, self.width(), gy)

        # 엣지 먼저
        self._draw_edges(painter)

        # 노드
        for n in self._nodes:
            self._draw_node(painter, n)

        # 드래그 연결 중 임시 선
        if self._connecting and self._conn_src:
            src_n = next((n for n in self._nodes
                          if n["id"] == self._conn_src), None)
            if src_n:
                p1 = self._port_pos(src_n, "bottom")
                p2 = self._conn_end
                path = QPainterPath()
                path.moveTo(p1)
                cy = (p1.y() + p2.y()) / 2
                path.cubicTo(QPointF(p1.x(), cy),
                             QPointF(p2.x(), cy), p2)
                pen = QPen(QColor(C['brand']), 2.5, Qt.PenStyle.DashLine)
                painter.setPen(pen)
                painter.setBrush(Qt.BrushStyle.NoBrush)
                painter.drawPath(path)

        painter.end()

    def _draw_edges(self, painter: QPainter):
        for e in self._edges:
            src = next((n for n in self._nodes if n["id"]==e["src"]), None)
            dst = next((n for n in self._nodes if n["id"]==e["dst"]), None)
            if not src or not dst: continue

            p1 = QPointF(self._port_pos(src, "bottom"))
            p2 = QPointF(self._port_pos(dst, "top"))

            # 베지어 곡선
            path = QPainterPath()
            path.moveTo(p1)
            cy = (p1.y() + p2.y()) / 2
            path.cubicTo(QPointF(p1.x(), cy),
                         QPointF(p2.x(), cy),
                         QPointF(p2.x(), p2.y()))

            painter.setPen(QPen(QColor(C["brand"]), 2.5))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawPath(path)

            # 화살표
            angle_pt = path.pointAtPercent(0.95)
            tip_pt   = path.pointAtPercent(1.0)
            dx = tip_pt.x() - angle_pt.x()
            dy = tip_pt.y() - angle_pt.y()
            import math
            length = math.sqrt(dx*dx + dy*dy) or 1
            dx /= length; dy /= length
            arr_len = 12; arr_w = 6
            lx = tip_pt.x() - arr_len*dx + arr_w*dy
            ly = tip_pt.y() - arr_len*dy - arr_w*dx
            rx = tip_pt.x() - arr_len*dx - arr_w*dy
            ry = tip_pt.y() - arr_len*dy + arr_w*dx
            arrow = QPainterPath()
            arrow.moveTo(tip_pt)
            arrow.lineTo(QPointF(lx,ly))
            arrow.lineTo(QPointF(rx,ry))
            arrow.closeSubpath()
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QBrush(QColor(C["brand"])))
            painter.drawPath(arrow)

            # 레이블 (YES/NO)
            lbl = e.get("label","")
            if lbl:
                mid = path.pointAtPercent(0.5)
                painter.setPen(QColor(C["t1"]))
                painter.setFont(QFont("Malgun Gothic", 10, QFont.Weight.Bold))
                painter.drawText(
                    QRect(int(mid.x())-20, int(mid.y())-14, 40, 20),
                    Qt.AlignmentFlag.AlignCenter, lbl)

    def _draw_node(self, painter: QPainter, n: dict):
        t    = n.get("type","process")
        meta = WF_NODE_TYPES.get(t, WF_NODE_TYPES["process"])
        ac   = QColor(meta["color"])
        r    = self._node_rect(n)
        is_sel = n["id"] == self._sel_id

        # 그림자
        shadow_r = r.adjusted(4, 4, 4, 4)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(QColor(0,0,0,30)))
        if t == "cond":
            self._draw_diamond(painter, shadow_r)
        elif t in ("start","end"):
            painter.drawEllipse(shadow_r)
        else:
            painter.drawRoundedRect(shadow_r, 12, 12)

        # 선택 강조
        if is_sel:
            sel_r = r.adjusted(-4,-4,4,4)
            painter.setPen(QPen(ac.lighter(140), 3))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            if t == "cond": self._draw_diamond(painter, sel_r)
            elif t in ("start","end"): painter.drawEllipse(sel_r)
            else: painter.drawRoundedRect(sel_r, 14, 14)

        # 본체 그라데이션
        grad = QLinearGradient(r.left(), r.top(), r.left(), r.bottom())
        grad.setColorAt(0, ac.lighter(125))
        grad.setColorAt(1, ac)
        painter.setPen(QPen(ac.darker(115), 1.5))
        painter.setBrush(QBrush(grad))

        if t == "cond":
            self._draw_diamond(painter, r)
        elif t in ("start","end"):
            painter.drawEllipse(r)
        else:
            painter.drawRoundedRect(r, 12, 12)

        # 텍스트
        painter.setPen(QColor("#FFFFFF"))
        # 아이콘
        f_icon = QFont("Malgun Gothic", 11, QFont.Weight.Bold)
        painter.setFont(f_icon)
        icon_r = QRect(r.left()+6, r.top(), 24, r.height())
        painter.drawText(icon_r, Qt.AlignmentFlag.AlignVCenter, meta["icon"])

        # 이름
        f_name = QFont("Malgun Gothic", 10, QFont.Weight.Bold)
        painter.setFont(f_name)
        name_r = QRect(r.left()+28, r.top(), r.width()-32, r.height()//2+4)
        painter.drawText(name_r,
            Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
            n.get("name",""))

        # 서브 텍스트
        macro_name = n.get("macro_name","")
        wait_sec   = n.get("wait_sec","")
        sub = ""
        if t == "process" and macro_name:
            sub = macro_name[:18] + ("..." if len(macro_name)>18 else "")
        elif t == "wait":
            sub = f"{wait_sec}초 대기"
        if sub:
            f_sub = QFont("Malgun Gothic", 8)
            painter.setFont(f_sub)
            painter.setPen(QColor(255,255,255,200))
            sub_r = QRect(r.left()+28, r.center().y()+2, r.width()-32, r.height()//2-4)
            painter.drawText(sub_r,
                Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft, sub)

        # 반복 횟수 배지 (process 노드에 repeat > 1 이면 우상단에 표시)
        if t == "process" and int(n.get("repeat", 1)) > 1:
            badge_txt = f"×{n['repeat']}"
            f_badge = QFont("Malgun Gothic", 8, QFont.Weight.Bold)
            painter.setFont(f_badge)
            fm = painter.fontMetrics()
            bw = fm.horizontalAdvance(badge_txt) + 8
            bh = 16
            badge_r = QRect(r.right()-bw-2, r.top()-bh//2, bw, bh)
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QBrush(QColor("#F43F5E")))
            painter.drawRoundedRect(badge_r, 8, 8)
            painter.setPen(QColor("#FFFFFF"))
            painter.drawText(badge_r, Qt.AlignmentFlag.AlignCenter, badge_txt)

        # 포트 표시 (선택된 노드)
        if is_sel:
            painter.setBrush(QBrush(QColor("#FFFFFF")))
            painter.setPen(QPen(ac, 2))
            for port in ["bottom", "top"]:
                pp = self._port_pos(n, port)
                painter.drawEllipse(int(pp.x())-5, int(pp.y())-5, 10, 10)

    def _draw_diamond(self, painter: QPainter, r: QRect):
        cx, cy = float(r.center().x()), float(r.center().y())
        path = QPainterPath()
        path.moveTo(cx, float(r.top()))
        path.lineTo(float(r.right()), cy)
        path.lineTo(cx, float(r.bottom()))
        path.lineTo(float(r.left()), cy)
        path.closeSubpath()
        painter.drawPath(path)


# =============================================
#  워크플로우 실행 워커
# =============================================
class WFRunWorker(QThread):
    node_started  = pyqtSignal(str)
    node_done     = pyqtSignal(str, bool)   # id, success
    wf_done       = pyqtSignal(bool, str)   # success, message
    status_update = pyqtSignal(str)
    repeat_update = pyqtSignal(int, int)    # 현재 반복회차, 총 반복횟수

    def __init__(self, nodes, edges, repeat=1, repeat_gap=0.0):
        super().__init__()
        self.nodes  = {n["id"]: n for n in nodes}
        self.edges  = edges
        self.running = True
        self.repeat     = max(1, repeat)
        self.repeat_gap = max(0.0, repeat_gap)
        self._child = None   # 현재 실행 중인 자식 PlayWorker (ESC 중단 전파용)

    def _next_nodes(self, nid: str, success=True) -> list:
        result = []
        for e in self.edges:
            if e["src"] != nid: continue
            lbl = e.get("label","")
            if lbl == "YES" and not success: continue
            if lbl == "NO"  and success:     continue
            dst = self.nodes.get(e["dst"])
            if dst: result.append(dst)
        return result

    def _find_start(self):
        for n in self.nodes.values():
            if n["type"] == "start": return n
        return list(self.nodes.values())[0] if self.nodes else None

    def _run_once(self):
        """그래프를 start→end 까지 한 번 실행. 반환: (success, message) 또는 None(계속 진행 가능 신호 없음)"""
        cur = self._find_start()
        if not cur:
            return (False, "시작 노드가 없습니다.")

        visited = set()
        while cur and self.running:
            nid = cur["id"]
            if nid in visited:
                return (False, "순환 참조 감지됨.")
            visited.add(nid)

            self.node_started.emit(nid)
            self.status_update.emit(f"실행 중: {cur['name']}")
            success = True

            t = cur.get("type","process")

            if t == "start":
                time.sleep(0.3)

            elif t == "end":
                self.node_done.emit(nid, True)
                LogEngine.add("프로세스", "완료", "성공", "")
                return (True, "프로세스 완료!")

            elif t == "process":
                path = cur.get("macro_path","")
                node_repeat     = max(1, int(cur.get("repeat", 1)))
                node_repeat_gap = float(cur.get("repeat_gap", 0.0))
                if path and Path(path).exists():
                    try:
                        with open(path, encoding="utf-8") as f:
                            d = json.load(f)
                        acts = d.get("actions",[])
                        retries = 3 if cur.get("on_fail") == "retry" else 1
                        for rep_idx in range(node_repeat):
                            if not self.running:
                                break
                            if node_repeat > 1:
                                self.status_update.emit(
                                    f"실행 중: {cur['name']}  ({rep_idx+1}/{node_repeat}회)")
                            for attempt in range(retries):
                                try:
                                    w = PlayWorker(acts, 1, 1.0)
                                    self._child = w   # ESC/stop 시 자식도 중단하기 위해 추적
                                    w.start(); w.wait()
                                    self._child = None
                                    if not self.running:
                                        # ESC 등으로 중단됨 - 반복/재시도 즉시 탈출
                                        success = False; break
                                    success = True; break
                                except Exception:
                                    success = False
                                    if attempt < retries-1: time.sleep(2)
                            if not self.running:
                                break
                            if not success and cur.get("on_fail") == "stop":
                                break
                            if rep_idx < node_repeat - 1 and node_repeat_gap > 0 and self.running:
                                _safe_wait(node_repeat_gap, running_check=lambda: self.running)
                        LogEngine.add("프로세스", cur["name"],
                                      "성공" if success else "실패",
                                      f"{path}  ({node_repeat}회 반복)" if node_repeat > 1 else path)
                    except Exception as e:
                        success = False
                        LogEngine.add("프로세스", cur["name"], "실패", str(e))
                else:
                    self.status_update.emit(
                        f"[주의] 매크로 없음: {cur['name']}")
                    time.sleep(0.5)

                if not success and cur.get("on_fail") == "stop":
                    self.node_done.emit(nid, False)
                    return (False, f"실패로 중단: {cur['name']}")

            elif t == "wait":
                secs = cur.get("wait_sec", 5)
                self.status_update.emit(f"대기 중: {secs}초")
                _safe_wait(float(secs), running_check=lambda: self.running)

            elif t == "cond":
                img_path = cur.get("cond_img","")
                if CV2_OK and img_path and Path(img_path).exists():
                    try:
                        loc = pyautogui.locateOnScreen(img_path, confidence=0.8)
                        success = loc is not None
                    except Exception:
                        success = False
                else:
                    success = True   # 조건 없으면 YES

            self.node_done.emit(nid, success)
            nexts = self._next_nodes(nid, success)
            cur   = nexts[0] if nexts else None

        if not self.running:
            return (False, "중지됨")
        return (True, "프로세스 완료!")

    def run(self):
        # 글로벌 ESC 감시 (PROC 반복 실행 중에도 ESC 로 즉시 전체 중단)
        _esc_watcher = _GlobalEscWatcher(on_esc=self.stop)
        _esc_watcher.start()
        try:
            for r in range(self.repeat):
                if not self.running:
                    break
                self.repeat_update.emit(r+1, self.repeat)
                if self.repeat > 1:
                    self.status_update.emit(f"반복 실행 ({r+1}/{self.repeat}회)")
                ok, msg = self._run_once()
                if not self.running:
                    self.wf_done.emit(False, "ESC 로 중단됨")
                    return
                if not ok:
                    self.wf_done.emit(False, msg)
                    return
                if r < self.repeat - 1 and self.repeat_gap > 0 and self.running:
                    self.status_update.emit(f"다음 반복까지 {self.repeat_gap:.0f}초 대기...")
                    _safe_wait(self.repeat_gap, running_check=lambda: self.running)

            if self.running:
                final_msg = (f"프로세스 완료! ({self.repeat}회 반복 완료)"
                             if self.repeat > 1 else "프로세스 완료!")
                self.wf_done.emit(True, final_msg)
            else:
                self.wf_done.emit(False, "ESC 로 중단됨")
        finally:
            _esc_watcher.stop()

    def stop(self):
        self.running = False
        # 실행 중인 자식 PlayWorker 도 함께 중단 (반복 중 ESC 즉시 반응)
        child = self._child
        if child:
            try: child.stop()
            except Exception: pass


# =============================================
#  WorkflowPage
# =============================================
class WorkflowPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._workflows   = load_workflows()
        self._cur_wf      = None
        self._worker      = None
        self._node_states = {}
        self._build()
        # 샘플 프로세스 자동 로드
        QTimer.singleShot(300, self._load_sample_if_empty)

    def _load_sample_if_empty(self):
        """프로세스가 없으면 샘플 자동 선택"""
        if self._workflows:
            self._cur_wf = self._workflows[0]
            self._load_to_canvas(self._workflows[0])
            # 목록 첫 번째 선택
            if self.wf_list.count() > 0:
                self.wf_list.setCurrentRow(0)

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0,0,0,0); root.setSpacing(0)

        # ── 툴바 ──
        toolbar = QFrame()
        toolbar.setFixedHeight(52)
        toolbar.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};border-bottom:1.5px solid {C['border']};}}
        """)
        tb = QHBoxLayout(toolbar); tb.setContentsMargins(16,0,16,0); tb.setSpacing(8)
        tb.addWidget(SectionHeader("PROCESS  /  프로세스 설계"))
        tb.addStretch()

        for lbl, color, fn in [
            ("새 프로세스",  C["brand"],  self._new_wf),
            ("불러오기",       C["t2"],     self._load_wf),
            ("저장",           C["play"],   self._save_wf),
        ]:
            b = GlowButton(lbl, color)
            b.setFixedHeight(34); b.clicked.connect(fn)
            tb.addWidget(b)
        root.addWidget(toolbar)

        # ── 메인 영역 ──
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setStyleSheet("QSplitter{background:transparent;}")

        # ── 좌측: 노드 팔레트 + 워크플로우 목록 ──
        left = QFrame()
        left.setFixedWidth(200)
        left.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};border-right:1px solid {C['border']};}}
        """)
        ll = QVBoxLayout(left); ll.setContentsMargins(10,12,10,12); ll.setSpacing(6)

        pal_title = QLabel("노드 추가")
        pal_title.setStyleSheet(f"color:{C['brand']};font-weight:900;font-size:11px;background:transparent;")
        ll.addWidget(pal_title)

        for ntype, meta in WF_NODE_TYPES.items():
            btn = QPushButton(f"{meta['icon']}  {meta['label']}")
            btn.setFixedHeight(36)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            ac = meta["color"]
            btn.setStyleSheet(f"""
                QPushButton{{background:{C['bg1']};color:{ac};
                    border:none;border-left:4px solid {ac};border-radius:8px;
                    font-size:11px;font-weight:bold;text-align:left;padding-left:10px;}}
                QPushButton:hover{{background:{ac}18;}}
            """)
            btn.clicked.connect(lambda _, t=ntype: self._add_node(t))
            ll.addWidget(btn)

        div = QFrame(); div.setFixedHeight(1)
        div.setStyleSheet(f"background:{C['border']};"); ll.addWidget(div)

        # 워크플로우 목록
        wf_title = QLabel("프로세스 목록")
        wf_title.setStyleSheet(f"color:{C['brand']};font-weight:900;font-size:11px;background:transparent;")
        ll.addWidget(wf_title)

        self.wf_list = QListWidget()
        self.wf_list.setStyleSheet(f"""
            QListWidget{{background:{C['bg2']};border:1px solid {C['border']};
                border-radius:8px;padding:2px;outline:none;font-size:12px;}}
            QListWidget::item{{padding:7px 10px;border-radius:4px;color:{C['t1']};}}
            QListWidget::item:selected{{background:{C['brand']};color:#FFFFFF;}}
            QListWidget::item:hover{{background:{C['bg4']};}}
        """)
        self.wf_list.itemClicked.connect(self._on_wf_select)
        ll.addWidget(self.wf_list)

        # 반복 실행 설정
        ll.addSpacing(4)
        rep_lbl = QLabel("전체 반복 실행")
        rep_lbl.setStyleSheet(f"color:{C['brand']};font-weight:900;font-size:11px;background:transparent;")
        rep_lbl.setToolTip("START~END 전체 흐름을 N회 반복\n(노드별 반복은 각 노드 더블클릭으로 설정)")
        ll.addWidget(rep_lbl)

        rep_row1 = QHBoxLayout(); rep_row1.setSpacing(4)
        rep_row1.addWidget(QLabel("횟수"))
        self.wf_repeat_spin = StyledSpin()
        self.wf_repeat_spin.setRange(1, 9999)
        self.wf_repeat_spin.setValue(1)
        self.wf_repeat_spin.setSuffix(" 회")
        self.wf_repeat_spin.setToolTip("전체 프로세스(START~END)를 N번 반복 실행")
        rep_row1.addWidget(self.wf_repeat_spin)
        ll.addLayout(rep_row1)

        rep_row2 = QHBoxLayout(); rep_row2.setSpacing(4)
        rep_row2.addWidget(QLabel("간격"))
        self.wf_repeat_gap = StyledDSpin()
        self.wf_repeat_gap.setRange(0, 600)
        self.wf_repeat_gap.setValue(0)
        self.wf_repeat_gap.setSuffix(" 초")
        self.wf_repeat_gap.setToolTip("반복 사이 대기 시간")
        rep_row2.addWidget(self.wf_repeat_gap)
        ll.addLayout(rep_row2)

        # 창 최소화 옵션
        self.chk_min_others = QCheckBox("배경창 최소화")
        self.chk_min_others.setChecked(True)
        self.chk_min_others.setToolTip("실행 전 다른 창 모두 최소화 (좌표 오작동 방지)")
        self.chk_min_others.setStyleSheet(f"color:{C['t2']};font-size:11px;background:transparent;")
        ll.addWidget(self.chk_min_others)

        self.chk_min_self = QCheckBox("RPA 창도 최소화")
        self.chk_min_self.setChecked(True)
        self.chk_min_self.setToolTip("5MRPA 창을 최소화하여 화면 가림 제거")
        self.chk_min_self.setStyleSheet(f"color:{C['t2']};font-size:11px;background:transparent;")
        ll.addWidget(self.chk_min_self)

        # 실행 버튼
        ll.addSpacing(4)
        self.btn_run = GlowButton("PLAY  실행", C["play"])
        self.btn_run.setFixedHeight(40)
        self.btn_run.clicked.connect(self._run_wf)
        self.btn_stop = GlowButton("STOP  중지", C["rec"])
        self.btn_stop.setFixedHeight(40)
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self._stop_wf)
        btn_del_wf = GlowButton("삭제", C["t3"])
        btn_del_wf.setFixedHeight(32)
        btn_del_wf.clicked.connect(self._delete_wf)
        ll.addWidget(self.btn_run)
        ll.addWidget(self.btn_stop)
        ll.addWidget(btn_del_wf)

        splitter.addWidget(left)

        # ── 중앙: 캔버스 ──
        canvas_wrap = QWidget()
        canvas_wrap.setStyleSheet(f"background:{C['bg0']};")
        cw_lay = QVBoxLayout(canvas_wrap); cw_lay.setContentsMargins(0,0,0,0); cw_lay.setSpacing(0)

        # 워크플로우 이름 바
        self.wf_name_bar = QFrame()
        self.wf_name_bar.setFixedHeight(36)
        self.wf_name_bar.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};border-bottom:1px solid {C['border']};}}
        """)
        nb_lay = QHBoxLayout(self.wf_name_bar); nb_lay.setContentsMargins(14,0,14,0)
        self.wf_name_lbl = QLabel("새 프로세스를 만들거나 불러오세요")
        self.wf_name_lbl.setStyleSheet(f"color:{C['t2']};font-size:12px;background:transparent;")
        nb_lay.addWidget(self.wf_name_lbl); nb_lay.addStretch()
        self.wf_status_lbl = QLabel("")
        self.wf_status_lbl.setStyleSheet(f"color:{C['play']};font-size:11px;font-weight:bold;background:transparent;")
        nb_lay.addWidget(self.wf_status_lbl)
        cw_lay.addWidget(self.wf_name_bar)

        # 안내 바
        hint_bar = QFrame(); hint_bar.setFixedHeight(28)
        hint_bar.setStyleSheet(f"QFrame{{background:{C['ai_dim']};border-bottom:1px solid {C['border']}}}")
        hb = QHBoxLayout(hint_bar); hb.setContentsMargins(14,0,14,0)
        hint_lbl = QLabel(
            "  노드 팔레트에서 노드 추가  |  "
            "노드 드래그 = 이동  |  "
            "노드 더블클릭 = 편집  |  "
            "우클릭 = 연결/삭제")
        hint_lbl.setStyleSheet(f"color:{C['ai']};font-size:11px;font-weight:bold;background:transparent;")
        hb.addWidget(hint_lbl)
        cw_lay.addWidget(hint_bar)

        # 캔버스 스크롤
        canvas_scroll = QScrollArea()
        canvas_scroll.setWidgetResizable(True)
        canvas_scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        self.canvas = WFCanvas()
        self.canvas.canvas_changed.connect(self._on_canvas_changed)
        canvas_scroll.setWidget(self.canvas)
        QShortcut(QKeySequence("Ctrl+Z"), self,
                  lambda: self.canvas.undo())
        cw_lay.addWidget(canvas_scroll)
        splitter.addWidget(canvas_wrap)

        # ── 우측: 노드 정보 패널 ──
        right = QFrame()
        right.setFixedWidth(220)
        right.setStyleSheet(f"""
            QFrame{{background:{C['bg1']};border-left:1px solid {C['border']};}}
        """)
        rl = QVBoxLayout(right); rl.setContentsMargins(12,12,12,12); rl.setSpacing(8)

        info_title = QLabel("노드 정보")
        info_title.setStyleSheet(f"color:{C['brand']};font-weight:900;font-size:12px;background:transparent;")
        rl.addWidget(info_title)

        self.node_info = QTextEdit()
        self.node_info.setReadOnly(True)
        self.node_info.setStyleSheet(f"""
            QTextEdit{{background:{C['bg2']};border:1px solid {C['border']};
                border-radius:8px;color:{C['t1']};padding:8px;font-size:12px;}}
        """)
        self.node_info.setPlaceholderText("노드를 클릭하면\n정보가 표시됩니다.")
        rl.addWidget(self.node_info)

        # 실행 로그
        log_title = QLabel("실행 로그")
        log_title.setStyleSheet(f"color:{C['brand']};font-weight:900;font-size:12px;background:transparent;")
        rl.addWidget(log_title)

        self.run_log = QListWidget()
        self.run_log.setStyleSheet(f"""
            QListWidget{{background:{C['bg2']};border:1px solid {C['border']};
                border-radius:8px;padding:4px;outline:none;
                font-family:Consolas,monospace;font-size:11px;}}
            QListWidget::item{{padding:4px 8px;color:{C['t2']};}}
        """)
        rl.addWidget(self.run_log)

        splitter.addWidget(right)
        splitter.setSizes([200, 580, 220])

        root.addWidget(splitter)

        self._refresh_wf_list()

    # ── 워크플로우 관리 ──
    def _new_wf(self):
        name, ok = QInputDialog.getText(self, "새 프로세스", "프로세스 이름:")
        if not ok or not name.strip(): return
        wf = {
            "id":    _new_id(),
            "name":  name.strip(),
            "nodes": [],
            "edges": [],
        }
        # 기본 START/END 노드 자동 추가
        start_node = {"id":_new_id(),"type":"start","name":"시작",
                      "x":300,"y":100,"macro_path":"","macro_name":"",
                      "wait_sec":5,"cond_img":"","on_fail":"stop"}
        end_node   = {"id":_new_id(),"type":"end","name":"종료",
                      "x":300,"y":450,"macro_path":"","macro_name":"",
                      "wait_sec":5,"cond_img":"","on_fail":"stop"}
        wf["nodes"] = [start_node, end_node]
        self._workflows.append(wf)
        save_workflows(self._workflows)
        self._cur_wf = wf
        self._refresh_wf_list()
        self._load_to_canvas(wf)

    def _load_wf(self):
        fp, _ = rpa_open_file(self, "워크플로우 불러오기", SAVE_DIR, ".wf.json")
        if not fp: return
        try:
            with open(fp, encoding="utf-8") as f:
                wf = json.load(f)
            # 기존 목록에 없으면 추가
            if not any(w["id"]==wf.get("id") for w in self._workflows):
                self._workflows.append(wf)
            self._cur_wf = wf
            self._refresh_wf_list()
            self._load_to_canvas(wf)
        except Exception as e:
            msg_error(self, "오류", str(e))

    def _save_wf(self):
        if not self._cur_wf:
            msg_info(self, "없음", "먼저 워크플로우를 만드세요."); return
        nodes, edges = self.canvas.get_data()
        self._cur_wf["nodes"] = nodes
        self._cur_wf["edges"] = edges
        # workflows.json 에 저장
        save_workflows(self._workflows)
        # 별도 파일로도 저장
        fname = re.sub(r'[\\/:*?"<>|]', "_", self._cur_wf["name"])
        fpath = SAVE_DIR / f"{fname}.wf.json"
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(self._cur_wf, f, ensure_ascii=False, indent=2)
        self.wf_status_lbl.setText(f"저장 완료: {fpath.name}")
        LogEngine.add("프로세스", self._cur_wf["name"], "성공", "저장")

    def _delete_wf(self):
        if not self._cur_wf: return
        r = msg_ask(self, "삭제", f"'{self._cur_wf['name']}' 삭제?")
        if r == True:
            self._workflows = [w for w in self._workflows
                               if w["id"] != self._cur_wf["id"]]
            save_workflows(self._workflows)
            self._cur_wf = None
            self.canvas.set_data([], [])
            self.wf_name_lbl.setText("새 프로세스를 만들거나 불러오세요")
            self._refresh_wf_list()

    def _refresh_wf_list(self):
        self.wf_list.clear()
        for wf in self._workflows:
            item = QListWidgetItem(f"  {wf['name']}")
            item.setData(Qt.ItemDataRole.UserRole, wf["id"])
            self.wf_list.addItem(item)

    def _on_wf_select(self, item):
        wid = item.data(Qt.ItemDataRole.UserRole)
        wf  = next((w for w in self._workflows if w["id"]==wid), None)
        if wf:
            self._cur_wf = wf
            self._load_to_canvas(wf)

    def _load_to_canvas(self, wf: dict):
        self.canvas.set_data(wf.get("nodes",[]), wf.get("edges",[]))
        self.wf_name_lbl.setText(f"  {wf['name']}")
        self.wf_name_lbl.setStyleSheet(
            f"color:{C['t1']};font-size:13px;font-weight:bold;background:transparent;")
        self.wf_status_lbl.setText("")

    def _on_canvas_changed(self):
        if self._cur_wf:
            nodes, edges = self.canvas.get_data()
            self._cur_wf["nodes"] = nodes
            self._cur_wf["edges"] = edges

    def _add_node(self, ntype: str):
        if not self._cur_wf:
            msg_info(self, "없음", "먼저 워크플로우를 만드세요."); return
        node = self.canvas.add_node(ntype)
        # 편집 다이얼로그 바로 열기
        dlg = WFNodeDialog(node, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            for i, n in enumerate(self.canvas._nodes):
                if n["id"] == node["id"]:
                    self.canvas._nodes[i] = dlg.node
                    break
            self.canvas.update()

    # ── 실행 ──
    def _run_wf(self):
        if not self._cur_wf:
            msg_info(self, "없음", "프로세스를 선택하세요."); return
        nodes, edges = self.canvas.get_data()
        if not nodes:
            msg_warn(self, "없음", "노드가 없습니다."); return

        self.run_log.clear()
        self._node_states.clear()
        self.btn_run.setEnabled(False)
        self.btn_stop.setEnabled(True)

        repeat = self.wf_repeat_spin.value()
        gap    = self.wf_repeat_gap.value()
        min_others = self.chk_min_others.isChecked()
        min_self   = self.chk_min_self.isChecked()

        self._worker = WFRunWorker(nodes, edges, repeat=repeat, repeat_gap=gap)
        self._worker.node_started.connect(self._on_node_start)
        self._worker.node_done.connect(self._on_node_done)
        self._worker.status_update.connect(self._on_status)
        self._worker.wf_done.connect(self._on_wf_done)
        self._worker.repeat_update.connect(self._on_repeat_update)

        if repeat > 1:
            self._log(f"워크플로우 시작: {self._cur_wf['name']}  ({repeat}회 반복)", C["brand"])
        else:
            self._log(f"워크플로우 시작: {self._cur_wf['name']}", C["brand"])

        def _start():
            if min_others:
                _minimize_all_windows()
            if min_self:
                win = self.window()
                if win: QTimer.singleShot(0, win.showMinimized)
            self._worker.start()

        QTimer.singleShot(3000, _start)
        self.wf_status_lbl.setText("3초 후 시작...")

    def _on_repeat_update(self, cur: int, total: int):
        if total > 1:
            self.wf_name_lbl.setText(
                f"{self._cur_wf['name'] if self._cur_wf else ''}  —  반복 {cur}/{total}회")

    def _stop_wf(self):
        if self._worker: self._worker.stop()
        self.btn_run.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.wf_status_lbl.setText("중지됨")
        self._log("사용자 중지", C["stop"])
        win = self.window()
        if win and win.isMinimized():
            win.showNormal(); win.activateWindow(); win.raise_()

    def _on_node_start(self, nid: str):
        self._node_states[nid] = "running"
        self.canvas.update()
        node = next((n for n in self.canvas._nodes if n["id"]==nid), None)
        if node: self._log(f"실행: {node['name']}", C["brand"])

    def _on_node_done(self, nid: str, success: bool):
        self._node_states[nid] = "ok" if success else "fail"
        # 캔버스에서 노드 상태 색상 반영
        for n in self.canvas._nodes:
            if n["id"] == nid:
                n["_state"] = "ok" if success else "fail"
        self.canvas.update()
        node = next((n for n in self.canvas._nodes if n["id"]==nid), None)
        if node:
            c = C["play"] if success else C["rec"]
            self._log(f"완료: {node['name']} ({'성공' if success else '실패'})", c)

    def _on_status(self, msg: str):
        self.wf_status_lbl.setText(msg)

    def _on_wf_done(self, success: bool, msg: str):
        self.btn_run.setEnabled(True)
        self.btn_stop.setEnabled(False)
        c = C["play"] if success else C["rec"]
        self.wf_status_lbl.setText(msg)
        self._log(msg, c)
        if self._cur_wf:
            self.wf_name_lbl.setText(self._cur_wf['name'])
        # RPA 창 복원
        win = self.window()
        if win and win.isMinimized():
            win.showNormal(); win.activateWindow(); win.raise_()
        # 트레이 알림
        win = self.window()
        if win and hasattr(win, "_tray"):
            win._tray.show_message(
                "5MRPA",
                f"워크플로우 {'완료' if success else '실패'}: "
                f"{self._cur_wf['name'] if self._cur_wf else ''}",
                duration_ms=3000)

    def _log(self, msg: str, color: str = None):
        ts   = datetime.datetime.now().strftime("%H:%M:%S")
        item = QListWidgetItem(f"  {ts}  {msg}")
        if color: item.setForeground(QColor(color))
        self.run_log.insertItem(0, item)
        if self.run_log.count() > 100:
            self.run_log.takeItem(self.run_log.count()-1)


# =============================================
#  시작프로그램 등록/해제 유틸리티
# =============================================

def _get_startup_path() -> str:
    """Windows 시작프로그램 폴더 경로"""
    import os
    return os.path.join(
        os.environ.get("APPDATA", ""),
        "Microsoft", "Windows", "Start Menu",
        "Programs", "Startup", "HD5MRPA7.lnk")


def register_startup() -> tuple:
    """
    시작프로그램 등록.
    COM 없이 순수 Python 으로 .lnk 파일 생성.
    반환: (성공여부, 메시지)
    """
    import subprocess, sys
    lnk_path = _get_startup_path()
    py_path  = sys.executable
    rpa_path = str(Path(__file__).resolve())

    # PowerShell 로 바로가기 생성 (COM 직접 호출 대신)
    ps_cmd = (
        f"$wsh = New-Object -ComObject WScript.Shell; "
        f"$sc = $wsh.CreateShortcut('{lnk_path}'); "
        f"$sc.TargetPath = '{py_path}'; "
        f"$sc.Arguments = '{rpa_path} --hidden'; "
        f"$sc.WorkingDirectory = '{Path(rpa_path).parent}'; "
        f"$sc.Description = '5MRPA 자동 시작'; "
        f"$sc.WindowStyle = 7; "
        f"$sc.Save()"
    )
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_cmd],
            capture_output=True, text=True, timeout=10,
            creationflags=subprocess.CREATE_NO_WINDOW
                if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
        if result.returncode == 0:
            return True, lnk_path
        else:
            return False, result.stderr.strip()
    except Exception as e:
        return False, str(e)


def unregister_startup() -> tuple:
    """시작프로그램 해제 - .lnk 파일 삭제"""
    lnk_path = _get_startup_path()
    try:
        p = Path(lnk_path)
        if p.exists():
            p.unlink()
            return True, "해제 완료"
        else:
            return False, "등록된 항목이 없습니다"
    except Exception as e:
        return False, str(e)


def is_startup_registered() -> bool:
    """시작프로그램 등록 여부 확인"""
    return Path(_get_startup_path()).exists()

# =============================================
#  HelpPage
# =============================================
class HelpPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24,20,24,20); root.setSpacing(0)
        root.addWidget(SectionHeader("HELP  /  도움말"))
        root.addSpacing(10)
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        inner = QWidget(); inner.setStyleSheet("background:transparent;")
        lay   = QVBoxLayout(inner); lay.setSpacing(10)

        sections = [
            ("핵심 사용 흐름  (이것만 알면 됩니다!)", C['play'], [
                ("1. 녹화","[REC] 탭에서 녹화 버튼 → 자동화할 동작 수행 → ESC 로 중지"),
                ("2. 프로세스 에디터로 보내기","녹화 탭의 [프로세스 에디터로 보내기] 버튼 클릭"),
                ("3. 시각적 편집","[FLOW] 탭에서 블록을 클릭해 수정, 추가, 삭제"),
                ("4. AI 활용","AI 프로세스 자동 생성, AI 프로세스 검토, 오류 자동 점검"),
                ("5. 재생","[PLAY] 버튼으로 실행"),
            ]),
            ("프로세스 에디터 사용법", C['brand'], [
                ("프로세스 클릭/더블클릭","해당 프로세스 편집 팝업 열기"),
                ("UP / DN 버튼","프로세스 순서 위아래 변경"),
                ("ED 버튼","프로세스 편집"),
                ("DEL 버튼","프로세스 삭제"),
                ("+ 플로우 추가","해당 위치에 새 프로세스 삽입"),
                ("팔레트 버튼","왼쪽 팔레트에서 원하는 프로세스 타입 선택해 맨 끝에 추가"),
            ]),
            ("AI 도우미 기능", C['ai'], [
                ("AI 프로세스 자동 생성","'엑셀 저장하고 닫아줘' 처럼 말로 입력하면 프로세스 자동 생성"),
                ("AI 프로세스 검토","전체 프로세스를 AI 가 분석하고 개선 제안"),
                ("오류 자동 점검","잘못된 설정 자동 감지 + AI 해결방법 안내"),
                ("블록 설명 자동 생성","각 블록에 AI 가 한글 설명 자동 추가"),
            ]),
            ("이미지 인식 활용", C['ai'], [
                ("언제 쓰나요?","버튼/메뉴 위치가 바뀌거나, 팝업이 뜰 때"),
                ("프로세스에서 사용","블록 팔레트에서 [IMG 이미지 클릭] 선택"),
                ("이미지 파일 준비","[IMG] 탭에서 화면 캡처 후 저장"),
                ("경로 연결","블록 편집에서 이미지 파일 경로 선택"),
            ]),
            ("비상 정지", C['rec'], [
                ("마우스를 화면 모서리로","빠르게 왼쪽 위 모서리로 이동하면 자동 중지"),
                ("STOP 버튼","앱의 중지 버튼 클릭"),
            ]),
        ]

        for title, color, items in sections:
            card = QFrame()
            card.setStyleSheet(f"""
                QFrame {{ background:{C['bg1']};
                    border:1px solid {C['border2']};
                    border-left:3px solid {color}; border-radius:10px; }}
            """)
            cl = QVBoxLayout(card); cl.setContentsMargins(16,12,16,12); cl.setSpacing(6)
            tl = QLabel(title)
            tl.setStyleSheet(f"color:{color}; font-size:13px; font-weight:bold; background:transparent;")
            cl.addWidget(tl)
            for sub,desc in items:
                sl = QLabel(f"  * {sub}")
                sl.setStyleSheet(f"color:{C['t1']}; font-weight:bold; font-size:12px; background:transparent;")
                dl = QLabel(f"     {desc}")
                dl.setStyleSheet(f"color:{C['t2']}; font-size:12px; background:transparent;")
                dl.setWordWrap(True)
                cl.addWidget(sl); cl.addWidget(dl)
            lay.addWidget(card)

        lay.addStretch()

        # ── PC 시작 시 자동 실행 등록 카드 ──
        startup_card = QFrame()
        startup_card.setStyleSheet(f"""
            QFrame {{
                background: {C['bg1']};
                border: 1.5px solid {C['brand']};
                border-left: 4px solid {C['brand']};
                border-radius: 12px;
            }}
        """)
        sc_lay = QVBoxLayout(startup_card)
        sc_lay.setContentsMargins(18, 14, 18, 14)
        sc_lay.setSpacing(10)

        sc_title = QLabel("PC 시작 시 자동 실행 설정")
        sc_title.setStyleSheet(
            f"color:{C['brand']}; font-size:14px; font-weight:bold; background:transparent;")
        sc_lay.addWidget(sc_title)

        sc_desc = QLabel(
            "등록하면 PC 켤 때 자동으로 트레이에 숨어서 시작됩니다.\n"
            "스케줄러도 자동으로 활성화됩니다.")
        sc_desc.setStyleSheet(
            f"color:{C['t2']}; font-size:12px; background:transparent;")
        sc_desc.setWordWrap(True)
        sc_lay.addWidget(sc_desc)

        # 현재 등록 상태 표시
        self.startup_status = QLabel("")
        self.startup_status.setStyleSheet(
            f"color:{C['t2']}; font-size:12px; background:transparent;")
        sc_lay.addWidget(self.startup_status)
        self._refresh_startup_status()

        # 버튼 행
        sc_btn_row = QHBoxLayout(); sc_btn_row.setSpacing(8)
        btn_register = GlowButton("PC 시작 시 자동 실행 등록", C['play'])
        btn_register.setFixedHeight(40)
        btn_register.clicked.connect(self._register_startup)

        btn_unregister = GlowButton("등록 해제", C['rec'])
        btn_unregister.setFixedHeight(40)
        btn_unregister.clicked.connect(self._unregister_startup)

        sc_btn_row.addWidget(btn_register)
        sc_btn_row.addWidget(btn_unregister)
        sc_btn_row.addStretch()
        sc_lay.addLayout(sc_btn_row)

        lay.addWidget(startup_card)

        scroll.setWidget(inner)
        root.addWidget(scroll)
        root.addSpacing(8)
        about_lbl = QLabel(
            f"5MRPA  {APP_VERSION}  {APP_REV}    |    Made by  박찬욱    |    ALL RIGHTS RESERVED  ©  2025")
        about_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        about_lbl.setStyleSheet(
            f"color:{C['t3']};font-size:10px;background:transparent;letter-spacing:1px;padding:6px;")
        root.addWidget(about_lbl)

    def _refresh_startup_status(self):
        if is_startup_registered():
            self.startup_status.setText("현재 상태: 등록됨 (PC 시작 시 자동 실행)")
            self.startup_status.setStyleSheet(
                f"color:{C['play']}; font-size:12px; font-weight:bold; background:transparent;")
        else:
            self.startup_status.setText("현재 상태: 미등록")
            self.startup_status.setStyleSheet(
                f"color:{C['t3']}; font-size:12px; background:transparent;")

    def _register_startup(self):
        ok, msg = register_startup()
        if ok:
            msg_info(self, "등록 완료", f"PC 자동 시작 등록 완료. 등록 위치: {msg}")
            LogEngine.add("설정", "시작프로그램", "성공", "PC 자동 시작 등록")
        else:
            msg_error(self, "등록 실패", f"등록에 실패했습니다: {msg}")
        self._refresh_startup_status()

    def _unregister_startup(self):
        ok, msg = unregister_startup()
        if ok:
            msg_info(self, "해제 완료", "자동 시작이 해제되었습니다.")
            LogEngine.add("설정", "시작프로그램", "성공", "PC 자동 시작 해제")
        else:
            msg_warn(self, "알림", msg)
        self._refresh_startup_status()


# =============================================
#  ScreenCaptureOverlay
# =============================================
class ScreenCaptureOverlay(QWidget):
    captured = pyqtSignal(QPixmap)

    def __init__(self, bg_pixmap: QPixmap, parent=None):
        super().__init__(parent)
        self._bg       = bg_pixmap
        self._start    = QPoint()
        self._end      = QPoint()
        self._dragging = False
        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint)
        self.setCursor(Qt.CursorShape.CrossCursor)
        self.setMouseTracking(True)

    def showEvent(self, event):
        screen = QApplication.primaryScreen().geometry()
        self.setGeometry(screen)
        self.activateWindow(); self.raise_()
        super().showEvent(event)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Escape:
            self.captured.emit(QPixmap()); self.close()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._start = self._end = event.pos()
            self._dragging = True; self.update()

    def mouseMoveEvent(self, event):
        if self._dragging: self._end = event.pos(); self.update()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton and self._dragging:
            self._dragging = False; self._end = event.pos(); self._finish()

    def _finish(self):
        x1=min(self._start.x(),self._end.x()); y1=min(self._start.y(),self._end.y())
        w=abs(self._end.x()-self._start.x()); h=abs(self._end.y()-self._start.y())
        self.close()
        if w>4 and h>4: self.captured.emit(self._bg.copy(x1,y1,w,h))
        else: self.captured.emit(QPixmap())

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(0,0,self._bg)
        painter.fillRect(self.rect(),QColor(0,0,0,140))
        if self._dragging:
            x1=min(self._start.x(),self._end.x()); y1=min(self._start.y(),self._end.y())
            w=abs(self._end.x()-self._start.x()); h=abs(self._end.y()-self._start.y())
            sel=QRect(x1,y1,w,h)
            painter.drawPixmap(sel,self._bg,sel)
            painter.setPen(QPen(QColor(C['brand']),2))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRect(sel)
            painter.setPen(QColor(C['brand']))
            painter.setFont(QFont("Malgun Gothic",11,QFont.Weight.Bold))
            painter.drawText(x1, y1-10 if y1>28 else y1+h+20, f"  {w}x{h}  (떼면 확정/ESC 취소)")
        else:
            painter.setPen(QColor(C['t1']))
            painter.setFont(QFont("Malgun Gothic",16,QFont.Weight.Bold))
            painter.drawText(self.rect(),Qt.AlignmentFlag.AlignCenter,
                             "드래그해서 캡처할 영역을 선택하세요\n\nESC = 취소")
        painter.end()


# =============================================
#  ImageAutoPage (간소화 - 핵심 기능만)
# =============================================
class ImageAutoPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._templates   = []
        self._captured_path = None
        self._running     = False
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24,20,24,20); root.setSpacing(12)
        root.addWidget(SectionHeader("IMAGE  /  이미지 인식 자동화"))

        if not CV2_OK:
            w=QLabel("  [주의]  pip install opencv-python  후 재시작")
            w.setStyleSheet(f"color:{C['stop']};background:{C['stop_dim']};border:1px solid {C['stop']};border-radius:8px;padding:10px;")
            root.addWidget(w)

        guide = QLabel(
            "사용법: [CAP] 캡처 → 이름 입력 → [목록 추가] → [실행]\n"
            "창 위치가 바뀌어도 이미지 생김새로 찾아 클릭합니다.")
        guide.setStyleSheet(f"color:{C['t2']};background:{C['bg1']};border:1.5px solid {C['border']};border-left:3px solid {C['brand']};border-radius:8px;padding:10px;")
        guide.setWordWrap(True); root.addWidget(guide)

        cap_card = QFrame()
        cap_card.setStyleSheet(f"QFrame{{background:{C['bg1']};border:1.5px solid {C['border']};border-radius:12px;}}")
        cg = QGridLayout(cap_card); cg.setContentsMargins(16,12,16,12); cg.setSpacing(8)

        cg.addWidget(QLabel("이름:"),0,0)
        self.img_name = StyledInput("예: 저장버튼, 확인팝업")
        cg.addWidget(self.img_name,0,1,1,3)

        cg.addWidget(QLabel("동작:"),1,0)
        self.act_combo = StyledCombo()
        self.act_combo.addItems(["왼쪽 클릭","오른쪽 클릭","더블 클릭","위치만 찾기"])
        cg.addWidget(self.act_combo,1,1)

        cg.addWidget(QLabel("정확도:"),1,2)
        self.conf_spin = StyledDSpin()
        self.conf_spin.setRange(0.5,1.0); self.conf_spin.setValue(0.85); self.conf_spin.setSingleStep(0.05)
        cg.addWidget(self.conf_spin,1,3)

        br = QHBoxLayout()
        self.btn_capture = GlowButton("[CAP] 캡처", C['brand']); self.btn_capture.setFixedHeight(36)
        self.btn_loadfile= GlowButton("파일 불러오기",C['t2']); self.btn_loadfile.setFixedHeight(36)
        self.btn_add     = GlowButton("목록에 추가",C['play']); self.btn_add.setFixedHeight(36)
        self.btn_capture.clicked.connect(self._capture)
        self.btn_loadfile.clicked.connect(self._load_img)
        self.btn_add.clicked.connect(self._add_template)
        br.addWidget(self.btn_capture); br.addWidget(self.btn_loadfile); br.addWidget(self.btn_add)
        cg.addLayout(br,2,0,1,4)

        self.preview_lbl = QLabel("캡처된 이미지 미리보기")
        self.preview_lbl.setFixedHeight(70); self.preview_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_lbl.setStyleSheet(f"color:{C['t3']};background:{C['bg2']};border:1px dashed {C['border']};border-radius:6px;")
        cg.addWidget(self.preview_lbl,3,0,1,4)
        root.addWidget(cap_card)

        self.tmpl_list = QListWidget()
        self.tmpl_list.setFixedHeight(100)
        self.tmpl_list.setStyleSheet(f"""
            QListWidget{{background:{C['bg1']};border:1.5px solid {C['border']};border-radius:8px;padding:4px;}}
            QListWidget::item{{padding:6px 12px;border-radius:4px;}}
            QListWidget::item:selected{{background:{C['brand']};color:#FFFFFF;}}
        """)
        root.addWidget(self.tmpl_list)

        run_row = QHBoxLayout(); run_row.setSpacing(8)
        run_row.addWidget(QLabel("반복:")); self.run_repeat=StyledSpin()
        self.run_repeat.setRange(1,9999); self.run_repeat.setValue(1); self.run_repeat.setSuffix(" 회")
        self.run_repeat.setFixedWidth(90); run_row.addWidget(self.run_repeat)
        run_row.addWidget(QLabel("간격:")); self.run_interval=StyledDSpin()
        self.run_interval.setRange(0.1,60); self.run_interval.setValue(1.0); self.run_interval.setSuffix(" 초")
        self.run_interval.setFixedWidth(100); run_row.addWidget(self.run_interval)
        run_row.addStretch()
        self.btn_run=GlowButton("[PLAY] 실행",C['play']); self.btn_run.setFixedHeight(40)
        self.btn_run.clicked.connect(self._run_sequence)
        btn_del=GlowButton("선택 삭제",C['rec']); btn_del.setFixedHeight(40)
        btn_del.clicked.connect(self._del_template)
        run_row.addWidget(self.btn_run); run_row.addWidget(btn_del)
        root.addLayout(run_row)

        self.img_status=QLabel("대기 중")
        self.img_status.setStyleSheet(f"color:{C['t2']};padding:6px 12px;background:{C['bg1']};border:1px solid {C['border']};border-radius:6px;")
        root.addWidget(self.img_status)

    def _capture(self):
        self.img_status.setText("3초 후 캡처 오버레이...")
        QTimer.singleShot(3000, self._open_overlay)

    def _open_overlay(self):
        self.img_status.setText("화면 캡처 중...")
        QApplication.processEvents()
        screenshot = pyautogui.screenshot()
        self._ss_buf = screenshot.tobytes("raw","RGB")
        qimg = QImage(self._ss_buf,screenshot.width,screenshot.height,
                      screenshot.width*3,QImage.Format.Format_RGB888)
        bg = QPixmap.fromImage(qimg)
        if bg.isNull(): self.img_status.setText("스크린샷 실패"); return
        self._overlay = ScreenCaptureOverlay(bg)
        self._overlay.captured.connect(self._on_captured)
        self._overlay.show(); self._overlay.showFullScreen()

    def _on_captured(self, pixmap):
        if pixmap.isNull(): self.img_status.setText("캡처 취소"); return
        sp = SAVE_DIR/f"tmpl_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        pixmap.save(str(sp),"PNG")
        self._captured_path=str(sp)
        px=pixmap.scaled(200,60,Qt.AspectRatioMode.KeepAspectRatio,Qt.TransformationMode.SmoothTransformation)
        self.preview_lbl.setPixmap(px)
        self.img_status.setText(f"캡처 완료: {sp.name}")

    def _load_img(self):
        fp,_=rpa_open_file(self,"이미지 선택",SAVE_DIR,".png")
        if fp:
            self._captured_path=fp
            px=QPixmap(fp).scaled(200,60,Qt.AspectRatioMode.KeepAspectRatio)
            self.preview_lbl.setPixmap(px)
            self.img_status.setText(f"불러옴: {Path(fp).name}")

    def _add_template(self):
        if not self._captured_path:
            msg_warn(self, "없음", "먼저 캡처하거나 파일을 불러오세요."); return
        name=self.img_name.text().strip() or Path(self._captured_path).stem
        entry={"name":name,"path":self._captured_path,"action":self.act_combo.currentText(),
               "confidence":self.conf_spin.value(),"offset_x":0,"offset_y":0}
        self._templates.append(entry)
        item=QListWidgetItem(f"  [{len(self._templates):02d}]  {name}  |  {entry['action']}  |  정확도 {entry['confidence']:.2f}")
        item.setForeground(QColor(C['brand']))
        self.tmpl_list.addItem(item)
        self.img_name.clear(); self._captured_path=None
        self.preview_lbl.setText("캡처된 이미지 미리보기"); self.preview_lbl.setPixmap(QPixmap())

    def _del_template(self):
        row=self.tmpl_list.currentRow()
        if row>=0: self._templates.pop(row); self.tmpl_list.takeItem(row)

    def _run_sequence(self):
        if not CV2_OK:
            msg_warn(self, "미설치", "pip install opencv-python"); return
        if not self._templates:
            msg_warn(self, "없음", "이미지를 먼저 등록하세요."); return
        self._running=True; self.btn_run.setEnabled(False)
        repeat=self.run_repeat.value(); interval=self.run_interval.value()
        templates=list(self._templates)
        def _w():
            for r in range(repeat):
                if not self._running: break
                for tmpl in templates:
                    if not self._running: break
                    QTimer.singleShot(0,lambda t=tmpl,rr=r:self.img_status.setText(f"찾는 중: [{t['name']}] ({rr+1}/{repeat}회)"))
                    try:
                        loc=pyautogui.locateOnScreen(tmpl['path'],confidence=tmpl['confidence'])
                        if loc:
                            cx=int(loc.left+loc.width/2); cy=int(loc.top+loc.height/2)
                            act=tmpl['action']
                            if act=="왼쪽 클릭": pyautogui.click(cx,cy)
                            elif act=="오른쪽 클릭": pyautogui.rightClick(cx,cy)
                            elif act=="더블 클릭": pyautogui.doubleClick(cx,cy)
                            QTimer.singleShot(0,lambda t=tmpl,x=cx,y=cy:self.img_status.setText(f"[OK] {t['name']} 발견 -> ({x},{y})"))
                        else:
                            QTimer.singleShot(0,lambda t=tmpl:self.img_status.setText(f"[--] {t['name']} 못 찾음"))
                    except Exception as e:
                        QTimer.singleShot(0,lambda err=e:self.img_status.setText(f"[오류] {err}"))
                    time.sleep(interval)
            self._running=False
            QTimer.singleShot(0,lambda:(self.btn_run.setEnabled(True),self.img_status.setText("실행 완료")))
        threading.Thread(target=_w,daemon=True).start()


# =============================================
#  MainWindow
# =============================================
class MainWindow(QMainWindow):
    def keyPressEvent(self, event):
        # ESC 키로 창이 닫히지 않도록 차단
        if event.key() == Qt.Key.Key_Escape:
            event.ignore()
            return
        super().keyPressEvent(event)
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"5MRPA  {APP_VERSION}  {APP_REV}")
        screen_geo = QApplication.primaryScreen().availableGeometry()
        sw, sh     = screen_geo.width(), screen_geo.height()
        self.setMinimumSize(int(sw*0.55), int(sh*0.55))
        self.resize(int(sw*0.78), int(sh*0.82))
        self.setStyleSheet(GLOBAL_QSS)
        self._build()
        self._sb_timer = QTimer()
        self._sb_timer.timeout.connect(self._update_sb)
        self._sb_timer.start(1500)

        # 트레이 초기화
        self._quit_confirmed = False
        self._tray = TrayManager(self)

        # 스케줄러 상태 1분마다 트레이에 반영
        self._tray_sched_timer = QTimer()
        self._tray_sched_timer.timeout.connect(self._update_tray_sched)
        self._tray_sched_timer.start(60000)
        self._update_tray_sched()

    def _build(self):
        root = QWidget(); root.setStyleSheet(f"background:{C['bg0']};")
        self.setCentralWidget(root)
        main_row = QHBoxLayout(root)
        main_row.setContentsMargins(0,0,0,0); main_row.setSpacing(0)

        # 사이드바
        sidebar = QFrame(); sidebar.setFixedWidth(88)
        sidebar.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 {C['side_top']}, stop:1 {C['side_bot']});
                border-right: none;
            }}
        """)
        sb = QVBoxLayout(sidebar); sb.setContentsMargins(4,12,4,12); sb.setSpacing(3)

        logo = QLabel("HD\nRPA"); logo.setFixedHeight(60)
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo.setStyleSheet("""
            color:#FFFFFF; font-size:11px; font-weight:900; letter-spacing:1px;
            background:transparent; border-bottom:1px solid rgba(255,255,255,0.15);
            padding-bottom:8px; text-align:center;
        """)
        sb.addWidget(logo); sb.addSpacing(4)

        nav_items = [
            ("REC",  "녹화",     C['rec']),
            ("PLAY", "재생",     C['play']),
            ("FLOW", "플로우",   C['side_act']),
            ("IMG",  "이미지",   C['ai']),
            ("PROC", "프로세스", C['tool']),
            ("DASH", "현황판",   C['play']),
            ("TRIG", "트리거",   C['ai']),
            ("SCH",  "스케줄",   C['stop']),
            ("LOG",  "로그",     C['tool']),
            ("LIST", "관리",     C['brand']),
            ("HELP", "도움말",   C['t3']),
        ]
        self._nav_btns = []
        for icon, label, color in nav_items:
            btn = NavButton(icon, label, color)
            btn.clicked.connect(lambda _, i=len(self._nav_btns): self._goto(i))
            self._nav_btns.append(btn)
            sb.addWidget(btn)

        sb.addStretch()
        copy_sb = QLabel("© 박찬욱"); copy_sb.setAlignment(Qt.AlignmentFlag.AlignCenter)
        copy_sb.setStyleSheet("color:rgba(255,255,255,0.35); font-size:9px; background:transparent;")
        sb.addWidget(copy_sb)
        ver = QLabel(f"{APP_VERSION}  {APP_REV}  5MRPA"); ver.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ver.setStyleSheet("color:rgba(255,255,255,0.45); font-size:9px; background:transparent; letter-spacing:1px;")
        sb.addWidget(ver)
        main_row.addWidget(sidebar)

        # 콘텐츠
        cw = QFrame(); cw.setStyleSheet(f"background:{C['bg2']}; border:none;")
        cwl = QVBoxLayout(cw); cwl.setContentsMargins(0,0,0,0); cwl.setSpacing(0)

        # 상단바
        topbar = QFrame(); topbar.setFixedHeight(50)
        topbar.setStyleSheet(f"QFrame{{background:{C['bg1']};border-bottom:1px solid {C['border']};}}")
        tb = QHBoxLayout(topbar); tb.setContentsMargins(20,0,20,0)
        self.page_title = QLabel("녹화")
        self.page_title.setStyleSheet(f"color:{C['brand2']};font-size:15px;font-weight:900;background:transparent;letter-spacing:0.5px;")
        tb.addWidget(self.page_title); tb.addStretch()

        # AI 상태
        self.ai_bar_lbl = QLabel("AI 확인 중...")
        self.ai_bar_lbl.setStyleSheet(f"color:{C['t3']};font-size:10px;background:transparent;")
        tb.addWidget(self.ai_bar_lbl)

        btn_ai_retry = QPushButton("AI 재확인")
        btn_ai_retry.setFixedHeight(24); btn_ai_retry.setFixedWidth(60)
        btn_ai_retry.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_ai_retry.setStyleSheet(f"""
            QPushButton {{
                background:transparent; color:{C['brand']};
                border:1px solid {C['brand']}; border-radius:4px; font-size:9px;
            }}
            QPushButton:hover {{ background:{C['bg4']}; }}
        """)
        btn_ai_retry.clicked.connect(lambda: (
            self.ai_bar_lbl.setText("확인 중..."),
            self.ai_bar_lbl.setStyleSheet(f"color:{C['t3']};font-size:10px;background:transparent;"),
            threading.Thread(target=self._check_ai_bar, daemon=True).start()
        ))
        tb.addWidget(btn_ai_retry)
        tb.addSpacing(8)

        for flag, name in [(PYNPUT_OK,"pynput"),(CV2_OK,"opencv")]:
            l = QLabel(f" {name} {'OK' if flag else '--'} ")
            l.setStyleSheet(f"color:{'#2F9E44' if flag else '#A0AEC0'};font-size:10px;background:transparent;")
            tb.addWidget(l)
        tb.addSpacing(8)

        self.pos_lbl = QLabel("X:---- Y:----")
        self.pos_lbl.setStyleSheet(f"color:{C['t2']};font-size:11px;font-family:Consolas;background:transparent;padding:3px 8px;")
        self.time_lbl = QLabel("")
        self.time_lbl.setStyleSheet(f"color:{C['t2']};font-size:11px;background:transparent;padding:3px 8px;")
        tb.addWidget(self.pos_lbl); tb.addWidget(self.time_lbl)

        btn_folder = GlowButton("저장 폴더",C['brand'])
        btn_folder.setFixedHeight(28); btn_folder.setFixedWidth(80)
        btn_folder.clicked.connect(lambda: os.startfile(str(SAVE_DIR)))
        tb.addSpacing(8); tb.addWidget(btn_folder)
        tb.addSpacing(12)
        copy_top = QLabel("ALL RIGHTS RESERVED  ©  박찬욱")
        copy_top.setStyleSheet(
            f"color:{C['t3']};font-size:9px;background:transparent;letter-spacing:0.5px;")
        tb.addWidget(copy_top)
        cwl.addWidget(topbar)

        # 페이지 스택
        self.stack = QStackedWidget()
        self.stack.setStyleSheet(f"background:{C['bg2']};")

        self.flow_page  = FlowEditorPage()
        self.sched_page = SchedulerPage()
        self.log_page   = LogPage()
        self.wf_page    = WorkflowPage()
        self.dash_page  = DashboardPage()
        self.trig_page  = TriggerPage()
        self.pages = [
            RecordPage(), PlayPage(),
            self.flow_page,
            ImageAutoPage(),
            self.wf_page,
            self.dash_page,
            self.trig_page,
            self.sched_page,
            self.log_page,
            ManagerPage(), HelpPage(),
        ]
        self.page_names = [
            "녹화", "재생", "플로우 에디터",
            "이미지 자동화", "프로세스",
            "모니터링 대시보드", "트리거",
            "스케줄러", "실행 로그",
            "매크로 관리", "도움말"
        ]
        for p in self.pages: self.stack.addWidget(p)
        cwl.addWidget(self.stack)
        main_row.addWidget(cw)

        self._goto(0)
        for i in range(11):
            QShortcut(QKeySequence(f"Ctrl+{i+1}"), self,
                      lambda _=None,idx=i: self._goto(idx))

        # AI 상태 확인 - 빌드 완료 후 500ms 딜레이
        QTimer.singleShot(500, lambda: threading.Thread(
            target=self._check_ai_bar, daemon=True).start())
        # 30초마다 주기적 재확인
        self._ai_check_timer = QTimer()
        self._ai_check_timer.timeout.connect(lambda: threading.Thread(
            target=self._check_ai_bar, daemon=True).start())
        self._ai_check_timer.start(30000)

    def _check_ai_bar(self):
        """AI 상태 확인 - 직접 UI 업데이트"""
        ok = GemmaEngine.is_online()
        # QTimer 없이 직접 람다로 메인스레드 전달
        if ok:
            QTimer.singleShot(0, self._set_ai_on)
            if not getattr(self, '_ai_warmed', False):
                self._ai_warmed = True
                # 모델을 미리 메모리에 올려 첫 실사용시 지연(콜드스타트) 완화
                threading.Thread(target=GemmaEngine.warmup, daemon=True).start()
        else:
            QTimer.singleShot(0, self._set_ai_off)
            # 30초 후 재시도
            QTimer.singleShot(30000, lambda: threading.Thread(
                target=self._check_ai_bar, daemon=True).start())

    def _set_ai_on(self):
        if hasattr(self, 'ai_bar_lbl'):
            self.ai_bar_lbl.setText(f"AI ON ({OLLAMA_MODEL})")
            self.ai_bar_lbl.setStyleSheet(
                "color:#2F9E44; font-size:10px; font-weight:bold; background:transparent;")

    def _set_ai_off(self):
        if hasattr(self, 'ai_bar_lbl'):
            self.ai_bar_lbl.setText("AI OFF (ollama serve / 모델설치 확인)")
            self.ai_bar_lbl.setStyleSheet(
                "color:#F43F5E; font-size:10px; background:transparent;")

    def send_to_flow(self, actions: list):
        """RecordPage 에서 플로우 에디터로 동작 전달"""
        self.flow_page.canvas.set_actions(actions)
        self.flow_page.flow_status.setText(f"{len(actions)}개 플로우를 불러왔습니다. 편집하세요!")
        self._goto(2)

    def _goto(self, idx):
        for i, btn in enumerate(self._nav_btns): btn.setActive(i==idx)
        self.stack.setCurrentIndex(idx)
        self.page_title.setText(self.page_names[idx])
        if idx == 9: self.pages[9].refresh()   # 매크로 관리

    def _update_sb(self):
        x,y = pyautogui.position()
        self.pos_lbl.setText(f"X:{x:4d} Y:{y:4d}")
        self.time_lbl.setText(datetime.datetime.now().strftime("%H:%M:%S"))

    def closeEvent(self, event):
        # 트레이 종료 버튼으로만 완전 종료
        # X 버튼은 트레이로 숨김
        if not self._quit_confirmed:
            event.ignore()
            self.hide()
            # 트레이 알림 (처음 숨길 때 한 번만)
            if not getattr(self, '_tray_notified', False):
                self._tray_notified = True
                self._tray.show_message(
                    "5MRPA",
                    "트레이로 숨겼습니다.\n"
                    "오른쪽 아래 아이콘을 클릭하면 다시 열립니다.",
                    duration_ms=4000)
            return

        # 완전 종료 (트레이 메뉴 → 종료)
        if hasattr(self, 'sched_page'):
            self.sched_page._engine.stop()
        for page in self.pages:
            w = getattr(page,'worker',None)
            if w and hasattr(w,'stop'): w.stop()
            if hasattr(page,'_running'): page._running = False
            w2 = getattr(page,'_worker',None)
            if w2 and hasattr(w2,'stop'): w2.stop()
        if hasattr(self, '_tray'):
            self._tray.hide_tray()
        event.accept()

    def _update_tray_sched(self):
        """트레이 메뉴 스케줄러 상태 업데이트"""
        if hasattr(self, 'sched_page') and hasattr(self, '_tray'):
            scheds = self.sched_page._schedules
            n_on   = sum(1 for s in scheds if s.get('enabled', True))
            self._tray.update_sched_status(n_on)



# =============================================
#  시스템 트레이 관리자
#  - PyQt6 QSystemTrayIcon 사용 (추가 설치 없음)
#  - 보안정책 안전: 시스템 표준 API만 사용
# =============================================

class TrayManager:
    """
    시스템 트레이 아이콘 관리.
    창 닫기(X) 시 트레이로 숨김.
    트레이 아이콘 클릭/메뉴로 창 복원.
    """
    def __init__(self, main_window):
        self._win    = main_window
        self._app    = QApplication.instance()
        # 부모를 None 으로 설정 - 창이 숨겨져도 트레이 독립 유지
        self._tray   = QSystemTrayIcon(None)

        # 트레이 아이콘
        self._tray.setIcon(self._make_icon())
        self._tray.setToolTip(f"5MRPA  {APP_VERSION}  {APP_REV}")

        # 트레이 메뉴 생성
        self._menu = self._build_menu()
        self._tray.setContextMenu(self._menu)

        # 트레이 아이콘 클릭 이벤트
        self._tray.activated.connect(self._on_tray_click)

        self._tray.show()

    def _make_icon(self) -> QIcon:
        """
        텍스트 기반 아이콘 생성 (외부 이미지 파일 불필요)
        회사 보안정책 안전 - 순수 PyQt6 렌더링
        """
        size   = 64   # 고해상도로 생성 후 축소
        pixmap = QPixmap(size, size)
        pixmap.fill(QColor(0, 0, 0, 0))

        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # 배경 - 네이비 그라데이션
        grad = QLinearGradient(0, 0, size, size)
        grad.setColorAt(0, QColor(C['side_top']))
        grad.setColorAt(1, QColor(C['side_bot']))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(grad))
        painter.drawRoundedRect(2, 2, size-4, size-4, 14, 14)

        # 밝은 테두리
        painter.setPen(QPen(QColor(C['brand']), 2))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRoundedRect(2, 2, size-4, size-4, 14, 14)

        # "함" 텍스트 (한 글자로 직관적으로)
        painter.setPen(QColor("#FFFFFF"))
        f = QFont("Malgun Gothic", 26, QFont.Weight.Bold)
        painter.setFont(f)
        painter.drawText(QRect(0, -4, size, size),
                         Qt.AlignmentFlag.AlignCenter, "함")

        # 하단 작은 "RPA" 텍스트
        painter.setPen(QColor(C['brand']))
        f2 = QFont("Malgun Gothic", 10, QFont.Weight.Bold)
        painter.setFont(f2)
        painter.drawText(QRect(0, size//2+6, size, size//2),
                         Qt.AlignmentFlag.AlignCenter, "RPA")

        painter.end()
        return QIcon(pixmap)

    def _build_menu(self):
        menu = QMenu()
        menu.setStyleSheet(f"""
            QMenu {{
                background: {C['bg1']};
                border: 1px solid {C['border']};
                border-radius: 8px;
                padding: 4px;
                color: {C['t1']};
                font-size: 13px;
            }}
            QMenu::item {{
                padding: 8px 20px;
                border-radius: 4px;
            }}
            QMenu::item:selected {{
                background: {C['brand']};
                color: #FFFFFF;
            }}
            QMenu::separator {{
                height: 1px;
                background: {C['border']};
                margin: 4px 8px;
            }}
        """)

        # 제목 (클릭 안 되는 라벨)
        title_act = QAction(f"5MRPA  {APP_VERSION}  {APP_REV}", menu)
        title_act.setEnabled(False)
        menu.addAction(title_act)
        menu.addSeparator()

        # 창 열기
        open_act = QAction("창 열기", menu)
        open_act.triggered.connect(self.show_window)
        menu.addAction(open_act)

        menu.addSeparator()

        # 스케줄러 상태
        self.sched_act = QAction("스케줄러: 확인 중", menu)
        self.sched_act.setEnabled(False)
        menu.addAction(self.sched_act)

        menu.addSeparator()

        # 종료
        quit_act = QAction("종료", menu)
        quit_act.triggered.connect(self.quit_app)
        menu.addAction(quit_act)

        return menu

    def update_sched_status(self, n_active: int):
        """스케줄러 상태 메뉴 업데이트"""
        if hasattr(self, 'sched_act'):
            if n_active > 0:
                self.sched_act.setText(f"스케줄러: 실행 중  ({n_active}개 활성)")
            else:
                self.sched_act.setText("스케줄러: 대기 중")

    def _on_tray_click(self, reason):
        """트레이 아이콘 클릭 처리"""
        if reason == QSystemTrayIcon.ActivationReason.DoubleClick:
            self.show_window()
        elif reason == QSystemTrayIcon.ActivationReason.Trigger:
            # 단일 클릭: 창 토글
            if self._win.isVisible():
                self._win.hide()
            else:
                self.show_window()

    def show_window(self):
        """창 복원 - 보안정책 안전한 방식"""
        self._win.showNormal()
        self._win.raise_()
        self._win.activateWindow()

    def show_message(self, title: str, msg: str,
                     icon=QSystemTrayIcon.MessageIcon.Information,
                     duration_ms: int = 3000):
        """트레이 알림 팝업"""
        self._tray.showMessage(title, msg, icon, duration_ms)

    def quit_app(self):
        """완전 종료"""
        self._win._quit_confirmed = True
        self._tray.hide()
        self._win.close()
        QApplication.instance().quit()

    def hide_tray(self):
        self._tray.hide()

# =============================================
#  스플래시
# =============================================
class Splash(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint|Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        sc  = QApplication.primaryScreen().geometry()
        dpi = QApplication.primaryScreen().logicalDotsPerInch()
        sf  = min(dpi/96.0, 1.6)
        w,h = int(460*sf), int(280*sf)
        self.setFixedSize(w,h)
        self.move((sc.width()-w)//2,(sc.height()-h)//2)

        lay=QVBoxLayout(self); lay.setContentsMargins(0,0,0,0)
        frame=QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:1,stop:0 #FFFFFF,stop:1 #F0F4FF);
                border: 2px solid {C['brand']}; border-radius: 18px;
            }}
        """)
        fl=QVBoxLayout(frame); fl.setContentsMargins(36,28,36,28); fl.setSpacing(8)

        title=QLabel("5MRPA"); title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet(f"color:{C['brand']};font-size:38px;font-weight:900;letter-spacing:8px;background:transparent;")
        sub=QLabel(f"5분이면 배우는 업무 자동화 도구  {APP_VERSION}  {APP_REV}")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub.setStyleSheet(f"color:{C['t2']};font-size:13px;background:transparent;")
        ai_lbl=QLabel(f"Powered by {OLLAMA_MODEL}")
        ai_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ai_lbl.setStyleSheet(f"color:{C['ai']};font-size:11px;font-weight:bold;background:transparent;")

        # 저작권 / 만든 사람
        copy_lbl = QLabel("Made by  박찬욱     |     ALL RIGHTS RESERVED  ©  2025")
        copy_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        copy_lbl.setStyleSheet(
            f"color:{C['t3']};font-size:10px;background:transparent;letter-spacing:1px;")

        self.prog=QProgressBar(); self.prog.setFixedHeight(6); self.prog.setTextVisible(False)
        self.prog.setStyleSheet(f"""
            QProgressBar{{background:{C['border']};border:none;border-radius:3px;}}
            QProgressBar::chunk{{
                background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 {C['brand']},stop:1 {C['ai']});
                border-radius:3px;
            }}
        """)
        self.msg=QLabel("초기화 중...")
        self.msg.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.msg.setStyleSheet(f"color:{C['t2']};font-size:11px;background:transparent;")

        fl.addWidget(title); fl.addWidget(sub); fl.addWidget(ai_lbl)
        fl.addWidget(copy_lbl)
        fl.addSpacing(14); fl.addWidget(self.prog); fl.addWidget(self.msg)
        lay.addWidget(frame)

        msgs=["PyAutoGUI 초기화...","UI 빌드 중...","AI 연결 확인...","시작합니다!"]
        self._v=0
        t=QTimer(self)
        def tick():
            self._v=min(self._v+4,100)
            self.prog.setValue(self._v)
            self.msg.setText(msgs[min(self._v//28,3)])
            if self._v>=100: t.stop()
        t.timeout.connect(tick); t.start(30)


# =============================================
#  진입점
# =============================================
def main():
    QApplication.setHighDpiScaleFactorRoundingPolicy(
        Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)

    app = QApplication(sys.argv)
    app.setApplicationName("5MRPA")
    # 창을 모두 닫아도 앱 종료 안 함 (트레이 유지)
    app.setQuitOnLastWindowClosed(False)

    screen  = app.primaryScreen()
    dpi     = screen.logicalDotsPerInch()
    base_pt = max(9, min(12, int(9*96/dpi*1.5)))
    app.setFont(QFont("Malgun Gothic", base_pt))

    splash = Splash(); splash.show(); app.processEvents()

    win = MainWindow()

    # --hidden 플래그: 창 없이 트레이로만 시작
    start_hidden = "--hidden" in sys.argv

    def _after_splash():
        splash.close()
        if start_hidden:
            win._tray.show_message(
                "5MRPA 시작",
                "트레이에서 실행 중입니다.\n"
                "아이콘을 클릭하면 창을 열 수 있습니다.",
                duration_ms=4000)
        else:
            win.show()
            win.raise_()

    QTimer.singleShot(1600, _after_splash)
    sys.exit(app.exec())


if __name__ == "__main__":
    main()